from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DATASET_DIR = ROOT / "agent_knowledge" / "hkt_product_tariffs"
SOURCE_REGISTRY_JSON = ROOT / "source_registry.json"

HKT_PRODUCT_SOURCES = [
    {
        "source_id": "csl_5g_tariff_en",
        "brand": "csl",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.hkcsl-5g.com/en/5g-tariff-plan/?5g=5g",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "1010_5g_service_plan",
        "brand": "1O1O",
        "product_category": "mobile_prestige_5g",
        "url": "https://www.1010.com.hk/5g-service-plan",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "1010_infinite_entertainment_5g_prestige",
        "brand": "1O1O",
        "product_category": "mobile_prestige_5g_entertainment",
        "url": "https://1010.com.hk/en/infinite-entertainment-5g-prestige-service",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "hkt_enterprise_5g_business_mobile",
        "brand": "HKT Enterprise",
        "product_category": "business_mobile_5g",
        "url": "https://www.hkt-enterprise.com/en/products-solutions/mobility/enterprise-mobile-solutions",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "hkt_enterprise_5g_business_broadband",
        "brand": "HKT Enterprise",
        "product_category": "business_broadband_5g",
        "url": "https://www.hkt-enterprise.com/en/products-solutions/data-connectivity/business-broadband/solution/5g-broadband",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "hkt_enterprise_business_broadband_overview",
        "brand": "HKT Enterprise",
        "product_category": "business_broadband",
        "url": "https://www.hkt-enterprise.com/en/products-solutions/data-connectivity/business-broadband",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "hkt_enterprise_local_business_telephone",
        "brand": "HKT Enterprise",
        "product_category": "business_fixed_voice",
        "url": "https://www.hkt-enterprise.com/en/cases-trends/local-business-telephone-services",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "hkt_enterprise_local_business_telephone_tc",
        "brand": "HKT Enterprise",
        "product_category": "business_fixed_voice",
        "url": "https://www.hkt-enterprise.com/tc/cases-trends/local-business-telephone-services",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "hkt_local_business_telephone_hktcom_en",
        "brand": "HKT Enterprise",
        "product_category": "business_fixed_voice",
        "url": "https://www.hkt.com/at-the-office/products-and-solutions/voice-and-unified-communications/local-business-telephone-services?locale=en",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "hkt_homephone_value_added_services_en",
        "brand": "HKT",
        "product_category": "consumer_fixed_voice_vas",
        "url": "http://hkt-homephone.com/vas?lang=eng",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "hkt_sme_5g_business_mobile",
        "brand": "HKT SME",
        "product_category": "business_mobile_5g",
        "url": "https://www.hkt-sme.com/en/5g-business-mobile/",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "hkt_sme_business_broadband",
        "brand": "HKT SME",
        "product_category": "business_broadband",
        "url": "https://www.hkt-sme.com/en/business-broadband/",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "netvigator_home_broadband_index",
        "brand": "NETVIGATOR",
        "product_category": "home_fibre_broadband",
        "url": "https://www.netvigator.com/eng/index.html",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "netvigator_5g_home_broadband_plus",
        "brand": "NETVIGATOR",
        "product_category": "home_5g_broadband",
        "url": "https://www.netvigator.com/eng/High-Speed-Home-Broadband-Plus.html",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "netvigator_csl_5g_home_internet_offer",
        "brand": "NETVIGATOR",
        "product_category": "home_5g_broadband",
        "url": "https://www.netvigator.com/chi/promotion/5G-Home-Internet.html",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "csl_netvigator_5g_home_internet_offer",
        "brand": "NETVIGATOR",
        "product_category": "home_5g_broadband",
        "url": "https://www.hkcsl.com/tc/cslxnetvigator-5g-home-internet/",
        "official_source_type": "official_public_product_page",
    },
    {
        "source_id": "netvigator_broadband_list_price",
        "brand": "NETVIGATOR",
        "product_category": "home_fibre_broadband",
        "url": "https://www.netvigator.com/eng/info/list-price.html",
        "official_source_type": "official_public_product_page",
    },
]

HISTORICAL_SNAPSHOT_SOURCES = [
    {
        "snapshot_id": "csl_data_voice_2016",
        "timestamp": "20160220192657",
        "original_url": "http://www.hkcsl.com:80/en/New-Data-and-Voice-Service-Plan/",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "csl_old_data_voice",
    },
    {
        "snapshot_id": "pccw_mobile_3g_tariff_20100107",
        "timestamp": "20100107113043",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/3G-tariff-plan.jsp",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_3g_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_3g_tariff_20110612",
        "timestamp": "20110612090602",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/3G-tariff-plan.jsp",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_3g_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_other_3g_tariff_20100107",
        "timestamp": "20100107225908",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Other-3G-Tariff-Plan.jsp",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_other_3g_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_other_3g_tariff_20110612",
        "timestamp": "20110612090633",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Other-3G-Tariff-Plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_other_3g_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_other_3g_tariff_20120117",
        "timestamp": "20120117221635",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Other-3G-Tariff-Plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_other_3g_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_web_talk_tariff_20100107",
        "timestamp": "20100107015253",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Web-Talk-tariff-plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_web_talk_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_web_talk_tariff_20110623",
        "timestamp": "20110623120610",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Web-Talk-tariff-plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_web_talk_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_web_talk_tariff_20120117",
        "timestamp": "20120117052939",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Web-Talk-tariff-plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_web_talk_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_tablet_data_tariff_20100712",
        "timestamp": "20100712111708",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Tablet-Mobile-Data-Tariff-Plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g_tablet",
        "parser": "pccw_mobile_tablet_data_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_free_to_go_sim_only_20111007",
        "timestamp": "20111007190459",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/NEW_Free-to-go_SIM_only_Plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_free_to_go_sim_only",
    },
    {
        "snapshot_id": "pccw_mobile_new_monthly_plan_20120119",
        "timestamp": "20120119212851",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/New_Monthly_Plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_new_monthly_plan",
    },
    {
        "snapshot_id": "pccw_mobile_ultimate_4g_smartphone_20120507",
        "timestamp": "20120507023051",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Ultimate_Mobility_monthly_plan_for_4G_smartphones.jsp",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "pccw_mobile_ultimate_4g_smartphone",
    },
    {
        "snapshot_id": "pccw_mobile_multi_smart_sims_20120514",
        "timestamp": "20120514024717",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Multi_Smart_SIMs_monthly_plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "pccw_mobile_multi_smart_sims",
    },
    {
        "snapshot_id": "pccw_mobile_2g_tariff_20091017",
        "timestamp": "20091017090534",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/2G_tariff_plan.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_2g",
        "parser": "pccw_mobile_2g_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_cdma_service_20100228",
        "timestamp": "20100228044935",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/2G_tariff_plan/CDMA_mobile_service.jsp",
        "brand": "csl",
        "product_category": "mobile_consumer_cdma",
        "parser": "pccw_mobile_cdma_service",
    },
    {
        "snapshot_id": "pccw_mobile_cdma_service_20100827",
        "timestamp": "20100827001430",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/2G_tariff_plan/CDMA_mobile_service.jsp?lang=en",
        "brand": "csl",
        "product_category": "mobile_consumer_cdma",
        "parser": "pccw_mobile_cdma_service",
    },
    {
        "snapshot_id": "pccw_mobile_cdma_service_20101221",
        "timestamp": "20101221064032",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/2G_tariff_plan/CDMA_mobile_service.jsp?lang=ch",
        "brand": "csl",
        "product_category": "mobile_consumer_cdma",
        "parser": "pccw_mobile_cdma_service",
    },
    {
        "snapshot_id": "pccw_mobile_concierge_20101017",
        "timestamp": "20101017051803",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/PCCW_Concierge_service.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_concierge",
    },
    {
        "snapshot_id": "pccw_mobile_netvigator_customer_offer_20110413",
        "timestamp": "20110413071106",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/Special_offer_for_Netvigator_customer.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_3g",
        "parser": "pccw_mobile_netvigator_customer_offer",
    },
    {
        "snapshot_id": "pccw_mobile_new_ultimate_smartphones_20120915",
        "timestamp": "20120915183828",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/New_Ultimate_Mobility_Monthly_Plan_for_Smartphones.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "pccw_mobile_new_ultimate_smartphones",
    },
    {
        "snapshot_id": "pccw_mobile_new_ultimate_tablets_20120914",
        "timestamp": "20120914184631",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/New_Ultimate_Mobility_Monthly_Plan_for_Tablets.jsp?",
        "brand": "csl",
        "product_category": "mobile_consumer_4g_tablet",
        "parser": "pccw_mobile_new_ultimate_tablets",
    },
    {
        "snapshot_id": "pccw_mobile_2g_99_tariff_20100107",
        "timestamp": "20100107204901",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/2G_99_Integrated_Minutes_Monthly_Tariff_Plan.jsp",
        "brand": "csl",
        "product_category": "mobile_consumer_2g",
        "parser": "pccw_mobile_2g_99_tariff",
    },
    {
        "snapshot_id": "pccw_mobile_2g_99_tariff_20110110",
        "timestamp": "20110110063736",
        "original_url": "http://www2.pccwmobile.com:80/portal/gen/WEB/home/Services_And_Pricing/tariff/2G_99_Integrated_Minutes_Monthly_Tariff_Plan.jsp?lang=ch&treeMenu=treeMenu_mainMenuID0&subMenu=subMenu_level_1ID0_1&parent=parent_level_1ID0_1&cid=level_2ID0_1_6",
        "brand": "csl",
        "product_category": "mobile_consumer_2g",
        "parser": "pccw_mobile_2g_99_tariff",
    },
    {
        "snapshot_id": "1010_kingking_voice_roaming_20151005",
        "timestamp": "20151005055017",
        "original_url": "http://1010.hkcsl.com/jsp/roaming_and_idd/kingking/kingking.jsp",
        "brand": "1O1O",
        "product_category": "mobile_voip_daily_pass",
        "parser": "1010_kingking_voice_roaming",
    },
    {
        "snapshot_id": "1010_kingking_voice_roaming_20160316",
        "timestamp": "20160316165945",
        "original_url": "http://1010.hkcsl.com/jsp/roaming_and_idd/kingking/kingking.jsp",
        "brand": "1O1O",
        "product_category": "mobile_voip_daily_pass",
        "parser": "1010_kingking_voice_roaming",
    },
    {
        "snapshot_id": "1010_3g_mobile_tv_20071230",
        "timestamp": "20071230054428",
        "original_url": "http://1010.hkcsl.com:80/jsp/3g_service_and_infotainment/3g_mobile_tv/charges_and_subscription/charges_and_subscription.jsp?language=eng",
        "brand": "1O1O",
        "product_category": "mobile_3g_tv_monthly_service",
        "parser": "1010_3g_mobile_tv",
    },
    {
        "snapshot_id": "1010_3g_mobile_tv_20090211",
        "timestamp": "20090211055534",
        "original_url": "http://1010.hkcsl.com:80/jsp/3g_service_and_infotainment/3g_mobile_tv/charges_and_subscription/charges_and_subscription.jsp?language=eng",
        "brand": "1O1O",
        "product_category": "mobile_3g_tv_monthly_service",
        "parser": "1010_3g_mobile_tv",
    },
    {
        "snapshot_id": "1010_football_service_20070223",
        "timestamp": "20070223072430",
        "original_url": "http://1010.hkcsl.com:80/jsp/3g_service_and_infotainment/football/charges_and_subscription/charges_and_subscription.jsp",
        "brand": "1O1O",
        "product_category": "mobile_football_content_service",
        "parser": "1010_football_service",
    },
    {
        "snapshot_id": "1010_football_service_20080212",
        "timestamp": "20080212083206",
        "original_url": "http://1010.hkcsl.com:80/jsp/3g_service_and_infotainment/football/charges_and_subscription/charges_and_subscription.jsp",
        "brand": "1O1O",
        "product_category": "mobile_football_content_service",
        "parser": "1010_football_service",
    },
    {
        "snapshot_id": "1010_music_service_20080430",
        "timestamp": "20080430190308",
        "original_url": "http://1010.hkcsl.com:80/jsp/3g_service_and_infotainment/music/charges_and_subscription/charges_and_subscription.jsp",
        "brand": "1O1O",
        "product_category": "mobile_music_content_service",
        "parser": "1010_music_service",
    },
    {
        "snapshot_id": "1010_music_service_20090122",
        "timestamp": "20090122112813",
        "original_url": "http://1010.hkcsl.com:80/jsp/3g_service_and_infotainment/music/charges_and_subscription/charges_and_subscription.jsp",
        "brand": "1O1O",
        "product_category": "mobile_music_content_service",
        "parser": "1010_music_service",
    },
    {
        "snapshot_id": "1010_anyplex_20130807_tc",
        "timestamp": "20130807114804",
        "original_url": "http://1010.hkcsl.com/jsp/3g_service_and_infotainment/anyplex/charges_and_subscription/charges_and_subscription.jsp?language=tch",
        "brand": "1O1O",
        "product_category": "mobile_movie_voucher_service",
        "parser": "1010_anyplex",
    },
    {
        "snapshot_id": "1010_anyplex_20140815_en",
        "timestamp": "20140815235737",
        "original_url": "http://1010.hkcsl.com/jsp/3g_service_and_infotainment/anyplex/charges_and_subscription/charges_and_subscription.jsp",
        "brand": "1O1O",
        "product_category": "mobile_movie_voucher_service",
        "parser": "1010_anyplex",
    },
    {
        "snapshot_id": "csl_data_voice_2017",
        "timestamp": "20170110013626",
        "original_url": "https://www.hkcsl.com/en/New-Data-and-Voice-Service-Plan/",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "csl_old_data_voice",
    },
    {
        "snapshot_id": "csl_data_voice_2018",
        "timestamp": "20181029200935",
        "original_url": "https://www.hkcsl.com/en/new-data-and-voice-service-plan/",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "csl_old_data_voice",
    },
    {
        "snapshot_id": "csl_data_voice_2019",
        "timestamp": "20190716104151",
        "original_url": "https://www.hkcsl.com/en/new-data-and-voice-service-plan/",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "csl_old_data_voice",
    },
    {
        "snapshot_id": "csl_data_voice_2020",
        "timestamp": "20200808094232",
        "original_url": "https://www.hkcsl.com/en/new-data-and-voice-service-plan/",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "csl_old_data_voice",
    },
    {
        "snapshot_id": "csl_data_voice_2022",
        "timestamp": "20220610173714",
        "original_url": "https://www.hkcsl.com/en/new-data-and-voice-service-plan/",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "csl_old_data_voice",
    },
    {
        "snapshot_id": "csl_data_voice_2023",
        "timestamp": "20230407214425",
        "original_url": "https://www.hkcsl.com/en/new-data-and-voice-service-plan/",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "csl_old_data_voice",
    },
    {
        "snapshot_id": "csl_data_voice_2024",
        "timestamp": "20240118074912",
        "original_url": "https://www.hkcsl.com/en/new-data-and-voice-service-plan/",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "csl_old_data_voice",
    },
    {
        "snapshot_id": "csl_data_voice_2025",
        "timestamp": "20250704204530",
        "original_url": "https://www.hkcsl.com/en/new-data-and-voice-service-plan/",
        "brand": "csl",
        "product_category": "mobile_consumer_4g",
        "parser": "csl_old_data_voice",
    },
    *[
        {
            "snapshot_id": f"csl_5g_{timestamp[:8]}",
            "timestamp": timestamp,
            "original_url": "https://www.hkcsl-5g.com/en/5g-tariff-plan/",
            "brand": "csl",
            "product_category": "mobile_consumer_5g",
            "parser": "csl_5g",
        }
        for timestamp in [
            "20200804084709",
            "20201020232041",
            "20210124093538",
            "20210417114529",
            "20210728142105",
            "20211016212632",
            "20220129015111",
            "20220331013930",
            "20220625142535",
            "20220802183825",
            "20220929020211",
            "20221230002628",
            "20230128083120",
            "20230408102543",
            "20240520063011",
            "20250213063913",
            "20250704204542",
            "20250815155607",
            "20251230090356",
            "20260217094350",
        ]
    ],
    *[
        {
            "snapshot_id": f"1010_5g_{timestamp[:8]}",
            "timestamp": timestamp,
            "original_url": "https://www.1010.com.hk/5g-service-plan",
            "brand": "1O1O",
            "product_category": "mobile_prestige_5g",
            "parser": "1010_5g",
        }
        for timestamp in [
            "20241113150756",
            "20250117113527",
            "20250514172906",
            "20250804154753",
            "20251017104945",
            "20251119124113",
            "20251229153110",
            "20260301002747",
            "20260412003735",
        ]
    ],
    *[
        {
            "snapshot_id": f"netvigator_list_price_{timestamp[:8]}",
            "timestamp": timestamp,
            "original_url": "https://www.netvigator.com/eng/info/list-price.html",
            "brand": "NETVIGATOR",
            "product_category": "home_fibre_broadband",
            "parser": "netvigator_list_price",
        }
        for timestamp in [
            "20180114101545",
            "20190825191114",
            "20200814104529",
            "20210303081333",
            "20220521034537",
            "20230204000227",
            "20240530135129",
            "20250124142216",
            "20260206092017",
        ]
    ],
]

OFFICIAL_TARIFF_DOCUMENT_SOURCES = [
    {
        "snapshot_id": "csl_mobile_2g_3g_tariff_20100603",
        "timestamp": "20100603",
        "original_url": "https://www.ofca.gov.hk/filemanager/ofca/en/share/tariff/mob_operators/U003-004.pdf",
        "brand": "csl",
        "product_category": "official_tariff_mobile_2g_3g",
        "parser": "official_tariff_linkout_source_gap",
        "published_on": "2010-06-03",
        "source_kind": "official_public_tariff_pdf",
    },
    {
        "snapshot_id": "hkt_business_broadband_tariff_20100401",
        "timestamp": "20100401",
        "original_url": "https://www.hkt.com/api-service/assets/F050_0049%20Business_Broadband_Services.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_business_broadband",
        "parser": "hkt_business_broadband_tariff_pdf",
        "published_on": "2010-04-01",
    },
    {
        "snapshot_id": "hkt_international_toll_free_tariff_20100331",
        "timestamp": "20100331",
        "original_url": "https://www.hkt.com/api-service/assets/F050_0051_Int_l_Toll_Free_Service.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_international_toll_free",
        "parser": "hkt_international_toll_free_tariff_pdf",
        "published_on": "2010-03-31",
    },
    {
        "snapshot_id": "hkt_homefax_1_tariff_20101216",
        "timestamp": "20101216",
        "original_url": "https://www.hkt.com/api-service/assets/U025-014%20_revised_Homefax_1_servic.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_home_fax_service",
        "parser": "hkt_homefax_1_tariff_pdf",
        "published_on": "2010-12-16",
    },
    {
        "snapshot_id": "hkt_super_hotline_tariff_20101029",
        "timestamp": "20101029",
        "original_url": "https://www.hkt.com/api-service/assets/U025-007_super_hotline.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_super_hotline",
        "parser": "hkt_super_hotline_tariff_pdf",
        "published_on": "2010-10-29",
    },
    {
        "snapshot_id": "hkt_faxline_100_tariff_20101216",
        "timestamp": "20101216",
        "original_url": "https://www.hkt.com/api-service/assets/U025-017_Faxline_100_service.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_business_faxline",
        "parser": "hkt_faxline_tariff_pdf",
        "published_on": "2010-12-16",
    },
    {
        "snapshot_id": "hkt_faxline_3_tariff_20101216",
        "timestamp": "20101216",
        "original_url": "https://www.hkt.com/api-service/assets/U025-016_Faxline_3_service.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_business_faxline",
        "parser": "hkt_faxline_tariff_pdf",
        "published_on": "2010-12-16",
    },
    {
        "snapshot_id": "hkt_faxline_2_tariff_20101216",
        "timestamp": "20101216",
        "original_url": "https://www.hkt.com/api-service/assets/U025-013_Faxline_2_service.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_business_faxline",
        "parser": "hkt_faxline_tariff_pdf",
        "published_on": "2010-12-16",
    },
    {
        "snapshot_id": "hkt_faxline_200_tariff_20101216",
        "timestamp": "20101216",
        "original_url": "https://www.hkt.com/api-service/assets/U025-015_Faxline_200_service.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_business_faxline",
        "parser": "hkt_faxline_tariff_pdf",
        "published_on": "2010-12-16",
    },
    {
        "snapshot_id": "csl_voip_monthly_pass_tariff_20150720",
        "timestamp": "20150720",
        "original_url": "https://www.hkt.com/api-service/assets/U0008-015-Jul2015-R%20Internet%20Proto.pdf",
        "brand": "csl",
        "product_category": "mobile_voip_service",
        "parser": "csl_voip_monthly_pass_tariff_pdf",
        "published_on": "2015-07-20",
        "source_kind": "official_public_tariff_pdf",
    },
    {
        "snapshot_id": "1010_ipad_pro_2020_en_official_product_page",
        "timestamp": "20200401",
        "original_url": "https://www.1010.com.hk/en/ipad_pro_2020",
        "brand": "1O1O",
        "product_category": "mobile_tablet_data_plan",
        "parser": "1010_ipad_pro_2020_product_page",
        "published_on": "2020-04-01",
        "source_kind": "official_public_product_page",
    },
    {
        "snapshot_id": "1010_ipad_pro_2020_tc_official_product_page",
        "timestamp": "20200401",
        "original_url": "https://www.1010.com.hk/tc/ipad_pro_2020",
        "brand": "1O1O",
        "product_category": "mobile_tablet_data_plan",
        "parser": "1010_ipad_pro_2020_product_page",
        "published_on": "2020-04-01",
        "source_kind": "official_public_product_page",
    },
    {
        "snapshot_id": "hkt_one_communications_tariff_20110708",
        "timestamp": "20110708",
        "original_url": "https://www.hkt.com/api-service/assets/U025-031_one_communications.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_one_communications",
        "parser": "hkt_one_communications_tariff_pdf",
        "published_on": "2011-07-08",
    },
    {
        "snapshot_id": "hkt_home_easywatch_tariff_20130401",
        "timestamp": "20130401",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-005-Apr2013-R%20Home%20Easywatch.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_home_monitoring",
        "parser": "hkt_home_easywatch_tariff_pdf",
        "published_on": "2013-04-01",
    },
    {
        "snapshot_id": "hkt_eye_communication_package_tariff_20130530",
        "timestamp": "20130530",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-013-May2013-R%20eye%20Communicat.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_eye_service",
        "parser": "hkt_eye_service_tariff_pdf",
        "published_on": "2013-05-30",
    },
    {
        "snapshot_id": "hkt_eye_home_smartphone_tariff_20120910",
        "timestamp": "20120910",
        "original_url": "https://www.hkt.com/api-service/assets/U025-013-Sep2012-R%20eye%20Home%20Smartp.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_eye_home_smartphone",
        "parser": "hkt_eye_home_smartphone_tariff_pdf",
        "published_on": "2012-09-10",
    },
    {
        "snapshot_id": "hkt_eye_home_smartphone_tariff_20111101",
        "timestamp": "20111101",
        "original_url": "https://www.hkt.com/api-service/assets/U025_032_eye_Home_Smartphone_Packa.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_eye_home_smartphone",
        "parser": "hkt_eye_home_smartphone_tariff_pdf",
        "published_on": "2011-11-01",
    },
    {
        "snapshot_id": "hkt_eye_multimedia_service_tariff_20120910",
        "timestamp": "20120910",
        "original_url": "https://www.hkt.com/api-service/assets/U025-011-Sep2012-R%20eye%20Multimedia%20.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_eye_multimedia_service",
        "parser": "hkt_eye_multimedia_service_tariff_pdf",
        "published_on": "2012-09-10",
    },
    {
        "snapshot_id": "hkt_easywatch_commercial_tariff_20140620",
        "timestamp": "20140620",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-014-Jun2014-R%20HKT%20EasyWatch%20.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_commercial_monitoring",
        "parser": "hkt_easywatch_commercial_tariff_pdf",
        "published_on": "2014-06-20",
    },
    {
        "snapshot_id": "hkt_ip_voice_tariff_20130531",
        "timestamp": "20130531",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-015-May2013-R%20Internet%20Proto.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_ip_voice",
        "parser": "hkt_ip_voice_tariff_pdf",
        "published_on": "2013-05-31",
    },
    {
        "snapshot_id": "hkt_customer_voice_hotline_tariff_20130601",
        "timestamp": "20130601",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-017-Jun2013-R%20Customer%20voice.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_customer_voice_hotline",
        "parser": "hkt_customer_voice_hotline_tariff_pdf",
        "published_on": "2013-06-01",
    },
    {
        "snapshot_id": "hkt_residential_cell_relay_tariff_20130601",
        "timestamp": "20130601",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-022-Jun2013-R%20CRS%20Services.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_residential_cell_relay",
        "parser": "hkt_residential_cell_relay_tariff_pdf",
        "published_on": "2013-06-01",
    },
    {
        "snapshot_id": "hkt_ip_net_tariff_20130320",
        "timestamp": "20130320",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-004-Mar2013-R%20IP-Net.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_ip_net",
        "parser": "hkt_ip_net_tariff_pdf",
        "published_on": "2013-03-20",
    },
    {
        "snapshot_id": "hkt_megalink_plus_tariff_20081128",
        "timestamp": "20081128",
        "original_url": "https://www.hkt.com/api-service/assets/F050-0007%20Megalink%20Plus.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_megalink_plus",
        "parser": "hkt_megalink_plus_tariff_pdf",
        "published_on": "2008-11-28",
    },
    {
        "snapshot_id": "hkt_megalink_plus_tariff_20130514",
        "timestamp": "20130514",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-011-May2013-R%20Megalink%20Plus.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_megalink_plus",
        "parser": "hkt_megalink_plus_tariff_pdf",
        "published_on": "2013-05-14",
    },
    {
        "snapshot_id": "hkt_freedome_network_safety_tariff_20161201",
        "timestamp": "20161201",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-019-Dec2016-N%20Freedome%20Netwo.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_security_software",
        "parser": "hkt_freedome_network_safety_tariff_pdf",
        "published_on": "2016-12-01",
    },
    {
        "snapshot_id": "hkt_severe_weather_warning_tariff_20110124",
        "timestamp": "20110124",
        "original_url": "https://www.hkt.com/api-service/assets/U025_020_Severe_Weather_Warning_se.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_severe_weather_warning",
        "parser": "hkt_severe_weather_warning_tariff_pdf",
        "published_on": "2011-01-24",
    },
    {
        "snapshot_id": "hkt_eye2_communication_package_tariff_20100927",
        "timestamp": "20100927",
        "original_url": "https://www.hkt.com/api-service/assets/U025-003%20eye2%20Communication%20Package.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_eye2_service",
        "parser": "hkt_eye2_communication_package_tariff_pdf",
        "published_on": "2010-09-27",
    },
    {
        "snapshot_id": "hkt_eye2_communication_package_tariff_20120910",
        "timestamp": "20120910",
        "original_url": "https://www.hkt.com/api-service/assets/U025-015-Sep2012-R%20eye2%20communicat.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_eye2_service",
        "parser": "hkt_eye2_communication_package_tariff_pdf",
        "published_on": "2012-09-10",
    },
    {
        "snapshot_id": "hkt_norton_secure_vpn_tariff_20200601",
        "timestamp": "20200601",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-004-Jun2020-R%20Norton%20Secure%20.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_security_software",
        "parser": "hkt_norton_secure_vpn_tariff_pdf",
        "published_on": "2020-06-01",
    },
    {
        "snapshot_id": "hkt_integrated_digital_access_tariff_20250721",
        "timestamp": "20250721",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-001-Jul2025-R_IDA_Service.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_integrated_digital_access",
        "parser": "hkt_integrated_digital_access_tariff_pdf",
        "published_on": "2025-07-18",
        "effective_from": "2025-07-21",
    },
    {
        "snapshot_id": "hkt_local_business_telephone_tariff_20130601",
        "timestamp": "20130601",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-020-Jun2013-R%20Local%20Business.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_local_business_telephone",
        "parser": "hkt_local_business_telephone_tariff_pdf",
        "published_on": "2013-06-01",
    },
    {
        "snapshot_id": "hkt_local_business_telephone_tariff_20140109",
        "timestamp": "20140109",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-001-Jan2014-R%20Local%20Business.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_local_business_telephone",
        "parser": "hkt_local_business_telephone_tariff_pdf",
        "published_on": "2014-01-09",
    },
    {
        "snapshot_id": "hkt_eye_service_tariff_20140516",
        "timestamp": "20140516",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-012-May2014-R%20eye%20Service.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_eye_service",
        "parser": "hkt_eye_service_tariff_pdf",
        "published_on": "2014-05-16",
    },
    {
        "snapshot_id": "hkt_local_telephone_tariff_20130509",
        "timestamp": "20130509",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-010-May2013-R%20Local%20Telephon.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_consumer_fixed_line",
        "parser": "hkt_consumer_fixed_line_tariff_pdf",
        "published_on": "2013-05-09",
    },
    {
        "snapshot_id": "hkt_local_telephone_tariff_20130601",
        "timestamp": "20130601",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-018-Jun2013-R%20Local%20Telephon.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_consumer_fixed_line",
        "parser": "hkt_consumer_fixed_line_tariff_pdf",
        "published_on": "2013-06-01",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20130201",
        "timestamp": "20130201",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-001-Feb2013-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2013-02-01",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20150210",
        "timestamp": "20150210",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-003-Feb2015-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2015-02-10",
    },
    {
        "snapshot_id": "csl_mobile_2g_3g_4g_tariff_20120503",
        "timestamp": "20120503",
        "original_url": "https://www.hkt.com/api-service/assets/U003-001-May2012-R.pdf",
        "brand": "csl",
        "product_category": "official_tariff_mobile_2g_3g_4g",
        "parser": "csl_2g_3g_4g_mobile_tariff_pdf",
        "published_on": "2012-05-03",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20120515",
        "timestamp": "20120515",
        "original_url": "https://www.hkt.com/api-service/assets/U025-003-May2012-R%20Internet%20Access.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2012-05-15",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20121001",
        "timestamp": "20121001",
        "original_url": "https://www.hkt.com/api-service/assets/U025-017-Oct2012-R%20Internet%20Access.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2012-10-01",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20130502",
        "timestamp": "20130502",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-009-May2013-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2013-05-02",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20130823",
        "timestamp": "20130823",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-029-Aug2013-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2013-08-23",
    },
    {
        "snapshot_id": "hkt_business_broadband_tariff_20131122",
        "timestamp": "20131122",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-032-Nov2013-R%20Business%20Broad.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_business_broadband",
        "parser": "hkt_business_broadband_tariff_pdf",
        "published_on": "2013-11-22",
    },
    {
        "snapshot_id": "hkt_premium_broadband_tariff_20131209",
        "timestamp": "20131209",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-035-Dec2013-R%20Premium%20Broadb.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_premium_broadband",
        "parser": "hkt_premium_broadband_tariff_pdf",
        "published_on": "2013-12-09",
        "effective_from": "2014-01-10",
    },
    {
        "snapshot_id": "hkt_megalink_service_tariff_20120703",
        "timestamp": "20120703",
        "original_url": "https://www.hkt.com/api-service/assets/U025-006-Jul2012-R%20Megalink%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_megalink",
        "parser": "hkt_megalink_service_tariff_pdf",
        "published_on": "2012-07-03",
        "effective_from": "2012-07-03",
    },
    {
        "snapshot_id": "hkt_megalink_service_tariff_20131209",
        "timestamp": "20131209",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-034-Dec2013-R%20Megalink%20Servi.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_megalink",
        "parser": "hkt_megalink_service_tariff_pdf",
        "published_on": "2013-12-09",
        "effective_from": "2014-01-10",
    },
    {
        "snapshot_id": "hkt_megalink_service_tariff_20161130",
        "timestamp": "20161130",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-016-Nov2016-R%20Megalink%20Servi.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_megalink",
        "parser": "hkt_megalink_service_tariff_pdf",
        "published_on": "2016-11-30",
        "effective_from": "2017-01-10",
    },
    {
        "snapshot_id": "hkt_megalink_service_tariff_20230919",
        "timestamp": "20230919",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-005-Sep2023-R%20Megalink%20Servi.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_megalink",
        "parser": "hkt_megalink_service_tariff_pdf",
        "published_on": "2023-09-19",
        "effective_from": "2023-09-19",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20120903",
        "timestamp": "20120903",
        "original_url": "https://www.hkt.com/api-service/assets/U025-008-Aug2012-R%20Datapak%20Service.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2012-09-03",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20121012",
        "timestamp": "20121012",
        "original_url": "https://www.hkt.com/api-service/assets/U025-018-Oct2012-R%20Datapak%20Service.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2012-10-12",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20131105",
        "timestamp": "20131105",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-031-Nov2013-R%20Datapak%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2013-11-05",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20140626",
        "timestamp": "20140626",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-019-Jun2014-R%20Datapak%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2014-06-26",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20150926",
        "timestamp": "20150926",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-013-Sep2015-R%20Datapak%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2015-09-26",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20151027",
        "timestamp": "20151027",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-014-Oct2015-R%20Datapak%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2015-10-27",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20161130",
        "timestamp": "20161130",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-018-Nov2016-R%20Datapak%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2016-11-30",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20180401",
        "timestamp": "20180401",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-003-Apr2018-R%20Datapak%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2018-04-01",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20221116",
        "timestamp": "20221116",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-002-Nov2022-R%20Datapak%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2022-11-16",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20230401",
        "timestamp": "20230401",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-002-Apr2023-R%20Datapak%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2023-04-01",
    },
    {
        "snapshot_id": "hkt_datapak_private_circuit_tariff_20230915",
        "timestamp": "20230915",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-004-Sep2023-R%20Datapak%20Servic.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_datapak_private_circuit",
        "parser": "hkt_datapak_private_circuit_tariff_pdf",
        "published_on": "2023-09-15",
    },
    {
        "snapshot_id": "hkt_local_telephone_tariff_20131231",
        "timestamp": "20131231",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-038-Dec2013-R%20Local%20Telephon.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_consumer_fixed_line",
        "parser": "hkt_consumer_fixed_line_tariff_pdf",
        "published_on": "2013-12-31",
    },
    {
        "snapshot_id": "csl_u_plan_tariff_20170301",
        "timestamp": "20170301",
        "original_url": "https://www.hkt.com/api-service/assets/U0008-001-Mar2017-N%20Postpaid%20Servi.pdf",
        "brand": "csl",
        "product_category": "official_tariff_student_mobile_plan",
        "parser": "csl_postpaid_service_plan_tariff_pdf",
        "published_on": "2017-03-01",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20141010",
        "timestamp": "20141010",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-022-Oct2014-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2014-10-10",
    },
    {
        "snapshot_id": "csl_mobile_2g_3g_4g_tariff_20140919",
        "timestamp": "20140919",
        "original_url": "https://www.hkt.com/api-service/assets/U0003-001-Sep2014-R%202G%203G%204G%20Mobil.pdf",
        "brand": "csl",
        "product_category": "official_tariff_mobile_2g_3g_4g",
        "parser": "csl_2g_3g_4g_mobile_tariff_pdf",
        "published_on": "2014-09-19",
    },
    {
        "snapshot_id": "csl_mobile_2g_3g_4g_tariff_20131021",
        "timestamp": "20131021",
        "original_url": "https://www.hkt.com/api-service/assets/U0003-007-Oct2013-R%202G%203G%204G%20Mobil.pdf",
        "brand": "csl",
        "product_category": "official_tariff_mobile_2g_3g_4g",
        "parser": "csl_2g_3g_4g_mobile_tariff_pdf",
        "published_on": "2013-10-21",
    },
    {
        "snapshot_id": "hkt_local_telephone_tariff_20140924",
        "timestamp": "20140924",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-021-Sep2014-R%20Local%20Telephon.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_consumer_fixed_line",
        "parser": "hkt_consumer_fixed_line_tariff_pdf",
        "published_on": "2014-09-24",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20150119",
        "timestamp": "20150119",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-001-Jan2015-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2015-01-19",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20150619",
        "timestamp": "20150619",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-008-Jun2015-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2015-06-19",
    },
    {
        "snapshot_id": "csl_mobile_2g_3g_4g_tariff_20150720",
        "timestamp": "20150720",
        "original_url": "https://www.hkt.com/api-service/assets/U0008-011-Jul2015-R%202G%2C%203G%20and%204G%20.pdf",
        "brand": "csl",
        "product_category": "official_tariff_mobile_2g_3g_4g",
        "parser": "csl_2g_3g_4g_mobile_tariff_pdf",
        "published_on": "2015-07-20",
    },
    {
        "snapshot_id": "hkt_business_broadband_tariff_20160229",
        "timestamp": "20160229",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-006-Feb2016-R%20Business%20Broad.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_business_broadband",
        "parser": "hkt_business_broadband_tariff_pdf",
        "published_on": "2016-02-29",
    },
    {
        "snapshot_id": "hkt_business_broadband_tariff_20221209",
        "timestamp": "20221209",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-003-Dec2022-R%20Business%20Broad.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_business_broadband",
        "parser": "hkt_business_broadband_tariff_pdf",
        "published_on": "2022-12-09",
    },
    {
        "snapshot_id": "hkt_metro_ip_service_tariff_20191115",
        "timestamp": "20191115",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-005-Nov2019-R%20Metro%20IP%20Servi.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_metro_ip",
        "parser": "hkt_metro_ip_service_tariff_pdf",
        "published_on": "2019-11-15",
    },
    {
        "snapshot_id": "hkt_flexible_bandwidth_service_tariff_20190131",
        "timestamp": "20190131",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-001-Jan2019-R%20Flexible%20Bandw.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_flexible_bandwidth",
        "parser": "hkt_flexible_bandwidth_service_tariff_pdf",
        "published_on": "2019-01-31",
    },
    {
        "snapshot_id": "hkt_telecommunications_backup_tariff_20130601",
        "timestamp": "20130601",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-024-Jun2013-RTelecommunicati.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_telecommunications_backup",
        "parser": "hkt_telecommunications_backup_tariff_pdf",
        "published_on": "2013-06-01",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20170531",
        "timestamp": "20170531",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-004-May2017-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2017-05-31",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20170705",
        "timestamp": "20170705",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-006-Jul2017-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2017-07-05",
    },
    {
        "snapshot_id": "csl_smart_pama_tariff_20170314",
        "timestamp": "20170314",
        "original_url": "https://www.hkt.com/api-service/assets/U0008-003-Mar2017-N%20Postpaid%20Servi.pdf",
        "brand": "csl",
        "product_category": "official_tariff_senior_mobile_plan",
        "parser": "csl_smart_pama_tariff_pdf",
        "published_on": "2017-03-14",
    },
    {
        "snapshot_id": "csl_club_sim_tariff_20170802",
        "timestamp": "20170802",
        "original_url": "https://www.hkt.com/api-service/assets/U0008-004-Aug2017-N%20Postpaid%20Servi.pdf",
        "brand": "csl",
        "product_category": "official_tariff_club_sim_mobile_plan",
        "parser": "csl_postpaid_service_plan_tariff_pdf",
        "published_on": "2017-08-02",
    },
    {
        "snapshot_id": "csl_club_sim_tariff_20170930",
        "timestamp": "20170930",
        "original_url": "https://www.hkt.com/api-service/assets/U0008-005-Sep2017-R%20Postpaid%20Servi.pdf",
        "brand": "csl",
        "product_category": "official_tariff_club_sim_mobile_plan",
        "parser": "csl_postpaid_service_plan_tariff_pdf",
        "published_on": "2017-09-30",
    },
    {
        "snapshot_id": "csl_club_sim_tariff_20171019",
        "timestamp": "20171019",
        "original_url": "https://www.hkt.com/api-service/assets/U0008-006-Oct2017-R%20Postpaid%20Servi.pdf",
        "brand": "csl",
        "product_category": "official_tariff_club_sim_mobile_plan",
        "parser": "csl_postpaid_service_plan_tariff_pdf",
        "published_on": "2017-10-19",
    },
    {
        "snapshot_id": "csl_1010_three_postpaid_tariff_20180822",
        "timestamp": "20180822",
        "original_url": "https://www.hkt.com/api-service/assets/U0008-002-Aug2018-R%20Three%20Postpaid.pdf",
        "brand": "csl / 1O1O",
        "product_category": "official_tariff_postpaid_mobile_plan",
        "parser": "csl_1010_postpaid_mobile_tariff_pdf",
        "published_on": "2018-08-22",
    },
    {
        "snapshot_id": "csl_1010_three_postpaid_tariff_20180731",
        "timestamp": "20180731",
        "original_url": "https://www.hkt.com/api-service/assets/U0008_001_Jul2018_N%20Three%20Postpaid.pdf",
        "brand": "csl / 1O1O",
        "product_category": "official_tariff_postpaid_mobile_plan",
        "parser": "csl_1010_postpaid_mobile_tariff_pdf",
        "published_on": "2018-07-31",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20210308",
        "timestamp": "20210308",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-003-Mar2021-R%20Internet%20Acces.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2021-03-08",
    },
    {
        "snapshot_id": "hkt_home_phone_tariff_20200210",
        "timestamp": "20200210",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-002-Feb2020-R%20Home%20Phone%20Ser.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_consumer_fixed_line",
        "parser": "hkt_consumer_fixed_line_tariff_pdf",
        "published_on": "2020-02-10",
    },
    {
        "snapshot_id": "hkt_internet_access_tariff_20260106",
        "timestamp": "20260106",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-001-Jan2026-R_Internet_Access_Services.pdf",
        "brand": "NETVIGATOR",
        "product_category": "official_tariff_internet_access",
        "parser": "hkt_internet_access_tariff_pdf",
        "published_on": "2026-01-06",
    },
    {
        "snapshot_id": "hkt_premium_broadband_tariff_20161130",
        "timestamp": "20161130",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-017-Nov2016-R%20Premium%20Broadb.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_premium_broadband",
        "parser": "hkt_premium_broadband_tariff_pdf",
        "published_on": "2016-11-30",
        "effective_from": "2017-01-10",
    },
    {
        "snapshot_id": "hkt_premium_broadband_tariff_20181105",
        "timestamp": "20181105",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-004-Nov2018-R%20Premium%20Broadb.pdf",
        "brand": "HKT Enterprise",
        "product_category": "official_tariff_premium_broadband",
        "parser": "hkt_premium_broadband_tariff_pdf",
        "published_on": "2018-11-05",
        "effective_from": "2018-11-05",
    },
    {
        "snapshot_id": "hkt_eye_service_tariff_20161223",
        "timestamp": "20161223",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-020-Dec2016-R%20eye%20Service.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_eye_service",
        "parser": "hkt_eye_service_tariff_pdf",
        "published_on": "2016-12-23",
    },
    {
        "snapshot_id": "hkt_eye_service_tariff_20190812",
        "timestamp": "20190812",
        "original_url": "https://www.hkt.com/api-service/assets/U0025-003-Aug2019-R%20eye%20Service.pdf",
        "brand": "HKT",
        "product_category": "official_tariff_eye_service",
        "parser": "hkt_eye_service_tariff_pdf",
        "published_on": "2019-08-12",
    },
]

CURRENT_SOURCES_WITH_INTENTIONAL_NO_STRUCTURED_ROWS = {
    "hkt_enterprise_business_broadband_overview",
    "netvigator_5g_home_broadband_plus",
}

PRODUCT_FIELDS = [
    "record_key",
    "captured_at_hkt",
    "first_seen_at_hkt",
    "last_seen_at_hkt",
    "brand",
    "product_category",
    "plan_name",
    "monthly_fee_hkd",
    "published_price_hkd",
    "price_billing_unit",
    "local_data_gb",
    "roaming_data_gb",
    "post_fup_speed_mbps",
    "contract_months",
    "local_voice",
    "add_on_charges_hkd",
    "source_id",
    "source_url",
    "final_url",
    "official_source_type",
    "http_status",
    "fetch_method",
    "extraction_status",
    "content_hash",
    "evidence_excerpt",
]

HISTORICAL_FIELDS = [
    *PRODUCT_FIELDS,
    "snapshot_id",
    "archive_timestamp",
    "archive_year",
    "archive_month",
    "archive_url",
    "source_kind",
]

CURRENT_STRUCTURED_FIELDS = [
    "as_of_hkt",
    "brand",
    "service_generation",
    "customer_segment",
    "plan_family",
    "plan_name",
    "monthly_fee_hkd",
    "published_price_hkd",
    "price_billing_unit",
    "local_data_gb",
    "roaming_data_gb",
    "post_fup_speed_mbps",
    "contract_months",
    "local_voice",
    "add_on_charges_hkd",
    "source_status",
    "source_id",
    "source_url",
    "evidence_excerpt",
    "record_key",
]

HISTORICAL_STRUCTURED_FIELDS = [
    "archive_year",
    "archive_month",
    "brand",
    "service_generation",
    "customer_segment",
    "plan_family",
    "plan_name",
    "monthly_fee_hkd",
    "published_price_hkd",
    "price_billing_unit",
    "local_data_gb",
    "roaming_data_gb",
    "post_fup_speed_mbps",
    "contract_months",
    "local_voice",
    "add_on_charges_hkd",
    "source_status",
    "snapshot_id",
    "source_url",
    "archive_url",
    "evidence_excerpt",
    "record_key",
]

SOURCE_GAP_FIELDS = [
    "gap_year",
    "brand",
    "product_category",
    "gap_type",
    "http_status",
    "snapshot_id",
    "source_id",
    "source_url",
    "archive_url",
    "reason",
    "evidence_excerpt",
]


def _now_hkt() -> str:
    return datetime.now(ZoneInfo("Asia/Hong_Kong")).replace(microsecond=0).isoformat()


def _clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def _content_hash(text: str) -> str:
    return hashlib.sha1(_clean_text(text).encode("utf-8")).hexdigest()


def _excerpt(text: str, start: int, end: int, radius: int = 220) -> str:
    clean = _clean_text(text)
    left = max(0, start - radius)
    right = min(len(clean), end + radius)
    return clean[left:right].strip()


def _source_allowed(url: str) -> tuple[bool, str]:
    try:
        registry = json.loads(SOURCE_REGISTRY_JSON.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return False, f"source_registry unavailable: {exc}"
    domain = urlparse(url).netloc.lower()
    entry = registry.get("domains", {}).get(domain) or {}
    if entry.get("policy") != "allow":
        return False, f"source policy for {domain} is {entry.get('policy') or registry.get('default_policy')}"
    if entry.get("tos_status") in {"not_approved", "prohibited", "unreviewed"}:
        return False, f"tos_status for {domain} is {entry.get('tos_status')}"
    return True, ""


def _html_to_text(raw: bytes, content_type: str) -> tuple[str, str]:
    if "pdf" in content_type.lower() or raw.startswith(b"%PDF"):
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(raw))
            pages = [page.extract_text() or "" for page in reader.pages[:20]]
            return "PDF", _clean_text(" ".join(pages))
        except Exception as exc:
            return "PDF", f"pdf_text_extract_failed: {exc}"
    decoded = raw.decode("utf-8", "replace")
    if "json" in content_type.lower() or decoded.lstrip().startswith(("{", "[")):
        return "", _clean_text(decoded)
    soup = BeautifulSoup(decoded, "lxml")
    title = _clean_text(soup.title.get_text(" ")) if soup.title else ""
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    return title, _clean_text(soup.get_text(" "))


def fetch_product_page(client: Any | None, url: str) -> Dict[str, Any]:
    allowed, skip_reason = _source_allowed(url)
    if not allowed:
        return {
            "url": url,
            "final_url": url,
            "status": 0,
            "content_type": "",
            "bytes": 0,
            "title": "",
            "text": "",
            "error": skip_reason,
            "method": "source_registry_skipped",
        }
    close_client = False
    if client is None:
        client = httpx.Client(
            follow_redirects=True,
            timeout=httpx.Timeout(20.0, connect=8.0),
            headers={
                "User-Agent": "CMHK-public-crawl/1.0 (+public official pages only)",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,application/pdf;q=0.8,*/*;q=0.7",
                "Accept-Language": "en,zh-CN;q=0.9,zh;q=0.8",
            },
            trust_env=True,
        )
        close_client = True
    try:
        response = client.get(url, timeout=httpx.Timeout(20.0, connect=8.0))
        raw = response.content
        content_type = response.headers.get("content-type", "")
        title, text = _html_to_text(raw, content_type)
        method = "hkt_product_httpx"
        if (
            (response.status_code in {403, 406} or len(text) < 200)
            and urlparse(url).netloc.lower() == "www.hkt.com"
            and url.lower().split("?", 1)[0].endswith(".pdf")
        ):
            browser_headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
                ),
                "Accept": "application/pdf,text/html,*/*",
                "Accept-Language": "en,zh-CN;q=0.9,zh;q=0.8",
                "Referer": "https://www.hkt.com/",
            }
            response = client.get(url, timeout=httpx.Timeout(20.0, connect=8.0), headers=browser_headers)
            raw = response.content
            content_type = response.headers.get("content-type", "")
            title, text = _html_to_text(raw, content_type)
            method = "hkt_product_httpx_browser_pdf"
        if len(text) < 1000 and urlparse(url).netloc.lower() in {"www.1010.com.hk", "www.hkcsl-5g.com"}:
            browser_headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
                ),
                "Accept-Language": "en,zh-CN;q=0.9,zh;q=0.8",
            }
            response = client.get(url, timeout=httpx.Timeout(20.0, connect=8.0), headers=browser_headers)
            raw = response.content
            content_type = response.headers.get("content-type", "")
            title, text = _html_to_text(raw, content_type)
            method = "hkt_product_httpx_browser_ua"
        return {
            "url": url,
            "final_url": str(response.url),
            "status": response.status_code,
            "content_type": content_type,
            "bytes": len(raw),
            "title": title,
            "text": text,
            "error": "",
            "method": method,
        }
    except Exception as exc:
        return {
            "url": url,
            "final_url": url,
            "status": 0,
            "content_type": "",
            "bytes": 0,
            "title": "",
            "text": "",
            "error": repr(exc),
            "method": "hkt_product_httpx_error",
        }
    finally:
        if close_client:
            client.close()


def _float_text(value: str | None) -> str:
    if value is None:
        return ""
    return value.replace(",", "").strip()


def _record_key(row: Dict[str, str]) -> str:
    key_parts = [
        row.get("brand", ""),
        row.get("product_category", ""),
        row.get("plan_name", ""),
        row.get("monthly_fee_hkd", ""),
        row.get("published_price_hkd", ""),
        row.get("price_billing_unit", ""),
        row.get("local_data_gb", ""),
        row.get("roaming_data_gb", ""),
        row.get("post_fup_speed_mbps", ""),
        row.get("contract_months", ""),
        row.get("source_id", ""),
    ]
    return hashlib.sha1("|".join(key_parts).encode("utf-8")).hexdigest()[:16]


def _base_row(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> Dict[str, str]:
    return {
        "record_key": "",
        "captured_at_hkt": captured_at,
        "first_seen_at_hkt": captured_at,
        "last_seen_at_hkt": captured_at,
        "brand": source["brand"],
        "product_category": source["product_category"],
        "plan_name": "",
        "monthly_fee_hkd": "",
        "published_price_hkd": "",
        "price_billing_unit": "",
        "local_data_gb": "",
        "roaming_data_gb": "",
        "post_fup_speed_mbps": "",
        "contract_months": "",
        "local_voice": "",
        "add_on_charges_hkd": "",
        "source_id": source["source_id"],
        "source_url": source["url"],
        "final_url": str(result.get("final_url") or result.get("url") or source["url"]),
        "official_source_type": source["official_source_type"],
        "http_status": str(result.get("status") or ""),
        "fetch_method": str(result.get("method") or ""),
        "extraction_status": "parsed",
        "content_hash": _content_hash(str(result.get("text") or "")),
        "evidence_excerpt": "",
    }


def _dedupe(rows: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    seen: set[str] = set()
    output: List[Dict[str, str]] = []
    for row in rows:
        row["record_key"] = row.get("record_key") or _record_key(row)
        if row["record_key"] in seen:
            continue
        seen.add(row["record_key"])
        output.append(row)
    return output


def _parse_csl(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    rows: List[Dict[str, str]] = []
    fee_matches = list(re.finditer(r"Monthly Plan Fee(?:\s*\(?\d*\)?)*\s*\$(?P<fee>\d{2,4})", text, re.I))
    if not fee_matches:
        fee_matches = list(re.finditer(r"\$(?P<fee>\d{2,4})\s+Local data usage", text, re.I))
    for index, match in enumerate(fee_matches):
        block_start = match.start()
        block_end = fee_matches[index + 1].start() if index + 1 < len(fee_matches) else min(len(text), match.end() + 1300)
        block = text[block_start:block_end]
        if "Local data usage" not in block:
            continue
        data_match = re.search(r"Local data usage(?:\s*\(?[\d\s()]*\)?)*\s+(?P<data>\d+(?:\.\d+)?)\s*GB", block, re.I)
        speed_match = re.search(r"up to\s*(?P<speed>\d+(?:\.\d+)?)\s*Mbps", block, re.I)
        roaming_match = re.search(
            r"(?:Mainland China\s*&\s*Macau|China).*?roaming data usage(?:\s*\(?[\d\s()]*\)?)*\s+(?P<roaming>\d+(?:\.\d+)?)\s*GB",
            block,
            re.I,
        )
        if not data_match:
            continue
        row = _base_row(source, result, captured_at)
        fee = _float_text(match.group("fee"))
        row.update(
            {
                "plan_name": f"csl 5G service plan HK${fee}",
                "monthly_fee_hkd": fee,
                "local_data_gb": _float_text(data_match.group("data")),
                "roaming_data_gb": _float_text(roaming_match.group("roaming") if roaming_match else ""),
                "post_fup_speed_mbps": _float_text(speed_match.group("speed") if speed_match else ""),
                "contract_months": "",
                "local_voice": "Unlimited local voice call mins",
                "add_on_charges_hkd": "Additional SIM monthly fee: 1st $120; 2nd $100; 3rd $60; 4th $20",
                "evidence_excerpt": block[:900],
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_1010(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    rows: List[Dict[str, str]] = []
    fee_matches = list(re.finditer(r"Monthly\s*1\s*HK\$(?P<fee>\d{2,4})", text, re.I))
    if not fee_matches:
        fee_matches = list(re.finditer(r"HK\$(?P<fee>\d{2,4})", text, re.I))
    for index, match in enumerate(fee_matches):
        block_start = match.start()
        block_end = fee_matches[index + 1].start() if index + 1 < len(fee_matches) else min(len(text), match.end() + 1400)
        block = text[block_start:block_end]
        if "Local mobile data" not in block or "roaming data usage" not in block:
            continue
        data_match = re.search(
            r"Local mobile data(?:\s+\d+(?!\s*GB))?\s+(?P<data>\d+(?:\.\d+)?)\s*GB",
            block,
            re.I,
        )
        speed_match = re.search(r"(?:Max|up to)\s*(?P<speed>\d+(?:\.\d+)?)\s*Mbps", block, re.I)
        roaming_match = re.search(
            r"Mainland China\s*&\s*Macau roaming data usage(?:\s+\d+(?!\s*GB))?\s+(?P<roaming>\d+(?:\.\d+)?)\s*GB",
            block,
            re.I,
        )
        contract_match = re.search(r"\b(?P<contract>24|30|36)\s*months\b", block, re.I)
        if not data_match or not speed_match or not roaming_match:
            continue
        row = _base_row(source, result, captured_at)
        fee = _float_text(match.group("fee"))
        row.update(
            {
                "plan_name": f"1O1O 5G service plan HK${fee}",
                "monthly_fee_hkd": fee,
                "local_data_gb": _float_text(data_match.group("data")),
                "roaming_data_gb": _float_text(roaming_match.group("roaming")),
                "post_fup_speed_mbps": _float_text(speed_match.group("speed")),
                "contract_months": _float_text(contract_match.group("contract") if contract_match else ""),
                "local_voice": "Unlimited local voice calling",
                "add_on_charges_hkd": "",
                "evidence_excerpt": block[:900],
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_1010_infinite_entertainment(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    rows: List[Dict[str, str]] = []
    pattern = re.compile(
        r"(?P<name>Infinite Entertainment 5G Prestige Service \((?:sports|entertainment) pack\))\s+"
        r"HK\$\s*(?P<fee>509|559)\s*\.00\s*/\s*month.*?"
        r"Commitment period:\s*(?P<contract>24\s*/\s*36)\s*months.*?"
        r"Monthly roaming mobile data usage.*?:\s*(?P<roaming>\d+(?:\.\d+)?)GB",
        re.I,
    )
    footnote_match = re.search(
        r"HK\$509/\$559 can enjoy (?P<first>\d+)TB of local mobile data usage per month for the first 6 months, "
        r"followed by (?P<after>\d+)GB of local mobile data usage per month thereafter",
        text,
        re.I,
    )
    local_data = footnote_match.group("after") if footnote_match else "500"
    add_on = "First 6 months 1TB local data, then 500GB/month; HK$18/month administration fee applies."
    for match in pattern.finditer(text):
        row = _base_row(source, result, captured_at)
        fee = _float_text(match.group("fee"))
        row.update(
            {
                "plan_name": f"1O1O {match.group('name')} HK${fee}",
                "monthly_fee_hkd": fee,
                "local_data_gb": local_data,
                "roaming_data_gb": _float_text(match.group("roaming")),
                "post_fup_speed_mbps": "",
                "contract_months": re.sub(r"\s+", "", match.group("contract")),
                "local_voice": "Unlimited local voice call minutes",
                "add_on_charges_hkd": add_on,
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_netvigator_current_offer(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    offer_match = re.search(r"Hot services Offers (?P<offers>.*?) What.?s more", text, re.I)
    offer_text = offer_match.group("offers") if offer_match else text
    offset = offer_match.start("offers") if offer_match else 0
    rows: List[Dict[str, str]] = []
    pattern = re.compile(
        r"(?P<name>(?:1000M Fibre-to-the-Home|2500M Super Broadband|5G Home Internet))\s+"
        r"(?P<benefits>(?:(?!1000M Fibre-to-the-Home|2500M Super Broadband|5G Home Internet).)*?)(?:From\s+)?HK\$\s*(?P<fee>\d{2,4})\s*/month\s+"
        r"(?P<contract>\d{2})\s*-?month commitment",
        re.I,
    )
    for match in pattern.finditer(offer_text):
        row = _base_row(source, result, captured_at)
        name = _clean_text(match.group("name"))
        fee = _float_text(match.group("fee"))
        speed_match = re.search(r"(\d[\d,]{2,6})M", name)
        row.update(
            {
                "plan_name": f"NETVIGATOR {name} from HK${fee}",
                "monthly_fee_hkd": fee,
                "contract_months": _float_text(match.group("contract")),
                "add_on_charges_hkd": _clean_text(match.group("benefits")),
                "evidence_excerpt": _excerpt(text, offset + match.start(), offset + match.end()),
            }
        )
        if speed_match:
            row["post_fup_speed_mbps"] = _float_text(speed_match.group(1))
        rows.append(row)
    return _dedupe(rows)


def _parse_netvigator_list_price(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    rows: List[Dict[str, str]] = []
    table_match = re.search(
        r"Broadband Service.*?Service Plan Commitment Period Monthly Fee (?P<table>.*?)Promotional Offers Notes:",
        text,
        re.I,
    )
    table = table_match.group("table") if table_match else text
    pattern = re.compile(
        r"(?P<name>"
        r"50,000M Fiber-to-the-Home Service|"
        r"10G Fiber-to-the-Home Service|"
        r"10,000M Fiber-to-the-Home Service|"
        r"4x1000M Multi-Use Broadband Service|"
        r"2500M \+ 1000M Multi-Use Broadband Service|"
        r"2500M Fiber-to-the-Home Service|"
        r"2x1000M Multi-Use Broadband Service|"
        r"1000M Fiber-to-the-Home Service \(incl\. LiKE1OO broadband service\)|"
        r"1000M Fiber-to-the-Home Service|"
        r"500M or below Fiber-to-the-Home / Basic Service \(incl\. LiKE1OO broadband service\)"
        r"|500M or below Fiber-to-the-Home / Basic Service"
        r")\s+(?P<contract>24/36|24)\s*months?\s+HK\$(?P<fee>[\d,]+)",
        re.I,
    )
    for match in pattern.finditer(table):
        row = _base_row(source, result, captured_at)
        name = _clean_text(match.group("name"))
        fee = _float_text(match.group("fee"))
        speed_match = re.search(r"(\d[\d,]{2,6})M", name)
        speed_value = speed_match.group(1).replace(",", "") if speed_match else ""
        if not speed_value:
            speed_g_match = re.search(r"(\d{1,3})G\b", name, re.I)
            if speed_g_match:
                speed_value = str(int(speed_g_match.group(1)) * 1000)
        row.update(
            {
                "plan_name": f"NETVIGATOR list price {name}",
                "monthly_fee_hkd": fee,
                "contract_months": match.group("contract"),
                "add_on_charges_hkd": "Official list price; promotional offers may differ by building or address.",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        if speed_value:
            row["post_fup_speed_mbps"] = _float_text(speed_value)
        rows.append(row)
    return _dedupe(rows)


def _parse_netvigator_csl_5g_home_internet(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if not (
        re.search(r"5G\s*(?:私.?家|Home).{0,30}(?:寬頻|Internet)|csl\s*x\s*(?:網上行|NETVIGATOR)", text, re.I)
        and re.search(r"\$\s*168\s*/\s*\$?\s*198|\$\s*168.*?\$\s*198", text, re.S)
        and re.search(r"250\s*GB", text, re.I)
    ):
        return []
    match = re.search(r"(?:月費|月费|monthly fee).{0,40}\$\s*168.*?\$\s*198.*?250\s*GB", text, re.I | re.S)
    if not match:
        match = re.search(r"250\s*GB.*?\$\s*168.*?\$\s*198", text, re.I | re.S)
    excerpt = _excerpt(text, match.start(), match.end(), radius=200) if match else text[:700]
    rows: List[Dict[str, str]] = []
    for fee in ("168", "198"):
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"csl x NETVIGATOR 5G Home Internet HK${fee} 250GB",
                "monthly_fee_hkd": fee,
                "local_data_gb": "250",
                "roaming_data_gb": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "HK$18 monthly administration fee waived during the designated commitment period.",
                "evidence_excerpt": excerpt[:850],
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_internet_access_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Internet Access Services" not in text or "Current Services" not in text:
        return []
    rows: List[Dict[str, str]] = []
    pattern = re.compile(
        r"(?P<name>"
        r"1\.5M to 100M Basic Plan|"
        r"8M to 1000M and 10G Fiber-to-the-home Plan|"
        r"2x1000M Multi-Use Broadband Service Plan|"
        r"4x1000M Multi-Use Broadband Service Plan|"
        r"2500M \+ 1000M / 2x1000M/4x1000M Multi-Use Broadband Service Plan|"
        r"8M to 50,000M Fiber-to-the-home Plan|"
        r"Additional NETVIGATOR Email Account"
        r")\s+\$(?P<fee>[\d,]+)(?:/email account)?",
        re.I,
    )
    for match in pattern.finditer(text):
        name = _clean_text(match.group("name"))
        fee = _float_text(match.group("fee"))
        row = _base_row(source, result, captured_at)
        speed_value = ""
        speed_matches = [int(value.replace(",", "")) for value in re.findall(r"(\d[\d,]*)M", name)]
        if speed_matches:
            speed_value = str(max(speed_matches))
        multi_port_matches = [int(value) * 1000 for value in re.findall(r"(\d)x1000M", name, re.I)]
        if multi_port_matches:
            speed_value = str(max([int(speed_value or 0), *multi_port_matches]))
        if "10G" in name:
            speed_value = str(max(int(speed_value or 0), 10000))
        row.update(
            {
                "plan_name": f"NETVIGATOR official Internet Access Services tariff {source.get('published_on', '')} - {name}",
                "monthly_fee_hkd": fee,
                "post_fup_speed_mbps": speed_value,
                "add_on_charges_hkd": "Official tariff/list price from HKT Internet Access Services PDF; not a promotional retail offer.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_business_broadband_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Business Broadband Services" not in text or "Monthly Rental" not in text:
        return []
    specs = [
        ("@WORK Broadband Ultra Line Single/Multi-Access", "1.5M/640K - 100M", "2500"),
        ("@WORK Broadband Premier Multi-Access", "1.5M/640K - 100M", "4000"),
        ("@WORK Broadband Business Broadband / Ultra Broadband", "1.5M - 1000M", "3000"),
        ("Always-On Broadband", "640K - 100M", "35000"),
        ("Dedicated Internet 128K-521K", "128K - 521K", "12000"),
        ("Dedicated Internet T1", "T1", "20000"),
        ("Dedicated Internet E1", "E1", "35000"),
        ("Dedicated Internet 2M-1000M", "2M - 1000M", "150000"),
        ("Metro-Internet", "2M - 1000M", "1500000"),
        ("ATM Internet", "2M - 60M", "350000"),
    ]
    table_match = re.search(r"Types of Services:(?P<table>.*?)Terms and Conditions:", text, re.I)
    table = table_match.group("table") if table_match else text
    rows: List[Dict[str, str]] = []
    for name, bandwidth, fee in specs:
        if name.split()[0].replace("@", "") not in table:
            continue
        row = _base_row(source, result, captured_at)
        speed_value = ""
        speed_matches = [int(value) for value in re.findall(r"(\d+)M", bandwidth)]
        if speed_matches:
            speed_value = str(max(speed_matches))
        row.update(
            {
                "plan_name": f"HKT Enterprise official Business Broadband tariff {source.get('published_on', '')} - {name}",
                "monthly_fee_hkd": fee,
                "post_fup_speed_mbps": speed_value,
                "add_on_charges_hkd": f"Official commercial tariff PDF; bandwidth range {bandwidth}; list/tariff monthly rental, not promotional retail offer.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _clean_text(table)[:900],
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_premium_broadband_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Premium Broadband Service" not in text or "Monthly Rental" not in text:
        return []
    specs = [
        ("Premium Commercial Broadband Access Service client end - Asymmetric 10/100Base-T", "up to 8M", "420", "Bandwidth charge HK$14 per 1M; installation and relocation charges excluded from monthly fee."),
        ("Premium Commercial Broadband Access Service client end - Asymmetric G.DMT", "up to 8M", "380", "Bandwidth charge HK$14 per 1M; installation and relocation charges excluded from monthly fee."),
        ("Premium Commercial Broadband Access Service client end - Symmetric 10/100Base-T", "up to 4M", "560", "Bandwidth charge HK$70 per 1M; installation and relocation charges excluded from monthly fee."),
        ("Premium Commercial Broadband Access Service client end - Symmetric 10/100Base-T", "higher than 4M and up to 10M", "610", "Bandwidth charge HK$65 per 1M; installation and relocation charges excluded from monthly fee."),
        ("Premium Commercial Broadband Access Service client end - Symmetric 10/100Base-T", "higher than 10M and up to 20M", "970", "Bandwidth charge HK$55 per 1M; installation and relocation charges excluded from monthly fee."),
        ("Premium Commercial Broadband Access Service client end - Symmetric 10/100Base-T", "higher than 20M and up to 100M", "1745", "Bandwidth charge HK$165 per 10M; installation and relocation charges excluded from monthly fee."),
        ("Premium Broadband Service server end - OC-3c or FE", "up to 100M", "52440", "First 15 km monthly rental; additional 20 km HK$12,420 and additional 500 sessions HK$45,540; installation and relocation charges excluded."),
        ("Premium Broadband Service server end - GE", "up to 1G", "120060", "First 15 km monthly rental; additional 20 km HK$13,800 and additional 500 sessions HK$45,540; installation and relocation charges excluded."),
        ("Multiple Domain Name", "domain name service", "2760", "Monthly rental for multiple domain name; installation/deletion/change charges excluded."),
        ("Premium Commercial Broadband VPN - Asymmetric 10/100Base-T", "up to 8M", "1390", "Bandwidth charge HK$220 per 1M; installation and relocation charges excluded from monthly fee."),
        ("Premium Commercial Broadband VPN - Asymmetric G.DMT", "up to 8M", "1250", "Bandwidth charge HK$220 per 1M; installation and relocation charges excluded from monthly fee."),
        ("Premium Commercial Broadband VPN - Symmetric 10/100Base-T", "up to 4M", "1660", "Bandwidth charge HK$550 per 1M; installation and relocation charges excluded from monthly fee."),
        ("Premium Commercial Broadband VPN - Symmetric 10/100Base-T", "higher than 4M and up to 100M", "3470", "Bandwidth charge HK$490 per 1M; installation and relocation charges excluded from monthly fee."),
        ("Premium Commercial Broadband VPN - Symmetric OC-3c", "up to 150M", "16560", "First 15 km monthly rental; additional 20 km HK$12,420 and bandwidth HK$2,760 per 10M; installation and relocation charges excluded."),
        ("Premium Commercial Broadband VPN - Symmetric GE", "up to 1G", "27600", "First 15 km monthly rental; additional 20 km HK$13,800 and bandwidth HK$6,210 per 50M; installation and relocation charges excluded."),
    ]
    rows: List[Dict[str, str]] = []
    effective_from = source.get("effective_from") or "2017-01-10"
    for name, bandwidth, fee, charge_note in specs:
        row = _base_row(source, result, captured_at)
        speed_value = ""
        speed_matches = [int(value) for value in re.findall(r"(\d+)M", bandwidth)]
        if speed_matches:
            speed_value = str(max(speed_matches))
        if "1G" in bandwidth:
            speed_value = "1000"
        row.update(
            {
                "plan_name": f"HKT Enterprise official Premium Broadband tariff effective {effective_from} - {name}",
                "monthly_fee_hkd": fee,
                "post_fup_speed_mbps": speed_value,
                "add_on_charges_hkd": f"Official Premium Broadband tariff component; {charge_note} Tariff/list price, not promotional retail offer.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, 0, min(len(text), 1200)),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_megalink_service_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "MegaLink Service" not in text or "Monthly Rental" not in text:
        return []
    sections = {
        "access_ce": re.search(r"A\.\s+MegaLink Access Client End \(CE\) Service(?P<section>.*?)B\.\s+MegaLink Access Server End", text, flags=re.I | re.S),
        "access_se": re.search(r"B\.\s+MegaLink Access Server End \(SE\) Service(?P<section>.*?)C\.\s+MegaLink VPN Client End", text, flags=re.I | re.S),
        "vpn_ce": re.search(r"C\.\s+MegaLink VPN Client End \(CE\) Service(?P<section>.*?)D\.\s+MegaLink VPN Server End", text, flags=re.I | re.S),
        "vpn_se": re.search(r"D\.\s+MegaLink VPN Server End \(SE\) Service(?P<section>.*?)The Services are subject", text, flags=re.I | re.S),
    }
    section_text = {
        key: match.group("section") if match else ""
        for key, match in sections.items()
    }
    specs = [
        ("access_ce", "MegaLink Access Client End - Asymmetric 10Base-T bridging or ATM-25", "1.5M", "1.5", "380", "Access CE monthly rental for 1.5M service."),
        ("access_ce", "MegaLink Access Client End - Asymmetric 10Base-T bridging or ATM-25", "6M", "6", "435", "Access CE monthly rental for 6M service."),
        ("access_ce", "MegaLink Access Client End - Asymmetric 10Base-T bridging or ATM-25", "8M", "8", "459", "Access CE monthly rental for 8M service."),
        ("access_ce", "MegaLink Access Client End - Asymmetric 10Base-T routing", "1.5M", "1.5", "474", "Access CE monthly rental for 1.5M routing service."),
        ("access_ce", "MegaLink Access Client End - Asymmetric 10Base-T routing", "6M", "6", "490", "Access CE monthly rental for 6M routing service."),
        ("access_ce", "MegaLink Access Client End - Asymmetric G.DMT", "1.5M", "1.5", "344", "Access CE monthly rental for 1.5M G.DMT service."),
        ("access_ce", "MegaLink Access Client End - Asymmetric G.DMT", "6M", "6", "399", "Access CE monthly rental for 6M G.DMT service."),
        ("access_ce", "MegaLink Access Client End - Asymmetric G.DMT", "8M", "8", "423", "Access CE monthly rental for 8M G.DMT service."),
        ("access_ce", "MegaLink Access Client End - Symmetric 10Base-T", "2M", "2", "765", "Access CE monthly rental for 2M symmetric service."),
        ("access_se", "MegaLink Access Server End - OC-3c or FE first 15 km", "up to 100M", "100", "57820", "Access SE monthly rental for first 15 km."),
        ("access_se", "MegaLink Access Server End - OC-3c or FE additional 20 km", "up to 100M", "100", "12240", "Access SE monthly rental for each additional 20 km."),
        ("access_se", "MegaLink Access Server End - OC-3c or FE additional 500 sessions", "up to 100M", "100", "45890", "Access SE monthly rental for each additional 500 sessions."),
        ("access_se", "MegaLink Access Server End - Multiple Domain Name", "per additional domain name", "", "3060", "Access SE monthly rental per additional domain name per server end."),
        ("vpn_ce", "MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "1.5M", "1.5", "1910", "VPN CE monthly rental for 1.5M service."),
        ("vpn_ce", "MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "3M", "3", "2220", "VPN CE monthly rental for 3M service."),
        ("vpn_ce", "MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "4.5M", "4.5", "2370", "VPN CE monthly rental for 4.5M service."),
        ("vpn_ce", "MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "6M", "6", "2880", "VPN CE monthly rental for 6M service."),
        ("vpn_ce", "MegaLink VPN Client End - Asymmetric G.DMT", "1.5M", "1.5", "1530", "VPN CE monthly rental for 1.5M G.DMT service."),
        ("vpn_ce", "MegaLink VPN Client End - Asymmetric G.DMT", "3M", "3", "1910", "VPN CE monthly rental for 3M G.DMT service."),
        ("vpn_ce", "MegaLink VPN Client End - Asymmetric G.DMT", "4.5M", "4.5", "2070", "VPN CE monthly rental for 4.5M G.DMT service."),
        ("vpn_ce", "MegaLink VPN Client End - Asymmetric G.DMT", "6M", "6", "2570", "VPN CE monthly rental for 6M G.DMT service."),
        ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "2M", "2", "2750", "VPN CE monthly rental for 2M symmetric service."),
        ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "4M", "4", "4280", "VPN CE monthly rental for 4M symmetric service."),
        ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "6M", "6", "6820", "VPN CE monthly rental for 6M symmetric service."),
        ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "10M", "10", "9300", "VPN CE monthly rental for 10M symmetric service."),
        ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "25M", "25", "16050", "VPN CE monthly rental for 25M symmetric service."),
        ("vpn_se", "MegaLink VPN Server End - Symmetric OC-3c first 15 km", "up to 150M", "150", "17600", "VPN SE monthly rental for first 15 km."),
        ("vpn_se", "MegaLink VPN Server End - Symmetric OC-3c additional 20 km", "up to 150M", "150", "12240", "VPN SE monthly rental for each additional 20 km."),
        ("vpn_se", "MegaLink VPN Server End - Symmetric OC-3c bandwidth per 10M", "10M bandwidth block", "10", "2600", "VPN SE bandwidth monthly rental per 10M."),
        ("vpn_se", "MegaLink VPN Server End - Symmetric GE first 15 km", "up to 1G", "1000", "31350", "VPN SE monthly rental for first 15 km."),
        ("vpn_se", "MegaLink VPN Server End - Symmetric GE additional 20 km", "up to 1G", "1000", "12240", "VPN SE monthly rental for each additional 20 km."),
        ("vpn_se", "MegaLink VPN Server End - Symmetric GE bandwidth per 50M", "50M bandwidth block", "50", "6500", "VPN SE bandwidth monthly rental per 50M."),
    ]
    fee_overrides_2012 = {
        ("MegaLink Access Client End - Asymmetric 10Base-T bridging or ATM-25", "1.5M"): "330",
        ("MegaLink Access Client End - Asymmetric 10Base-T bridging or ATM-25", "6M"): "378",
        ("MegaLink Access Client End - Asymmetric 10Base-T bridging or ATM-25", "8M"): "399",
        ("MegaLink Access Client End - Asymmetric 10Base-T routing", "1.5M"): "412",
        ("MegaLink Access Client End - Asymmetric 10Base-T routing", "6M"): "426",
        ("MegaLink Access Client End - Asymmetric G.DMT", "1.5M"): "299",
        ("MegaLink Access Client End - Asymmetric G.DMT", "6M"): "347",
        ("MegaLink Access Client End - Asymmetric G.DMT", "8M"): "368",
        ("MegaLink Access Client End - Symmetric 10Base-T", "2M"): "665",
        ("MegaLink Access Server End - OC-3c or FE first 15 km", "up to 100M"): "50274",
        ("MegaLink Access Server End - OC-3c or FE additional 20 km", "up to 100M"): "10640",
        ("MegaLink Access Server End - OC-3c or FE additional 500 sessions", "up to 100M"): "39900",
        ("MegaLink Access Server End - Multiple Domain Name", "per additional domain name"): "2660",
        ("MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "1.5M"): "1663",
        ("MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "3M"): "1930",
        ("MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "4.5M"): "2060",
        ("MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "6M"): "2500",
        ("MegaLink VPN Client End - Asymmetric G.DMT", "1.5M"): "1330",
        ("MegaLink VPN Client End - Asymmetric G.DMT", "3M"): "1660",
        ("MegaLink VPN Client End - Asymmetric G.DMT", "4.5M"): "1800",
        ("MegaLink VPN Client End - Asymmetric G.DMT", "6M"): "2235",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "2M"): "2390",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "4M"): "3725",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "6M"): "5930",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "10M"): "8090",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "25M"): "13960",
        ("MegaLink VPN Server End - Symmetric OC-3c first 15 km", "up to 150M"): "15300",
        ("MegaLink VPN Server End - Symmetric OC-3c additional 20 km", "up to 150M"): "10640",
        ("MegaLink VPN Server End - Symmetric OC-3c bandwidth per 10M", "10M bandwidth block"): "2260",
        ("MegaLink VPN Server End - Symmetric GE first 15 km", "up to 1G"): "27260",
        ("MegaLink VPN Server End - Symmetric GE additional 20 km", "up to 1G"): "10640",
        ("MegaLink VPN Server End - Symmetric GE bandwidth per 50M", "50M bandwidth block"): "5650",
    }
    fee_overrides_2017 = {
        ("MegaLink Access Client End - Asymmetric 10Base-T bridging or ATM-25", "1.5M"): "460",
        ("MegaLink Access Client End - Asymmetric 10Base-T bridging or ATM-25", "6M"): "530",
        ("MegaLink Access Client End - Asymmetric 10Base-T bridging or ATM-25", "8M"): "560",
        ("MegaLink Access Client End - Asymmetric 10Base-T routing", "1.5M"): "570",
        ("MegaLink Access Client End - Asymmetric 10Base-T routing", "6M"): "590",
        ("MegaLink Access Client End - Asymmetric G.DMT", "1.5M"): "420",
        ("MegaLink Access Client End - Asymmetric G.DMT", "6M"): "480",
        ("MegaLink Access Client End - Asymmetric G.DMT", "8M"): "510",
        ("MegaLink Access Client End - Symmetric 10Base-T", "2M"): "920",
        ("MegaLink Access Server End - OC-3c or FE first 15 km", "up to 100M"): "69390",
        ("MegaLink Access Server End - OC-3c or FE additional 20 km", "up to 100M"): "14690",
        ("MegaLink Access Server End - OC-3c or FE additional 500 sessions", "up to 100M"): "55070",
        ("MegaLink Access Server End - Multiple Domain Name", "per additional domain name"): "3680",
        ("MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "1.5M"): "2300",
        ("MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "3M"): "2670",
        ("MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "4.5M"): "2850",
        ("MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "6M"): "3460",
        ("MegaLink VPN Client End - Asymmetric G.DMT", "1.5M"): "1840",
        ("MegaLink VPN Client End - Asymmetric G.DMT", "3M"): "2300",
        ("MegaLink VPN Client End - Asymmetric G.DMT", "4.5M"): "2490",
        ("MegaLink VPN Client End - Asymmetric G.DMT", "6M"): "3090",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "2M"): "3300",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "4M"): "5140",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "6M"): "8190",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "10M"): "11160",
        ("MegaLink VPN Client End - Symmetric 10/100Base-T", "25M"): "19260",
        ("MegaLink VPN Server End - Symmetric OC-3c first 15 km", "up to 150M"): "21120",
        ("MegaLink VPN Server End - Symmetric OC-3c additional 20 km", "up to 150M"): "14690",
        ("MegaLink VPN Server End - Symmetric OC-3c bandwidth per 10M", "10M bandwidth block"): "3120",
        ("MegaLink VPN Server End - Symmetric GE first 15 km", "up to 1G"): "37620",
        ("MegaLink VPN Server End - Symmetric GE additional 20 km", "up to 1G"): "14690",
        ("MegaLink VPN Server End - Symmetric GE bandwidth per 50M", "50M bandwidth block"): "7800",
    }
    effective_from = source.get("effective_from") or "2014-01-10"
    specs_to_parse = specs
    if effective_from == "2023-09-19" or source.get("published_on") == "2023-09-19":
        specs_to_parse = [
            spec for spec in specs
            if spec[0] != "vpn_ce"
        ] + [
            ("vpn_ce", "MegaLink VPN Client End - Asymmetric 10Base-T or ATM-25", "up to 6M flat monthly rental", "6", "20000", "VPN CE monthly rental for Asymmetric 10Base-T or ATM-25 up to 6M."),
            ("vpn_ce", "MegaLink VPN Client End - Asymmetric G.DMT", "up to 6M flat monthly rental", "6", "20000", "VPN CE monthly rental for Asymmetric G.DMT up to 6M."),
            ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "2M", "2", "10000", "VPN CE monthly rental for 2M symmetric service."),
            ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "4M", "4", "15000", "VPN CE monthly rental for 4M symmetric service."),
            ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "6M", "6", "20000", "VPN CE monthly rental for 6M symmetric service."),
            ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "10M", "10", "30000", "VPN CE monthly rental for 10M symmetric service."),
            ("vpn_ce", "MegaLink VPN Client End - Symmetric 10/100Base-T", "25M", "25", "50000", "VPN CE monthly rental for 25M symmetric service."),
        ]
    rows: List[Dict[str, str]] = []
    for section_key, service_type, bandwidth, speed, fee, voice in specs_to_parse:
        if effective_from == "2012-07-03" or source.get("published_on") == "2012-07-03":
            fee = fee_overrides_2012.get((service_type, bandwidth), fee)
        elif effective_from == "2017-01-10" or source.get("published_on") == "2016-11-30":
            fee = fee_overrides_2017.get((service_type, bandwidth), fee)
        search_section = section_text.get(section_key, "").replace(",", "")
        if not re.search(rf"\$?\s*{fee}\b", search_section):
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Enterprise official MegaLink Service effective {effective_from} - {service_type} {bandwidth}",
                "monthly_fee_hkd": fee,
                "post_fup_speed_mbps": speed,
                "contract_months": "12",
                "local_voice": voice,
                "add_on_charges_hkd": "Official MegaLink Service monthly rental/list price component; installation, internal/external relocation, reconfiguration, deletion, domain-name change and non-standard works charges are disclosed separately and excluded from monthly_fee_hkd.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(section_text.get(section_key, "") or text, 0, min(len(section_text.get(section_key, "") or text), 1000)),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_datapak_private_circuit_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Datapak Services and Private Circuit Service" not in text or "Monthly Rental" not in text:
        return []
    fiberline_match = re.search(r"4\.\s+Datapak Fiberline Service(?P<section>.*?)Remarks for Datapak Fiberline Service", text, flags=re.I | re.S)
    internal_match = re.search(r"9\.\s+Internal Leased Circuit(?P<section>.*?)Note for Internal Leased Circuit", text, flags=re.I | re.S)
    fiberline_section = fiberline_match.group("section") if fiberline_match else ""
    internal_section = internal_match.group("section") if internal_match else ""
    specs = [
        ("fiberline", "Datapak Fiberline Service - Fast Ethernet", "18800", "100", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Ethernet Private Line 50M", "12000", "50", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Ethernet Private Line 30M", "10200", "30", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Ethernet Private Line 20M", "9300", "20", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Ethernet Private Line 15M", "8625", "15", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Ethernet Private Line 10M", "7950", "10", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Gigabit Ethernet", "68000", "1000", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Gigabit Ethernet 200M", "28200", "200", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Fibre Channel 1G", "68000", "1000", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Fibre Channel 2G", "78000", "2000", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Fibre Channel 4G", "88000", "4000", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - 10 Gigabit Ethernet", "140000", "10000", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - STM1/OC3", "52000", "155", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - STM1/OC3", "70000", "155", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - T3", "29325", "45", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Channelized T3", "43000", "45", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Multi-Drop T3", "43000", "45", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Fiberline Connect", "25000", "", "Fiberline Connect monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Fiberline FDDI", "13250", "100", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Fiberline Coupling", "40000", "", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Fiberline Cellular Dual Band 2-Fibre", "16000", "", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Fiberline Cellular Dual Band 3-fibre", "20000", "", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline Service - Fiberline Wavelength 1", "18800", "", "Fiberline monthly rental per end."),
        ("fiberline", "Datapak Fiberline DWDM - 4 interfaces and bandwidth of 2.5G", "70000", "2500", "DWDM monthly rental per location."),
        ("fiberline", "Datapak Fiberline DWDM - additional interface and bandwidth of 1.25G", "5000", "1250", "DWDM additional interface monthly rental."),
        ("fiberline", "Datapak Fiberline CWDM - first 2 interfaces", "40000", "", "CWDM monthly rental per location."),
        ("fiberline", "Datapak Fiberline CWDM - additional interface", "5000", "", "CWDM additional interface monthly rental."),
        ("internal", "Internal Leased Circuit - private circuit within same building 2-wire", "117.8", "", "Internal leased circuit monthly rental for standard 2-wire circuit."),
        ("internal", "Internal Leased Circuit - private circuit within same building 4-wire", "245.8", "", "Internal leased circuit monthly rental for standard 4-wire circuit."),
        ("internal", "Internal Leased Circuit - private circuit within same building 2-wire", "215", "", "Internal leased circuit monthly rental for standard 2-wire circuit."),
        ("internal", "Internal Leased Circuit - private circuit within same building 4-wire", "446", "", "Internal leased circuit monthly rental for standard 4-wire circuit."),
    ]
    sections = {"fiberline": fiberline_section, "internal": internal_section}
    rows: List[Dict[str, str]] = []
    for section_key, service_type, fee, speed, voice in specs:
        section = sections.get(section_key, "")
        fee_pattern = re.escape(f"{float(fee):,.1f}") if "." in fee else re.escape(f"{int(fee):,}.0")
        if not re.search(rf"\b{fee_pattern}\b", section):
            continue
        if service_type.endswith("Gigabit Ethernet 200M") and "Gigabit Ethernet (200M)" not in section:
            continue
        if service_type.endswith("STM1/OC3") and fee == "52000" and "STM1/OC3 52,000.0" not in section:
            continue
        if service_type.endswith("STM1/OC3") and fee == "70000" and "STM1/OC3 70,000.0" not in section:
            continue
        if service_type.endswith("Channelized T3") and "Channelized T3" not in section:
            continue
        if service_type.endswith("Multi-Drop T3") and "Multi-Drop T3" not in section:
            continue
        if service_type.endswith("Fiberline Wavelength 1") and "Fiberline Wavelength 1" not in section:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Enterprise official Datapak/Private Circuit tariff {source.get('published_on', '')} - {service_type}",
                "monthly_fee_hkd": fee,
                "post_fup_speed_mbps": speed,
                "local_voice": voice,
                "add_on_charges_hkd": "Official Datapak Services and Private Circuit Service tariff/list price; installation, external/internal removal or relocation, reconfiguration, usage charges, diversity distance multipliers and non-standard provision costs are disclosed separately and excluded from monthly_fee_hkd.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(section, 0, min(len(section), 1000)),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_metro_ip_service_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Metro IP Service Packages" not in text or "Rental/month" not in text:
        return []
    specs = {
        "Silver": [
            ("512K~10Mbps", "10", "20000"),
            ("11~20", "20", "25960"),
            ("21~30", "30", "31240"),
            ("31~40", "40", "36520"),
            ("41~50", "50", "41800"),
            ("51~60", "60", "47080"),
            ("61~70", "70", "52360"),
            ("71~80", "80", "57640"),
            ("81~90", "90", "62920"),
            ("91~100", "100", "68200"),
            ("101~200", "200", "90200"),
            ("201~300", "300", "107800"),
            ("301~400", "400", "125400"),
            ("401~500", "500", "143000"),
            ("501~600", "600", "160600"),
            ("601~700", "700", "178200"),
            ("701~800", "800", "195800"),
            ("801~900", "900", "213400"),
            ("901~1000", "1000", "231000"),
        ],
        "Gold": [
            ("1~10", "10", "33000"),
            ("11~20", "20", "51840"),
            ("21~30", "30", "61340"),
            ("31~40", "40", "71720"),
            ("41~50", "50", "82080"),
            ("51~60", "60", "92440"),
            ("61~70", "70", "102820"),
            ("71~80", "80", "113180"),
            ("81~90", "90", "123560"),
            ("91~100", "100", "133920"),
            ("101~200", "200", "167280"),
            ("201~300", "300", "199920"),
            ("301~400", "400", "232560"),
            ("401~500", "500", "265200"),
            ("501~600", "600", "297840"),
            ("601~700", "700", "330480"),
            ("701~800", "800", "363120"),
            ("801~900", "900", "395760"),
            ("901~1000", "1000", "428400"),
        ],
        "Diamond": [
            ("1~10", "10", "35100"),
            ("11~20", "20", "54120"),
            ("21~30", "30", "61340"),
            ("31~40", "40", "71720"),
            ("41~50", "50", "82080"),
            ("51~60", "60", "92440"),
            ("61~70", "70", "102820"),
            ("71~80", "80", "113180"),
            ("81~90", "90", "123560"),
            ("91~100", "100", "133920"),
            ("101~200", "200", "167280"),
            ("201~300", "300", "199920"),
            ("301~400", "400", "232560"),
            ("401~500", "500", "265200"),
            ("501~600", "600", "297840"),
            ("601~700", "700", "330480"),
            ("701~800", "800", "363120"),
            ("801~900", "900", "395760"),
            ("901~1000", "1000", "428400"),
        ],
    }
    rows: List[Dict[str, str]] = []
    for plan, plan_specs in specs.items():
        plan_match = re.search(rf"Metro IP Service - {plan} Plan.*?(?:Page \d+|$)", text, flags=re.I | re.S)
        plan_text = plan_match.group(0) if plan_match else text
        if f"{plan} Plan" not in text:
            continue
        for bandwidth, speed, fee in plan_specs:
            if not re.search(rf"{re.escape(bandwidth)}\s+\$?[\d,]+\s+\$?{int(fee):,}", plan_text, flags=re.I):
                continue
            row = _base_row(source, result, captured_at)
            row.update(
                {
                    "plan_name": f"HKT Enterprise official Metro IP Service {plan} Plan {bandwidth} tariff {source.get('published_on', '')}",
                    "monthly_fee_hkd": fee,
                    "post_fup_speed_mbps": speed,
                    "contract_months": "12",
                    "local_voice": f"Metro IP Service {plan} Plan bandwidth range {bandwidth}.",
                    "add_on_charges_hkd": "Official Metro IP tariff/list price; setup, internal/external relocation and reconfiguration charges are disclosed separately and excluded from monthly_fee_hkd.",
                    "official_source_type": "official_public_tariff_pdf",
                    "extraction_status": "parsed_official_tariff_pdf",
                    "evidence_excerpt": _excerpt(plan_text, 0, min(len(plan_text), 1000)),
                }
            )
            rows.append(row)
    return _dedupe(rows)


def _parse_hkt_flexible_bandwidth_service_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Flexible Bandwidth Service" not in text or "Monthly Recurrent Charges" not in text:
        return []
    monthly_section = re.search(r"1\.3\.2\s+Monthly Recurrent Charges(?P<section>.*?)1\.3\.3\s+Diversity Option", text, flags=re.I | re.S)
    if not monthly_section:
        return []
    section = monthly_section.group("section")
    server_section = re.search(r"Server End\s*:(?P<section>.*?)Client End\s*:", section, flags=re.I | re.S)
    client_section = re.search(r"Client End\s*:(?P<section>.*?)Point-to-Point Service", section, flags=re.I | re.S)
    point_to_point_section = re.search(r"Point-to-Point Service(?P<section>.*)", section, flags=re.I | re.S)
    sections_by_kind = {
        "server": server_section.group("section") if server_section else "",
        "client": client_section.group("section") if client_section else "",
        "point_to_point": point_to_point_section.group("section") if point_to_point_section else "",
    }
    specs = [
        ("server", "Point-to-Multipoint Server End", "1-50 Mbps", "50", "15000"),
        ("server", "Point-to-Multipoint Server End", "51-100 Mbps", "100", "25000"),
        ("server", "Point-to-Multipoint Server End", "101-500 Mbps", "500", "39000"),
        ("server", "Point-to-Multipoint Server End GE interface", "501-1000 Mbps", "1000", "51000"),
        ("server", "Point-to-Multipoint Server End 10GE interface", "501-1000 Mbps", "1000", "180000"),
        ("server", "Point-to-Multipoint Server End", "1001-2000 Mbps", "2000", "210000"),
        ("server", "Point-to-Multipoint Server End", "2001-3000 Mbps", "3000", "240000"),
        ("server", "Point-to-Multipoint Server End", "3001-4000 Mbps", "4000", "270000"),
        ("server", "Point-to-Multipoint Server End", "4001-5000 Mbps", "5000", "300000"),
        ("server", "Point-to-Multipoint Server End", "5001-10000 Mbps", "10000", "450000"),
        ("client", "Point-to-Multipoint Client End", "1 Mbps", "1", "3600"),
        ("client", "Point-to-Multipoint Client End", "2 Mbps", "2", "4200"),
        ("client", "Point-to-Multipoint Client End", "3 Mbps", "3", "4800"),
        ("client", "Point-to-Multipoint Client End", "4 Mbps", "4", "5400"),
        ("client", "Point-to-Multipoint Client End", "5 Mbps", "5", "6000"),
        ("client", "Point-to-Multipoint Client End", "6 Mbps", "6", "6600"),
        ("client", "Point-to-Multipoint Client End", "7 Mbps", "7", "7200"),
        ("client", "Point-to-Multipoint Client End", "8 Mbps", "8", "7800"),
        ("client", "Point-to-Multipoint Client End", "9 Mbps", "9", "8400"),
        ("client", "Point-to-Multipoint Client End", "10 Mbps", "10", "9000"),
        ("client", "Point-to-Multipoint Client End", "15 Mbps", "15", "10500"),
        ("client", "Point-to-Multipoint Client End", "20 Mbps", "20", "11000"),
        ("client", "Point-to-Multipoint Client End", "25 Mbps", "25", "11500"),
        ("client", "Point-to-Multipoint Client End", "30 Mbps", "30", "12000"),
        ("client", "Point-to-Multipoint Client End", "35 Mbps", "35", "12500"),
        ("client", "Point-to-Multipoint Client End", "40 Mbps", "40", "13000"),
        ("client", "Point-to-Multipoint Client End", "45 Mbps", "45", "13500"),
        ("client", "Point-to-Multipoint Client End", "50 Mbps", "50", "14000"),
        ("client", "Point-to-Multipoint Client End", "60 Mbps", "60", "15000"),
        ("client", "Point-to-Multipoint Client End", "70 Mbps", "70", "16000"),
        ("client", "Point-to-Multipoint Client End", "80 Mbps", "80", "17000"),
        ("client", "Point-to-Multipoint Client End", "90 Mbps", "90", "18000"),
        ("client", "Point-to-Multipoint Client End", "100 Mbps", "100", "32000"),
        ("client", "Point-to-Multipoint Client End", "200 Mbps", "200", "34000"),
        ("client", "Point-to-Multipoint Client End", "300 Mbps", "300", "36000"),
        ("client", "Point-to-Multipoint Client End", "400 Mbps", "400", "38000"),
        ("client", "Point-to-Multipoint Client End", "500 Mbps", "500", "40000"),
        ("client", "Point-to-Multipoint Client End GE interface", "501-1000 Mbps", "1000", "50000"),
        ("point_to_point", "Point-to-Point Service", "1 Mbps", "1", "3600"),
        ("point_to_point", "Point-to-Point Service", "2 Mbps", "2", "4200"),
        ("point_to_point", "Point-to-Point Service", "3 Mbps", "3", "4800"),
        ("point_to_point", "Point-to-Point Service", "4 Mbps", "4", "5400"),
        ("point_to_point", "Point-to-Point Service", "5 Mbps", "5", "6000"),
        ("point_to_point", "Point-to-Point Service", "6 Mbps", "6", "6600"),
        ("point_to_point", "Point-to-Point Service", "7 Mbps", "7", "7200"),
        ("point_to_point", "Point-to-Point Service", "8 Mbps", "8", "7800"),
        ("point_to_point", "Point-to-Point Service", "9 Mbps", "9", "8400"),
        ("point_to_point", "Point-to-Point Service", "10 Mbps", "10", "9000"),
        ("point_to_point", "Point-to-Point Service", "15 Mbps", "15", "10500"),
        ("point_to_point", "Point-to-Point Service", "20 Mbps", "20", "11000"),
        ("point_to_point", "Point-to-Point Service", "25 Mbps", "25", "11500"),
        ("point_to_point", "Point-to-Point Service", "30 Mbps", "30", "12000"),
        ("point_to_point", "Point-to-Point Service", "35 Mbps", "35", "12500"),
        ("point_to_point", "Point-to-Point Service", "40 Mbps", "40", "13000"),
        ("point_to_point", "Point-to-Point Service", "45 Mbps", "45", "13500"),
        ("point_to_point", "Point-to-Point Service", "50 Mbps", "50", "14000"),
        ("point_to_point", "Point-to-Point Service", "60 Mbps", "60", "15000"),
        ("point_to_point", "Point-to-Point Service", "70 Mbps", "70", "16000"),
        ("point_to_point", "Point-to-Point Service", "80 Mbps", "80", "17000"),
        ("point_to_point", "Point-to-Point Service", "90 Mbps", "90", "18000"),
        ("point_to_point", "Point-to-Point Service", "100 Mbps", "100", "32000"),
        ("point_to_point", "Point-to-Point Service", "1000 Mbps", "1000", "50000"),
        ("point_to_point", "Point-to-Point Service", "10 Gbps", "10000", "450000"),
        ("point_to_point", "Point-to-Point Service", "100 Gbps", "100000", "1500000"),
    ]
    rows: List[Dict[str, str]] = []
    for section_key, service_type, bandwidth, speed, fee in specs:
        search_section = sections_by_kind.get(section_key, "").replace(",", "").replace("–", "-")
        bandwidth_pattern = re.escape(bandwidth).replace(r"\-", r"\s*[\-–]\s*")
        if not re.search(rf"{bandwidth_pattern}(?:\s+\([^)]*\))?\s+{fee}", search_section, flags=re.I):
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Enterprise official Flexible Bandwidth Service {service_type} {bandwidth} tariff {source.get('published_on', '')}",
                "monthly_fee_hkd": fee,
                "post_fup_speed_mbps": speed,
                "contract_months": "12",
                "local_voice": f"Flexible Bandwidth Service monthly recurrent charge: {service_type}, {bandwidth}.",
                "add_on_charges_hkd": "Official Flexible Bandwidth Service monthly recurrent tariff/list price; one-off installation, relocation, reconfiguration, diversity option and extra-distance port charges are disclosed separately and excluded from monthly_fee_hkd.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(section, 0, min(len(section), 1000)),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_telecommunications_backup_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Telecommunications Backup Service for Commercial Customers" not in text or "Monthly Charge" not in text:
        return []
    specs = [
        ("Continuity Plan CPC 1-50 Lines", "2500", r"1-50 Lines\s+HK\$?2,500", "Continuity Plan Configuration monthly charge for 1-50 lines."),
        ("Continuity Plan CPC 51-100 Lines", "3000", r"51-100 Lines\s+HK\$?3,000", "Continuity Plan Configuration monthly charge for 51-100 lines."),
        ("Continuity Plan CPC 101-200 Lines", "4000", r"101-200 Lines\s+HK\$?4,000", "Continuity Plan Configuration monthly charge for 101-200 lines."),
        ("Continuity Plan CPC 201-300 Lines", "6000", r"201-300 Lines\s+HK\$?6,000", "Continuity Plan Configuration monthly charge for 201-300 lines."),
        ("Continuity Plan CPC every additional 100 Lines", "2000", r"Thereafter every\s+HK\$?2,000\s+additional 100 Lines", "Continuity Plan Configuration monthly charge for each additional 100 lines."),
        ("IDA-P Mutual Backup Service per IDA-P Line", "2000", r"Monthly charge:\s+HK\$?2,000\s*/\s*IDA-P Line", "IDA-P Mutual Backup monthly charge per IDA-P Line."),
        ("Smartline iBCP 1-25 Lines", "5000", r"1-25 Lines\s+HK\$?5,000", "Smartline iBCP monthly charge for 1-25 lines."),
        ("Smartline iBCP 26-50 Lines", "6000", r"26-50 Lines\s+HK\$?6,000", "Smartline iBCP monthly charge for 26-50 lines."),
        ("Smartline iBCP every additional 25 Lines", "6000", r"Thereafter every\s+HK\$?6,000\s+additional 25Lines", "Smartline iBCP monthly charge for each additional 25 lines."),
    ]
    rows: List[Dict[str, str]] = []
    monthly_section = re.search(r"\(b\)\s+Monthly Charge.*?Remarks:", text, flags=re.I | re.S)
    search_text = monthly_section.group(0) if monthly_section else text
    for label, fee, pattern, voice in specs:
        match = re.search(pattern, search_text, flags=re.I | re.S)
        if not match:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Enterprise official Telecommunications Backup Service - {label} tariff {source.get('published_on', '')}",
                "monthly_fee_hkd": fee,
                "local_voice": voice,
                "add_on_charges_hkd": "Official telecommunications backup tariff/list price; setup, external removal, extra activation and extra switch-over charges are disclosed separately and excluded from monthly_fee_hkd.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_one_communications_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Name of Tariff: one communications" not in text or "Monthly Service" not in text:
        return []
    specs = [
        ("A", "Staff", "voice and broadband service", "258"),
        ("B", "Executive", "Staff user plan with FMC features", "298"),
        ("C", "Boss/Secretary", "Executive user plan with Boss/Sec", "398"),
        ("D", "Boss (Lite)/Secretary (Lite)", "Executive user plan with Boss/Sec", "368"),
        ("E", "Operator", "Boss/Secretary user plan with operator features", "638"),
        ("F", "Enterprise Centrex Executive (1:1)", "voice service only with FMC features; line-to-user ratio 1:1", "238"),
        ("G", "Enterprise Centrex Boss/Secretary (1:1)", "EC-Executive user plan with Boss/Sec; line-to-user ratio 1:1", "338"),
        ("H", "Enterprise Centrex Boss (Lite)/Secretary (Lite) (1:1)", "EC-Executive user plan with Boss/Sec; line-to-user ratio 1:1", "308"),
        ("I", "Enterprise Centrex Operator (1:1)", "EC Boss/Secretary user plan with operator features; line-to-user ratio 1:1", "578"),
        ("J", "Enterprise Centrex Executive (1:3)", "voice service only with FMC features; line-to-user ratio 1:3", "208"),
        ("K", "Enterprise Centrex Boss/Secretary (1:3)", "EC-Executive user plan with Boss/Sec; line-to-user ratio 1:3", "308"),
        ("L", "Enterprise Centrex Boss (Lite)/Secretary (Lite) (1:3)", "EC-Executive user plan with Boss/Sec; line-to-user ratio 1:3", "278"),
        ("M", "Enterprise Centrex Operator (1:3)", "EC Boss/Secretary user plan with operator features; line-to-user ratio 1:3", "548"),
    ]
    rows: List[Dict[str, str]] = []
    for code, name, service_note, fee in specs:
        if code not in text or fee not in text:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Enterprise official one communications tariff {source.get('published_on', '')} - Plan {code} {name}",
                "monthly_fee_hkd": fee,
                "local_voice": service_note,
                "add_on_charges_hkd": "Official one communications tariff/list price; FMC features HK$98/user/month, Office Anywhere HK$98/user/month and multi-site intercom HK$100/site/month are add-on charges, not main plan monthly fee.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, 0, min(len(text), 1600)),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_eye_service_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Name of Tariff:" not in text or "eye" not in text or "core service charge" not in text:
        return []
    match = re.search(r"eye(?: Communication Package| Service)? core\s+service charge.*?888\s+per month", text, re.I)
    if not match:
        return []
    service_label = "eye Communication Package" if "eye Communication Package" in text else "eye Service"
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"HKT official {service_label} core service tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": "888",
            "local_voice": f"Voice call, Infotainment Services access, SMS and local video call features under {service_label} core service.",
            "add_on_charges_hkd": (
                f"Official {service_label} tariff/list price; SMS, local video call, infotainment services, "
                "installation and replacement device charges are usage/add-on/one-off items and are not written as main monthly fee."
            ),
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_hkt_international_toll_free_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "International Toll Free Service" not in text:
        return []
    match = re.search(r"\(b\)\s*Service per month\s+\$?\s*400", text, flags=re.I | re.S)
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"HKT Enterprise official International Toll Free Service tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": "400",
            "local_voice": "International direct-dial voice line for which the registered Hong Kong customer pays the call charges.",
            "add_on_charges_hkd": "Official tariff/list price. Registration is HK$400 per line and international calls are charged under the applicable IDD service-rate table; neither is included in the monthly service fee.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_hkt_super_hotline_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Super Hotline" not in text:
        return []
    match = re.search(r"Rental:\s*\$?\s*1000\s*/\s*port\s*/\s*month", text, flags=re.I)
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update({
        "plan_name": f"HKT Enterprise official Super Hotline tariff {source.get('published_on', '')}",
        "monthly_fee_hkd": "1000",
        "local_voice": "Super Hotline supports agreed routing of telephone numbers, call forwarding and simultaneous calls through multiple ports.",
        "add_on_charges_hkd": "Official tariff/list price of HK$1,000 per port per month. Setup is HK$8,000 per man-day and IVRS features carry additional charges; neither is included in the monthly rental.",
        "official_source_type": "official_public_tariff_pdf",
        "extraction_status": "parsed_official_tariff_pdf",
        "evidence_excerpt": _excerpt(text, match.start(), match.end()),
    })
    return [row]


def _parse_hkt_faxline_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    specs = (("Faxline 100", "31"), ("Faxline 200", "48"), ("Faxline 3", "48"), ("Faxline 2", "31"))
    service, fee = next(((name, amount) for name, amount in specs if f"{name} service" in text), ("", ""))
    match = re.search(rf"Charges:\s*HK\$\s*{fee}\s+per month", text, flags=re.I | re.S) if fee else None
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update({
        "plan_name": f"HKT Enterprise official {service} tariff {source.get('published_on', '')}",
        "monthly_fee_hkd": fee,
        "local_voice": f"Dedicated business fax service under the official {service} tariff.",
        "add_on_charges_hkd": "Official tariff/list price. A business telephone line is required separately; Faxline 100 also requires fax-support service. International Follow Me usage is charged at applicable IDD rates. These separate charges are excluded from the monthly fee.",
        "official_source_type": "official_public_tariff_pdf",
        "extraction_status": "parsed_official_tariff_pdf",
        "evidence_excerpt": _excerpt(text, match.start(), match.end()),
    })
    return [row]


def _parse_hkt_homefax_1_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Homefax 1 service" not in text:
        return []
    match = re.search(r"Charges:\s*HK\$\s*30\s+per month", text, flags=re.I | re.S)
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"HKT official Homefax 1 service tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": "30",
            "local_voice": "Premium dedicated home fax service with Follow Me, Call Security, Fax Support Service and Hong Kong Fax Directory listing.",
            "add_on_charges_hkd": "Official tariff/list price. Customers must separately subscribe to a residential telephone exchange line; that underlying line rental is not included in this Homefax service monthly fee.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_1010_ipad_pro_2020_product_page(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    is_english = "Mobile fee" in text and "Local Mobile Data Usage" in text
    is_chinese = "4G服務計劃" in text and "本地流動數據用量" in text
    if not (is_english or is_chinese):
        return []

    table_label = "Mobile fee" if is_english else "月費"
    data_label = "Local Mobile Data Usage" if is_english else "本地流動數據用量"
    table_start = text.find(table_label)
    table_end = text.find("iPad Pro 11", table_start)
    block = text[table_start:table_end if table_end > table_start else table_start + 1800]
    fees = re.search(r"\$529\s+\$609\s+\$799", block)
    data = re.search(r"(?:6GB\s*\+\s*2GB).*?(?:6GB\s*\+\s*6GB).*?(?:10GB\s*\+\s*10GB)", block, flags=re.I | re.S)
    if not fees or not data or data_label not in block:
        return []

    contract_match = re.search(
        r"(?:24-month commitment period|commitment period of 24 months|24個月承諾期)",
        text,
        flags=re.I,
    )
    rows: List[Dict[str, str]] = []
    for fee, base_data in (("529", "6"), ("609", "6"), ("799", "10")):
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"1O1O iPad Pro 2020 4G service plan HK${fee}",
                "monthly_fee_hkd": fee,
                "local_data_gb": base_data,
                "contract_months": "24" if contract_match else "",
                "local_voice": "Unlimited local basic voice minutes.",
                "add_on_charges_hkd": "The page displays local data as a base entitlement plus Club member bonus data. local_data_gb records only the directly stated base entitlement (6GB, 6GB or 10GB); member bonus data, handset price, prepayment and the HK$18 monthly licence/admin charge are excluded from the base monthly fee.",
                "official_source_type": "official_public_product_page",
                "extraction_status": "parsed_official_product_page",
                "evidence_excerpt": _excerpt(text, table_start, min(len(text), table_start + 1000)),
            }
        )
        rows.append(row)
    return rows


def _parse_hkt_eye_home_smartphone_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "eye Home Smartphone Package" not in text:
        return []
    core_match = re.search(r"eye Home\s+Smartphone.*?\$?\s*(258|888)\s+or more per month", text, flags=re.I | re.S)
    extension_match = re.search(r"Parallel\s+(?:extension|phone).*?Monthly rental.*?\$?\s*(110|298)", text, flags=re.I | re.S)
    if not core_match:
        return []
    core_fee = core_match.group(1)
    specs = [
        (
            "core service",
            core_fee,
            core_match,
            "Voice call, access to Infotainment Services, SMS and local video call under the eye Home Smartphone Package.",
            "Official tariff/list price lower bound; actual monthly fee varies with subscribed information-service offers. SMS, video calls, installation, relocation and reconnection are usage or one-off charges and excluded from the main monthly fee.",
        ),
    ]
    if extension_match:
        extension_fee = extension_match.group(1)
        specs.append(
            (
                "parallel extension phone line",
                extension_fee,
                extension_match,
                "Monthly rental for a parallel extension phone line where applicable to an eye Home Smartphone Package.",
                "Official tariff/list price. Installation, relocation and other line changes are one-off charges and excluded from this monthly rental.",
            )
        )
    rows: List[Dict[str, str]] = []
    for label, fee, match, local_voice, notes in specs:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT official eye Home Smartphone Package {label} tariff {source.get('published_on', '')}",
                "monthly_fee_hkd": fee,
                "local_voice": local_voice,
                "add_on_charges_hkd": notes,
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return rows


def _parse_hkt_eye_multimedia_service_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "eye Multimedia Service" not in text:
        return []
    core_match = re.search(r"eye Multimedia\s+Service.*?888\s+per month", text, flags=re.I | re.S)
    extension_match = re.search(r"Parallel extension.*?Monthly rental.*?298", text, flags=re.I | re.S)
    if not core_match:
        return []
    specs = [
        (
            "core service",
            "888",
            core_match,
            "Voice call, Infotainment Services access, SMS and local video call under the eye Multimedia Service.",
            "Official tariff/list price. SMS, video calls, information services, installation, relocation, reconnection and paper-bill charges are usage, optional or one-off charges and excluded from the main monthly fee.",
        ),
    ]
    if extension_match:
        specs.append(
            (
                "parallel extension phone line",
                "298",
                extension_match,
                "Monthly rental for a parallel extension phone line where applicable to the eye Multimedia Service.",
                "Official tariff/list price. Installation, relocation and other line changes are one-off charges and excluded from this monthly rental.",
            )
        )
    rows: List[Dict[str, str]] = []
    for label, fee, match, local_voice, notes in specs:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT official eye Multimedia Service {label} tariff {source.get('published_on', '')}",
                "monthly_fee_hkd": fee,
                "local_voice": local_voice,
                "add_on_charges_hkd": notes,
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return rows


def _parse_hkt_home_easywatch_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "PCCW Home EasyWatch" not in text or "Tariff Table" not in text:
        return []
    match = re.search(r"Service Fee including\s+8\s+hours storage\s+498\s+per month", text, flags=re.I)
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"HKT official PCCW Home EasyWatch tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": "498",
            "local_voice": "Remote home monitoring service with 8 hours storage; service ceased for new customers from 2013-04-01 per tariff.",
            "add_on_charges_hkd": "Official Home EasyWatch tariff/list price; installation/relocation fee HK$2,000 and video-call/data access charges are one-off or usage charges, not main monthly fee.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_hkt_easywatch_commercial_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "HKT EasyWatch Commercial Service" not in text or "Service Fee" not in text:
        return []
    match = re.search(r"Service Fee\s+1,500\s+per month", text, flags=re.I)
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"HKT Enterprise official EasyWatch Commercial Service tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": "1500",
            "local_voice": "Commercial broadband Internet access service dedicated for connection to IP cameras under HKT EasyWatch Plus / PCCW EasyWatch Enterprise Solution.",
            "add_on_charges_hkd": "Official EasyWatch Commercial tariff/list price; installation/relocation fee HK$1,600 and storage fee HK$5/GB/month are one-off/add-on items, not the main service fee.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_hkt_ip_voice_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Internet Protocol Voice Service" not in text or "Service fee" not in text:
        return []
    match = re.search(r"Service fee.*?Unlimited voice calls to any Hong Kong telephone number.*?298\s*/\s*month", text, flags=re.I | re.S)
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"HKT official Internet Protocol Voice Service tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": "298",
            "local_voice": "Unlimited voice calls to Hong Kong telephone numbers except chargeable Infoline service.",
            "add_on_charges_hkd": "Official IP Voice tariff/list price; SMS, local video calls, IDD, installation/relocation, device and cancellation charges are usage/add-on/one-off items and are not written as main monthly fee.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_csl_voip_monthly_pass_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if not re.search(r"Internet Protocol\s+V\s*oice Service", text, flags=re.I):
        return []
    match = re.search(r"Basic Charges:.*?Day Pass:\s*\$15/day.*?Monthly Pass:\s*\$100/month", text, flags=re.I | re.S)
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"csl official Internet Protocol Voice Monthly Pass tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": "100",
            "local_voice": "VoIP service for eligible CSL mobile subscribers through a designated smartphone application over Wi-Fi or mobile data; the tariff does not state an included voice-minute quota.",
            "add_on_charges_hkd": "Official tariff/list price. The separately listed Day Pass is HK$15/day and is not recorded as a monthly plan. Underlying mobile service, Wi-Fi, roaming data, international SMS and local mobile usage charges remain separate where applicable.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_hkt_freedome_network_safety_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Freedome Network Safety Software" not in text or "HK$38 per month per device" not in text:
        return []
    match = re.search(r"HK\$38\s+per month per device\s+with a commitment period of\s+24\s+months", text, flags=re.I)
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"NETVIGATOR official Freedome Network Safety Software tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": "38",
            "contract_months": "24",
            "local_voice": "Network safety software add-on for NETVIGATOR broadband customers; per-device monthly fee.",
            "add_on_charges_hkd": "Official Freedome tariff/list price; service availability and terms apply to NETVIGATOR broadband customers.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_hkt_norton_secure_vpn_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Norton Secure VPN" not in text or "HK$28 per month per device" not in text:
        return []
    match = re.search(r"HK\$28\s+per month per device\s+with a commitment period of\s+24\s+months", text, flags=re.I)
    if not match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"NETVIGATOR official Norton Secure VPN tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": "28",
            "contract_months": "24",
            "local_voice": "Norton Secure VPN add-on for NETVIGATOR broadband customers; per-device monthly fee.",
            "add_on_charges_hkd": "Official Norton Secure VPN tariff/list price; service availability and terms apply to NETVIGATOR broadband customers.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_hkt_eye2_communication_package_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "eye2 Communication Package" not in text:
        return []
    match = re.search(r"eye2\s+Communication\s+Package.*?\$?\s*(278|888)\s+or more per month", text, flags=re.I | re.S)
    if not match:
        return []
    fee = match.group(1)
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"HKT official eye2 Communication Package tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": fee,
            "local_voice": f"eye2 Telephone Line service and access to Information Services; official tariff states HK${fee} or more per month.",
            "add_on_charges_hkd": "Official eye2 tariff/list price lower bound; final fee varies by Information Service provider offers. SMS, video call, installation and usage charges are not written as main monthly fee.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [row]


def _parse_hkt_severe_weather_warning_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Severe Weather Warning Service" not in text or "Service charge" not in text:
        return []
    specs = [
        ("tropical cyclone warning announcement", "90"),
        ("thunderstorm warning announcement", "220"),
        ("flood warning announcement", "90"),
    ]
    rows: List[Dict[str, str]] = []
    for label, fee in specs:
        pattern = rf"{re.escape(label)}.*?Registration\s+HK\$\s*300\s+Service charge\s+HK\$\s*{fee}\s+per month"
        match = re.search(pattern, text, flags=re.I | re.S)
        if not match:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT official Severe Weather Warning Service - {label} HK${fee}",
                "monthly_fee_hkd": fee,
                "contract_months": "12",
                "local_voice": f"Telephone severe-weather warning announcement service: {label}.",
                "add_on_charges_hkd": "Official Severe Weather Warning tariff/list price; registration HK$300 is one-off and not written as main monthly fee.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return rows


def _parse_hkt_integrated_digital_access_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Integrated Digital Access Business Telephone Service" not in text or "Line Rental" not in text:
        return []
    specs = [
        ("IDA-P Line", "20000"),
        ("IDA-M Line", "20000"),
        ("Priority IDA line", "20000"),
    ]
    rows: List[Dict[str, str]] = []
    for label, fee in specs:
        match = re.search(rf"{re.escape(label)}\s+{int(fee):,}", text, flags=re.I)
        if not match:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Enterprise official IDA Service - {label} tariff {source.get('effective_from') or source.get('published_on', '')}",
                "monthly_fee_hkd": fee,
                "local_voice": f"Integrated Digital Access Business Telephone Service line rental: {label}.",
                "add_on_charges_hkd": "Official IDA tariff/list price; VAS monthly charges, local call charges, DDI rental and one-off/service-specific fees are retained in source notes and not written as the main line rental.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return rows


def _parse_hkt_customer_voice_hotline_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Customer Voice Hotline Management Service" not in text or "Monthly charge" not in text:
        return []
    specs = [
        (
            "Service port/user or telephone line",
            "2500",
            r"Monthly\s+charge:\s*HK\$?\s*2,500\s+on a per port/user and/or per telephone line basis",
            "Customer Voice Hotline Management Service monthly charge on a per port/user and/or per telephone line basis.",
            "Official customer voice hotline tariff/list price; setup charge HK$32,000 and service-feature charges are retained in source notes and not written as this monthly fee.",
        ),
        (
            "Call Queue per call queue",
            "2800",
            r"Rates for Call Queue:.*?Monthly\s+charge:\s*HK\$?\s*2,800\s*/\s*call queue",
            "Customer Voice Hotline Management Service Call Queue monthly charge per call queue.",
            "Official customer voice hotline Call Queue tariff/list price; setup/installation charge HK$2,000 and other service-feature charges are retained in source notes.",
        ),
        (
            "Call Queue per user",
            "1000",
            r"Rates for Call Queue:.*?HK\$?\s*1,000\s*/\s*user",
            "Customer Voice Hotline Management Service Call Queue monthly charge per user.",
            "Official customer voice hotline Call Queue tariff/list price; setup/installation charge HK$2,000 and other service-feature charges are retained in source notes.",
        ),
    ]
    rows: List[Dict[str, str]] = []
    for label, fee, pattern, local_voice, notes in specs:
        match = re.search(pattern, text, flags=re.I | re.S)
        if not match:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Enterprise official Customer Voice Hotline Management Service - {label} tariff {source.get('published_on', '')}",
                "monthly_fee_hkd": fee,
                "local_voice": local_voice,
                "add_on_charges_hkd": notes,
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return rows


def _parse_hkt_residential_cell_relay_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if not (
        re.search(r"Residential Cell Relay Services", text, flags=re.I)
        and re.search(r"CRS Customer Access", text, flags=re.I)
        and re.search(r"CRS Service Provider End", text, flags=re.I)
    ):
        return []
    specs = [
        ("CRS CA C Asymmetric 10Base-T single session", "1.5M", "150"),
        ("CRS CA C Asymmetric 10Base-T single session", "3M", "158"),
        ("CRS CA C Asymmetric 10Base-T single session", "6M", "166"),
        ("CRS CA C Asymmetric 10Base-T single session", "8M", "171"),
        ("CRS CA C Asymmetric G.DMT single session", "1.5M", "110"),
        ("CRS CA C Asymmetric G.DMT single session", "3M", "118"),
        ("CRS CA C Asymmetric G.DMT single session", "6M", "126"),
        ("CRS CA C Asymmetric G.DMT single session", "8M", "131"),
        ("CRS CA C Asymmetric 10Base-T up to 4 sessions", "3M", "171"),
        ("CRS CA C Asymmetric 10Base-T up to 4 sessions", "6M", "179"),
        ("CRS CA C Asymmetric 10Base-T up to 4 sessions", "8M", "184"),
        ("CRS CA C Asymmetric G.DMT up to 4 sessions", "3M", "128"),
        ("CRS CA C Asymmetric G.DMT up to 4 sessions", "6M", "136"),
        ("CRS CA C Asymmetric G.DMT up to 4 sessions", "8M", "141"),
        ("CRS Service Provider End CRS SP", "155M", "15000"),
        ("CRS Service Provider End CRS SP-FE", "100M", "34500"),
        ("CRS Service Provider End CRS SP-GE", "1000M", "78000"),
        ("CRS Service Provider End multiple domain name", "per server end", "2000"),
    ]
    rows: List[Dict[str, str]] = []
    for label, bandwidth, fee in specs:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT official Residential Cell Relay Services - {label} {bandwidth} tariff 2013-06-01",
                "monthly_fee_hkd": fee,
                "post_fup_speed_mbps": bandwidth.removesuffix("M") if bandwidth.endswith("M") else "",
                "local_voice": "Residential Cell Relay client-server broadband transport service for service providers.",
                "add_on_charges_hkd": "Official CRS tariff/list price. Installation, internal/external relocation and reconfiguration charges are one-off or change charges and excluded from the main monthly rental.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": text[:850],
            }
        )
        rows.append(row)
    return rows


def _parse_hkt_megalink_plus_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if not re.search(r"Megalink Plus", text, flags=re.I):
        return []
    if source.get("published_on") == "2008-11-28":
        specs = [
            *( ("Host End", f"{bandwidth}M", fee) for bandwidth, fee in [
                ("2", "5000"), ("3", "6500"), ("4", "8000"), ("6", "11000"), ("8", "14000"),
                ("10", "17000"), ("15", "20000"), ("20", "25000"), ("25", "30000"), ("30", "35000"),
                ("40", "40000"), ("50", "45000"), ("100", "55000"),
            ]),
            ("Client End", "1.5M/640K", "1500"),
            ("Client End", "6M/640K", "2000"),
            ("Client End", "2M/2M", "3000"),
        ]
    else:
        specs = [
            ("Host End", "1-10M", "22100"),
            ("Host End", "11-45M", "58500"),
            ("Host End", "50-500M", "123500"),
            ("Host End", "550-1000M", "162500"),
            ("Client End asymmetric", "up to 8M/640K", "5000"),
            ("Client End symmetric", "up to 10M/10M", "10000"),
        ]
    rows: List[Dict[str, str]] = []
    for endpoint, bandwidth, fee in specs:
        row = _base_row(source, result, captured_at)
        max_speed = re.search(r"(\d+(?:\.\d+)?)M", bandwidth)
        row.update(
            {
                "plan_name": f"HKT official MegaLink Plus {endpoint} {bandwidth} tariff 2013-05-14",
                "monthly_fee_hkd": fee,
                "post_fup_speed_mbps": max_speed.group(1) if max_speed else "",
                "local_voice": "MegaLink Plus commercial VPN service over shared IP infrastructure.",
                "add_on_charges_hkd": "Official MegaLink Plus tariff/list price. Setup, internal/external relocation and reconfiguration are not monthly rental and are excluded from the main fee.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": text[:850],
            }
        )
        rows.append(row)
    return rows


def _parse_hkt_ip_net_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if not (re.search(r"IP-Net", text, flags=re.I) and re.search(r"Monthly Rental", text, flags=re.I)):
        return []
    specs = [
        ("1-10M", "9100"), ("11-20M", "14300"), ("25-30M", "18200"), ("35-40M", "22100"),
        ("45-50M", "24700"), ("55-60M", "28600"), ("65-70M", "32500"), ("75-80M", "35100"),
        ("85-90M", "39000"), ("95-100M", "42900"), ("150-200M", "57200"), ("250-300M", "67600"),
        ("350-400M", "79300"), ("450-500M", "91000"), ("550-600M", "102700"), ("650-700M", "114400"),
        ("750-800M", "124800"), ("850-900M", "136500"), ("950-1000M", "148200"),
    ]
    rows: List[Dict[str, str]] = []
    for bandwidth, fee in specs:
        max_speed = re.search(r"(\d+)M$", bandwidth)
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT official IP-Net {bandwidth} tariff 2013-03-20",
                "monthly_fee_hkd": fee,
                "post_fup_speed_mbps": max_speed.group(1) if max_speed else "",
                "local_voice": "IP-Net single-port enterprise IP network service with guaranteed bandwidth per port.",
                "add_on_charges_hkd": "Official IP-Net tariff/list price. Setup, internal/external relocation and reconfiguration are excluded because they are not recurring monthly rental.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": text[:850],
            }
        )
        rows.append(row)
    return rows


def _parse_hkt_consumer_fixed_line_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Rates table" not in text or "Line rental" not in text or "298 / month" not in text:
        return []
    rows: List[Dict[str, str]] = []
    specs = [
        (
            "Residential line rental",
            "298",
            "Leasing of a line/channel enabling one simultaneous call for using a consumer fixed-line Service.",
            r"Line rental.*?298\s*/\s*month",
        ),
        (
            "Consumer fixed-line VAS per feature",
            "50",
            "Value-added service feature offered individually or in a package; HK$50/month per feature.",
            r"VAS.*?50\s*/\s*month\s*per feature",
        ),
    ]
    for label, fee, service_note, pattern in specs:
        match = re.search(pattern, text, re.I)
        if not match:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT official consumer fixed-line tariff {source.get('published_on', '')} - {label}",
                "monthly_fee_hkd": fee,
                "local_voice": service_note,
                "add_on_charges_hkd": (
                    "Official consumer fixed-line tariff/list price; fixed-line SMS, installation, relocation, "
                    "reconnection, cancellation, number porting and delivery charges are usage/one-off items and are not written as main monthly fee."
                ),
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_local_business_telephone_tariff_pdf(
    source: Dict[str, str], result: Dict[str, Any], captured_at: str
) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if (
        "Local Business Telephone Service" not in text
        or "Charge / mth" not in text
        or "Line Rental" not in text
    ):
        return []
    specs = [
        (
            "Business Telephone Line and Business Select Series",
            "300",
            r"Business Telephone Line and Business Select.*?Series\s+300\s*/\s*line",
        ),
        (
            "Business Faxline and Datel Series",
            "300",
            r"Business Faxline and Datel.*?Series\s+300\s*/\s*line",
        ),
        (
            "Business Hunting Line Series",
            "300",
            r"Business Hunting Line.*?Series\s+300\s*/\s*line",
        ),
        (
            "Business Citinet Line Series",
            "550",
            r"Business Citinet Line.*?Series\s+550\s*/\s*line",
        ),
        (
            "DDI Line Series",
            "550",
            r"Direct-Dialing-In.*?DDI.*?Line\s+Series\s*6?\s+550\s*/\s*line",
        ),
        (
            "Next Generation Business Fixed Line Series",
            "300",
            r"Next Generation Business Fixed Line.*?NGBFL.*?Series\s*7?\s+300\s*/\s*line",
        ),
        (
            "Value-added Services per feature",
            "40",
            r"Abbreviated dialing.*?40",
        ),
    ]
    rows: List[Dict[str, str]] = []
    for label, fee, pattern in specs:
        match = re.search(pattern, text, re.I)
        if not match:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Enterprise official Local Business Telephone tariff {source.get('published_on', '')} - {label}",
                "monthly_fee_hkd": fee,
                "local_voice": "Commercial fixed voice / facsimile / PABX line service under official Local Business Telephone tariff.",
                "add_on_charges_hkd": (
                    "Official Local Business Telephone tariff/list price; monthly_fee_hkd keeps line rental or high-level VAS recurring charge only. "
                    "Installation, relocation, reconnection and other one-off/service-change charges are excluded from monthly_fee_hkd."
                ),
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_csl_2g_3g_4g_mobile_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "2G" not in text or "3G" not in text or "4G" not in text:
        return []

    def section_between(start_pattern: str, end_patterns: Iterable[str]) -> str:
        match = re.search(start_pattern, text, flags=re.I)
        if not match:
            return ""
        start = match.start()
        end = len(text)
        for end_pattern in end_patterns:
            end_match = re.search(end_pattern, text[match.end() :], flags=re.I)
            if end_match:
                end = min(end, match.end() + end_match.start())
        return text[start:end]

    def data_values(block: str, count: int) -> List[str]:
        values = re.findall(r"Unlimited Wi-?Fi\s*(?:,|\+)\s*(\d+(?:\.\d+)?)\s*GB\+?", block, flags=re.I)
        if not values:
            values = re.findall(r"Local Data Usage\s+Unlimited Wi-?Fi,\s*(\d+(?:\.\d+)?)\s*GB", block, flags=re.I)
        if not values and re.search(r"100\s*MB", block, flags=re.I):
            values = ["0.1"] * count
        if not values:
            values = [""] * count
        if len(values) < count:
            values.extend([values[-1] if values else ""] * (count - len(values)))
        return values[:count]

    plan_sections = [
        (
            "SIM-only Plan",
            section_between(r"SIM[- ]only Plans?", [r"\b2\)\s*3G Multi", r"\b2\)\s*Handset", r"Handset/Tablet Plans?", r"\bThereafter Charges", r"\bEffective date"]),
        ),
        (
            "3G Multi Smart SIM Plan",
            section_between(r"3G Multi Smart SIM Plans?", [r"\b3\)\s*Handset", r"Handset/Tablet Plans?", r"\bThereafter Charges", r"\bEffective date"]),
        ),
        (
            "Handset/Tablet Plan",
            section_between(r"Handset/Tablet Plans?", [r"\bThereafter Charges", r"\b4\)\s*2G", r"\bEffective date"]),
        ),
    ]
    rows: List[Dict[str, str]] = []
    for plan_name, block in plan_sections:
        if not block:
            continue
        fees = re.findall(r"(?:HK\$|\$)\s*(\d[\d,]*)\s+per month", block, flags=re.I)
        if not fees:
            fees = re.findall(r"(?:HK\$|\$)\s*(\d[\d,]*)\s+per month per SIM", block, flags=re.I)
        local_data_values = data_values(block, len(fees))
        for idx, fee in enumerate(fees):
            local_data = local_data_values[idx] if idx < len(local_data_values) else ""
            row = _base_row(source, result, captured_at)
            plan_suffix = f"{plan_name} {local_data}GB" if local_data and local_data != "0.1" else plan_name
            row.update(
                {
                    "plan_name": f"csl official 2G/3G/4G mobile tariff {source.get('published_on', '')} - {plan_suffix}",
                    "monthly_fee_hkd": _float_text(fee),
                    "local_data_gb": local_data,
                    "local_voice": "Official tariff voice/data allowance as disclosed in source PDF.",
                    "add_on_charges_hkd": "Official csl 2G/3G/4G mobile tariff/list price; MTR/tunnel/mobile service licence and administration charges apply separately. Usage after allowance can include charges such as HK$40/100MB or HK$40/200MB depending on tariff version; PCCW-HKT Wi-Fi, SMS/MMS and Multi-SIM charges are not written as main monthly fee.",
                    "official_source_type": "official_public_tariff_pdf",
                    "extraction_status": "parsed_official_tariff_pdf",
                    "evidence_excerpt": _excerpt(text, 0, min(len(text), 1500)),
                }
            )
            rows.append(row)
    if rows:
        return _dedupe(rows)

    plan_specs = [
        ("SIM-only Plan", r"SIM-only Plans.*?HK\$(?P<fee>[\d,]+)\s+per month per SIM"),
        ("Handset/Tablet Plan", r"Handset/Tablet Plans.*?HK\$(?P<fee>[\d,]+)\s+per month per SIM"),
    ]
    for plan_name, pattern in plan_specs:
        match = re.search(pattern, text, flags=re.I)
        if not match:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"csl official 2G/3G/4G mobile tariff {source.get('published_on', '')} - {plan_name}",
                "monthly_fee_hkd": _float_text(match.group("fee")),
                "local_data_gb": "0.1",
                "local_voice": "1,000 local voice minutes; 100MB mobile data; unlimited csl Wi-Fi where covered by tariff.",
                "add_on_charges_hkd": "Official csl 2G/3G/4G mobile tariff/list price; MTR/tunnel/mobile service licence and administration charges apply separately. Usage after allowance includes HK$40/100MB mobile data, HK$1/min local voice and chargeable SMS/MMS items.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, 0, min(len(text), 1500)),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_csl_smart_pama_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Smart Pama Service Plan" not in text or "Monthly Local Data" not in text:
        return []
    fee_match = re.search(r"Monthly Fee\s*\$?(?P<fee>\d+)", text, flags=re.I)
    data_match = re.search(r"Monthly Local Data\s*(?P<data>\d+)\s*GB", text, flags=re.I)
    if not fee_match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"csl official Smart Pama senior postpaid tariff {source.get('published_on', '')}",
            "monthly_fee_hkd": _float_text(fee_match.group("fee")),
            "local_data_gb": data_match.group("data") if data_match else "6",
            "local_voice": "Unlimited monthly local airtime; local intra-network SMS and local MMS allowances disclosed in tariff.",
            "add_on_charges_hkd": "Official Smart Pama postpaid service plan for Hong Kong residents aged 65 or above; monthly fee discount and HK$18 administration fee are disclosed separately; local data thereafter is chargeable under the tariff.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, 0, min(len(text), 1400)),
        }
    )
    return _dedupe([row])


def _parse_csl_postpaid_service_plan_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "Postpaid Service Plan" not in text or "Monthly Fee" not in text:
        return []
    if "U-plan" in text:
        plan_name = "csl official U-plan student postpaid tariff"
        category_note = "U-plan student postpaid service plan."
    elif "The Club SIM" in text or "Club SIM" in text:
        plan_name = "csl official The Club SIM postpaid tariff"
        category_note = "The Club SIM postpaid service plan."
    else:
        return []
    availability_note = ""
    if re.search(r"no longer available for subscription", text, re.I):
        availability_note = " Source states this tariff is no longer available for subscription."
    elif re.search(r"until further notice", text, re.I):
        availability_note = " Source states the effective period continues until further notice."
    fee_match = re.search(r"Monthly Fee\s*\$?(?P<fee>\d+)", text, flags=re.I)
    data_match = re.search(r"(?:Monthly\s+)?Local Data\s*(?P<data>\d+)\s*(?P<unit>GB|MB)", text, flags=re.I)
    voice_match = re.search(
        r"(?:Monthly\s+)?Local Airtime(?:\s*\([^)]*\))?\s*(?P<voice>Unlimited|[\d,]+\s*minutes?)",
        text,
        flags=re.I,
    )
    if not fee_match:
        return []
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"{plan_name} {source.get('published_on', '')}",
            "monthly_fee_hkd": _float_text(fee_match.group("fee")),
            "local_data_gb": (
                data_match.group("data")
                if data_match and data_match.group("unit").upper() == "GB"
                else str(round(float(data_match.group("data")) / 1000, 4)).rstrip("0").rstrip(".")
                if data_match
                else ""
            ),
            "local_voice": f"{voice_match.group('voice')} monthly local airtime" if voice_match else "",
            "add_on_charges_hkd": f"Official csl postpaid service plan tariff/list price. {category_note}{availability_note} Monthly fee discount, administration fee, local data thereafter charges and eligibility conditions are disclosed separately in the source PDF and are not treated as the main plan monthly fee.",
            "official_source_type": "official_public_tariff_pdf",
            "extraction_status": "parsed_official_tariff_pdf",
            "evidence_excerpt": _excerpt(text, 0, min(len(text), 1400)),
        }
    )
    return _dedupe([row])


def _parse_csl_1010_postpaid_mobile_tariff_pdf(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if (
        "Three Postpaid Local Data Service Plans" not in text
        or "Greater China Data Service Plans" not in text
        or "Monthly Fee" not in text
    ):
        return []
    specs = [
        {
            "brand": "csl",
            "category": "official_tariff_local_data_mobile_plan",
            "plan": "csl official Postpaid Local Data Service Plan 5GB",
            "fee": "138",
            "data": "5",
            "voice": "Unlimited",
            "note": "Official csl local data postpaid tariff/list price.",
        },
        {
            "brand": "csl",
            "category": "official_tariff_local_data_mobile_plan",
            "plan": "csl official Postpaid Local Data Service Plan 8GB",
            "fee": "198",
            "data": "8",
            "voice": "Unlimited",
            "note": "Official csl local data postpaid tariff/list price.",
        },
        {
            "brand": "csl",
            "category": "official_tariff_local_data_mobile_plan",
            "plan": "csl official Postpaid Local Data Service Plan 6GB capacity data package",
            "fee": "220",
            "data": "6",
            "voice": "Unlimited",
            "note": "Official csl local data postpaid tariff/list price.",
        },
        {
            "brand": "1O1O",
            "category": "official_tariff_greater_china_mobile_plan",
            "plan": "1O1O official Greater China Service Plan 6GB",
            "fee": "677",
            "data": "6",
            "roaming": "6",
            "voice": "Unlimited",
            "note": "Official 1O1O Greater China tariff/list price.",
        },
        {
            "brand": "1O1O",
            "category": "official_tariff_greater_china_mobile_plan",
            "plan": "1O1O official Greater China Service Plan 10GB",
            "fee": "877",
            "data": "10",
            "roaming": "10",
            "voice": "Unlimited",
            "note": "Official 1O1O Greater China tariff/list price.",
        },
        {
            "brand": "csl",
            "category": "official_tariff_greater_china_mobile_plan",
            "plan": "csl official Greater China Service Plan SIM-only 6GB",
            "fee": "418",
            "data": "6",
            "roaming": "6",
            "voice": "Unlimited",
            "note": "Official csl Greater China tariff/list price.",
        },
        {
            "brand": "csl",
            "category": "official_tariff_greater_china_mobile_plan",
            "plan": "csl official Greater China Service Plan SIM-only 10GB",
            "fee": "618",
            "data": "10",
            "roaming": "10",
            "voice": "Unlimited",
            "note": "Official csl Greater China tariff/list price.",
        },
        {
            "brand": "csl",
            "category": "official_tariff_greater_china_mobile_plan",
            "plan": "csl official Greater China Service Plan Handset/Tablet 6GB",
            "fee": "618",
            "data": "6",
            "roaming": "6",
            "voice": "Unlimited",
            "note": "Official csl Greater China tariff/list price.",
        },
        {
            "brand": "csl",
            "category": "official_tariff_greater_china_mobile_plan",
            "plan": "csl official Greater China Service Plan Handset/Tablet 10GB",
            "fee": "818",
            "data": "10",
            "roaming": "10",
            "voice": "Unlimited",
            "note": "Official csl Greater China tariff/list price.",
        },
    ]
    rows: List[Dict[str, str]] = []
    for spec in specs:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "brand": spec["brand"],
                "product_category": spec["category"],
                "plan_name": f"{spec['plan']} {source.get('published_on', '')}",
                "monthly_fee_hkd": spec["fee"],
                "local_data_gb": spec["data"],
                "roaming_data_gb": spec.get("roaming", ""),
                "local_voice": spec["voice"],
                "add_on_charges_hkd": f"{spec['note']} Monthly fee discount, MTR/tunnel/mobile licence/administration fees, IDD deposit and usage-after-allowance charges are disclosed separately in the source PDF and are not treated as main plan monthly fee.",
                "official_source_type": "official_public_tariff_pdf",
                "extraction_status": "parsed_official_tariff_pdf",
                "evidence_excerpt": _excerpt(text, 0, min(len(text), 1800)),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_sme_5g_business_mobile(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    plan_specs = [
        {
            "label": "5G Basic Plan",
            "anchor": "5G Basic Plan",
            "fee": "128",
            "local_data_gb": "30",
            "roaming_data_gb": "3",
            "speed": "1",
            "voice": "Unlimited local voice call minutes",
        },
        {
            "label": "Chinese Mainland Roaming Data Plan",
            "anchor": "Chinese Mainland Roaming Data Plan",
            "fee": "199",
            "local_data_gb": "30",
            "roaming_data_gb": "25",
            "speed": "",
            "voice": "Unlimited local voice call minutes",
        },
        {
            "label": "Samsung Smartphones Subscription Plan",
            "anchor": "Samsung Smartphones Subscription Plan",
            "fee": "104",
            "local_data_gb": "1",
            "roaming_data_gb": "",
            "speed": "0.128",
            "voice": "2,000mins local voice call",
        },
    ]
    rows: List[Dict[str, str]] = []
    for spec in plan_specs:
        start = text.find(spec["anchor"])
        if start < 0:
            continue
        end_candidates = [text.find(next_spec["anchor"], start + len(spec["anchor"])) for next_spec in plan_specs]
        end_candidates = [end for end in end_candidates if end > start]
        end = min(end_candidates) if end_candidates else min(len(text), start + 1500)
        block = text[start:end]
        if not re.search(rf"(?:Starting from\s*)?\$?{re.escape(spec['fee'])}/month", block, re.I):
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT SME {spec['label']} HK${spec['fee']}",
                "monthly_fee_hkd": spec["fee"],
                "local_data_gb": spec["local_data_gb"],
                "roaming_data_gb": spec["roaming_data_gb"],
                "post_fup_speed_mbps": spec["speed"],
                "contract_months": "",
                "local_voice": spec["voice"],
                "add_on_charges_hkd": "Waived HK$18/month tunnel fee; official HKT SME public page.",
                "evidence_excerpt": _excerpt(text, start, end),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_sme_business_broadband(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    specs = [
        {
            "label": "Ultra-fast Business Broadband",
            "category": "business_broadband_ftto",
            "anchor": "Ultra-fast Business Broadband A reliable network for your business",
            "end_anchor": "5G Business Broadband Flexible to meet your business needs",
            "post_fup_speed_mbps": "100000",
            "add_on": "Official HKT SME public page states network ranging from 300M to 100G; starting price only.",
        },
        {
            "label": "5G Business Broadband",
            "category": "business_broadband_5g",
            "anchor": "5G Business Broadband Flexible to meet your business needs",
            "end_anchor": "Macbook for SMEs",
            "post_fup_speed_mbps": "",
            "add_on": "Official HKT SME public page; 5G service is subject to designated locations and devices; starting price only.",
        },
    ]
    rows: List[Dict[str, str]] = []
    for spec in specs:
        start = text.find(spec["anchor"])
        if start < 0:
            continue
        end = text.find(spec["end_anchor"], start + len(spec["anchor"]))
        if end < 0:
            end = min(len(text), start + 1800)
        block = text[start:end]
        if not re.search(r"Starting from\s+HK\$198\s*/\s*Month", block, re.I):
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "product_category": spec["category"],
                "plan_name": f"HKT SME {spec['label']} starting from HK$198",
                "monthly_fee_hkd": "198",
                "post_fup_speed_mbps": spec["post_fup_speed_mbps"],
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": spec["add_on"],
                "evidence_excerpt": _excerpt(text, start, end),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_enterprise_local_business_telephone(
    source: Dict[str, str], result: Dict[str, Any], captured_at: str
) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    specs = [
        (
            "Business Telephone line",
            "189.80",
            [
                "Business Telephone line Monthly Charge $189.80/line",
                "商業電話線 - 月費 $189.80",
                "標準價目： 安裝費 $600 月費 $189.80",
            ],
        ),
        (
            "Business Select line",
            "237",
            ["Business Select line Monthly Charge $237/line", "商業智選 - 月費 $237"],
        ),
        (
            "Tone DDI line",
            "281",
            [
                "Tone DDI line Installation $600/line Monthly Charge $281/line",
                "音頻直通內線 - 安裝費 $600（每條） - 月費 $281",
            ],
        ),
        (
            "Tone DDI Diversity with Adjacent Exchange Diversity",
            "379",
            ["DDI Diversity (with Adjacent Exchange Diversity) $379/line", "直通內線分途（與附近機樓分途） $379"],
        ),
        (
            "Pulse DDI line",
            "344",
            [
                "Pulse DDI line Installation $600/line Monthly Charge $344/line",
                "脈衝直通內線 - 安裝費 $600（每條） - 月費 $344",
            ],
        ),
        (
            "Pulse DDI Diversity with Adjacent Exchange Diversity",
            "442",
            [
                "DDI Diversity (with Adjacent Exchange Diversity) $442/line Business Telephone line",
                "直通內線分途（與附近機樓分途） $442",
            ],
        ),
        (
            "Datel",
            "198.80",
            ["Monthly Charge: $198.80 (including telephone line charge)", "價目：月費（包括租用電話線）$198.80"],
        ),
        (
            "IDA T1",
            "3950",
            [
                "IDA (T1) Installation $3,600/line Monthly rental $3,950/line",
                "IDA (T1) - 安裝費 $3,600（每條） - 月費 $3,950",
                "IDA (T1) - 安裝費 $3,600",
            ],
        ),
        ("Hunting Line", "209", ["Hunting Line Monthly Charge $209", "自動跳駁線月費 $209"]),
        ("Premium Hunting", "237", ["Premium Hunting Monthly Charge $237", "商務通1月費 $237"]),
        (
            "Basic Fax",
            "198.8",
            [
                "Basic Fax - Installation $600/line - Monthly Charge $198.8/line",
                "基本傳真 - 安裝費 $600（每條） - 月費 $198.8",
                "商用傳真 - 安裝費 $600（每條） - 月費 $198.8",
            ],
        ),
        (
            "Faxline 100",
            "229.8",
            [
                "Faxline 100 - Installation $600/line - Monthly Charge $229.8/line",
                "商用傳真100 - 安裝費 $600（每條） - 月費 $229.8",
            ],
        ),
        (
            "Faxline 3",
            "237.8",
            ["Faxline 3 - Installation $600/line - Monthly Charge $237.8/line", "商用傳真3 - 安裝費 $600（每條） - 月費 $237.8"],
        ),
        (
            "Faxline 100 Hunting",
            "249",
            [
                "Faxline 100 - Hunting - Installation $600/line - Monthly Charge $249/line",
                "Faxline 100 - Hunting - 安裝費 $600（每條） - 月費 $249",
            ],
        ),
        (
            "Citinet",
            "213",
            ["Citinet - Installation charge $600 - Monthly Charge $213", "城訊通 - 安裝費 $600 - 月費 $213"],
        ),
        (
            "Caller Display phone",
            "35",
            [
                "Caller Display phone monthly charge - PANASONIC KX-TSC11MXW $35/unit",
                "來電顯示電話月租 - PANASONIC KX-TSC11MXW電話 $35",
            ],
        ),
        (
            "Call Management Pack",
            "48",
            ["Caller Display + PhoneMail for only $48/month", "每月僅需$48"],
        ),
    ]
    rows: List[Dict[str, str]] = []
    for label, fee, evidence_options in specs:
        start = -1
        for evidence in evidence_options:
            anchor = evidence[: min(len(evidence), 70)]
            start = text.find(anchor)
            if start >= 0:
                break
        if start < 0 and source["source_id"].endswith("_tc"):
            continue
        if start < 0:
            start = text.find(label)
        if start < 0:
            continue
        end = min(len(text), start + 520)
        if fee not in text[start:end]:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Enterprise Local Business Telephone {label} HK${fee}",
                "monthly_fee_hkd": fee,
                "local_data_gb": "",
                "roaming_data_gb": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "Business fixed voice / telephone service",
                "add_on_charges_hkd": "Installation and service change charges are disclosed separately on the official page; monthly_fee_hkd keeps recurring charge only.",
                "evidence_excerpt": _excerpt(text, start, end),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_hkt_homephone_value_added_services(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if "HKT Home Phone Service" not in text or "Calling Features" not in text or "Service Fee per Month" not in text:
        return []
    specs = [
        ("Deluxe Package", "27", "Deluxe Package"),
        ("Abbreviated Dialing", "16", "Abbreviated Dialing"),
        ("Appointment Service", "16", "Appointment Service"),
        ("Block-the-Blocker", "8", "Block-the-Blocker"),
        ("Call Forwarding", "20", "Call Forwarding"),
        ("Conference Calling", "17", "Conference Calling"),
        ("Music on Hold", "3", "Music on Hold"),
        ("OneCall", "38", "OneCall"),
        ("PhoneMail", "25", "PhoneMail"),
        ("Smart Care Voice Reminder Service", "48", "Smart Care Voice Reminder Service"),
    ]
    rows: List[Dict[str, str]] = []
    for service_name, fee, evidence_name in specs:
        pattern = rf"{re.escape(evidence_name)}.*?(?:HK\$|\$)\s*{re.escape(fee)}(?:[\s#(]|$)"
        match = re.search(pattern, text, flags=re.I | re.S)
        if not match:
            continue
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"HKT Home Phone {service_name} HK${fee}",
                "monthly_fee_hkd": fee,
                "local_data_gb": "",
                "roaming_data_gb": "",
                "contract_months": "",
                "local_voice": "HKT Home Phone value-added calling feature",
                "add_on_charges_hkd": "Fees are charged in addition to the standard monthly service fee for HKT Telephone Line Service; page states fees are subject to change.",
                "official_source_type": "official_public_product_page",
                "extraction_status": "parsed",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180),
            }
        )
        rows.append(row)
    return _dedupe(rows)


def _parse_generic(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    rows: List[Dict[str, str]] = []
    price_pattern = re.compile(r"(?:HK\$|\$)\s*(?P<fee>\d{2,4}).{0,700}", re.I)
    for match in price_pattern.finditer(text):
        snippet = match.group(0)
        if not re.search(r"\b(?:5G|mobile|data|plan|service|GB|monthly)\b", snippet, re.I):
            continue
        row = _base_row(source, result, captured_at)
        fee = _float_text(match.group("fee"))
        data_match = re.search(r"(\d+(?:\.\d+)?)\s*GB", snippet, re.I)
        contract_match = re.search(r"\b(24|30|36)\s*months?\b", snippet, re.I)
        row.update(
            {
                "plan_name": f"{source['brand']} public product mention HK${fee}",
                "monthly_fee_hkd": fee,
                "local_data_gb": _float_text(data_match.group(1) if data_match else ""),
                "contract_months": _float_text(contract_match.group(1) if contract_match else ""),
                "extraction_status": "review_required",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(row)
    return _dedupe(rows[:12])


def _historical_row(row: Dict[str, str], snapshot: Dict[str, str]) -> Dict[str, str]:
    output = dict(row)
    timestamp = snapshot["timestamp"]
    output.update(
        {
            "snapshot_id": snapshot["snapshot_id"],
            "archive_timestamp": timestamp,
            "archive_year": timestamp[:4],
            "archive_month": timestamp[4:6],
            "archive_url": snapshot.get("archive_url") or f"https://web.archive.org/web/{timestamp}id_/{snapshot['original_url']}",
            "source_kind": snapshot.get("source_kind") or "wayback_official_page_snapshot",
        }
    )
    output["record_key"] = hashlib.sha1(
        "|".join(
            [
                output.get("brand", ""),
                output.get("product_category", ""),
                output.get("plan_name", ""),
                output.get("monthly_fee_hkd", ""),
                output.get("published_price_hkd", ""),
                output.get("price_billing_unit", ""),
                output.get("local_data_gb", ""),
                output.get("roaming_data_gb", ""),
                output.get("archive_year", ""),
                output.get("snapshot_id", ""),
            ]
        ).encode("utf-8")
    ).hexdigest()[:16]
    return output


def _parse_csl_old_data_voice(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    rows: List[Dict[str, str]] = []
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": snapshot.get("source_kind") or "wayback_official_page_snapshot",
    }

    def data_value(raw: str) -> str:
        values = [float(value) for value in re.findall(r"\d+(?:\.\d+)?", raw)]
        if not values:
            return ""
        total = sum(values)
        return str(int(total)) if total.is_integer() else str(total)

    def add_row(plan_name: str, fee: str, local_data: str, excerpt_start: int, excerpt_end: int, roaming: str = "") -> None:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"{plan_name} HK${fee}",
                "monthly_fee_hkd": fee,
                "local_data_gb": local_data,
                "roaming_data_gb": roaming,
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "Unlimited voice call mins",
                "extraction_status": "parsed_historical_archive",
                "evidence_excerpt": _excerpt(text, excerpt_start, excerpt_end),
            }
        )
        rows.append(_historical_row(row, snapshot))

    ultra_block = re.search(
        r"(?P<family>csl Ultra \d+)\*?\s+Service Plan\s+"
        r"Handset Plan Monthly Fee(?:\s*\([^)]*\))*\s*(?P<handset>(?:\$\s*\d{2,4}\s*){2,}).*?"
        r"SIM Only Plan Monthly Fee(?:\s*\([^)]*\))*\s*(?P<sim>(?:\$\s*\d{2,4}\s*){2,}).*?"
        r"Local Data usage(?:\s*\([^)]*\))*\s*(?P<data>(?:\d+(?:\.\d+)?GB\s*){2,})",
        text,
        re.I,
    )
    if ultra_block:
        family = _clean_text(ultra_block.group("family"))
        handset_fees = re.findall(r"\d{2,4}", ultra_block.group("handset"))
        sim_fees = re.findall(r"\d{2,4}", ultra_block.group("sim"))
        data_cells = re.findall(r"\d+(?:\.\d+)?GB", ultra_block.group("data"), re.I)
        for index, fee in enumerate(handset_fees):
            if index < len(data_cells):
                add_row(
                    f"{family} historical handset plan",
                    fee,
                    data_value(data_cells[index]),
                    ultra_block.start(),
                    ultra_block.end(),
                )
        for index, fee in enumerate(sim_fees):
            if index < len(data_cells):
                add_row(
                    f"{family} historical SIM only plan",
                    fee,
                    data_value(data_cells[index]),
                    ultra_block.start(),
                    ultra_block.end(),
                )
        if rows:
            return rows

    fee_block = re.search(
        r"Monthly SIM [–-]?only plan fee.*?(?P<sim>(?:HK\$|\$)\s*\d+(?:\s+(?:HK\$|\$)\s*\d+){1,}).*?"
        r"Monthly handset plan fee.*?(?P<handset>(?:-|(?:HK\$|\$)\s*\d+)(?:\s+(?:-|(?:HK\$|\$)\s*\d+)){1,}).*?"
        r"Local data usage.*?(?P<data>(?:\d+(?:\.\d+)?GB(?:\+\d+(?:\.\d+)?GB)?(?:\s+\([^)]+\))?(?:\s+4G\s+21Mbps\s+\+\s+Unlimited Data.*?)?\s*){2,})"
        r"(?:China roaming|Mainland China|Voice call|SMS|Local Mobile data Top-up Plan)",
        text,
        re.I,
    )
    if fee_block:
        sim_fees = re.findall(r"\d+", fee_block.group("sim"))
        handset_fees = re.findall(r"\d+|-", fee_block.group("handset"))
        data_cells = re.findall(r"\d+(?:\.\d+)?GB(?:\+\d+(?:\.\d+)?GB)?", fee_block.group("data"), re.I)
        for index, fee in enumerate(sim_fees):
            if index < len(data_cells):
                add_row("csl historical SIM only plan", fee, data_value(data_cells[index]), fee_block.start(), fee_block.end())
        for index, fee in enumerate(handset_fees):
            if fee != "-" and index < len(data_cells):
                add_row("csl historical handset plan", fee, data_value(data_cells[index]), fee_block.start(), fee_block.end())
        if rows:
            return rows

    segment_pattern = re.compile(
        r"Monthly SIM only plan fee.*?\$(?P<sim>\d{2,4}).*?"
        r"Monthly handset plan fee.*?\$(?P<handset>\d{2,4}).*?"
        r"Local data usage.*?(?P<data>\d+(?:\.\d+)?GB(?:\+\d+(?:\.\d+)?GB)?).*?"
        r"Voice call mins.*?Unlimited(?P<rest>.*?)(?=Monthly SIM only plan fee|Other services|Things to know before you buy|$)",
        re.I,
    )
    for match in segment_pattern.finditer(text):
        rest = match.group("rest")
        roaming_match = re.search(r"(?:Mainland China\s*&\s*Macau|China) roaming data usage\s+(?P<roaming>\d+(?:\.\d+)?)GB", rest, re.I)
        local_data = data_value(match.group("data"))
        roaming = _float_text(roaming_match.group("roaming") if roaming_match else "")
        add_row("csl historical SIM only plan", _float_text(match.group("sim")), local_data, match.start(), match.end(), roaming)
        add_row("csl historical handset plan", _float_text(match.group("handset")), local_data, match.start(), match.end(), roaming)
    return rows


def _parse_pccw_mobile_3g_tariff(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }
    block = re.search(
        r"(?:One 3G tariff plan meets all your needs|一個3G月費\s*盡享一切).*?"
        r"(?P<table>\$138\s+600.*?\$198\s+2,000.*?\$298\s+4,000.*?\$498\s+10,000.*?"
        r"(?:VAS included|附送增值服務))",
        text,
        re.I,
    )
    if not block:
        return []
    rows: List[Dict[str, str]] = []
    specs = [
        ("138", "600", "60", "60"),
        ("198", "2000", "200", "200"),
        ("298", "4000", "400", "400"),
        ("498", "10000", "1000", "1000"),
    ]
    for fee, voice_minutes, mms_count, sms_count in specs:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"PCCW mobile historical 3G monthly tariff HK${fee}",
                "monthly_fee_hkd": fee,
                "local_data_gb": "",
                "roaming_data_gb": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": f"{voice_minutes} voice call mins; {mms_count} intra-network MMS; {sms_count} intra-network SMS; unlimited PCCW Channel local browsing",
                "add_on_charges_hkd": "Thereafter/local inter-network charges and monthly mobile licence/related fee are excluded from monthly_fee_hkd.",
                "extraction_status": "parsed_historical_archive",
                "evidence_excerpt": _excerpt(text, block.start(), block.end()),
            }
        )
        rows.append(_historical_row(row, snapshot))
    return _dedupe(rows)


def _parse_pccw_mobile_other_3g_tariff(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if not re.search(r"(?:Other 3G Tariff Plan|其他3G月費計劃)", text, re.I):
        return []
    block = re.search(
        r"(?:For only \$98 per month|月費\$98).*?"
        r"(?P<table>(?:For Fun Seekers|啱傾啱玩型).*?(?:VAS included|附送增值服務))",
        text,
        re.I,
    )
    if not block:
        return []
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }
    rows: List[Dict[str, str]] = []
    specs = [
        (
            "PCCW mobile historical Other 3G tariff - For Fun Seekers",
            "3000 voice-call mins; normal 2000 mins and intra-network 1000 mins; unlimited intra-network SMS",
            "",
            "Free now-TV 300 mins or one selected VAS; Free Zone included. Monthly mobile licence/admin fee excluded from monthly_fee_hkd.",
        ),
        (
            "PCCW mobile historical Other 3G tariff - For Web Surfers",
            "1600 voice-call mins; normal 1000 mins and intra-network 600 mins; unlimited intra-network SMS",
            "0.1",
            "Includes 100MB mobile data/mobileWeb and unlimited Wi-Fi; optional VAS at up to half price. Monthly mobile licence/admin fee excluded from monthly_fee_hkd.",
        ),
    ]
    for plan_name, local_voice, local_data_gb, add_on in specs:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": plan_name,
                "monthly_fee_hkd": "98",
                "local_data_gb": local_data_gb,
                "roaming_data_gb": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": local_voice,
                "add_on_charges_hkd": add_on,
                "extraction_status": "parsed_historical_archive",
                "evidence_excerpt": _excerpt(text, block.start(), block.end()),
            }
        )
        rows.append(_historical_row(row, snapshot))
    return _dedupe(rows)


def _parse_pccw_mobile_web_talk_tariff(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if not re.search(r"(?:Web\s*&\s*Talk\s*月費計劃|Web\s*&\s*Talk)", text, re.I):
        return []
    block = re.search(
        r"(?:Web\s*&\s*Talk\s*月費計劃|Web\s*&\s*Talk).*?"
        r"(?:客戶須繳付每月|每月流動電話服務牌照|列印|返回頁首)",
        text,
        re.I,
    )
    if not block:
        return []

    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }
    year = snapshot["timestamp"][:4]
    if year == "2010":
        specs = [
            {"fee": "98", "data": "0.1", "voice": "600 normal + 600 intra-network voice-call mins; MNP +400 mins where applicable", "video": "60", "mms": "60", "promo": ""},
            {"fee": "149", "data": "0.1", "voice": "1200 normal voice-call mins", "video": "60", "mms": "60", "promo": ""},
            {"fee": "199", "data": "0.2", "voice": "1600 normal voice-call mins", "video": "100", "mms": "100", "promo": ""},
            {"fee": "299", "data": "0.5", "voice": "1800 normal voice-call mins", "video": "200", "mms": "200", "promo": ""},
            {"fee": "399", "data": "", "voice": "3000 normal voice-call mins; unlimited data shown in source", "video": "300", "mms": "300", "promo": ""},
            {"fee": "478", "data": "", "voice": "4000 normal + 4000 intra-network voice-call mins; MNP +400 mins where applicable; unlimited data shown in source", "video": "400", "mms": "400", "promo": ""},
        ]
    elif year == "2011":
        specs = [
            {"fee": "149", "data": "0.1", "voice": "1200 voice-call mins", "video": "60", "mms": "60", "promo": "promotional monthly fee HK$99 with designated handset"},
            {"fee": "199", "data": "0.2", "voice": "1600 voice-call mins", "video": "100", "mms": "100", "promo": "promotional monthly fee HK$116 with designated handset"},
            {"fee": "299", "data": "0.5", "voice": "1800 voice-call mins", "video": "200", "mms": "200", "promo": "promotional monthly fee HK$160 with designated handset"},
            {"fee": "399", "data": "", "voice": "3000 voice-call mins; unlimited data shown in source", "video": "300", "mms": "300", "promo": "promotional monthly fee HK$232 with designated handset"},
            {"fee": "478", "data": "", "voice": "8000 voice-call mins (4000 normal + 4000 intra-network); MNP +400 mins where applicable; unlimited data shown in source", "video": "400", "mms": "400", "promo": "promotional monthly fee HK$253.5 with designated handset"},
            {"fee": "499", "data": "", "voice": "4000 voice-call mins; China-HK one-number service 200 mins/offer noted in source; unlimited data shown in source", "video": "400", "mms": "400", "promo": "promotional monthly fee HK$275 with designated handset"},
            {"fee": "98", "data": "0.1", "voice": "1200 voice-call mins (600 normal + 600 intra-network); MNP +400 mins where applicable", "video": "60", "mms": "60", "promo": "SIM-only / bring-your-own handset plan"},
            {"fee": "198", "data": "", "voice": "1600 voice-call mins (1000 normal + 600 intra-network); MNP +400 mins where applicable; unlimited data shown in source", "video": "60", "mms": "60", "promo": "SIM-only / bring-your-own handset plan; data tethering excluded by source"},
        ]
    else:
        specs = [
            {"fee": "149", "data": "0.1", "voice": "1200 voice-call mins", "video": "60", "mms": "60", "promo": "promotional monthly fee HK$99 with designated handset"},
            {"fee": "199", "data": "0.2", "voice": "1600 voice-call mins", "video": "100", "mms": "100", "promo": "promotional monthly fee HK$116 with designated handset"},
            {"fee": "299", "data": "0.5", "voice": "1800 voice-call mins", "video": "200", "mms": "200", "promo": "promotional monthly fee HK$160 with designated handset"},
            {"fee": "339", "data": "5", "voice": "1800 voice-call mins", "video": "200", "mms": "200", "promo": "promotional monthly fee HK$199 with designated handset"},
            {"fee": "399", "data": "", "voice": "3000 voice-call mins; unlimited data shown in source", "video": "300", "mms": "300", "promo": "promotional monthly fee HK$232 with designated handset"},
            {"fee": "478", "data": "", "voice": "8000 voice-call mins (4000 normal + 4000 intra-network); MNP +400 mins where applicable; unlimited data shown in source", "video": "400", "mms": "400", "promo": "promotional monthly fee HK$253.5 with designated handset"},
            {"fee": "499", "data": "", "voice": "4000 voice-call mins; China-HK one-number service 200 mins/offer noted in source; unlimited data shown in source", "video": "400", "mms": "400", "promo": "promotional monthly fee HK$275 with designated handset"},
            {"fee": "98", "data": "0.1", "voice": "1200 voice-call mins (600 normal + 600 intra-network); MNP +400 mins where applicable", "video": "60", "mms": "60", "promo": "SIM-only / bring-your-own handset plan"},
            {"fee": "198", "data": "", "voice": "1200 voice-call mins (600 normal + 600 intra-network); MNP +400 mins where applicable; unlimited data shown in source", "video": "60", "mms": "60", "promo": "SIM-only / bring-your-own handset plan; data tethering excluded by source"},
        ]

    rows: List[Dict[str, str]] = []
    for spec in specs:
        row = _base_row(source, result, captured_at)
        add_on_parts = [
            "Includes PCCW Wi-Fi / Auto Connect where shown by source.",
            "Intra-network SMS shown as unlimited where disclosed.",
            f"Video call mins {spec['video']} and intra-network MMS {spec['mms']} per month.",
            "Monthly mobile licence/MTR/tunnel/admin fee HK$12 is excluded from monthly_fee_hkd.",
            "Unlimited data rows keep local_data_gb blank because the numeric structured field only stores GB allowances; fair-use and thereafter charges are retained in notes.",
        ]
        if spec["promo"]:
            add_on_parts.insert(0, spec["promo"])
        row.update(
            {
                "plan_name": f"PCCW mobile historical Web & Talk monthly tariff HK${spec['fee']}",
                "monthly_fee_hkd": spec["fee"],
                "local_data_gb": spec["data"],
                "roaming_data_gb": "",
                "post_fup_speed_mbps": "",
                "contract_months": "12" if year == "2010" else "",
                "local_voice": spec["voice"],
                "add_on_charges_hkd": " ".join(add_on_parts),
                "extraction_status": "parsed_historical_archive",
                "evidence_excerpt": _excerpt(text, block.start(), block.end()),
            }
        )
        rows.append(_historical_row(row, snapshot))
    return _dedupe(rows)


def _pccw_snapshot_source(snapshot: Dict[str, str]) -> Dict[str, str]:
    return {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }


def _pccw_static_rows(
    snapshot: Dict[str, str],
    result: Dict[str, Any],
    captured_at: str,
    block: re.Match[str],
    specs: List[Dict[str, str]],
    default_note: str,
) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    source = _pccw_snapshot_source(snapshot)
    rows: List[Dict[str, str]] = []
    for spec in specs:
        row = _base_row(source, result, captured_at)
        notes = [spec.get("note", ""), default_note]
        row.update(
            {
                "plan_name": spec["plan_name"],
                "monthly_fee_hkd": spec["fee"],
                "local_data_gb": spec.get("data", ""),
                "roaming_data_gb": "",
                "post_fup_speed_mbps": "",
                "contract_months": spec.get("contract", ""),
                "local_voice": spec.get("voice", ""),
                "add_on_charges_hkd": " ".join(note for note in notes if note).strip(),
                "extraction_status": "parsed_historical_archive",
                "evidence_excerpt": _excerpt(text, block.start(), block.end()),
            }
        )
        rows.append(_historical_row(row, snapshot))
    return _dedupe(rows)


def _parse_pccw_mobile_tablet_data_tariff(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"Tablet mobile data tariff plan.*?Unlimited services are provided subject to PCCW", text, re.I)
    if not block:
        return []
    specs = [
        {
            "plan_name": "PCCW mobile historical Tablet Mobile Data Tariff Plan HK$198",
            "fee": "198",
            "data": "",
            "contract": "12",
            "voice": "Tablet mobile data plan; unlimited local data shown in source; max downlink 3.6Mbps / uplink 2.0Mbps",
            "note": "HK$300 supermarket coupon contract offer disclosed; $50 5-day usage line skipped because it is not a clear monthly data plan.",
        },
        {
            "plan_name": "PCCW mobile historical Tablet Mobile Data Tariff Plan HK$328",
            "fee": "328",
            "data": "",
            "contract": "24",
            "voice": "Tablet mobile data plan; unlimited local data shown in source; max downlink 7.2Mbps / uplink 5.76Mbps",
            "note": "HK$2200 rebate contract offer disclosed.",
        },
    ]
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Monthly mobile licence/related fee HK$12 is excluded from monthly_fee_hkd; unlimited data rows keep local_data_gb blank because the numeric structured field only stores GB allowances.",
    )


def _parse_pccw_mobile_free_to_go_sim_only(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"Free-to-go SIM Only Plan.*?Services are provided subject to PCCW mobile", text, re.I)
    if not block:
        return []
    specs = [
        {
            "plan_name": "PCCW mobile historical Free-to-go SIM Only Plan HK$232",
            "fee": "232",
            "data": "",
            "voice": "3000 normal voice-call mins; unlimited local data and unlimited Wi-Fi shown in source",
            "note": "Includes uHub 25GB, RoamSave trial, HD now TV 300 mins and MOOV offers where disclosed.",
        }
    ]
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Monthly mobile service licence/related charge HK$12 is excluded from monthly_fee_hkd; unlimited data kept in notes rather than numeric local_data_gb.",
    )


def _parse_pccw_mobile_new_monthly_plan(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"New Monthly Plan.*?VAS included", text, re.I)
    if not block:
        return []
    specs = [
        {"plan_name": "PCCW mobile historical New Monthly smartphone handset plan HK$149", "fee": "149", "data": "0.1", "voice": "1800 inter/intra-network voice-call mins; 60 video mins; unlimited intra-network SMS; 60 MMS"},
        {"plan_name": "PCCW mobile historical New Monthly smartphone handset plan HK$299", "fee": "299", "data": "0.8", "voice": "3000 inter/intra-network voice-call mins; 200 video mins; unlimited intra-network SMS; 200 MMS"},
        {"plan_name": "PCCW mobile historical New Monthly smartphone handset plan HK$389", "fee": "389", "data": "5", "voice": "3000 inter/intra-network voice-call mins; 300 video mins; unlimited intra-network SMS; 300 MMS", "note": "5GB is disclosed as promotion offer; standard entitlement 2.5GB also shown."},
        {"plan_name": "PCCW mobile historical New Monthly smartphone SIM special plan HK$98", "fee": "98", "data": "0.1", "voice": "1800 inter/intra-network voice-call mins; 60 video mins; unlimited intra-network SMS; 60 MMS", "note": "SIM subscription special offer monthly fee."},
        {"plan_name": "PCCW mobile historical New Monthly smartphone SIM special plan HK$169", "fee": "169", "data": "0.8", "voice": "3000 inter/intra-network voice-call mins; 200 video mins; unlimited intra-network SMS; 200 MMS", "note": "SIM subscription special offer monthly fee."},
        {"plan_name": "PCCW mobile historical New Monthly smartphone SIM special plan HK$182", "fee": "182", "data": "5", "voice": "3000 inter/intra-network voice-call mins; 300 video mins; unlimited intra-network SMS; 300 MMS", "note": "SIM subscription special offer monthly fee; 5GB is disclosed as promotion offer, standard entitlement 2.5GB also shown."},
        {"plan_name": "PCCW mobile historical New Monthly tablet handset plan HK$119", "fee": "119", "data": "0.1", "voice": "Tablet data plan; unlimited intra-network SMS; 60 MMS"},
        {"plan_name": "PCCW mobile historical New Monthly tablet handset plan HK$269", "fee": "269", "data": "0.8", "voice": "Tablet data plan; unlimited intra-network SMS; 200 MMS"},
        {"plan_name": "PCCW mobile historical New Monthly tablet handset plan HK$359", "fee": "359", "data": "5", "voice": "Tablet data plan; unlimited intra-network SMS; 300 MMS", "note": "5GB is disclosed as promotion offer; standard entitlement 2.5GB also shown."},
        {"plan_name": "PCCW mobile historical New Monthly tablet SIM special plan HK$139", "fee": "139", "data": "0.8", "voice": "Tablet data plan; unlimited intra-network SMS; 200 MMS", "note": "SIM subscription special offer monthly fee."},
        {"plan_name": "PCCW mobile historical New Monthly tablet SIM special plan HK$169", "fee": "169", "data": "5", "voice": "Tablet data plan; unlimited intra-network SMS; 300 MMS", "note": "SIM subscription special offer monthly fee; 5GB is disclosed as promotion offer, standard entitlement 2.5GB also shown."},
    ]
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Monthly MTR/tunnel/mobile service licence and related charge HK$12 is excluded from monthly_fee_hkd; optional voice top-up HK$18/HK$50 and data top-up HK$40/HK$90 are retained only in notes/source evidence.",
    )


def _parse_pccw_mobile_ultimate_4g_smartphone(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"Ultimate Mobility monthly plan for 4G smartphone.*?VAS included", text, re.I)
    if not block:
        return []
    specs = [
        {"plan_name": "PCCW mobile historical Ultimate Mobility 4G smartphone handset plan HK$389", "fee": "389", "data": "5", "voice": "3000 inter/intra-network voice-call mins; unlimited intra-network SMS; 300 MMS; 300 video mins", "note": "5GB is disclosed as promotion offer; standard entitlement 2.5GB also shown."},
        {"plan_name": "PCCW mobile historical Ultimate Mobility 4G smartphone handset plan HK$429", "fee": "429", "data": "10", "voice": "3000 inter/intra-network voice-call mins; unlimited intra-network SMS; 300 MMS; 300 video mins", "note": "10GB is disclosed as promotion offer; standard entitlement 5GB also shown."},
        {"plan_name": "PCCW mobile historical Ultimate Mobility 4G smartphone SIM special plan HK$174", "fee": "174", "data": "5", "voice": "3000 inter/intra-network voice-call mins; unlimited intra-network SMS; 300 MMS; 300 video mins", "note": "SIM subscription special monthly fee; 5GB is disclosed as promotion offer, standard 2.5GB also shown."},
        {"plan_name": "PCCW mobile historical Ultimate Mobility 4G smartphone Free-to-go plan HK$199", "fee": "199", "data": "5", "voice": "3000 inter/intra-network voice-call mins; unlimited intra-network SMS; 300 MMS; 300 video mins", "note": "Free-to-go SIM offer; 5GB is disclosed as promotion offer, standard 2.5GB also shown."},
        {"plan_name": "PCCW mobile historical Ultimate Mobility 4G smartphone SIM special plan HK$190", "fee": "190", "data": "10", "voice": "3000 inter/intra-network voice-call mins; unlimited intra-network SMS; 300 MMS; 300 video mins", "note": "SIM subscription special monthly fee; 10GB is disclosed as promotion offer, standard 5GB also shown."},
        {"plan_name": "PCCW mobile historical Ultimate Mobility 4G smartphone Free-to-go plan HK$238", "fee": "238", "data": "10", "voice": "3000 inter/intra-network voice-call mins; unlimited intra-network SMS; 300 MMS; 300 video mins", "note": "Free-to-go SIM offer; 10GB is disclosed as promotion offer, standard 5GB also shown."},
    ]
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Monthly MTR/tunnel/mobile service licence/admin charge HK$12 and MNP rebates are excluded from monthly_fee_hkd.",
    )


def _parse_pccw_mobile_multi_smart_sims(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"Multi Smart SIMs monthly plan.*?Services are provided subject to PCCW mobile", text, re.I)
    if not block:
        return []
    specs = [
        {
            "plan_name": "PCCW mobile historical Multi Smart SIMs monthly plan HK$268",
            "fee": "268",
            "data": "5",
            "contract": "24",
            "voice": "1 primary SIM and 2 secondary SIMs; 3000 voice-call mins for primary SIM; 300 MMS; unlimited intra-network SMS; 3 PCCW Wi-Fi accounts",
            "note": "5GB is disclosed as promotion offer shared among 3 SIMs; standard entitlement 2.5GB also shown; free HK$38/month Smart Value Service Package disclosed.",
        }
    ]
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Monthly MTR/tunnel/mobile service licence/admin charge HK$12 and data top-up charges are excluded from monthly_fee_hkd.",
    )


def _parse_pccw_mobile_2g_tariff(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"2G tariff plan.*?Monthly\s*fee.*?\$88.*?1200.*?Monthly mobile service license", text, re.I)
    if not block:
        return []
    specs = [
        {
            "plan_name": "PCCW mobile historical 2G tariff plan HK$88",
            "fee": "88",
            "data": "",
            "voice": "1200 normal local voice-call mins; MNP +200 mins where applicable",
            "note": "12-month waiver of voice mail/call forwarding bundle/prime features package disclosed.",
        }
    ]
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Monthly mobile service licence/related charge HK$12 and new subscription/provisional charges are excluded from monthly_fee_hkd.",
    )


def _parse_pccw_mobile_cdma_service(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"CDMA mobile service.*?CDMA service plan.*?\$3,888.*?Unlimited.*?Unlimited.*?(?:Notes|註)", text, re.I)
    if not block:
        block = re.search(r"CDMA流動通訊服務.*?CDMA服務計劃.*?\$3,888.*?無限.*?無限.*?註", text, re.I)
    if not block:
        return []
    specs = [
        {
            "plan_name": "PCCW mobile historical CDMA service plan HK$3888",
            "fee": "3888",
            "data": "",
            "contract": "24",
            "voice": "Unlimited local voice-call minutes; unlimited local mobile data service disclosed in source",
            "note": "Unlimited local mobile data is disclosed, but numeric local_data_gb is kept blank because the source does not disclose a GB allowance.",
        }
    ]
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "First-two-month prepayment, early termination charges and unsupported VAS limitations are notes only; not included in monthly_fee_hkd.",
    )


def _parse_pccw_mobile_concierge(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"PCCW Concierge Tariff Plan.*?Customers are required to pay monthly mobile service license", text, re.I)
    if not block:
        return []
    specs = [
        {"plan_name": "PCCW mobile historical Concierge Tariff Plan HK$149", "fee": "149", "data": "0.1", "voice": "1200 voice-call mins; 60 video-call mins; unlimited intra-network SMS; 60 MMS"},
        {"plan_name": "PCCW mobile historical Concierge Tariff Plan HK$199", "fee": "199", "data": "0.2", "voice": "1600 voice-call mins; 100 video-call mins; unlimited intra-network SMS; 100 MMS"},
        {"plan_name": "PCCW mobile historical Concierge Tariff Plan HK$249", "fee": "249", "data": "0.5", "voice": "1600 voice-call mins; 100 video-call mins; unlimited intra-network SMS; 100 MMS"},
        {"plan_name": "PCCW mobile historical Concierge Tariff Plan HK$299", "fee": "299", "data": "0.5", "voice": "1800 voice-call mins; 200 video-call mins; unlimited intra-network SMS; 200 MMS"},
        {"plan_name": "PCCW mobile historical Concierge Tariff Plan HK$399", "fee": "399", "data": "", "voice": "3000 voice-call mins; unlimited local data shown in source; 300 video-call mins; 300 MMS"},
        {"plan_name": "PCCW mobile historical Concierge Tariff Plan HK$478", "fee": "478", "data": "", "voice": "8000 voice-call mins (4000 normal + 4000 intra-network); MNP +400 mins where applicable; unlimited local data shown in source; 400 video-call mins; 400 MMS"},
        {"plan_name": "PCCW mobile historical Concierge Tariff Plan HK$499", "fee": "499", "data": "", "voice": "4000 voice-call mins; China-HK 1-Card-2-Number 200 mins/special data charge shown in source; 400 video-call mins; 400 MMS", "note": "Data allowance column is not clear enough for this tariff; kept blank rather than estimating."},
    ]
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Extra handset rebate and HK$300 supermarket voucher are promotional notes only; monthly mobile licence/related charge HK$12 is excluded from monthly_fee_hkd. Unlimited data rows keep local_data_gb blank because the numeric field only stores GB allowances.",
    )


def _parse_pccw_mobile_netvigator_customer_offer(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"Special offer for Netvigator customer.*?Local data usage charge", text, re.I)
    if not block:
        return []
    specs = [
        {"plan_name": "PCCW mobile historical Netvigator customer special offer HK$256 36m", "fee": "256", "data": "1", "contract": "36", "voice": "1GB local data plus unlimited PCCW Wi-Fi; optional voice upgrade disclosed separately"},
        {"plan_name": "PCCW mobile historical Netvigator customer special offer HK$288 30m", "fee": "288", "data": "1", "contract": "30", "voice": "1GB local data plus unlimited PCCW Wi-Fi; optional voice upgrade disclosed separately"},
        {"plan_name": "PCCW mobile historical Netvigator customer special offer HK$318 24m", "fee": "318", "data": "1", "contract": "24", "voice": "1GB local data plus unlimited PCCW Wi-Fi; optional voice upgrade disclosed separately"},
    ]
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Upgrade data packages HK$20/HK$50, voice upgrade HK$50 and data overage charges are optional add-ons, not included in monthly_fee_hkd.",
    )


def _parse_pccw_mobile_new_ultimate_smartphones(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"New Ultimate Mobility\s+Service Plan for Smartphones.*?VAS included", text, re.I)
    if not block:
        return []
    tiers = [
        ("149", "1", "100MB standard entitlement also shown", "60"),
        ("299", "2.5", "800MB standard entitlement also shown", "200"),
        ("389", "5", "2.5GB standard entitlement also shown", "300"),
        ("429", "10", "5GB standard entitlement also shown", "300"),
        ("489", "20", "10GB standard entitlement also shown", "300"),
    ]
    specs: List[Dict[str, str]] = []
    for fee, data, note, media in tiers:
        specs.append({"plan_name": f"PCCW mobile historical New Ultimate Mobility smartphone handset plan HK${fee}", "fee": fee, "data": data, "voice": f"Unlimited inter/intra-network voice-call mins; unlimited intra-network SMS; {media} MMS; {media} video-call mins", "note": note})
    sim_tiers = [
        ("119", "1", "SIM subscription special monthly fee; 100MB standard entitlement also shown"),
        ("149", "1", "Free-to-go smartphone offer; 100MB standard entitlement also shown"),
        ("149", "2.5", "SIM subscription special monthly fee; 800MB standard entitlement also shown"),
        ("169", "2.5", "Free-to-go smartphone offer; 800MB standard entitlement also shown"),
        ("174", "5", "SIM subscription special monthly fee; 2.5GB standard entitlement also shown"),
        ("199", "5", "Free-to-go smartphone offer; 2.5GB standard entitlement also shown"),
        ("190", "10", "SIM subscription special monthly fee; 5GB standard entitlement also shown"),
        ("238", "10", "Free-to-go smartphone offer; 5GB standard entitlement also shown"),
        ("238", "20", "SIM subscription special monthly fee; 10GB standard entitlement also shown"),
        ("299", "20", "Free-to-go smartphone offer; 10GB standard entitlement also shown"),
    ]
    for fee, data, note in sim_tiers:
        specs.append({"plan_name": f"PCCW mobile historical New Ultimate Mobility smartphone {note.split(';')[0]} HK${fee}", "fee": fee, "data": data, "voice": "Unlimited inter/intra-network voice-call mins; unlimited intra-network SMS", "note": note})
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Monthly MTR/tunnel/mobile service licence/admin charge HK$12, rebates, roaming offers and Asia Miles are excluded from monthly_fee_hkd.",
    )


def _parse_pccw_mobile_new_ultimate_tablets(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    block = re.search(r"New Ultimate Mobility Monthly Plan for Tablets.*?VAS included", text, re.I)
    if not block:
        return []
    tiers = [
        ("269", "2.5", "800MB standard entitlement also shown", "60"),
        ("359", "5", "2.5GB standard entitlement also shown", "200"),
        ("399", "10", "5GB standard entitlement also shown", "300"),
        ("459", "20", "10GB standard entitlement also shown", "300"),
    ]
    specs: List[Dict[str, str]] = []
    for fee, data, note, media in tiers:
        specs.append({"plan_name": f"PCCW mobile historical New Ultimate Mobility tablet handset plan HK${fee}", "fee": fee, "data": data, "voice": f"Tablet data plan; unlimited intra-network SMS; {media} MMS", "note": note})
    sim_tiers = [
        ("129", "2.5", "SIM subscription special monthly fee; 800MB standard entitlement also shown"),
        ("148", "5", "SIM subscription special monthly fee; 2.5GB standard entitlement also shown"),
        ("166", "10", "SIM subscription special monthly fee; 5GB standard entitlement also shown"),
        ("208", "10", "Free-to-go tablet offer; 5GB standard entitlement also shown"),
        ("215", "20", "SIM subscription special monthly fee; 10GB standard entitlement also shown"),
        ("269", "20", "Free-to-go tablet offer; 10GB standard entitlement also shown"),
    ]
    for fee, data, note in sim_tiers:
        specs.append({"plan_name": f"PCCW mobile historical New Ultimate Mobility tablet {note.split(';')[0]} HK${fee}", "fee": fee, "data": data, "voice": "Tablet data plan; unlimited intra-network SMS", "note": note})
    return _pccw_static_rows(
        snapshot,
        result,
        captured_at,
        block,
        specs,
        "Monthly MTR/tunnel/mobile service licence/admin charge HK$12, rebates, roaming offers and Asia Miles are excluded from monthly_fee_hkd.",
    )


def _parse_pccw_mobile_2g_99_tariff(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if not re.search(r"(?:2G \$99 Integrated Minutes Monthly Tariff Plan|2G \$99綜合分鐘月費計劃)", text, re.I):
        return []
    block = re.search(
        r"(?:Monthly fee|月費).*?(?P<table>\$99\s+999\s+9,000\s+\$1\.0.*?(?:12-month fee waiver|每月流動電話服務牌照))",
        text,
        re.I,
    )
    if not block:
        return []
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": "PCCW mobile historical 2G $99 Integrated Minutes Monthly Tariff Plan",
            "monthly_fee_hkd": "99",
            "local_data_gb": "",
            "roaming_data_gb": "",
            "post_fup_speed_mbps": "",
            "contract_months": "12",
            "local_voice": "999 peak-hour local voice mins; 9000 off-peak local voice mins",
            "add_on_charges_hkd": "Thereafter charge HK$1/min and monthly mobile licence/related fee are excluded from monthly_fee_hkd; rebates and VAS waivers are kept as evidence only.",
            "extraction_status": "parsed_historical_archive",
            "evidence_excerpt": _excerpt(text, block.start(), block.end()),
        }
    )
    return [_historical_row(row, snapshot)]


def _parse_1010_kingking_voice_roaming(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    match = re.search(
        r"For \$149 above Service Plans\s+Free.*?For \$148 or below Service Plans\s+\$8/day",
        text,
        flags=re.I | re.S,
    )
    if not match or "KingKing" not in text:
        return []
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": "1O1O historical KingKing Voice Roaming Service daily pass",
            "published_price_hkd": "8",
            "price_billing_unit": "day",
            "local_voice": "VoIP service over Wi-Fi: calls to Hong Kong numbers, inbound calls worldwide and KingKing-to-KingKing video calls under the stated eligibility conditions.",
            "add_on_charges_hkd": "Official daily-pass tariff. Free for service plans HK$149 or above; HK$8/day for service plans HK$148 or below. IDD, SMS, Wi-Fi connectivity and applicable local voice-minute charges remain separate. This is a day charge and is intentionally not written to monthly_fee_hkd.",
            "extraction_status": "parsed_historical_archive",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [_historical_row(row, snapshot)]


def _parse_1010_3g_mobile_tv(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    match = re.search(r"3G Mobile TV(?: Package)?\s+Monthly (?:Fee )?Package.*?\$30\s+(?:50|100)mins.*?\$1\s*/\s*min", text, flags=re.I | re.S)
    if not match:
        return []
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": "1O1O historical 3G Mobile TV Package HK$30",
            "monthly_fee_hkd": "30",
            "local_voice": "3G Mobile TV service with 50 included viewing minutes per month under the standard package.",
            "add_on_charges_hkd": "Official tariff/list price. HK$1/min thereafter. The page states a promotional doubling to 100 minutes for the first 12 months; this temporary promotion is not recorded as the standard included entitlement. Adult stations carry separately listed monthly charges and are excluded from the base package fee.",
            "extraction_status": "parsed_historical_archive",
            "evidence_excerpt": _excerpt(text, match.start(), match.end()),
        }
    )
    return [_historical_row(row, snapshot)]


def _parse_1010_football_service(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    channel_match = re.search(r"Football Channel Package.*?\$33\s*/\s*month", text, flags=re.I | re.S)
    league_match = re.search(r"League Package.*?\$28", text, flags=re.I | re.S)
    team_match = re.search(r"Team Package.*?\$8", text, flags=re.I | re.S)
    if not (channel_match and league_match and team_match):
        return []
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }
    specs = [
        ("Football Channel Package", "33", channel_match, "Unlimited browsing of specified football video/text content and mobile online betting service under the stated fair-use conditions."),
        ("Football League Package", "28", league_match, "League match schedule, odds, news, half-time updates and results summaries."),
        ("Football Team Package", "8", team_match, "Team-focused football information package under the stated service terms."),
    ]
    rows: List[Dict[str, str]] = []
    for label, fee, match, description in specs:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"1O1O historical {label} HK${fee}",
                "monthly_fee_hkd": fee,
                "local_voice": description,
                "add_on_charges_hkd": "Official tariff/list price. Pay-as-you-go content, live-score/SMS charges, roaming charges and other separately listed content charges are excluded from the monthly package fee.",
                "extraction_status": "parsed_historical_archive",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(_historical_row(row, snapshot))
    return rows


def _parse_1010_music_service(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    specs = [
        ("Music Channel Package", "20", r"Music Channel Package Including\s+\$20\s*/\s*month", "20 MV viewings, one MP3 ringtone and one colour wallpaper for selected content."),
        ("RingMaster Poly/Mono Ringtone Package", "15", r"Poly/Mono Ringtone \+ Ringmaster Special Offer#\s+\$15", "Two designated poly/mono ringtones and Ringmaster special offer."),
        ("RingMaster MP3 Ringtone Package", "25", r"MP3 Ringtone \+ Ringmaster Special Offer#\s+\$25", "Two designated MP3 ringtones and Ringmaster special offer."),
        ("RingMaster Power Station 1 Channel", "15", r'"RingMaster" Power Station#\s+\$15\s+1 Channel', "One RingMaster Power Station channel."),
        ("RingMaster Power Station 2+ Channels", "28", r"\$28\s+2 or above Channels", "Two or more RingMaster Power Station channels."),
        ("Self-selected Ringmaster Package", "20", r"Self- selected Ringmaster Package#\s+\$20", "Three self-selected connecting tones per month."),
        ("Full Song Download Package", "20", r"Full Song Download Package\s+\$20\s+3 designated Full Song downloads", "Three designated full-song downloads and 20 MV viewings."),
    ]
    matches = [(label, fee, re.search(pattern, text, flags=re.I | re.S), description) for label, fee, pattern, description in specs]
    if any(match is None for _, _, match, _ in matches):
        return []
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }
    rows: List[Dict[str, str]] = []
    for label, fee, match, description in matches:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"1O1O historical {label} HK${fee}",
                "monthly_fee_hkd": fee,
                "local_voice": description,
                "add_on_charges_hkd": "Official tariff/list price. Pay-as-you-go downloads, previews, extra connecting tones, mobile data/WAP and roaming charges are separately listed and excluded from the monthly package fee.",
                "extraction_status": "parsed_historical_archive",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(_historical_row(row, snapshot))
    return rows


def _parse_1010_anyplex(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if snapshot["snapshot_id"].endswith("_en"):
        matches = [
            ("12", re.search(r"Anyplex.*?12 Months\s+\$38\s+3 movie vouchers per month", text, flags=re.I | re.S)),
            ("24", re.search(r"24 Months\s+\$38.*?extra movie voucher", text, flags=re.I | re.S)),
        ]
    else:
        # The old Traditional Chinese page is decoded with legacy-character replacement,
        # but the two commitment terms and HK$38 price cells remain directly readable.
        matches = [
            ("12", re.search(r"12[^$]{0,80}\$38", text, flags=re.S)),
            ("24", re.search(r"24[^$]{0,80}\$38", text, flags=re.S)),
        ]
    if any(match is None for _, match in matches):
        return []
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": "wayback_official_page_snapshot",
    }
    rows: List[Dict[str, str]] = []
    for contract_months, match in matches:
        row = _base_row(source, result, captured_at)
        row.update(
            {
                "plan_name": f"1O1O historical Anyplex movie-voucher service HK$38 {contract_months} months",
                "monthly_fee_hkd": "38",
                "contract_months": contract_months,
                "local_voice": "Three Anyplex movie vouchers per month under the stated service plan.",
                "add_on_charges_hkd": "Official tariff/list price. Extra vouchers stated for early months and post-commitment service extensions are promotional benefits; mobile-data and roaming-data charges for downloading the app remain separate.",
                "extraction_status": "parsed_historical_archive",
                "evidence_excerpt": _excerpt(text, match.start(), match.end()),
            }
        )
        rows.append(_historical_row(row, snapshot))
    return rows


def _fetch_wayback_snapshot(client: httpx.Client, snapshot: Dict[str, str]) -> Dict[str, Any]:
    archive_url = f"https://web.archive.org/web/{snapshot['timestamp']}id_/{snapshot['original_url']}"
    last_error = ""
    last_status = 0
    for attempt in range(3):
        try:
            response = client.get(
                archive_url,
                timeout=httpx.Timeout(25.0, connect=8.0),
                headers={
                    "User-Agent": (
                        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
                    ),
                    "Accept-Language": "en,zh-CN;q=0.9,zh;q=0.8",
                },
            )
            raw = response.content
            content_type = response.headers.get("content-type", "")
            title, text = _html_to_text(raw, content_type)
            last_status = response.status_code
            if response.status_code != 503 and len(text) >= 120:
                return {
                    "url": archive_url,
                    "final_url": str(response.url),
                    "status": response.status_code,
                    "content_type": content_type,
                    "bytes": len(raw),
                    "title": title,
                    "text": text,
                    "error": "",
                    "method": f"wayback_httpx_attempt_{attempt + 1}",
                }
            last_error = f"short_or_unavailable_snapshot status={response.status_code} text_chars={len(text)}"
        except Exception as exc:
            last_error = repr(exc)
        time.sleep(1.5 * (attempt + 1))
    return {
        "url": archive_url,
        "final_url": archive_url,
        "status": last_status,
        "content_type": "",
        "bytes": 0,
        "title": "",
        "text": "",
        "error": last_error,
        "method": "wayback_httpx_error",
    }


def _parse_historical_snapshot(snapshot: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    source = {
        "source_id": snapshot["snapshot_id"],
        "brand": snapshot["brand"],
        "product_category": snapshot["product_category"],
        "url": snapshot["original_url"],
        "official_source_type": snapshot.get("source_kind") or "wayback_official_page_snapshot",
    }
    if snapshot.get("published_on"):
        source["published_on"] = snapshot["published_on"]
    if snapshot["parser"] == "csl_old_data_voice":
        rows = _parse_csl_old_data_voice(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_3g_tariff":
        rows = _parse_pccw_mobile_3g_tariff(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_other_3g_tariff":
        rows = _parse_pccw_mobile_other_3g_tariff(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_web_talk_tariff":
        rows = _parse_pccw_mobile_web_talk_tariff(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_tablet_data_tariff":
        rows = _parse_pccw_mobile_tablet_data_tariff(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_free_to_go_sim_only":
        rows = _parse_pccw_mobile_free_to_go_sim_only(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_new_monthly_plan":
        rows = _parse_pccw_mobile_new_monthly_plan(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_ultimate_4g_smartphone":
        rows = _parse_pccw_mobile_ultimate_4g_smartphone(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_multi_smart_sims":
        rows = _parse_pccw_mobile_multi_smart_sims(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_2g_tariff":
        rows = _parse_pccw_mobile_2g_tariff(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_cdma_service":
        rows = _parse_pccw_mobile_cdma_service(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_concierge":
        rows = _parse_pccw_mobile_concierge(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_netvigator_customer_offer":
        rows = _parse_pccw_mobile_netvigator_customer_offer(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_new_ultimate_smartphones":
        rows = _parse_pccw_mobile_new_ultimate_smartphones(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_new_ultimate_tablets":
        rows = _parse_pccw_mobile_new_ultimate_tablets(snapshot, result, captured_at)
    elif snapshot["parser"] == "pccw_mobile_2g_99_tariff":
        rows = _parse_pccw_mobile_2g_99_tariff(snapshot, result, captured_at)
    elif snapshot["parser"] == "1010_kingking_voice_roaming":
        rows = _parse_1010_kingking_voice_roaming(snapshot, result, captured_at)
    elif snapshot["parser"] == "1010_3g_mobile_tv":
        rows = _parse_1010_3g_mobile_tv(snapshot, result, captured_at)
    elif snapshot["parser"] == "1010_football_service":
        rows = _parse_1010_football_service(snapshot, result, captured_at)
    elif snapshot["parser"] == "1010_music_service":
        rows = _parse_1010_music_service(snapshot, result, captured_at)
    elif snapshot["parser"] == "1010_anyplex":
        rows = _parse_1010_anyplex(snapshot, result, captured_at)
    elif snapshot["parser"] == "csl_5g":
        rows = [_historical_row(row, snapshot) for row in _parse_csl(source, result, captured_at)]
        for row in rows:
            row["extraction_status"] = "parsed_historical_archive"
    elif snapshot["parser"] == "1010_5g":
        rows = [_historical_row(row, snapshot) for row in _parse_1010(source, result, captured_at)]
        for row in rows:
            row["extraction_status"] = "parsed_historical_archive"
    elif snapshot["parser"] == "netvigator_list_price":
        rows = [_historical_row(row, snapshot) for row in _parse_netvigator_list_price(source, result, captured_at)]
        for row in rows:
            row["extraction_status"] = "parsed_historical_archive"
    elif snapshot["parser"] == "hkt_internet_access_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_internet_access_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_business_broadband_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_business_broadband_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_premium_broadband_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        source["effective_from"] = snapshot.get("effective_from", "")
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_premium_broadband_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_megalink_service_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        source["effective_from"] = snapshot.get("effective_from", "")
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_megalink_service_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_datapak_private_circuit_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_datapak_private_circuit_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_metro_ip_service_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_metro_ip_service_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_flexible_bandwidth_service_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_flexible_bandwidth_service_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_telecommunications_backup_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_telecommunications_backup_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_one_communications_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_one_communications_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_eye_service_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_eye_service_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_international_toll_free_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_international_toll_free_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_super_hotline_tariff_pdf":
        document_snapshot = {**snapshot, "archive_url": snapshot["original_url"], "source_kind": "official_public_tariff_pdf"}
        rows = [_historical_row(row, document_snapshot) for row in _parse_hkt_super_hotline_tariff_pdf(source, result, captured_at)]
    elif snapshot["parser"] == "hkt_faxline_tariff_pdf":
        document_snapshot = {**snapshot, "archive_url": snapshot["original_url"], "source_kind": "official_public_tariff_pdf"}
        rows = [_historical_row(row, document_snapshot) for row in _parse_hkt_faxline_tariff_pdf(source, result, captured_at)]
    elif snapshot["parser"] == "hkt_homefax_1_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_homefax_1_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "1010_ipad_pro_2020_product_page":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": snapshot.get("source_kind") or "official_public_product_page",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_1010_ipad_pro_2020_product_page(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_eye_home_smartphone_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_eye_home_smartphone_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_eye_multimedia_service_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_eye_multimedia_service_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_home_easywatch_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_home_easywatch_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_easywatch_commercial_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_easywatch_commercial_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_ip_voice_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_ip_voice_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "csl_voip_monthly_pass_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_csl_voip_monthly_pass_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_customer_voice_hotline_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_customer_voice_hotline_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_residential_cell_relay_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_residential_cell_relay_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_megalink_plus_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_megalink_plus_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_ip_net_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_ip_net_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_freedome_network_safety_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_freedome_network_safety_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_norton_secure_vpn_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_norton_secure_vpn_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_eye2_communication_package_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_eye2_communication_package_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_severe_weather_warning_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_severe_weather_warning_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_integrated_digital_access_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        source["effective_from"] = snapshot.get("effective_from", "")
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_integrated_digital_access_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_consumer_fixed_line_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_consumer_fixed_line_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "hkt_local_business_telephone_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_hkt_local_business_telephone_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "csl_2g_3g_4g_mobile_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_csl_2g_3g_4g_mobile_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "csl_smart_pama_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_csl_smart_pama_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "csl_postpaid_service_plan_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_csl_postpaid_service_plan_tariff_pdf(source, result, captured_at)
        ]
    elif snapshot["parser"] == "csl_1010_postpaid_mobile_tariff_pdf":
        document_snapshot = {
            **snapshot,
            "archive_url": snapshot["original_url"],
            "source_kind": "official_public_tariff_pdf",
        }
        rows = [
            _historical_row(row, document_snapshot)
            for row in _parse_csl_1010_postpaid_mobile_tariff_pdf(source, result, captured_at)
        ]
    else:
        rows = []
    if rows:
        return _dedupe(rows)
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"{snapshot['brand']} {snapshot['timestamp'][:4]} source gap",
            "extraction_status": "archive_fetched_no_plan_rows",
            "evidence_excerpt": _clean_text(str(result.get("text") or ""))[:500],
        }
    )
    return [_historical_row(row, snapshot)]


def _fetch_official_tariff_document(client: Any, snapshot: Dict[str, str]) -> Dict[str, Any]:
    return fetch_product_page(client, snapshot["original_url"])


def _parse_source(source: Dict[str, str], result: Dict[str, Any], captured_at: str) -> List[Dict[str, str]]:
    if source["source_id"] in CURRENT_SOURCES_WITH_INTENTIONAL_NO_STRUCTURED_ROWS:
        rows = []
    elif source["source_id"] == "csl_5g_tariff_en":
        rows = _parse_csl(source, result, captured_at)
    elif source["source_id"] == "1010_5g_service_plan":
        rows = _parse_1010(source, result, captured_at)
    elif source["source_id"] == "1010_infinite_entertainment_5g_prestige":
        rows = _parse_1010_infinite_entertainment(source, result, captured_at)
    elif source["source_id"] == "netvigator_home_broadband_index":
        rows = _parse_netvigator_current_offer(source, result, captured_at)
    elif source["source_id"] == "netvigator_broadband_list_price":
        rows = _parse_netvigator_list_price(source, result, captured_at)
    elif source["source_id"] in {"netvigator_csl_5g_home_internet_offer", "csl_netvigator_5g_home_internet_offer"}:
        rows = _parse_netvigator_csl_5g_home_internet(source, result, captured_at)
    elif source["source_id"] == "hkt_sme_5g_business_mobile":
        rows = _parse_hkt_sme_5g_business_mobile(source, result, captured_at)
    elif source["source_id"] == "hkt_sme_business_broadband":
        rows = _parse_hkt_sme_business_broadband(source, result, captured_at)
    elif source["source_id"] in {
        "hkt_enterprise_local_business_telephone",
        "hkt_enterprise_local_business_telephone_tc",
        "hkt_local_business_telephone_hktcom_en",
    }:
        rows = _parse_hkt_enterprise_local_business_telephone(source, result, captured_at)
    elif source["source_id"] == "hkt_homephone_value_added_services_en":
        rows = _parse_hkt_homephone_value_added_services(source, result, captured_at)
    else:
        rows = _parse_generic(source, result, captured_at)
    if rows:
        return rows
    row = _base_row(source, result, captured_at)
    row.update(
        {
            "plan_name": f"{source['brand']} source snapshot",
            "extraction_status": "source_fetched_no_plan_rows",
            "evidence_excerpt": _clean_text(str(result.get("text") or ""))[:500],
        }
    )
    row["record_key"] = _record_key(row)
    return [row]


def _read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [dict(row) for row in csv.DictReader(fh)]


def _write_csv(path: Path, rows: List[Dict[str, str]], fields: List[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _merge_history(history_path: Path, current_rows: List[Dict[str, str]], captured_at: str) -> List[Dict[str, str]]:
    existing = {row.get("record_key", ""): row for row in _read_csv(history_path) if row.get("record_key")}
    current_source_ids = {row.get("source_id", "") for row in current_rows if row.get("source_id")}
    current_keys = {row.get("record_key", "") for row in current_rows if row.get("record_key")}
    if current_source_ids:
        existing = {
            key: row
            for key, row in existing.items()
            if row.get("source_id") not in current_source_ids or key in current_keys
        }
    for row in current_rows:
        key = row["record_key"]
        if key in existing:
            merged = dict(existing[key])
            merged.update({k: v for k, v in row.items() if v})
            merged["first_seen_at_hkt"] = existing[key].get("first_seen_at_hkt") or captured_at
            merged["last_seen_at_hkt"] = captured_at
            existing[key] = merged
        else:
            existing[key] = dict(row)
    return sorted(existing.values(), key=lambda item: (item.get("brand", ""), item.get("monthly_fee_hkd", "")))


def _fallback_previous_success_rows(
    previous_rows: List[Dict[str, str]],
    source_id: str,
    captured_at: str,
) -> List[Dict[str, str]]:
    fallback_rows: List[Dict[str, str]] = []
    for row in previous_rows:
        if row.get("source_id") != source_id or row.get("extraction_status") != "parsed":
            continue
        restored = dict(row)
        restored["captured_at_hkt"] = captured_at
        restored["last_seen_at_hkt"] = captured_at
        restored["fetch_method"] = "previous_success_fallback"
        restored["evidence_excerpt"] = (
            "本轮页面抓取或解析临时失败，沿用上一轮同 source_id 的已解析公开记录；"
            f"原证据：{row.get('evidence_excerpt', '')}"
        )[:900]
        fallback_rows.append(restored)
    return fallback_rows


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_readme(dataset_dir: Path, current_rows: List[Dict[str, str]], snapshots: List[Dict[str, Any]], captured_at: str) -> None:
    brands = sorted({row["brand"] for row in current_rows})
    parsed = [row for row in current_rows if row.get("extraction_status") == "parsed"]
    historical_rows = _read_csv(dataset_dir / "historical_tariffs.csv")
    historical_parsed = [row for row in historical_rows if str(row.get("extraction_status", "")).startswith("parsed")]
    historical_snapshots = _read_csv(dataset_dir / "historical_source_snapshots.csv")
    structured_gaps = _read_csv(dataset_dir / "structured_source_gaps.csv")
    historical_years = sorted({row.get("archive_year", "") for row in historical_parsed if row.get("archive_year")})
    lines = [
        "# HKT/csl/1O1O/HKT SME/NETVIGATOR 产品资费监测数据库",
        "",
        f"更新时间（HKT）：{captured_at}",
        "",
        "## 覆盖范围",
        "",
        "- csl 5G 消费者套餐公开资费页",
        "- 1O1O 5G 高端/多用户套餐公开资费页",
        "- HKT SME 5G Business Mobile 公开资费页",
        "- NETVIGATOR 家庭宽频当前优惠页和官方标准价目表",
        "- HKT/csl 官方 Internet Access、Business Broadband、Premium Broadband、MegaLink、Datapak/Private Circuit、Metro IP、Flexible Bandwidth、Telecommunications Backup、one communications、eye Service、Home Phone/Local Telephone、2G/3G/4G Mobile、U-plan、Smart Pama、The Club SIM 与 csl/1O1O 2018 postpaid PDF 资费表（作为官方 tariff/list price，不等同促销价）",
        "- HKT Enterprise 5G 企业移动/宽频产品公开页面（若页面只披露产品描述，则标记为 review_required 或 source_fetched_no_plan_rows）",
        "",
        "## 当前抓取概览",
        "",
        f"- 覆盖品牌：{', '.join(brands)}",
        f"- 当前记录：{len(current_rows)} 条",
        f"- 已结构化解析：{len(parsed)} 条",
        f"- 来源页面：{len(snapshots)} 个",
        f"- 官方历史快照：{len(historical_snapshots)} 个",
        f"- 历史记录：{len(historical_rows)} 条，其中已结构化 {len(historical_parsed)} 条",
        f"- 历史覆盖年份：{', '.join(historical_years) if historical_years else '待补充'}",
        f"- 来源缺口：{len(structured_gaps)} 条；缺口不估算。",
        "",
        "## 主要文件",
        "",
        "- `latest_products.csv`：本次抓取的最新套餐/产品结构化表。",
        "- `historical_tariffs.csv`：公开官方归档快照和官方 PDF 资费表解析出的历史套餐表。",
        "- `structured_current_plans.csv`：面向分析和人工审核的当前套餐宽表。",
        "- `structured_historical_plans.csv`：面向分析和人工审核的历史套餐宽表。",
        "- `structured_source_gaps.csv`：可访问但不能结构化、或抓取失败的缺口清单，禁止估算。",
        "- `hkt_product_tariffs_structured.xlsx`：包含当前套餐、历史套餐、来源缺口和字段说明的人工审核工作簿。",
        "- `product_history.csv`：按产品特征去重后的历史快照，保留 first_seen/last_seen。",
        "- `source_snapshots.json`：每个官方页面的抓取状态、哈希和证据摘录。",
        "- `change_log.md`：相对上一次运行的新增/消失记录。",
        "",
        "## 合规口径",
        "",
        "仅抓取公开官方页面；不登录、不绕过权限、不抓取个人数据；沿用主爬虫的 source_registry、robots、代理和敏感信息过滤流程。",
    ]
    (dataset_dir / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    (dataset_dir / "summary.md").write_text("\n".join(lines[:22]) + "\n", encoding="utf-8")


def _write_change_log(dataset_dir: Path, previous_rows: List[Dict[str, str]], current_rows: List[Dict[str, str]], captured_at: str) -> None:
    previous_keys = {row.get("record_key", "") for row in previous_rows}
    current_keys = {row.get("record_key", "") for row in current_rows}
    new_rows = [row for row in current_rows if row["record_key"] not in previous_keys]
    removed_rows = [row for row in previous_rows if row.get("record_key") not in current_keys]
    lines = [
        "# HKT 产品资费变化记录",
        "",
        f"运行时间（HKT）：{captured_at}",
        "",
        f"- 新增记录：{len(new_rows)}",
        f"- 本次未再出现记录：{len(removed_rows)}",
        "",
        "## 新增记录",
        "",
    ]
    if new_rows:
        for row in new_rows:
            lines.append(f"- {row['brand']}｜{row['plan_name']}｜HK${row.get('monthly_fee_hkd') or 'N/A'}｜{row['source_url']}")
    else:
        lines.append("- 无")
    lines.extend(["", "## 本次未再出现记录", ""])
    if removed_rows:
        for row in removed_rows:
            lines.append(f"- {row.get('brand')}｜{row.get('plan_name')}｜HK${row.get('monthly_fee_hkd') or 'N/A'}｜{row.get('source_url')}")
    else:
        lines.append("- 无")
    (dataset_dir / "change_log.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def _service_generation(product_category: str, plan_name: str = "") -> str:
    text = f"{product_category} {plan_name}".lower()
    if "5g" in text:
        return "5G"
    if "4.5g" in text:
        return "4.5G"
    if "4g" in text:
        return "4G"
    return ""


def _customer_segment(product_category: str, brand: str = "") -> str:
    text = f"{product_category} {brand}".lower()
    if "enterprise" in text or "business" in text:
        return "企业客户"
    if "prestige" in text or "1010" in text or "1o1o" in text:
        return "高端个人客户"
    if "broadband" in text or "fibre" in text or "fiber" in text:
        return "个人客户"
    if "consumer" in text or "mobile" in text:
        return "个人客户"
    return ""


def _plan_family(row: Dict[str, str]) -> str:
    plan_name = row.get("plan_name", "")
    if "Ultra 300" in plan_name:
        return "csl Ultra 300"
    if row.get("brand") == "1O1O":
        return "1O1O 5G"
    if row.get("brand") == "csl" and _service_generation(row.get("product_category", ""), plan_name) == "5G":
        return "csl 5G"
    if row.get("brand") == "csl":
        return "csl 4G/4.5G"
    if row.get("brand") == "NETVIGATOR":
        return "NETVIGATOR Home Broadband"
    return row.get("brand", "")


def _is_plan_row(row: Dict[str, str]) -> bool:
    return bool(
        row.get("monthly_fee_hkd")
        or row.get("published_price_hkd")
        or row.get("local_data_gb")
        or row.get("roaming_data_gb")
    )


def _structured_current_row(row: Dict[str, str]) -> Dict[str, str]:
    return {
        "as_of_hkt": row.get("captured_at_hkt", ""),
        "brand": row.get("brand", ""),
        "service_generation": _service_generation(row.get("product_category", ""), row.get("plan_name", "")),
        "customer_segment": _customer_segment(row.get("product_category", ""), row.get("brand", "")),
        "plan_family": _plan_family(row),
        "plan_name": row.get("plan_name", ""),
        "monthly_fee_hkd": row.get("monthly_fee_hkd", ""),
        "published_price_hkd": row.get("published_price_hkd", ""),
        "price_billing_unit": row.get("price_billing_unit", ""),
        "local_data_gb": row.get("local_data_gb", ""),
        "roaming_data_gb": row.get("roaming_data_gb", ""),
        "post_fup_speed_mbps": row.get("post_fup_speed_mbps", ""),
        "contract_months": row.get("contract_months", ""),
        "local_voice": row.get("local_voice", ""),
        "add_on_charges_hkd": row.get("add_on_charges_hkd", ""),
        "source_status": row.get("extraction_status", ""),
        "source_id": row.get("source_id", ""),
        "source_url": row.get("source_url", ""),
        "evidence_excerpt": row.get("evidence_excerpt", ""),
        "record_key": row.get("record_key", ""),
    }


def _structured_historical_row(row: Dict[str, str]) -> Dict[str, str]:
    structured = _structured_current_row(row)
    structured.update(
        {
            "archive_year": row.get("archive_year", ""),
            "archive_month": row.get("archive_month", ""),
            "snapshot_id": row.get("snapshot_id", ""),
            "archive_url": row.get("archive_url", ""),
        }
    )
    return {field: structured.get(field, "") for field in HISTORICAL_STRUCTURED_FIELDS}


def _source_gap_row(row: Dict[str, str]) -> Dict[str, str]:
    status = row.get("extraction_status", "")
    reason = {
        "archive_fetched_no_plan_rows": "页面/快照可访问，但未解析到结构化套餐价格行；不能估算。",
        "source_fetched_no_plan_rows": "当前官方页面可访问，但未披露可结构化的套餐价格行；不能估算。",
    }.get(status, "来源未形成可用结构化套餐行；不能估算。")
    return {
        "gap_year": row.get("archive_year", ""),
        "brand": row.get("brand", ""),
        "product_category": row.get("product_category", ""),
        "gap_type": status,
        "http_status": row.get("http_status", ""),
        "snapshot_id": row.get("snapshot_id", ""),
        "source_id": row.get("source_id", ""),
        "source_url": row.get("source_url", ""),
        "archive_url": row.get("archive_url", ""),
        "reason": reason,
        "evidence_excerpt": row.get("evidence_excerpt", ""),
    }


def _autosize_sheet(ws) -> None:
    for column_cells in ws.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max((len(value) for value in values), default=10) + 2, 48)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="0F2742")
        cell.fill = PatternFill("solid", fgColor="EAF4FF")


def _write_workbook(path: Path, sheets: Dict[str, tuple[List[Dict[str, str]], List[str]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for title, (rows, fields) in sheets.items():
        ws = wb.create_sheet(title[:31])
        ws.append(fields)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        _autosize_sheet(ws)
    wb.save(path)


def _write_structured_views(dataset_dir: Path, current_rows: List[Dict[str, str]]) -> None:
    historical_rows = _read_csv(dataset_dir / "historical_tariffs.csv")
    current_plan_rows = [_structured_current_row(row) for row in current_rows if _is_plan_row(row)]
    historical_plan_rows = [_structured_historical_row(row) for row in historical_rows if _is_plan_row(row)]
    gap_rows = [_source_gap_row(row) for row in [*current_rows, *historical_rows] if not _is_plan_row(row)]
    dictionary_rows = [
        {"field": "brand", "description": "品牌，例如 csl、1O1O、HKT Enterprise、HKT SME、NETVIGATOR"},
        {"field": "service_generation", "description": "套餐网络世代，按产品分类/套餐名称归纳为 4G、4.5G 或 5G"},
        {"field": "customer_segment", "description": "客户分层：个人客户、高端个人客户或企业客户"},
        {"field": "plan_family", "description": "便于分析的套餐族，例如 csl 5G、1O1O 5G、csl Ultra 300"},
        {"field": "monthly_fee_hkd", "description": "公开月费，单位 HKD"},
        {"field": "published_price_hkd", "description": "非月费产品的公开价格，单位 HKD；不替代月费字段"},
        {"field": "price_billing_unit", "description": "published_price_hkd 的计费单位，例如 day"},
        {"field": "local_data_gb", "description": "本地高速数据量，单位 GB；空值表示官方页面未披露或不适用"},
        {"field": "roaming_data_gb", "description": "中国内地/澳门等漫游数据量，单位 GB；空值表示官方页面未披露或不适用"},
        {"field": "post_fup_speed_mbps", "description": "高速数据用完后的限速，单位 Mbps"},
        {"field": "source_status", "description": "parsed/parsed_historical_archive 表示已结构化；source gap 类状态不能估算"},
        {"field": "evidence_excerpt", "description": "从公开来源抽取的证据摘录，用于人工复核"},
    ]
    _write_csv(dataset_dir / "structured_current_plans.csv", current_plan_rows, CURRENT_STRUCTURED_FIELDS)
    _write_csv(dataset_dir / "structured_historical_plans.csv", historical_plan_rows, HISTORICAL_STRUCTURED_FIELDS)
    _write_csv(dataset_dir / "structured_source_gaps.csv", gap_rows, SOURCE_GAP_FIELDS)
    _write_csv(dataset_dir / "structured_data_dictionary.csv", dictionary_rows, ["field", "description"])
    _write_workbook(
        dataset_dir / "hkt_product_tariffs_structured.xlsx",
        {
            "当前套餐": (current_plan_rows, CURRENT_STRUCTURED_FIELDS),
            "历史套餐": (historical_plan_rows, HISTORICAL_STRUCTURED_FIELDS),
            "来源缺口": (gap_rows, SOURCE_GAP_FIELDS),
            "字段说明": (dictionary_rows, ["field", "description"]),
        },
    )


def _write_manifest(dataset_dir: Path, current_rows: List[Dict[str, str]], captured_at: str) -> None:
    parsed_count = sum(1 for row in current_rows if row.get("extraction_status") == "parsed")
    historical_rows = _read_csv(dataset_dir / "historical_tariffs.csv")
    historical_parsed = [row for row in historical_rows if str(row.get("extraction_status", "")).startswith("parsed")]
    historical_snapshots = _read_csv(dataset_dir / "historical_source_snapshots.csv")
    structured_gaps = _read_csv(dataset_dir / "structured_source_gaps.csv")
    historical_years = sorted({row.get("archive_year", "") for row in historical_parsed if row.get("archive_year")})
    manifest = {
        "id": "hkt_product_tariffs",
        "title": "HKT 产品资费",
        "summary": "面向用户的 HKT/csl/1O1O/HKT SME/NETVIGATOR 产品资费监测入口，覆盖公开 5G 套餐、企业/SME 移动套餐、家庭宽频、月费、本地数据、漫游数据、合约期、来源页面和变化记录。",
        "source_type": "official_public_product_pages",
        "visibility": "hidden",
        "superseded_by": "competitor_product_tariffs",
        "superseded_note": "作为底层产品子库保留；前端统一显示 competitor_product_tariffs，后端选中该合并库时会自动展开读取本目录。",
        "scope": "HKT/csl/1O1O/HKT SME/NETVIGATOR 公开产品与资费页面；不含登录、个性化报价或非公开渠道资料。",
        "tags": ["HKT", "HKT SME", "csl", "1O1O", "NETVIGATOR", "资费", "套餐", "宽频", "5G", "product", "tariff", "pricing"],
        "keywords": ["HKT", "HKT SME", "csl", "1O1O", "1010", "NETVIGATOR", "Netvigator", "宽频", "家居宽频", "5G", "套餐", "资费", "月费", "本地数据", "漫游数据", "合约期", "产品信息"],
        "entrypoints": [
            "README.md",
            "summary.md",
            "latest_products.csv",
            "latest_products.json",
            "historical_tariffs.csv",
            "historical_tariffs.json",
            "historical_source_snapshots.json",
            "structured_current_plans.csv",
            "structured_historical_plans.csv",
            "structured_source_gaps.csv",
            "structured_data_dictionary.csv",
            "hkt_product_tariffs_structured.xlsx",
            "product_history.csv",
            "product_history.json",
            "source_snapshots.json",
            "change_log.md",
        ],
        "updated_at": captured_at,
        "quality": {
            "status": "official_public_product_pages_parsed",
            "row_count": len(current_rows),
            "parsed_count": parsed_count,
            "historical_row_count": len(historical_rows),
            "historical_parsed_count": len(historical_parsed),
            "historical_snapshot_count": len(historical_snapshots),
            "source_gap_count": len(structured_gaps),
            "historical_years": historical_years,
            "notes": [
                "结构化字段来自公开官方产品页面解析。",
                "历史数据来自 HKT/csl/1O1O 官方公开页面的 Wayback 快照；同一年多快照保留日期级 snapshot_id；PCCW mobile 2010/2011 2G/3G tariff plan 来自 OFCA U003-0004 指向的官方网页快照。",
                "HKT/csl 官方 Internet Access、Business Broadband、Premium Broadband、MegaLink、Datapak/Private Circuit、Metro IP、Flexible Bandwidth、Telecommunications Backup、one communications、eye Service、Home Phone/Local Telephone、2G/3G/4G Mobile、U-plan、Smart Pama、The Club SIM 与 csl/1O1O 2018 postpaid PDF 资费表作为 official_public_tariff_pdf 记录，口径为官方 tariff/list price，不等同促销价；OFCA U003-0004 PDF 本身只作为 linkout/source-gap 保留。",
                "review_required 表示页面已抓取但需要人工确认具体资费口径。",
                "product_history.csv 以套餐特征去重并保留 first_seen/last_seen。",
            ],
        },
        "version": captured_at,
        "built_at": captured_at,
        "row_count": len(current_rows),
    }
    _write_json(dataset_dir / "manifest.json", manifest)


def crawl_hkt_historical_tariffs(
    *,
    output_dir: Path = DATASET_DIR,
    force: bool = False,
) -> Dict[str, Any]:
    captured_at = _now_hkt()
    output_dir.mkdir(parents=True, exist_ok=True)
    historical_csv = output_dir / "historical_tariffs.csv"
    if historical_csv.exists() and not force:
        existing_rows = _read_csv(historical_csv)
        return {
            "ok": True,
            "skipped": True,
            "captured_at_hkt": captured_at,
            "dataset_dir": str(output_dir),
            "historical_count": len(existing_rows),
            "parsed_count": sum(1 for row in existing_rows if str(row.get("extraction_status", "")).startswith("parsed")),
            "snapshots": len(_read_csv(output_dir / "historical_source_snapshots.csv")) if (output_dir / "historical_source_snapshots.csv").exists() else 0,
        }

    historical_rows: List[Dict[str, str]] = []
    snapshot_rows: List[Dict[str, Any]] = []
    with httpx.Client(follow_redirects=True, timeout=httpx.Timeout(25.0, connect=8.0), trust_env=False) as client:
        for snapshot in HISTORICAL_SNAPSHOT_SOURCES:
            result = _fetch_wayback_snapshot(client, snapshot)
            text = str(result.get("text") or "")
            snapshot_rows.append(
                {
                    "snapshot_id": snapshot["snapshot_id"],
                    "brand": snapshot["brand"],
                    "product_category": snapshot["product_category"],
                    "archive_timestamp": snapshot["timestamp"],
                    "archive_year": snapshot["timestamp"][:4],
                    "original_url": snapshot["original_url"],
                    "archive_url": f"https://web.archive.org/web/{snapshot['timestamp']}id_/{snapshot['original_url']}",
                    "status": result.get("status"),
                    "method": result.get("method"),
                    "text_chars": len(text),
                    "title": result.get("title"),
                    "content_hash": _content_hash(text),
                    "error": result.get("error", ""),
                    "captured_at_hkt": captured_at,
                    "evidence_excerpt": _clean_text(text)[:800],
                }
            )
            historical_rows.extend(_parse_historical_snapshot(snapshot, result, captured_at))
        for snapshot in OFFICIAL_TARIFF_DOCUMENT_SOURCES:
            result = _fetch_official_tariff_document(client, snapshot)
            text = str(result.get("text") or "")
            snapshot_rows.append(
                {
                    "snapshot_id": snapshot["snapshot_id"],
                    "brand": snapshot["brand"],
                    "product_category": snapshot["product_category"],
                    "archive_timestamp": snapshot["timestamp"],
                    "archive_year": snapshot["timestamp"][:4],
                    "original_url": snapshot["original_url"],
                    "archive_url": snapshot["original_url"],
                    "status": result.get("status"),
                    "method": result.get("method"),
                    "text_chars": len(text),
                    "title": result.get("title"),
                    "content_hash": _content_hash(text),
                    "error": result.get("error", ""),
                    "captured_at_hkt": captured_at,
                    "evidence_excerpt": _clean_text(text)[:800],
                }
            )
            historical_rows.extend(_parse_historical_snapshot(snapshot, result, captured_at))

    historical_rows = _dedupe(historical_rows)
    historical_rows = sorted(
        historical_rows,
        key=lambda item: (
            item.get("archive_year", ""),
            item.get("brand", ""),
            item.get("product_category", ""),
            item.get("monthly_fee_hkd", ""),
        ),
    )
    _write_csv(output_dir / "historical_tariffs.csv", historical_rows, HISTORICAL_FIELDS)
    _write_json(output_dir / "historical_tariffs.json", historical_rows)
    _write_json(output_dir / "historical_source_snapshots.json", snapshot_rows)
    snapshot_fields = [
        "snapshot_id",
        "brand",
        "product_category",
        "archive_timestamp",
        "archive_year",
        "original_url",
        "archive_url",
        "status",
        "method",
        "text_chars",
        "title",
        "content_hash",
        "error",
        "captured_at_hkt",
        "evidence_excerpt",
    ]
    _write_csv(output_dir / "historical_source_snapshots.csv", snapshot_rows, snapshot_fields)
    status = {
        "ok": True,
        "skipped": False,
        "captured_at_hkt": captured_at,
        "dataset_dir": str(output_dir),
        "historical_count": len(historical_rows),
        "parsed_count": sum(1 for row in historical_rows if str(row.get("extraction_status", "")).startswith("parsed")),
        "snapshots": len(snapshot_rows),
    }
    _write_json(output_dir / "historical_run_status.json", status)
    return status


def crawl_hkt_products(
    *,
    client: Any | None = None,
    fetcher: Callable[[Any, str], Dict[str, Any]] | None = None,
    output_dir: Path = DATASET_DIR,
    include_historical: bool = True,
) -> Dict[str, Any]:
    captured_at = _now_hkt()
    output_dir.mkdir(parents=True, exist_ok=True)
    if fetcher is None:
        fetcher = fetch_product_page

    previous_rows = _read_csv(output_dir / "latest_products.csv")
    snapshots: List[Dict[str, Any]] = []
    current_rows: List[Dict[str, str]] = []
    for source in HKT_PRODUCT_SOURCES:
        result = fetcher(client, source["url"])
        text = str(result.get("text") or "")
        snapshots.append(
            {
                "source_id": source["source_id"],
                "brand": source["brand"],
                "url": source["url"],
                "final_url": result.get("final_url") or result.get("url") or source["url"],
                "status": result.get("status"),
                "method": result.get("method"),
                "content_type": result.get("content_type"),
                "bytes": result.get("bytes"),
                "text_chars": len(text),
                "title": result.get("title"),
                "content_hash": _content_hash(text),
                "error": result.get("error", ""),
                "captured_at_hkt": captured_at,
                "evidence_excerpt": _clean_text(text)[:800],
                "fetch_attempts": result.get("fetch_attempts", []),
            }
        )
        parsed_rows = _parse_source(source, result, captured_at)
        if (
            not any(row.get("extraction_status") == "parsed" for row in parsed_rows)
            and previous_rows
        ):
            fallback_rows = _fallback_previous_success_rows(previous_rows, source["source_id"], captured_at)
            if fallback_rows:
                parsed_rows = fallback_rows
        current_rows.extend(parsed_rows)

    current_rows = _dedupe(current_rows)
    history_rows = _merge_history(output_dir / "product_history.csv", current_rows, captured_at)
    _write_csv(output_dir / "latest_products.csv", current_rows, PRODUCT_FIELDS)
    _write_csv(output_dir / "product_history.csv", history_rows, PRODUCT_FIELDS)
    _write_json(output_dir / "latest_products.json", current_rows)
    _write_json(output_dir / "product_history.json", history_rows)
    _write_json(output_dir / "source_snapshots.json", snapshots)
    _write_change_log(output_dir, previous_rows, current_rows, captured_at)
    if include_historical:
        historical_status = crawl_hkt_historical_tariffs(output_dir=output_dir, force=False)
    else:
        historical_status = {
            "ok": True,
            "skipped": True,
            "skip_reason": "include_historical_false",
            "historical_count": 0,
            "parsed_count": 0,
            "snapshots": 0,
        }
    _write_structured_views(output_dir, current_rows)
    _write_readme(output_dir, current_rows, snapshots, captured_at)
    _write_manifest(output_dir, current_rows, captured_at)
    status = {
        "ok": True,
        "captured_at_hkt": captured_at,
        "dataset_dir": str(output_dir),
        "latest_count": len(current_rows),
        "history_count": len(history_rows),
        "parsed_count": sum(1 for row in current_rows if row.get("extraction_status") == "parsed"),
        "sources": len(snapshots),
        "historical": historical_status,
    }
    _write_json(output_dir / "latest_run_status.json", status)
    return status


def main() -> None:
    result = crawl_hkt_products()
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
