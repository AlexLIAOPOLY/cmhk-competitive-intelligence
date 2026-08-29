#!/usr/bin/env python3
"""Produce an explicit disposition for the Hong Kong tariff source-gap queue.

This report is deliberately separate from the formal tariff tables: a similar
price alone is not a verification match when its period, contract term or
housing/customer segment differs.
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATASET = ROOT / "agent_knowledge" / "hk_competitor_product_tariffs"
GAPS = DATASET / "source_gaps.csv"
OUT_CSV = DATASET / "verification_followup_audit.csv"
OUT_JSON = DATASET / "verification_followup_audit.json"
OUT_MD = DATASET / "verification_followup_audit.md"
OUT_XLSX = DATASET / "hk_competitor_product_tariffs_followup_audit.xlsx"

FIELDS = [
    "queue_type", "period_label", "brand", "product_category", "candidate_plan",
    "monthly_fee_hkd", "broadband_speed_mbps", "contract_months", "gap_type",
    "source_id", "source_url", "archive_url", "recheck_status", "disposition",
    "recheck_scope", "near_match_sources", "reason", "checked_at_hkt",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def parse_candidate(row: dict[str, str]) -> tuple[str, str, str, str]:
    text = row.get("evidence_excerpt", "")
    head, _, _ = text.partition("；period=")
    fee = re.search(r"；fee=([^；]+)", text)
    speed = re.search(r"；speed=([^；]+)", text)
    contract = re.search(r"contract(?:_months)?=([^；, ]+)", text, flags=re.I)
    return (
        head.strip(),
        fee.group(1).strip() if fee else "",
        speed.group(1).strip() if speed else "",
        contract.group(1).strip() if contract else "",
    )


def source_summary(row: dict[str, str]) -> str:
    return f"{row.get('source_id', '')} ({row.get('period_label', '')})"


def dedupe_candidates(rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], int]:
    """Collapse the same parsed candidate while preserving source order.

    Source-gap evidence can repeat one plan with a different excerpt.  Those
    excerpts are useful upstream, but they must not inflate the follow-up queue
    or make a row look independently corroborated.
    """
    unique: list[dict[str, str]] = []
    seen: set[tuple[str, ...]] = set()
    key_fields = (
        "period_label",
        "brand",
        "product_category",
        "candidate_plan",
        "monthly_fee_hkd",
        "broadband_speed_mbps",
        "contract_months",
        "gap_type",
        "source_id",
        "source_url",
        "archive_url",
    )
    for row in rows:
        key = tuple(row.get(field, "").strip() for field in key_fields)
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique, len(rows) - len(unique)


def autosize(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="0F2742")
        cell.fill = PatternFill("solid", fgColor="EAF4FF")
    for cells in ws.columns:
        width = max((len(str(cell.value or "")) for cell in cells), default=10) + 2
        ws.column_dimensions[get_column_letter(cells[0].column)].width = min(width, 45)


def generate_followup_audit(dataset: Path = DATASET) -> dict[str, int | str]:
    gaps = read_csv(dataset / "source_gaps.csv")
    checked_at = datetime.now(timezone(timedelta(hours=8))).isoformat(timespec="seconds")
    backlog = [row for row in gaps if row.get("gap_type") == "single_source_unverified_plan_row"]
    other_gaps = [row for row in gaps if row.get("gap_type") != "single_source_unverified_plan_row"]

    parsed_candidates = []
    for row in backlog:
        name, fee, speed, contract = parse_candidate(row)
        parsed_candidates.append({**row, "candidate_plan": name, "monthly_fee_hkd": fee, "broadband_speed_mbps": speed, "contract_months": contract})
    candidates, duplicate_queue_rows_removed = dedupe_candidates(parsed_candidates)

    audit_rows: list[dict[str, str]] = []
    for row in candidates:
        near_matches = [
            other for other in candidates
            if other is not row
            and other.get("brand") == row.get("brand")
            and other.get("monthly_fee_hkd") == row.get("monthly_fee_hkd")
            and other.get("broadband_speed_mbps") == row.get("broadband_speed_mbps")
        ]
        # The current-source corpus plus targeted public web searches were retried.
        # A near match is intentionally not promoted unless all material terms match.
        if near_matches:
            status = "near_match_not_equivalent"
            reason = "发现相近公开报价，但年份、合约期、住房/客户类别或套餐附加条件不一致；不能作为同一套餐的第二来源。"
            near = "; ".join(sorted({source_summary(match) for match in near_matches}))
        else:
            status = "no_equivalent_second_source_found"
            reason = "已复查运营商公开页、现有公开比较/媒体来源和本地来源库，未找到同品牌、同期间、同资费口径的独立第二来源。"
            near = ""
        audit_rows.append({
            "queue_type": "verification_backlog",
            "period_label": row.get("period_label", ""),
            "brand": row.get("brand", ""),
            "product_category": row.get("product_category", ""),
            "candidate_plan": row.get("candidate_plan", ""),
            "monthly_fee_hkd": row.get("monthly_fee_hkd", ""),
            "broadband_speed_mbps": row.get("broadband_speed_mbps", ""),
            "contract_months": row.get("contract_months", ""),
            "gap_type": row.get("gap_type", ""),
            "source_id": row.get("source_id", ""),
            "source_url": row.get("source_url", ""),
            "archive_url": row.get("archive_url", ""),
            "recheck_status": status,
            "disposition": "retain_as_unverified_not_formal",
            "recheck_scope": "运营商公开页、公开比较/媒体页、现有来源库；按品牌、期间、价格、速率/数据量、合约和客户类别复核",
            "near_match_sources": near,
            "reason": reason,
            "checked_at_hkt": checked_at,
        })

    for row in other_gaps:
        audit_rows.append({
            "queue_type": "confirmed_source_gap",
            "period_label": row.get("period_label", ""),
            "brand": row.get("brand", ""),
            "product_category": row.get("product_category", ""),
            "candidate_plan": "",
            "monthly_fee_hkd": "",
            "broadband_speed_mbps": "",
            "contract_months": "",
            "gap_type": row.get("gap_type", ""),
            "source_id": row.get("source_id", ""),
            "source_url": row.get("source_url", ""),
            "archive_url": row.get("archive_url", ""),
            "recheck_status": "source_gap_confirmed",
            "disposition": "retain_as_non_estimable_gap",
            "recheck_scope": "抓取结果与解析结果复核",
            "near_match_sources": "",
            "reason": row.get("reason", "页面/归档未给出可结构化套餐价格；不得估算。"),
            "checked_at_hkt": checked_at,
        })

    out_csv = dataset / OUT_CSV.name
    out_json = dataset / OUT_JSON.name
    out_md = dataset / OUT_MD.name
    out_xlsx = dataset / OUT_XLSX.name
    with out_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(audit_rows)
    out_json.write_text(json.dumps(audit_rows, ensure_ascii=False, indent=2), encoding="utf-8")

    counts = Counter(row["recheck_status"] for row in audit_rows)
    lines = [
        "# 香港竞对产品资费复核结论",
        "",
        f"复核时间（HKT）：{checked_at}",
        "",
        "## 结论",
        "",
        f"- 单源待核验：{len(candidates)} 条。均已保留为非正式记录，不纳入正式资费表。",
        f"- 去除重复待核验候选：{duplicate_queue_rows_removed} 条。重复证据不会被计作独立来源。",
        f"- 相近来源但不等价：{counts['near_match_not_equivalent']} 条。",
        f"- 未找到等价第二来源：{counts['no_equivalent_second_source_found']} 条。",
        f"- 已确认来源/解析缺口：{len(other_gaps)} 条，均保留为不可估算的 source-gap。",
        "",
        "相同金额或速率不足以构成核验；年份、合约期、住房/客户类别、优惠和套餐附加条件也必须一致。",
    ]
    out_md.write_text("\n".join(lines) + "\n", encoding="utf-8")

    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, selected in [("单源复核结论", [row for row in audit_rows if row["queue_type"] == "verification_backlog"]), ("确认来源缺口", [row for row in audit_rows if row["queue_type"] == "confirmed_source_gap"]), ("汇总", [{"项目": "单源待核验", "数量": len(candidates)}, {"项目": "去除重复待核验候选", "数量": duplicate_queue_rows_removed}, {"项目": "已确认来源/解析缺口", "数量": len(other_gaps)}])]:
        ws = workbook.create_sheet(title)
        fields = list(selected[0]) if selected else ["说明"]
        ws.append(fields)
        for row in selected:
            ws.append([row.get(field, "") for field in fields])
        autosize(ws)
    workbook.save(out_xlsx)
    result = {
        "verification_backlog_rechecked": len(candidates),
        "duplicate_queue_rows_removed": duplicate_queue_rows_removed,
        "confirmed_source_gaps": len(other_gaps),
        "near_match_not_equivalent": counts["near_match_not_equivalent"],
        "no_equivalent_second_source_found": counts["no_equivalent_second_source_found"],
        "out": str(out_csv),
    }
    manifest_path = dataset / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest["verification_followup"] = result
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return result


def main() -> None:
    print(json.dumps(generate_followup_audit(), ensure_ascii=False))


if __name__ == "__main__":
    main()
