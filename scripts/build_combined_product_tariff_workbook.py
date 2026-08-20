#!/usr/bin/env python3
"""Build one review-ready workbook for every Hong Kong product tariff subpackage."""

from __future__ import annotations

import csv
from collections import Counter
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
HKT = ROOT / "agent_knowledge" / "hkt_product_tariffs"
OTHERS = ROOT / "agent_knowledge" / "hk_competitor_product_tariffs"
OUTPUT = ROOT / "agent_knowledge" / "competitor_product_tariffs" / "香港竞对产品资费_完整人读版.xlsx"
AGENT_FORMAL_CSV = OUTPUT.parent / "product_tariffs_formal_agent_records.csv"
AGENT_GAPS_CSV = OUTPUT.parent / "product_tariffs_source_gaps_agent_records.csv"
AGENT_FOLLOWUP_CSV = OUTPUT.parent / "product_tariffs_followup_agent_records.csv"
AGENT_CONTEXT_MD = OUTPUT.parent / "product_tariffs_agent_context.md"

PLAN_FIELDS = [
    "数据子库", "时间类型", "期间", "期间月份", "抓取/生效时间", "品牌", "产品类别", "网络代际", "客户分段", "产品系列", "套餐名称",
    "月费_HKD", "平均月费_HKD", "公开价格_HKD", "计价单位", "本地数据_GB", "漫游数据_GB", "宽频速度_Mbps", "限速_Mbps",
    "合约月数", "本地语音", "附加收费_HKD", "资费类型", "来源状态", "核验次数", "核验状态", "来源ID", "快照ID", "来源URL", "归档URL", "证据摘录", "记录键",
]
GAP_FIELDS = [
    "数据子库", "期间", "品牌", "产品类别", "缺口类型", "HTTP状态", "来源ID", "快照ID", "来源URL", "归档URL", "缺口原因", "证据摘录",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def hkt_plan_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for kind, name in [("当前", "structured_current_plans.csv"), ("历史", "structured_historical_plans.csv")]:
        for row in read_csv(HKT / name):
            rows.append({
                "数据子库": "HKT / csl / 1O1O / NETVIGATOR",
                "时间类型": kind,
                "期间": row.get("archive_year", "当前") if kind == "历史" else "当前",
                "期间月份": row.get("archive_month", "") if kind == "历史" else "",
                "抓取/生效时间": row.get("as_of_hkt", ""),
                "品牌": row.get("brand", ""),
                "产品类别": row.get("plan_family", ""),
                "网络代际": row.get("service_generation", ""),
                "客户分段": row.get("customer_segment", ""),
                "产品系列": row.get("plan_family", ""),
                "套餐名称": row.get("plan_name", ""),
                "月费_HKD": row.get("monthly_fee_hkd", ""),
                "平均月费_HKD": "",
                "公开价格_HKD": row.get("published_price_hkd", ""),
                "计价单位": row.get("price_billing_unit", ""),
                "本地数据_GB": row.get("local_data_gb", ""),
                "漫游数据_GB": row.get("roaming_data_gb", ""),
                "宽频速度_Mbps": "",
                "限速_Mbps": row.get("post_fup_speed_mbps", ""),
                "合约月数": row.get("contract_months", ""),
                "本地语音": row.get("local_voice", ""),
                "附加收费_HKD": row.get("add_on_charges_hkd", ""),
                "资费类型": "official_tariff_or_public_plan",
                "来源状态": row.get("source_status", ""),
                "核验次数": "",
                "核验状态": "official_public_source_structured",
                "来源ID": row.get("source_id", ""),
                "快照ID": row.get("snapshot_id", ""),
                "来源URL": row.get("source_url", ""),
                "归档URL": row.get("archive_url", ""),
                "证据摘录": row.get("evidence_excerpt", ""),
                "记录键": row.get("record_key", ""),
            })
    return rows


def other_plan_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for kind, name in [("当前", "current_plans.csv"), ("历史", "historical_plans.csv")]:
        for row in read_csv(OTHERS / name):
            rows.append({
                "数据子库": "3HK / SmarTone / HKBN / HGC / i-CABLE",
                "时间类型": kind,
                "期间": row.get("period_label", ""),
                "期间月份": "",
                "抓取/生效时间": row.get("captured_at_hkt", ""),
                "品牌": row.get("brand", ""),
                "产品类别": row.get("product_category", ""),
                "网络代际": row.get("service_generation", ""),
                "客户分段": row.get("customer_segment", ""),
                "产品系列": row.get("plan_family", ""),
                "套餐名称": row.get("plan_name", ""),
                "月费_HKD": row.get("monthly_fee_hkd", ""),
                "平均月费_HKD": row.get("average_monthly_fee_hkd", ""),
                "公开价格_HKD": "",
                "计价单位": "",
                "本地数据_GB": row.get("local_data_gb", ""),
                "漫游数据_GB": row.get("roaming_data_gb", ""),
                "宽频速度_Mbps": row.get("broadband_speed_mbps", ""),
                "限速_Mbps": row.get("post_fup_speed_mbps", ""),
                "合约月数": row.get("contract_months", ""),
                "本地语音": row.get("local_voice", ""),
                "附加收费_HKD": row.get("add_on_charges_hkd", ""),
                "资费类型": row.get("tariff_type", ""),
                "来源状态": row.get("source_status", ""),
                "核验次数": row.get("verification_count", ""),
                "核验状态": row.get("verification_status", ""),
                "来源ID": row.get("source_id", ""),
                "快照ID": "",
                "来源URL": row.get("source_url", ""),
                "归档URL": row.get("archive_url", ""),
                "证据摘录": row.get("evidence_excerpt", ""),
                "记录键": row.get("record_key", ""),
            })
    return rows


def gap_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in read_csv(HKT / "structured_source_gaps.csv"):
        rows.append({
            "数据子库": "HKT / csl / 1O1O / NETVIGATOR", "期间": row.get("gap_year", ""), "品牌": row.get("brand", ""),
            "产品类别": row.get("product_category", ""), "缺口类型": row.get("gap_type", ""), "HTTP状态": row.get("http_status", ""),
            "来源ID": row.get("source_id", ""), "快照ID": row.get("snapshot_id", ""), "来源URL": row.get("source_url", ""),
            "归档URL": row.get("archive_url", ""), "缺口原因": row.get("reason", ""), "证据摘录": row.get("evidence_excerpt", ""),
        })
    for row in read_csv(OTHERS / "source_gaps.csv"):
        rows.append({
            "数据子库": "3HK / SmarTone / HKBN / HGC / i-CABLE", "期间": row.get("period_label", ""), "品牌": row.get("brand", ""),
            "产品类别": row.get("product_category", ""), "缺口类型": row.get("gap_type", ""), "HTTP状态": row.get("http_status", ""),
            "来源ID": row.get("source_id", ""), "快照ID": "", "来源URL": row.get("source_url", ""),
            "归档URL": row.get("archive_url", ""), "缺口原因": row.get("reason", ""), "证据摘录": row.get("evidence_excerpt", ""),
        })
    return rows


def style_sheet(ws, widths: dict[str, int] | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="EAF4FF")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="0F2742")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, cell in enumerate(ws[1], 1):
        width = (widths or {}).get(cell.value, 18)
        ws.column_dimensions[get_column_letter(index)].width = width


def write_table(wb: Workbook, title: str, fields: list[str], rows: list[dict[str, str]], widths: dict[str, int] | None = None) -> None:
    ws = wb.create_sheet(title)
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    style_sheet(ws, widths)
    for url_field in ("来源URL", "归档URL", "source_url", "archive_url"):
        if url_field not in fields:
            continue
        column = fields.index(url_field) + 1
        for cell in next(ws.iter_cols(min_col=column, max_col=column, min_row=2), ()):
            value = cell.value
            if isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.style = "Hyperlink"


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def build() -> dict[str, int | str]:
    plans = [*hkt_plan_rows(), *other_plan_rows()]
    gaps = gap_rows()
    followups = read_csv(OTHERS / "verification_followup_audit.csv")
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    agent_plans = [{"record_class": "formal_product_tariff", **row} for row in plans]
    agent_gaps = [{"record_class": "source_gap", **row} for row in gaps]
    agent_followups = [{"record_class": "source_gap_followup", **row} for row in followups]
    write_csv(AGENT_FORMAL_CSV, ["record_class", *PLAN_FIELDS], agent_plans)
    write_csv(AGENT_GAPS_CSV, ["record_class", *GAP_FIELDS], agent_gaps)
    write_csv(AGENT_FOLLOWUP_CSV, ["record_class", *list(followups[0])], agent_followups)
    AGENT_CONTEXT_MD.write_text(
        "\n".join(
            [
                "# 香港竞对产品资费 Agent 读取说明",
                "",
                "## 数据用途",
                "",
                "本数据集用于查询和比较香港运营商公开产品、套餐、月费、数据量、宽频速率、合约期与历史资费变化。它不是公司财务、用户数或季度经营指标数据集，不能替代财务趋势预测数据。",
                "",
                "## Agent 读取顺序",
                "",
                "1. 先读本文件，确认口径和边界。",
                "2. 价格、套餐和历史比较只读取 `product_tariffs_formal_agent_records.csv`。其中每行 `record_class=formal_product_tariff`。",
                "3. 用户问缺口、完整性、来源可得性或为何没有某套餐时，读取 `product_tariffs_source_gaps_agent_records.csv`。其中 `record_class=source_gap`，不得将其作为价格事实、趋势样本或预测输入。",
                "4. 用户问单源候选是否复核、为何没有转为正式套餐时，读取 `product_tariffs_followup_agent_records.csv`；必须引用 `recheck_status`、`disposition` 和 `reason`。",
                "5. 需要追溯时，优先给出同一行的 `来源URL` / `归档URL`、`来源ID`、`快照ID` 和 `证据摘录`。",
                "",
                "## 正式口径",
                "",
                "- HKT/csl/1O1O/NETVIGATOR 行：`核验状态=official_public_source_structured`，表示由官方公开页面、价目表或官方归档结构化；不应伪称每行已有两个独立来源。",
                "- 3HK/SmarTone/HKBN/HGC/i-CABLE 行：正式表仅包含 `核验状态=multi_source_or_multi_snapshot_verified` 的记录。",
                "- `月费_HKD` 是月度套餐费。`公开价格_HKD` 与 `计价单位` 用于日费或其他不能转写为月费的公开价格，严禁自行换算成月费。",
                "- 套餐互比必须同时核对品牌、期间、产品类别、客户分段、合约期、数据量/宽频速率及附加条件。相同价格不代表同一套餐。",
                "",
                "## 覆盖与限制",
                "",
                f"- 当前正式套餐：{len(plans)} 条；来源缺口：{len(gaps)} 条；缺口复核结论：{len(followups)} 条。",
                "- HKT 历史可用范围覆盖 2007-2026；其他品牌和产品线的年份覆盖不同，不能声称每个品牌已有完整十年连续资费。",
                "- 任何 source-gap、单源候选、不同合约期/客户分段或价格口径冲突的记录，均不能用于估算、补数或价格预测。",
            ]
        ) + "\n",
        encoding="utf-8",
    )
    wb = Workbook()
    wb.remove(wb.active)
    summary = wb.create_sheet("说明与统计")
    now = datetime.now(timezone(timedelta(hours=8))).isoformat(timespec="seconds")
    summary_rows = [
        ("工作簿名称", "香港竞对产品资费完整人读版"),
        ("生成时间（HKT）", now),
        ("正式套餐总数", len(plans)),
        ("HKT/csl/1O1O/NETVIGATOR 正式套餐", sum(1 for row in plans if row["数据子库"].startswith("HKT"))),
        ("其他香港竞对正式套餐", sum(1 for row in plans if row["数据子库"].startswith("3HK"))),
        ("全部来源缺口", len(gaps)),
        ("单源复核结论", sum(1 for row in followups if row.get("queue_type") == "verification_backlog")),
        ("已确认来源/解析缺口复核", sum(1 for row in followups if row.get("queue_type") == "confirmed_source_gap")),
        ("使用说明", "正式套餐只含符合各子库正式口径的记录。来源缺口和复核结论不参与价格统计、趋势判断或预测。"),
        ("注释保留规则", "全部保留来源ID、来源/归档URL、快照ID、来源状态、核验状态、证据摘录、缺口原因和二次复核说明。"),
    ]
    summary.append(["项目", "内容"])
    for row in summary_rows:
        summary.append(row)
    style_sheet(summary, {"项目": 30, "内容": 110})

    widths = {"套餐名称": 42, "证据摘录": 80, "来源URL": 55, "归档URL": 55, "记录键": 42, "产品类别": 28, "来源ID": 42, "快照ID": 28}
    write_table(wb, "全部正式套餐", PLAN_FIELDS, plans, widths)
    write_table(wb, "来源缺口", GAP_FIELDS, gaps, {"缺口原因": 70, "证据摘录": 80, "来源URL": 55, "归档URL": 55, "来源ID": 42})
    followup_fields = list(followups[0]) if followups else ["说明"]
    write_table(wb, "缺口复核结论", followup_fields, followups, {"candidate_plan": 42, "reason": 70, "recheck_scope": 60, "near_match_sources": 50, "source_url": 55, "archive_url": 55, "source_id": 42})

    sources: dict[tuple[str, str], dict[str, str]] = {}
    for row in [*plans, *gaps]:
        source_id = row.get("来源ID", "")
        source_url = row.get("来源URL", "")
        if source_id or source_url:
            sources[(source_id, source_url)] = {"来源ID": source_id, "来源URL": source_url, "数据子库": row.get("数据子库", ""), "来源状态": row.get("来源状态", ""), "示例证据摘录": row.get("证据摘录", "")}
    write_table(wb, "来源索引", ["数据子库", "来源ID", "来源URL", "来源状态", "示例证据摘录"], list(sources.values()), {"来源URL": 60, "来源ID": 42, "示例证据摘录": 75})

    dictionary = [
        {"字段": "月费_HKD", "说明": "月度套餐费；一次性、日费、安装和迁移收费不混入该字段。"},
        {"字段": "公开价格_HKD / 计价单位", "说明": "不适合转写为月费的公开价格，例如 day；不得按月推算。"},
        {"字段": "核验状态", "说明": "其他竞对正式表只保留多来源/多快照验证；HKT 表为已结构化的官方公开来源。"},
        {"字段": "证据摘录", "说明": "来源页面中支持字段的短证据，保留供审核。"},
        {"字段": "来源缺口", "说明": "页面/归档无可结构化资费，或仅单来源且未通过同口径复核；不得估算。"},
        {"字段": "缺口复核结论", "说明": "记录二次检索结果；近似价格不等于同一套餐，需同时匹配期间、合约、客户类别和附加条件。"},
    ]
    write_table(wb, "字段说明", ["字段", "说明"], dictionary, {"字段": 32, "说明": 110})
    wb.save(OUTPUT)
    return {"output": str(OUTPUT), "formal_plans": len(plans), "source_gaps": len(gaps), "followups": len(followups), "sources": len(sources), "agent_context": str(AGENT_CONTEXT_MD)}


if __name__ == "__main__":
    import json
    print(json.dumps(build(), ensure_ascii=False))
