from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CORE_DIR = ROOT / "agent_knowledge" / "core_company_metrics_2026-06-16"
CLOUD_DIR = ROOT / "agent_knowledge" / "cloud_vendor_metrics_2026-06-17"


CORE_METRIC_ORDER = [
    "revenue",
    "revenue_growth_yoy",
    "gross_profit",
    "gross_margin",
    "operating_income",
    "operating_margin",
    "net_income",
    "net_margin",
    "ebitda",
    "ebitda_margin",
    "operating_cash_flow",
    "capital_expenditures",
    "free_cash_flow",
    "cash_and_equivalents",
    "total_assets",
    "total_debt",
]

CLOUD_METRIC_ORDER = [
    "cloud_revenue",
    "proxy_segment_revenue",
    "server_products_cloud_services_revenue",
    "cloud_services_license_support_revenue",
    "cloud_and_license_revenue",
    "revenue_yoy",
    "operating_income",
    "operating_margin",
    "adjusted_ebita",
    "adjusted_ebita_margin",
    "proxy_segment_gross_profit",
    "proxy_segment_gross_margin",
    "cloud_and_license_margin",
    "cloud_including_other_segments_revenue",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def format_value(value: str, unit: str, currency: str = "") -> str:
    if value in {"", "None", "null"}:
        return ""
    if unit == "percent":
        return f"{value}%"
    if currency:
        return f"{value} {currency} million"
    return value


def build_core_human_rows() -> list[dict[str, Any]]:
    rows = read_csv(CORE_DIR / "core_metrics_2023_2025.csv")
    sources_payload = json.loads((CORE_DIR / "sources.json").read_text(encoding="utf-8"))
    sources = sources_payload.get("sources_by_company", sources_payload)
    quality_notes = sources_payload.get("quality_notes_by_company", {})
    source_map: dict[str, str] = {}
    for subject, items in sources.items():
        labels = []
        for item in items:
            if not isinstance(item, dict):
                continue
            if item.get("type") == "broker_view":
                continue
            label = item.get("label") or item.get("type") or "source"
            url = item.get("url") or ""
            labels.append(f"{label}: {url}" if url else label)
        source_map[subject] = "；".join(labels[:3])

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row["subject"], row["metric_key"])
        item = grouped.setdefault(
            key,
            {
                "主体": row["subject"],
                "公司全称": row["legal_name"],
                "股票代码": row["ticker"],
                "指标": row["metric_zh"],
                "指标代码": row["metric_key"],
                "FY2023": "",
                "FY2024": "",
                "FY2025": "",
                "单位": row["unit"],
                "口径说明": "；".join(quality_notes.get(row["subject"], []))
                or "标准化三年财务表；HGC 未填造缺失完整三年财务数据。",
                "主要来源": source_map.get(row["subject"], ""),
            },
        )
        item[f"FY{row['year']}"] = row["value"]
    order = {metric: i for i, metric in enumerate(CORE_METRIC_ORDER)}
    return sorted(grouped.values(), key=lambda r: (r["主体"], order.get(r["指标代码"], 999), r["指标代码"]))


def build_cloud_human_rows() -> list[dict[str, Any]]:
    rows = read_csv(CLOUD_DIR / "cloud_vendor_metrics_2023_2025.csv")
    sources = json.loads((CLOUD_DIR / "sources.json").read_text(encoding="utf-8"))
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row["vendor"], row["metric_key"])
        source_labels = []
        for source_id in row["source_ids"].split(";"):
            source = sources.get(source_id)
            if source:
                source_labels.append(f"{source.get('label')}: {source.get('url')}")
        item = grouped.setdefault(
            key,
            {
                "云厂商/口径": row["vendor"],
                "公司全称": row["legal_name"],
                "股票代码": row["ticker"],
                "指标": row["metric_zh"],
                "指标代码": row["metric_key"],
                "FY2023": "",
                "FY2024": "",
                "FY2025": "",
                "单位": "%" if row["unit"] == "percent" else f"{row['currency']} million".strip(),
                "披露口径": row["disclosure_quality"],
                "口径说明": row["quality_note"],
                "主要来源": "；".join(source_labels[:3]),
            },
        )
        item[f"FY{row['fiscal_year']}"] = format_value(row["value"], row["unit"], row["currency"])
    order = {metric: i for i, metric in enumerate(CLOUD_METRIC_ORDER)}
    return sorted(grouped.values(), key=lambda r: (r["云厂商/口径"], order.get(r["指标代码"], 999), r["指标代码"]))


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], fields: list[str]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 24,
        "B": 34,
        "C": 14,
        "D": 24,
        "E": 24,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 22,
        "K": 56,
        "L": 72,
    }
    for idx, _field in enumerate(fields, 1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(letter, 18)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(path: Path, core_rows: list[dict[str, Any]], cloud_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    core_fields = ["主体", "公司全称", "股票代码", "指标", "FY2023", "FY2024", "FY2025", "单位", "口径说明", "主要来源"]
    cloud_fields = ["云厂商/口径", "公司全称", "股票代码", "指标", "FY2023", "FY2024", "FY2025", "单位", "披露口径", "口径说明", "主要来源"]
    add_sheet(wb, "竞对核心数据", core_rows, core_fields)
    add_sheet(wb, "重点云厂商", cloud_rows, cloud_fields)
    wb.save(path)


def main() -> None:
    core_rows = build_core_human_rows()
    cloud_rows = build_cloud_human_rows()
    core_fields = ["主体", "公司全称", "股票代码", "指标", "FY2023", "FY2024", "FY2025", "单位", "口径说明", "主要来源"]
    cloud_fields = ["云厂商/口径", "公司全称", "股票代码", "指标", "FY2023", "FY2024", "FY2025", "单位", "披露口径", "口径说明", "主要来源"]

    write_csv(CORE_DIR / "core_metrics_human_readable.csv", core_rows, core_fields)
    write_csv(CLOUD_DIR / "cloud_vendor_metrics_human_readable.csv", cloud_rows, cloud_fields)
    write_workbook(ROOT / "agent_knowledge" / "cmhk_competitor_cloud_metrics_human_readable.xlsx", core_rows, cloud_rows)
    print(f"core rows: {len(core_rows)}")
    print(f"cloud rows: {len(cloud_rows)}")
    print("wrote human-readable CSV and XLSX")


if __name__ == "__main__":
    main()
