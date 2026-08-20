from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import subprocess
import sys
import time
from collections import Counter
from contextlib import nullcontext
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List
from urllib.parse import quote, urlparse
from zoneinfo import ZoneInfo

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DATASET_DIR = ROOT / "agent_knowledge" / "hk_competitor_product_tariffs"
SOURCE_REGISTRY_JSON = ROOT / "source_registry.json"
CACHE_DIR = DATASET_DIR / "crawl_cache"
ARCHIVE_TIMEOUT_SECONDS = int(os.environ.get("HK_COMPETITOR_ARCHIVE_TIMEOUT", "30"))
ARCHIVE_CURL_TIMEOUT_SECONDS = int(os.environ.get("HK_COMPETITOR_CURL_TIMEOUT", "45"))
ARCHIVE_AVAILABILITY_TIMEOUT_SECONDS = int(
    os.environ.get("HK_COMPETITOR_AVAILABLE_TIMEOUT", os.environ.get("HK_COMPETITOR_ARCHIVE_TIMEOUT", "12"))
)

CURRENT_SOURCES = [
    {
        "source_id": "3hk_3ree_broadband_20100722_en_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp100722",
        "period_label": "2010",
    },
    {
        "source_id": "3hk_3ree_broadband_20100722_tc_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hthkh.com/tc/media/press.php?prid=%2Fpress%2Fcp100722",
        "period_label": "2010",
    },
    {
        "source_id": "3home_entertainment_super_pack_20160427_group_press",
        "brand": "3HK / Hutchison",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp160427",
        "period_label": "2016",
    },
    {
        "source_id": "3home_entertainment_super_pack_20160427_product_pdf",
        "brand": "3HK / Hutchison",
        "product_category": "home_fibre_broadband",
        "url": "https://web.three.com.hk/pressrelease/20160427e.pdf",
        "period_label": "2016",
    },
    {
        "source_id": "3hk_3gamer_20171207_en_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp171207",
        "period_label": "2017",
    },
    {
        "source_id": "3hk_3gamer_20171207_tc_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.hthkh.com/tc/media/press.php?prid=%2Fpress%2Fcp171207",
        "period_label": "2017",
    },
    {
        "source_id": "3hk_whatsapp_premium_20150506_group_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_value_added_service",
        "url": "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp150506",
        "period_label": "2015",
    },
    {
        "source_id": "3hk_whatsapp_premium_20150506_product_pdf",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_value_added_service",
        "url": "https://web.three.com.hk/pressrelease/20150506e.pdf",
        "period_label": "2015",
    },
    {
        "source_id": "3hk_greater_china_plan_20180720_official_pdf",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_4g",
        "url": "https://web.three.com.hk/files/three/20180720/20180720_HZMB_PressRelease_c_Final.pdf",
        "period_label": "2018",
    },
    {
        "source_id": "3hk_greater_china_plan_20180720_qooah",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_4g",
        "url": "https://qooah.com/2018/07/20/3-hong-kong-zhuhai-macao-bridge-45g/",
        "period_label": "2018",
    },
    {
        "source_id": "3hk_ofca_smartphone_super_plan_20140211",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.ofca.gov.hk/filemanager/ofca/en/share/tariff/apps/U0004-002-FEB2014-R.pdf",
        "period_label": "2014",
    },
    {
        "source_id": "3hk_jolla_smartphone_super_plan_20140812",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp140812",
        "period_label": "2014",
    },
    {
        "source_id": "3hk_ipair_20131127_en_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_value_added_service",
        "url": "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp131127",
        "period_label": "2013",
    },
    {
        "source_id": "3hk_ipair_20131127_tc_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_value_added_service",
        "url": "https://www.hthkh.com/tc/media/press.php?prid=%2Fpress%2Fcp131127",
        "period_label": "2013",
    },
    {
        "source_id": "3hk_samsung_galaxy_s2_20110610_en_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_3g",
        "url": "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp110610",
        "period_label": "2011",
    },
    {
        "source_id": "3hk_samsung_galaxy_s2_20110610_tc_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_3g",
        "url": "https://www.hthkh.com/tc/media/press.php?prid=%2Fpress%2Fcp110610",
        "period_label": "2011",
    },
    {
        "source_id": "3hk_anyplex_htc_20111101_en_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_value_added_service",
        "url": "https://www.hthkh.com/en/media/press.php?prid=%2Fpress%2Fp111101",
        "period_label": "2011",
    },
    {
        "source_id": "3hk_anyplex_htc_20111101_tc_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_value_added_service",
        "url": "https://www.hthkh.com/tc/media/press.php?prid=%2Fpress%2Fcp111101",
        "period_label": "2011",
    },
    {
        "source_id": "3hk_skype_unlimited_world_20140819_en_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_value_added_service",
        "url": "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp140819",
        "period_label": "2014",
    },
    {
        "source_id": "3hk_skype_unlimited_world_20140819_tc_official_press",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_value_added_service",
        "url": "https://www.hthkh.com/tc/media/press.php?prid=%2Fpress%2Fcp140819",
        "period_label": "2014",
    },
    {
        "source_id": "3hk_ofca_4g_smartphone_plan_20120503",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.ofca.gov.hk/filemanager/ofca/en/share/tariff/mob_operators/U004-011-May2012-N.pdf",
        "period_label": "2012",
    },
    {
        "source_id": "3hk_ofca_4g_smartphone_plan_20120530",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.ofca.gov.hk/filemanager/ofca/en/share/tariff/mob_operators/U004-013-May2012-R.pdf",
        "period_label": "2012",
    },
    {
        "source_id": "3hk_5g_sim_plan",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_5g",
        "url": "https://web.three.com.hk/plans/5g/index-en.html",
    },
    {
        "source_id": "3hk_5g_sim_plan_tc",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_5g",
        "url": "https://web.three.com.hk/plans/5g/index.html",
    },
    {
        "source_id": "3hk_world_plan",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_roaming",
        "url": "https://web.three.com.hk/3hkworld/index-en.html",
    },
    {
        "source_id": "3hk_world_plan_tc",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_roaming",
        "url": "https://web.three.com.hk/3hkworld/index.html",
    },
    {
        "source_id": "3hk_world_plan_alt",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_roaming",
        "url": "https://web.three.com.hk/3hkworld/index2-en.html",
    },
    {
        "source_id": "3hk_world_plan_alt_tc",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_roaming",
        "url": "https://web.three.com.hk/3hkworld/index2.html",
    },
    {
        "source_id": "3hk_business_5g",
        "brand": "3HK / Hutchison",
        "product_category": "business_mobile_5g",
        "url": "https://web.three.com.hk/plans/3business5g/index-en.html",
    },
    {
        "source_id": "3hk_business_5g_tc",
        "brand": "3HK / Hutchison",
        "product_category": "business_mobile_5g",
        "url": "https://web.three.com.hk/plans/3business5g/index.html",
    },
    {
        "source_id": "3hk_sosim_local",
        "brand": "3HK / Hutchison",
        "product_category": "prepaid_mobile",
        "url": "https://www.sosimhk.com/en/local/data-service.html",
    },
    {
        "source_id": "3hk_sosim_local_tc",
        "brand": "3HK / Hutchison",
        "product_category": "prepaid_mobile",
        "url": "https://www.sosimhk.com/tc/local/data-service.html",
    },
    {
        "source_id": "3hk_promo5g_2024_pdf",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_5g",
        "url": "https://web.three.com.hk/tnc/240625/tnc-promo5g-en.pdf",
        "period_label": "2024",
    },
    {
        "source_id": "3hk_extrabux_2025_product_guide",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.extrabux.com/chs/guide/8516563",
        "period_label": "2025",
    },
    {
        "source_id": "3hk_thriftyhk_2025_mobile_plan_comparison",
        "brand": "3HK / Hutchison",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.thriftyhk.com/post/best-5g-mobile-plan-hongkong",
        "period_label": "2025",
    },
    {
        "source_id": "smartone_5g_listing",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.smartone.com/en/home/mobile-service-plans/5G-listing/",
    },
    {
        "source_id": "smartone_5g_listing_tc",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.smartone.com/tc/home/mobile-service-plans/5G-listing/",
    },
    {
        "source_id": "smartone_subscription_offers",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.smartone.com/en/mobile_and_price_plans/subscription-offers/",
    },
    {
        "source_id": "smartone_5g_110g_30m_239_current",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.smartone.com/en/home/mobile-service-plans/5G-listing/detail/?group=5g_travel&plan=5g_110g_30m_239_travel",
    },
    {
        "source_id": "smartone_5g_110g_24m_299_current",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.smartone.com/en/home/mobile-service-plans/5G-listing/detail/?group=5g_travel&plan=5g_110g_24m_travel",
    },
    {
        "source_id": "smartone_moneysmart_2026_mobile_plan_comparison",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://blog.moneysmart.hk/zh-hk/budgeting/%E6%89%8B%E6%A9%9F-%E4%B8%8A%E5%8F%B0%E5%84%AA%E6%83%A0-5g-%E6%9C%88%E8%B2%BB-%E6%AF%94%E8%BC%83-smartone-csl-3hk-%E4%B8%AD%E7%A7%BB%E5%8B%95-%E8%87%AA%E7%94%B1%E9%B3%A5",
        "period_label": "2026",
    },
    {
        "source_id": "smartone_quoquo_2026_mobile_plan_comparison",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.quoquoapp.com/index.php?id=1493&route=module%2Fapp_news1",
        "period_label": "2026",
    },
    {
        "source_id": "hkbn_moneysmart_broadband_comparison_2026",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://blog.moneysmart.hk/zh-hk/credit-cards/%E5%AF%AC%E9%A0%BB-%E4%B8%8A%E7%B6%B2-%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E6%AF%94%E8%BC%83",
        "period_label": "2026",
    },
    {
        "source_id": "smartone_moneysmart_broadband_comparison_2026",
        "brand": "SmarTone",
        "product_category": "home_fibre_broadband",
        "url": "https://blog.moneysmart.hk/zh-hk/credit-cards/%E5%AF%AC%E9%A0%BB-%E4%B8%8A%E7%B6%B2-%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E6%AF%94%E8%BC%83",
        "period_label": "2026",
    },
    {
        "source_id": "icable_moneysmart_broadband_comparison_2026",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://blog.moneysmart.hk/zh-hk/credit-cards/%E5%AF%AC%E9%A0%BB-%E4%B8%8A%E7%B6%B2-%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E6%AF%94%E8%BC%83",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_moneysmart_broadband_comparison_2026",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://blog.moneysmart.hk/zh-hk/credit-cards/%E5%AF%AC%E9%A0%BB-%E4%B8%8A%E7%B6%B2-%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E6%AF%94%E8%BC%83",
        "period_label": "2026",
    },
    {
        "source_id": "icable_fibrehk_isp_comparison_2026",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://fibrebroadbandprice.com/en/blog/hkbn-vs-hgc-vs-smartone-comparison",
        "period_label": "2026",
    },
    {
        "source_id": "hkbn_yahoo_broadband_comparison_2026",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://hk.news.yahoo.com/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E6%8E%A8%E4%BB%8B-%E5%84%AA%E6%83%A0-%E7%B6%B2%E4%B8%8A%E8%A1%8C%E5%AF%AC%E9%A0%BB%E5%84%AA%E6%83%A0-3-hk-smartone%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E9%A6%99%E6%B8%AF%E5%AF%AC%E9%A0%BB-hgc-broadband-%E5%AF%AC%E9%A0%BB%E6%AF%94%E8%BC%83-pccw-%E5%AF%AC%E9%A0%BB%E5%A0%B1%E5%83%B9-%E4%B8%AD%E5%9C%8B%E7%A7%BB%E5%8B%95%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-034804257.html",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_yahoo_broadband_comparison_2026",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://hk.news.yahoo.com/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E6%8E%A8%E4%BB%8B-%E5%84%AA%E6%83%A0-%E7%B6%B2%E4%B8%8A%E8%A1%8C%E5%AF%AC%E9%A0%BB%E5%84%AA%E6%83%A0-3-hk-smartone%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E9%A6%99%E6%B8%AF%E5%AF%AC%E9%A0%BB-hgc-broadband-%E5%AF%AC%E9%A0%BB%E6%AF%94%E8%BC%83-pccw-%E5%AF%AC%E9%A0%BB%E5%A0%B1%E5%83%B9-%E4%B8%AD%E5%9C%8B%E7%A7%BB%E5%8B%95%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-034804257.html",
        "period_label": "2026",
    },
    {
        "source_id": "smartone_yahoo_broadband_comparison_2026",
        "brand": "SmarTone",
        "product_category": "home_fibre_broadband",
        "url": "https://hk.news.yahoo.com/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E6%8E%A8%E4%BB%8B-%E5%84%AA%E6%83%A0-%E7%B6%B2%E4%B8%8A%E8%A1%8C%E5%AF%AC%E9%A0%BB%E5%84%AA%E6%83%A0-3-hk-smartone%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E9%A6%99%E6%B8%AF%E5%AF%AC%E9%A0%BB-hgc-broadband-%E5%AF%AC%E9%A0%BB%E6%AF%94%E8%BC%83-pccw-%E5%AF%AC%E9%A0%BB%E5%A0%B1%E5%83%B9-%E4%B8%AD%E5%9C%8B%E7%A7%BB%E5%8B%95%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-034804257.html",
        "period_label": "2026",
    },
    {
        "source_id": "icable_yahoo_broadband_comparison_2026",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://hk.news.yahoo.com/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E6%8E%A8%E4%BB%8B-%E5%84%AA%E6%83%A0-%E7%B6%B2%E4%B8%8A%E8%A1%8C%E5%AF%AC%E9%A0%BB%E5%84%AA%E6%83%A0-3-hk-smartone%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E9%A6%99%E6%B8%AF%E5%AF%AC%E9%A0%BB-hgc-broadband-%E5%AF%AC%E9%A0%BB%E6%AF%94%E8%BC%83-pccw-%E5%AF%AC%E9%A0%BB%E5%A0%B1%E5%83%B9-%E4%B8%AD%E5%9C%8B%E7%A7%BB%E5%8B%95%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-034804257.html",
        "period_label": "2026",
    },
    {
        "source_id": "hkbn_shangtaika_broadband_comparison_2026",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://kuan.shangtaika.com/",
        "period_label": "2026",
    },
    {
        "source_id": "smartone_shangtaika_broadband_comparison_2026",
        "brand": "SmarTone",
        "product_category": "home_fibre_broadband",
        "url": "https://kuan.shangtaika.com/",
        "period_label": "2026",
    },
    {
        "source_id": "smartone_quoquo_broadband_comparison_2026",
        "brand": "SmarTone",
        "product_category": "home_fibre_broadband",
        "url": "https://www.quoquoapp.com/index.php?id=1479&route=module%2Fapp_news1",
        "period_label": "2026",
    },
    {
        "source_id": "icable_shangtaika_broadband_comparison_2026",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://kuan.shangtaika.com/",
        "period_label": "2026",
    },
    {
        "source_id": "icable_shangtaika_brand_page_2026",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://kuan.shangtaika.com/icable",
        "period_label": "2026",
    },
    {
        "source_id": "smartone_home_5g_broadband",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartone.com/en/Home5GBroadband/",
    },
    {
        "source_id": "smartone_home_5g_flexi_combo_current",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartone.com/en/mobile_and_price_plans/offer_detail/11-flexi-combo/4483/",
    },
    {
        "source_id": "smartone_home_5g_disney_offer_current",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartone.com/en/mobile_and_price_plans/offer_detail/disney-plus-special-offer/4883/",
    },
    {
        "source_id": "smartone_home_5g_disney_offer_current_tc",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartone.com/tc/mobile_and_price_plans/offer_detail/disney-plus-special-offer/4883/",
    },
    {
        "source_id": "smartone_roaming_pack",
        "brand": "SmarTone",
        "product_category": "roaming_data_pack",
        "url": "https://5g.smartone.com/en/mobile_and_price_plans/roaming/apac_worldwide_roaming_data_pack/charges.jsp",
    },
    {
        "source_id": "smartone_roaming_pack_tc",
        "brand": "SmarTone",
        "product_category": "roaming_data_pack",
        "url": "https://5g.smartone.com/tc/mobile_and_price_plans/roaming/apac_worldwide_roaming_data_pack/charges.jsp",
    },
    {
        "source_id": "smartone_st_protect_2016_pdf",
        "brand": "SmarTone",
        "product_category": "mobile_value_added_service",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2016/06/ST_Protect_press_release_eng.pdf",
        "period_label": "2016",
    },
    {
        "source_id": "smartone_st_protect_2016_pdf_chi",
        "brand": "SmarTone",
        "product_category": "mobile_value_added_service",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2016/06/ST_Protect_press_release_chi.pdf",
        "period_label": "2016",
    },
    {
        "source_id": "smartone_aquos_s2_supercare_2017_pdf",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2017/11/2017_11_23_401.pdf",
        "period_label": "2017",
    },
    {
        "source_id": "smartone_aquos_s2_supercare_2017_pdf_chi",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2017/11/2017_11_23_401_chi.pdf",
        "period_label": "2017",
    },
    {
        "source_id": "smartone_aquos_s2_2017_pcm_report",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.pcmarket.com.hk/%E6%8B%8E%E8%88%8Asharp%E6%89%8B%E6%A9%9F%E5%8F%AF%E4%B8%8A%E5%8F%B0%E5%8D%8A%E5%83%B9%E8%B2%B7aquos-s2/",
        "period_label": "2017",
    },
    {
        "source_id": "smartone_kono_magazine_2018_pdf",
        "brand": "SmarTone",
        "product_category": "mobile_value_added_service",
        "url": "https://www.smartone.com/other/english/en_V123_e.pdf",
        "period_label": "2018",
    },
    {
        "source_id": "smartone_kono_magazine_2019_pdf",
        "brand": "SmarTone",
        "product_category": "mobile_value_added_service",
        "url": "https://www.smartone.com/other/english/tc_V123_e.pdf",
        "period_label": "2019",
    },
    {
        "source_id": "smartone_5g_launch_2020_pdf",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2020/05/2020_05_26_431.pdf",
        "period_label": "2020",
    },
    {
        "source_id": "smartone_5g_launch_2020_pdf_chi",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2020/05/2020_05_26_431_chi.pdf",
        "period_label": "2020",
    },
    {
        "source_id": "smartone_gamergizer_2020_pdf",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2020/07/2020_07_14_432.pdf",
        "period_label": "2020",
    },
    {
        "source_id": "smartone_gamergizer_2020_pdf_chi",
        "brand": "SmarTone",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2020/07/2020_07_14_432_chi.pdf",
        "period_label": "2020",
    },
    {
        "source_id": "smartone_home_5g_launch_2020_pdf",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2020/09/2020_09_14_437.pdf",
        "period_label": "2020",
    },
    {
        "source_id": "smartone_home_5g_2020_ezone_review",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://ezone.hk/article/2753352/SmarTone-5G-%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-148-%E6%9C%89%E5%BE%97%E7%8E%A9-%E5%B8%82%E5%8D%80-%E6%9D%91%E5%B1%8B%E5%AF%A6%E6%B8%AC",
        "period_label": "2020",
    },
    {
        "source_id": "smartone_home_5g_2021_pdf",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2021/09/2021_09_14_448.pdf",
        "period_label": "2021",
    },
    {
        "source_id": "smartone_home_5g_2021_pdf_chi",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2021/09/2021_09_14_448_chi.pdf",
        "period_label": "2021",
    },
    {
        "source_id": "smartone_learning_support_2022_pdf",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2022/04/2022_04_28_460.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "smartone_learning_support_2022_pdf_chi",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2022/04/2022_04_28_460_chi.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "smartone_1c2n_2024_pdf",
        "brand": "SmarTone",
        "product_category": "mobile_value_added_service",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2024/04/2024_04_19_504.pdf",
        "period_label": "2024",
    },
    {
        "source_id": "smartone_1c2n_2024_pdf_chi",
        "brand": "SmarTone",
        "product_category": "mobile_value_added_service",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2024/04/2024_04_19_504_chi.pdf",
        "period_label": "2024",
    },
    {
        "source_id": "smartone_roaming_multiday_2023_pdf",
        "brand": "SmarTone",
        "product_category": "roaming_data_pack",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2023/04/2023_04_01_488.pdf",
        "period_label": "2023",
    },
    {
        "source_id": "smartone_roaming_multiday_2023_linkedin",
        "brand": "SmarTone",
        "product_category": "roaming_data_pack",
        "url": "https://www.linkedin.com/posts/smartone_smartone-5g-roaming-activity-7047852825325817856-TVKF",
        "period_label": "2023",
    },
    {
        "source_id": "smartone_roaming_yas_2023_pdf",
        "brand": "SmarTone",
        "product_category": "roaming_data_pack",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2023/03/2023_03_02_487.pdf",
        "period_label": "2023",
    },
    {
        "source_id": "smartone_roaming_yas_2023_pdf_chi",
        "brand": "SmarTone",
        "product_category": "roaming_data_pack",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2023/03/2023_03_02_487_chi.pdf",
        "period_label": "2023",
    },
    {
        "source_id": "smartone_home_5g_wifi7_2025_pdf",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2025/01/2025_01_15_514.pdf",
        "period_label": "2025",
    },
    {
        "source_id": "smartone_home_5g_wifi7_2025_pdf_chi",
        "brand": "SmarTone",
        "product_category": "home_5g_broadband",
        "url": "https://www.smartoneholdings.com/about/media_centre/press_release/press/2025/01/2025_01_15_514_chi.pdf",
        "period_label": "2025",
    },
    {
        "source_id": "hkbn_home_broadband_offer",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hkbn.net/personal/onlineexclusive/en/select-plan/oexbn",
    },
    {
        "source_id": "hkbn_oexbn_2026_official_api",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hkbn.net/personal/PurchaseFlow/GetAcqPlansByTag/en/oexbn",
        "period_label": "2026",
    },
    {
        "source_id": "hkbn_fhkpuaa_2025_member_offer_pdf",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.fhkpuaa.org.hk/image/data/FHKPUAA%20Cardholders%E2%80%99%20Special%20Offer%20%20Handbill_JAN25_ENG.pdf",
        "period_label": "2025",
    },
    {
        "source_id": "hkbn_hkis_2025_member_offer_pdf",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hkis.org.hk/uploads/editor/MemWelfare/202502_HKIS%20Members%20Special%20Offer%20Handbill_JAN25_ENG.pdf",
        "period_label": "2025",
    },
    {
        "source_id": "hkbn_fhkpuaa_2024_sep_member_offer_pdf",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.fhkpuaa.org.hk/image/data/FHKPUAA%20Cardholders%E2%80%99%20Special%20Offer%20%20Handbill_Sep24_Eng.pdf",
        "period_label": "2024",
    },
    {
        "source_id": "hkbn_pay_tv_bundle_2019_mediaoutreach",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.media-outreach.com/news/hong-kong/2019/05/09/8713/hkbn-launches-mind-blowing-offer-to-all-pay-tv-customers/",
        "period_label": "2019",
    },
    {
        "source_id": "hkbn_pay_tv_bundle_2019_official_pdf",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/20190509-HKBN-Launches-Mind-blowing-Offer-to-All-Pay-TV-Customers-EN-web.pdf",
        "period_label": "2019",
    },
    {
        "source_id": "hkbn_fibre_broadband",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hkbn.net/personal/broadband/en",
    },
    {
        "source_id": "hkbn_5g_home_broadband",
        "brand": "HKBN",
        "product_category": "home_5g_broadband",
        "url": "https://www.hkbn.net/personal/5g-home-broadband/en",
    },
    {
        "source_id": "hkbn_5g_home_broadband_terms_pdf",
        "brand": "HKBN",
        "product_category": "home_5g_broadband",
        "url": "https://www.hkbn.net/personal/cmsdata/content/queryPdfContents/6FRs5S4DySv8i5UozrSTeZQtiZm951BEMaSuWenOo.pdf",
    },
    {
        "source_id": "hkbn_5g_home_broadband_terms_pdf_202405",
        "brand": "HKBN",
        "product_category": "home_5g_broadband",
        "url": "https://images.hkbn.net/apply/orpres/broadband/tc_pdf/MS_5G%20BN_3HK_%24118%20SIM%20Only%20PlanT%26C_%20ENG_202405.pdf",
    },
    {
        "source_id": "hkbn_mobile_launch_2016_prnewswire",
        "brand": "HKBN",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.prnewswire.com/news-releases/hkbn-launches-all-new-mobiles-services-300326838.html",
        "period_label": "2016",
    },
    {
        "source_id": "hkbn_mobile_launch_2016_official_html",
        "brand": "HKBN",
        "product_category": "mobile_consumer_4g",
        "url": "https://reg.hkbn.net/WwwCMS/upload/web/en/Engagement-news-20160913-mobile-services-official-launch-web.html",
        "period_label": "2016",
    },
    {
        "source_id": "hkbn_mobile_trial_2016_pdf",
        "brand": "HKBN",
        "product_category": "mobile_consumer_4g",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/20160804_press_release_HKBNMobile_Services_E_final.pdf",
        "period_label": "2016",
    },
    {
        "source_id": "hkbn_4g_mobile_bundle_2018_pdf",
        "brand": "HKBN",
        "product_category": "mobile_consumer_4g",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/HKBN%20Delivers%20Bang%20for%20the%20Buck%20with%204G%20%2478month%20Mobile%20Bundle_web.pdf",
        "period_label": "2018",
    },
    {
        "source_id": "hkbn_4g_mobile_bundle_2018_pdf_chi",
        "brand": "HKBN",
        "product_category": "mobile_consumer_4g",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/tc/HKBN%20Delivers%20Bang%20for%20the%20Buck%20with%204G%20%2478month%20Mobile%20Bundle_CHI_web.pdf",
        "period_label": "2018",
    },
    {
        "source_id": "hkbn_high_usage_mobile_2017_mediaoutreach",
        "brand": "HKBN",
        "product_category": "mobile_consumer_4g",
        "url": "https://www.media-outreach.com/news/hong-kong/2017/08/30/3859/hkbn-rolls-out-4-5g-full-speed-high-usage-mobile-bundles/",
        "period_label": "2017",
    },
    {
        "source_id": "hkbn_4g_mobile_bundle_2017_official_pdf",
        "brand": "HKBN",
        "product_category": "mobile_consumer_4g",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/HKBN%20Announces%20All-new%20Disruptive%204G%20Mobile%20Services%20Bundle_web.pdf",
        "period_label": "2017",
    },
    {
        "source_id": "hkbn_wifi_concierge_2017_official_html",
        "brand": "HKBN",
        "product_category": "home_telephone_wifi_bundle",
        "url": "https://reg.hkbn.net/WwwCMS/upload/web/en/Engagement-news-201710-Wi-Fi-concierge-web.html",
        "period_label": "2017",
    },
    {
        "source_id": "hkbn_wifi_concierge_2017_official_pdf_en",
        "brand": "HKBN",
        "product_category": "home_telephone_wifi_bundle",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/HKBN-Launches-All-new-Home-Telephone-and-Wi-Fi-Concierge-Service-EN-web.pdf",
        "period_label": "2017",
    },
    {
        "source_id": "hkbn_wifi_concierge_2017_official_pdf_tc",
        "brand": "HKBN",
        "product_category": "home_telephone_wifi_bundle",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/tc/HKBN-Launches-All-new-Home-Telephone-and-Wi-Fi-Concierge-Service-TC-web.pdf",
        "period_label": "2017",
    },
    {
        "source_id": "hkbn_travel_pocket_wifi_2018_pdf",
        "brand": "HKBN",
        "product_category": "roaming_wifi",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/20180305_Press%20release_HKBN%20Travel%20Pocket%20Wi-Fi_Eng_web.pdf",
        "period_label": "2018",
    },
    {
        "source_id": "hkbn_travel_pocket_wifi_2018_official_html",
        "brand": "HKBN",
        "product_category": "roaming_wifi",
        "url": "https://reg.hkbn.net/WwwCMS/upload/web/en/20180305_HKBN_announcements-web.html",
        "period_label": "2018",
    },
    {
        "source_id": "hkbn_momax_smart_home_2020_pdf",
        "brand": "HKBN",
        "product_category": "smart_home_bundle",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/20201218_PressRelease_MOMAXxHKBN_EN_web.pdf",
        "period_label": "2020",
    },
    {
        "source_id": "hkbn_momax_smart_home_2020_pdf_tc",
        "brand": "HKBN",
        "product_category": "smart_home_bundle",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/tc/20201218_PressRelease_MOMAXxHKBN_TC_web.pdf",
        "period_label": "2020",
    },
    {
        "source_id": "hkbn_homeplus_5g_2021_pdf",
        "brand": "HKBN",
        "product_category": "mobile_consumer_5g",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/20210423_PressRelease_5G_Mobile_EN_web.pdf",
        "period_label": "2021",
    },
    {
        "source_id": "hkbn_homeplus_5g_2021_techent_syndication",
        "brand": "HKBN",
        "product_category": "mobile_consumer_5g",
        "url": "https://techent.tv/2021/04/23/hkbn-and-home-join-forces-to-deliver-breakthrough-shopping-rewards-to-5g-mobile-services-customers/",
        "period_label": "2021",
    },
    {
        "source_id": "hkbn_100gb_5g_2022_official_pdf_en",
        "brand": "HKBN",
        "product_category": "mobile_consumer_5g",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/20220210_PressRelease_New_5G_Plan_EN_web.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "hkbn_100gb_5g_2022_official_pdf_tc",
        "brand": "HKBN",
        "product_category": "mobile_consumer_5g",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/tc/20220210_PressRelease_New_5G_Plan_TC_web.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "hkbn_fy22_interim_5g_market_snapshot",
        "brand": "HKBN",
        "product_category": "mobile_consumer_5g",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/FY2022_HKBN_Interim_Results_Presentation.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "hkbn_68_10gb_2022_mobilemagazine",
        "brand": "HKBN",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.mobilemagazinehk.com/2022/02/10gb-5ghk68.html",
        "period_label": "2022",
    },
    {
        "source_id": "icable_hkbn_fy22_fixed_broadband_market_snapshot",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/FY2022_HKBN_Interim_Results_Presentation.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "hgc_hkbn_fy22_fixed_broadband_market_snapshot",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/FY2022_HKBN_Interim_Results_Presentation.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "icable_hkbn_fy22_fixed_broadband_market_snapshot_irasia",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://webcast.irasia.com/hkbn/interim/2022/archived/documents/pre_i.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "icable_mytv_bundle_2021_tvb_press",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://corporate.tvb.com/article/799a5b7839f606a43d71c42f096e50e7.html",
        "period_label": "2021",
    },
    {
        "source_id": "icable_mytv_2021_mytvsuper_service_fee",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://promo.mytvsuper.com/en/service-fee/i-cable",
        "period_label": "2021",
    },
    {
        "source_id": "icable_broadbandqueen_2024_public_offer",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadbandqueen.com/en/post/i-cable-%E5%AF%9B%E9%A0%BB%E5%84%AA%E6%83%A0/",
        "period_label": "2024",
    },
    {
        "source_id": "icable_broadbandqueen_2024_public_offer_tc",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadbandqueen.com/post/i-cable-%E5%AF%9B%E9%A0%BB%E5%84%AA%E6%83%A0/",
        "period_label": "2024",
    },
    {
        "source_id": "hkbn_moneyhero_broadband_comparison_2025",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.moneyhero.com.hk/zh/credit-card/blog/%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2-%E5%83%B9%E9%8C%A2%E6%AF%94%E8%BC%83",
        "period_label": "2025",
    },
    {
        "source_id": "hkbn_broadband_roadshowoffer_2500m_148_2025",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-roadshowoffer.com/blog",
        "period_label": "2025",
    },
    {
        "source_id": "hkbn_findplanking_2500m_149_2026",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://findplanking.com/broadband/-Ma4eTavqkvwIHSFy5_i",
        "period_label": "2026",
    },
    {
        "source_id": "hkbn_pricequote_posts_2500m_149_2026",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-pricequote.com/posts",
        "period_label": "2026",
    },
    {
        "source_id": "icable_moneyhero_broadband_comparison_2025",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.moneyhero.com.hk/zh/credit-card/blog/%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2-%E5%83%B9%E9%8C%A2%E6%AF%94%E8%BC%83",
        "period_label": "2025",
    },
    {
        "source_id": "hgc_moneyhero_broadband_comparison_2025",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.moneyhero.com.hk/zh/credit-card/blog/%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2-%E5%83%B9%E9%8C%A2%E6%AF%94%E8%BC%83",
        "period_label": "2025",
    },
    {
        "source_id": "smartone_moneyhero_broadband_comparison_2025",
        "brand": "SmarTone",
        "product_category": "home_fibre_broadband",
        "url": "https://www.moneyhero.com.hk/zh/credit-card/blog/%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2-%E5%83%B9%E9%8C%A2%E6%AF%94%E8%BC%83",
        "period_label": "2025",
    },
    {
        "source_id": "smartone_investbrother_broadband_comparison_2025",
        "brand": "SmarTone",
        "product_category": "home_fibre_broadband",
        "url": "https://www.investbrother.com/brother-academy/broadband-comparison/",
        "period_label": "2025",
    },
    {
        "source_id": "hkbn_investbrother_broadband_comparison_2025",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.investbrother.com/brother-academy/broadband-comparison/",
        "period_label": "2025",
    },
    {
        "source_id": "hgc_investbrother_broadband_comparison_2025",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.investbrother.com/brother-academy/broadband-comparison/",
        "period_label": "2025",
    },
    {
        "source_id": "icable_investbrother_broadband_comparison_2025",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.investbrother.com/brother-academy/broadband-comparison/",
        "period_label": "2025",
    },
    {
        "source_id": "hkbn_hk01_broadband_comparison_2026",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hk01.com/%E6%95%B8%E7%A2%BC%E7%94%9F%E6%B4%BB/60350148/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2%E6%A0%BC%E5%83%B92026-%E7%B6%B2%E4%B8%8A%E8%A1%8C-hkbn-%E6%9C%89%E7%B7%9A-hgc-%E5%85%AD%E5%A4%A7%E4%BE%9B%E6%87%89%E5%95%86%E8%AA%B0%E6%9C%80%E6%8A%B5",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_hk01_broadband_comparison_2026",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hk01.com/%E6%95%B8%E7%A2%BC%E7%94%9F%E6%B4%BB/60350148/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2%E6%A0%BC%E5%83%B92026-%E7%B6%B2%E4%B8%8A%E8%A1%8C-hkbn-%E6%9C%89%E7%B7%9A-hgc-%E5%85%AD%E5%A4%A7%E4%BE%9B%E6%87%89%E5%95%86%E8%AA%B0%E6%9C%80%E6%8A%B5",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_hk01_public_housing_109_broadband_comparison_2026",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hk01.com/%E6%95%B8%E7%A2%BC%E7%94%9F%E6%B4%BB/60350148/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2%E6%A0%BC%E5%83%B92026-%E7%B6%B2%E4%B8%8A%E8%A1%8C-hkbn-%E6%9C%89%E7%B7%9A-hgc-%E5%85%AD%E5%A4%A7%E4%BE%9B%E6%87%89%E5%95%86%E8%AA%B0%E6%9C%80%E6%8A%B5",
        "period_label": "2026",
    },
    {
        "source_id": "smartone_hk01_broadband_comparison_2026",
        "brand": "SmarTone",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hk01.com/%E6%95%B8%E7%A2%BC%E7%94%9F%E6%B4%BB/60350148/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2%E6%A0%BC%E5%83%B92026-%E7%B6%B2%E4%B8%8A%E8%A1%8C-hkbn-%E6%9C%89%E7%B7%9A-hgc-%E5%85%AD%E5%A4%A7%E4%BE%9B%E6%87%89%E5%95%86%E8%AA%B0%E6%9C%80%E6%8A%B5",
        "period_label": "2026",
    },
    {
        "source_id": "icable_hk01_broadband_comparison_2026",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hk01.com/%E6%95%B8%E7%A2%BC%E7%94%9F%E6%B4%BB/60350148/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2%E6%A0%BC%E5%83%B92026-%E7%B6%B2%E4%B8%8A%E8%A1%8C-hkbn-%E6%9C%89%E7%B7%9A-hgc-%E5%85%AD%E5%A4%A7%E4%BE%9B%E6%87%89%E5%95%86%E8%AA%B0%E6%9C%80%E6%8A%B5",
        "period_label": "2026",
    },
    {
        "source_id": "hkbn_booga_broadband_comparison_2025",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://booga.com.hk/en/blog/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB%E5%84%AA%E6%83%A0",
        "period_label": "2025",
    },
    {
        "source_id": "icable_booga_broadband_comparison_2025",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://booga.com.hk/en/blog/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB%E5%84%AA%E6%83%A0",
        "period_label": "2025",
    },
    {
        "source_id": "icable_booga_2026_current_broadband_offer",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://booga.com.hk/zh-HK/blog/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB/icable",
        "period_label": "2026",
    },
    {
        "source_id": "icable_service_wifi_2026_official",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://service.i-cable.com/tc/wifi",
        "period_label": "2026",
    },
    {
        "source_id": "icable_telcoquo_1000m_89_2026",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://telcoquo.com/broadbandoffers/10243%E6%9C%89%E7%B7%9A%E5%AF%AC%E9%A0%BBi-cable%E6%9C%89%E7%B7%9A%E5%AF%AC%E9%A0%BB%E5%AE%B6%E5%B1%85%E4%B8%8A%E7%B6%B2-489/",
        "period_label": "2026",
    },
    {
        "source_id": "icable_telcoquo_1000m_118_2026",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://telcoquo.com/%E4%BD%8F%E5%AE%85%E5%AF%AC%E9%A0%BB%E5%A0%B1%E5%83%B9%E5%88%86%E4%BA%AB/",
        "period_label": "2026",
    },
    {
        "source_id": "icable_broadband_quote_2026_offer_listing",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-quote.com/%E6%9C%89%E7%B7%9A%E5%AF%AC%E9%A0%BB/%E6%9C%89%E7%B7%9A%E5%AF%AC%E9%A0%BB-%E6%9C%80%E6%96%B0%E5%84%AA%E6%83%A0/",
        "period_label": "2026",
    },
    {
        "source_id": "icable_broadband_pricequote_1000m_68_2023",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-pricequote.com/post/icable1000m-1023",
        "period_label": "2023",
    },
    {
        "source_id": "icable_booga_public_housing_broadband_comparison_2025",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://booga.com.hk/zh-HK/blog/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB/2024%E5%85%AC%E5%B1%8B%E5%B1%85%E5%B1%8B%E6%9C%80%E6%8A%B5%E5%85%89%E7%BA%96%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2%E6%AF%94%E8%BC%83",
        "period_label": "2025",
    },
    {
        "source_id": "hgc_booga_broadband_comparison_2025",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://booga.com.hk/en/blog/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB%E5%84%AA%E6%83%A0",
        "period_label": "2025",
    },
    {
        "source_id": "hkbn_hktechreview_broadband_comparison_2026",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://hktechreview.com/broadband-compare/",
        "period_label": "2026",
    },
    {
        "source_id": "smartone_hktechreview_broadband_comparison_2026",
        "brand": "SmarTone",
        "product_category": "home_fibre_broadband",
        "url": "https://hktechreview.com/broadband-compare/",
        "period_label": "2026",
    },
    {
        "source_id": "icable_hktechreview_broadband_comparison_2026",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://hktechreview.com/broadband-compare/",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_hktechreview_broadband_comparison_2025",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://hktechreview.com/broadband-compare/",
        "period_label": "2025",
    },
    {
        "source_id": "icable_kennechu_2020_home_broadband_guide",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.kennechu.info/2020/03/Broadband-Service-home.html",
        "period_label": "2020",
    },
    {
        "source_id": "hgc_kennechu_2020_home_broadband_guide",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.kennechu.info/2020/03/Broadband-Service-home.html",
        "period_label": "2020",
    },
    {
        "source_id": "icable_hkepc_2019_forum_market_observation",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hkepc.com/forum/viewthread.php?fid=12&page=5&tid=2484516",
        "period_label": "2019",
    },
    {
        "source_id": "icable_discuss_2017_forum_market_observation",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.discuss.com.hk/archiver/?tid-26638103.html=",
        "period_label": "2017",
    },
    {
        "source_id": "icable_appledaily_2017_broadband_market_comparison",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://collection.news/appledaily/articles/HIYXCTNYJS2EW5SIIGINT5GXRA",
        "period_label": "2017",
    },
    {
        "source_id": "hgc_hkbn_fy22_fixed_broadband_market_snapshot_irasia",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://webcast.irasia.com/hkbn/interim/2022/archived/documents/pre_i.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "hgc_ezone_2019_2g_broadband_comparison",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://ezone.hk/article/2245977/%E6%A5%B5%E9%80%9F-2Gbps-%E5%AE%B6%E7%94%A8%E5%AF%AC%E9%A0%BB-%E4%B8%89%E5%A4%A7-ISP-%E6%9C%8D%E5%8B%99%E8%A8%88%E5%8A%83%E6%A0%BC%E5%83%B9",
        "period_label": "2019",
    },
    {
        "source_id": "hgc_smart_home_living_2020_press",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/press-releases/hgc-broadband-launches-smart-home-living-offer-for-building-your-ideal-smart-home",
        "period_label": "2020",
    },
    {
        "source_id": "hgc_smart_home_living_2020_official_pdf",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/assets/images/2020_10_23_Eng.pdf",
        "period_label": "2020",
    },
    {
        "source_id": "hkbn_crossborder_5g_2023_prnewswire",
        "brand": "HKBN",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.prnewswire.com/apac/news-releases/hkbn-launches-cross-border-5g-local--1gb-gba-data-plans-301834440.html",
        "period_label": "2023",
    },
    {
        "source_id": "hkbn_crossborder_5g_2023_official",
        "brand": "HKBN",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.hkbn.net/group/en/newsroom/press-releases/20230525_FY23_HKBN_Launches_Cross-border_5G_Local_1GB_GBA_Data_Plans_Disruptive_Mobile_Service_Plans",
        "period_label": "2023",
    },
    {
        "source_id": "hkbn_iqiyi_vip_2023_pdf",
        "brand": "HKBN",
        "product_category": "ott_addon_bundle",
        "url": "https://reg.hkbn.net/WwwCMS/upload/pdf/en/20230802_PressRelease_iQIYI_HKBN_Eng.pdf",
        "period_label": "2023",
    },
    {
        "source_id": "hkbn_iqiyi_vip_2023_official",
        "brand": "HKBN",
        "product_category": "ott_addon_bundle",
        "url": "https://www.hkbn.net/group/en/newsroom/press-releases/20230802_iQIYI_HKBN",
        "period_label": "2023",
    },
    {
        "source_id": "hkbn_n_mobile_2023_official",
        "brand": "HKBN",
        "product_category": "mobile_travel_lifestyle",
        "url": "https://www.hkbn.net/group/en/newsroom/press-releases/20231206_Nmobile",
        "period_label": "2023",
    },
    {
        "source_id": "hkbn_n_mobile_2023_official_tc",
        "brand": "HKBN",
        "product_category": "mobile_travel_lifestyle",
        "url": "https://www.hkbn.net/group/tc/newsroom/press-releases/20231206_Nmobile",
        "period_label": "2023",
    },
    {
        "source_id": "hkbn_gigafast_tplink_2024_official",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hkbn.net/group/en/newsroom/press-releases/20241128-TPLink",
        "period_label": "2024",
    },
    {
        "source_id": "hkbn_gigafast_tplink_2024_official_tc",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hkbn.net/group/tc/newsroom/press-releases/20241128-TPLink",
        "period_label": "2024",
    },
    {
        "source_id": "hkbn_enterprise_mobile_current",
        "brand": "HKBN",
        "product_category": "business_mobile_5g",
        "url": "https://www.hkbnes.com/web/sme/solutions/broadband/mobile-solutions/",
    },
    {
        "source_id": "hkbn_enterprise_mobile_current_tc",
        "brand": "HKBN",
        "product_category": "business_mobile_5g",
        "url": "https://www.hkbnes.com/web/tc/sme/solutions/broadband/mobile-solutions/",
    },
    {
        "source_id": "hkbn_enterprise_mobile_current_sc",
        "brand": "HKBN",
        "product_category": "business_mobile_5g",
        "url": "https://www.hkbnes.com/web/sc/sme/solutions/broadband/mobile-solutions/",
    },
    {
        "source_id": "hkbn_enterprise_mobile_4g_offer",
        "brand": "HKBN",
        "product_category": "business_mobile_4g",
        "url": "https://www.hkbnes.net/form/ec-quad-offer-en.jsp",
    },
    {
        "source_id": "hkbn_enterprise_mobile_4g_offer_tc",
        "brand": "HKBN",
        "product_category": "business_mobile_4g",
        "url": "https://www.hkbnes.net/form/ec-quad-offer-tc.jsp",
    },
    {
        "source_id": "hkbn_enterprise_mobile_4g_offer_sc",
        "brand": "HKBN",
        "product_category": "business_mobile_4g",
        "url": "https://www.hkbnes.net/form/ec-quad-offer-sc.jsp",
    },
    {
        "source_id": "hgc_on_air_wifi",
        "brand": "HGC",
        "product_category": "wifi_pass",
        "url": "https://hub.hgc.com.hk/HGConAir/hgconairOnlineEN.html",
    },
    {
        "source_id": "hgc_on_air_plan_select",
        "brand": "HGC",
        "product_category": "wifi_pass",
        "url": "https://hub.hgc.com.hk/HGConAir/Plan/select.do?planSelect=",
    },
    {
        "source_id": "hgc_on_air_plan_en",
        "brand": "HGC",
        "product_category": "wifi_pass",
        "url": "https://hub.hgc.com.hk/HGConAir/Plan.do?language=EN&ClearS=1",
    },
    {
        "source_id": "hgc_on_air_plan_tc",
        "brand": "HGC",
        "product_category": "wifi_pass",
        "url": "https://hub.hgc.com.hk/HGConAir/Plan.do?language=TC&ClearS=1",
    },
    {
        "source_id": "hgc_2g_broadband_2023_press",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgcbroadband.com/en/broadband/fibre-to-home",
    },
    {
        "source_id": "hgc_home_broadband_standard_monthly_2026",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgcbroadband.com/en/broadband/fibre-to-home",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_home_broadband_standard_monthly_2026_tc",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgcbroadband.com/tc/broadband/fibre-to-home",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_home_broadband_terms_standard_monthly_2026",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgcbroadband.com/en/pages/terms-conditions",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_home_broadband_terms_standard_monthly_2026_tc",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgcbroadband.com/tc/pages/terms-conditions",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_broadband_pricequote_2500m_2025",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-pricequote.com/post/hgc2500m042025",
        "period_label": "2025",
    },
    {
        "source_id": "hgc_broadband_pricequote_category_2500m_2025",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-pricequote.com/broadband-plan/categories/hgc",
        "period_label": "2025",
    },
    {
        "source_id": "hgc_broadband_pricequote_2000m_x50poe_2025",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-pricequote.com/post/hgcx50poe-0725",
        "period_label": "2025",
    },
    {
        "source_id": "hgc_broadband_pricequote_2200m_2026",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-pricequote.com/post/hgc2200m",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_broadband_quote_2026_offer_listing",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-quote.com/hgc/hgc%E6%9C%80%E6%96%B0%E5%84%AA%E6%83%A0/",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_mobile_launch_2026_press",
        "brand": "HGC",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.hgc.com.hk/press-releases/hgc-announces-the-launch-of-hgc-mobile-expanding-mobile-connectivity-footprint-with-enhanced-network-on-the-go-experience",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_mobile_launch_2026_etnet",
        "brand": "HGC",
        "product_category": "mobile_consumer_5g",
        "url": "https://www.etnet.com.hk/www/tc/news/news-article.php?category=mediaoutreach&newsid=459236&section=index",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_mobile_launch_2026_telecomramblings",
        "brand": "HGC",
        "product_category": "mobile_consumer_5g",
        "url": "https://newswire.telecomramblings.com/2026/04/hgc-announces-the-launch-of-hgc-mobile-expanding-mobile-connectivity-footprint-with-enhanced-network-on-the-go-experience/",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_wifi6_2022_press",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://enmobile.prnasia.com/releases/apac/hgc-broadband-offers-higher-speed-internet-enriched-service-experiences-and-waives-installation-fees-to-grab-opportunities-in-the-work-from-home-economy-352523.shtml",
        "period_label": "2022",
    },
    {
        "source_id": "hgc_wifi6_router_2022_prnasia",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://enmobile.prnasia.com/releases/apac/hgc-broadband-launches-wi-fi-6-router-service-for-hong-kong-households-355741.shtml",
        "period_label": "2022",
    },
    {
        "source_id": "hgc_wifi6_router_2022_official",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/press-releases/hgc-%E7%92%B0%E9%9B%BB%E5%AF%AC%E9%A0%BB%E6%8E%A8%E5%87%BA%E5%85%A8%E6%96%B0%E5%AE%B6%E5%B1%85-wifi-6%E6%9C%8D%E5%8B%99",
        "period_label": "2022",
    },
    {
        "source_id": "hgc_wifi6_router_2022_official_pdf",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/corp/assets/2022_03_24_EN.pdf",
        "period_label": "2022",
    },
    {
        "source_id": "hgc_25g_2023_official",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/press-releases/hgc-%E7%92%B0%E9%9B%BB%E5%AF%AC%E9%A0%BB%E6%8E%A8%E5%87%BA2-5g-2500mbps%E5%85%89%E7%BA%96%E5%AF%AC%E9%A0%BB%E6%9C%8D%E5%8B%99",
        "period_label": "2023",
    },
    {
        "source_id": "hgc_25g_2023_official_pdf",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/assets/images/2023_01_19_EN.pdf",
        "period_label": "2023",
    },
    {
        "source_id": "hgc_2g_2023_official_pdf",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/assets/images/For-immediate-release-HGC-Broadband-Launches-2G-2000Mbps-Broadband-Service.pdf",
        "period_label": "2023",
    },
    {
        "source_id": "hgc_2g_2023_telecomreviewamericas",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.telecomreviewamericas.com/articles/wholesale-and-capacity/hgc-broadband-launches-superfast-2-gbps-broadband-service/",
        "period_label": "2023",
    },
    {
        "source_id": "hgc_10g_25g_2024_official",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/press-releases/hgc-diversified-broadband-service-plans-lead-new-era-of-ultra-fast-internet-access",
        "period_label": "2024",
    },
    {
        "source_id": "hgc_10g_25g_2024_prnewswire",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.prnewswire.com/apac/news-releases/hgc-diversified-broadband-service-plans-lead-new-era-of-ultra-fast-internet-access-302235794.html",
        "period_label": "2024",
    },
    {
        "source_id": "hgc_findplanking_2026_offer_listing",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://findplanking.com/broadband/1134",
        "period_label": "2026",
    },
    {
        "source_id": "hgc_mytv_2025_broadband_pricequote",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-pricequote.com/post/%E3%80%90hgc%E5%85%89%E7%BA%8E%E5%AF%9B%E9%A0%BB%E9%80%A3mytv-gold%E3%80%91%E6%9C%88%E8%B2%BB-%EF%BC%91%EF%BC%99%EF%BC%98-%E3%80%90mytv-super%E5%9F%BA%E6%9C%AC%E7%89%88%E3%80%91%E6%9C%88%E8%B2%BB-%EF%BC%91%EF%BC%90%EF%BC%99",
        "period_label": "2025",
    },
    {
        "source_id": "hgc_mytv_2025_broadband_pricequote_post",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.broadband-pricequote.com/posts/%E3%80%90hgc%E5%85%89%E7%BA%8E%E5%AF%9B%E9%A0%BB%E9%80%A3mytv-gold%E3%80%91%E6%9C%88%E8%B2%BB%EF%B9%A9%EF%BC%91%EF%BC%99%EF%BC%98%2F%E3%80%90mytv-super%E5%9F%BA%E6%9C%AC%E7%89%88%E3%80%91%E6%9C%88%E8%B2%BB%EF%B9%A9%EF%BC%91%EF%BC%90%EF%BC%99",
        "period_label": "2025",
    },
    {
        "source_id": "hgc_mytv_2025_mytvsuper_service_fee",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://promo.mytvsuper.com/tc/service-fee/hgc",
        "period_label": "2025",
    },
    {
        "source_id": "hgc_mytv_1g_2021_press",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/press-releases/mytv-super-hgc-broadband-1g-home-broadband-mytv-gold-concurrent-viewing-on-3-devices-198-mth",
        "period_label": "2021",
    },
    {
        "source_id": "hgc_mytv_1g_2021_official_pdf",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/assets/images/2021_0426_ENG.pdf",
        "period_label": "2021",
    },
    {
        "source_id": "hgc_line_for_four_2016_press",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://en.prnasia.com/releases/apac/HGC_Launches_New_Line_for_Four_Service_Customers_Can_Now_Enjoy_High_Speed_Broadband_NowTV_Telephone_Line_and_UniFi_Service_on_One_Line_-162077.shtml",
        "period_label": "2016",
    },
    {
        "source_id": "hgc_super_fun_2016_official_press",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hthkh.com/tc/ir/press.php?prid=/press/cp160314",
        "period_label": "2016",
    },
    {
        "source_id": "hgc_super_fun_2016_official_press_en",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hthkh.com/en/ir/press.php?prid=/press/cp160314",
        "period_label": "2016",
    },
    {
        "source_id": "hgc_broadband_pro_2017_referral_case",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://broadband-pro.weebly.com/case2.html",
        "period_label": "2017",
    },
    {
        "source_id": "hgc_smart_home_2018_press",
        "brand": "HGC",
        "product_category": "smart_home_bundle",
        "url": "https://www.hgc.com.hk/press-releases/hgc-broadband-offers-smart-home-entertainment-privilege/",
        "period_label": "2018",
    },
    {
        "source_id": "hgc_line_for_four_2018_official",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/press-releases/hgc-broadband-launches-line-for-four-2-2g-fibre-broadband-and-wi-fi-360-services",
        "period_label": "2018",
    },
    {
        "source_id": "hgc_line_for_four_2018_official_tc",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/zh/press-releases/hgc-broadband-launches-line-for-four-2-2g-fibre-broadband-and-wi-fi-360-services",
        "period_label": "2018",
    },
    {
        "source_id": "hgc_line_for_four_2018_official_pdf",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hgc.com.hk/corp/assets/documents/press/English/2018-05-28_ENG.pdf",
        "period_label": "2018",
    },
    {
        "source_id": "hgc_line_for_four_2018_prnasia",
        "brand": "HGC",
        "product_category": "home_fibre_broadband",
        "url": "https://en.prnasia.com/releases/apac/hgc-broadband-launches-line-for-four-2-2g-fibre-broadband-and-wi-fi-360-services-212121.shtml",
        "period_label": "2018",
    },
    {
        "source_id": "icable_broadband_offer",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.i-cablebroadband-offer.com/",
    },
    {
        "source_id": "icable_broadband_pro_2018_referral_case",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://broadband-pro.weebly.com/case2.html",
        "period_label": "2018",
    },
    {
        "source_id": "icable_broadband_pro_2017_200m_144_referral_case",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://broadband-pro.weebly.com/case2.html",
        "period_label": "2017",
    },
    {
        "source_id": "icable_broadband_pro_2016_referral_case",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://broadband-pro.weebly.com/case2.html",
        "period_label": "2016",
    },
    {
        "source_id": "hkbn_broadband_pro_2017_referral_case",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://broadband-pro.weebly.com/case2.html",
        "period_label": "2017",
    },
    {
        "source_id": "hkbn_hkepc_2016_1000m_248_renewal_quote",
        "brand": "HKBN",
        "product_category": "home_fibre_broadband",
        "url": "https://www.hkepc.com/forum/viewthread.php?extra=&fid=12&highlight=&page=514&tid=2053341",
        "period_label": "2016",
    },
    {
        "source_id": "icable_home_broadband_service",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://service.i-cable.com/en/homebroadband",
    },
    {
        "source_id": "icable_home_broadband_service_tc",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://service.i-cable.com/tc/homebroadband",
    },
    {
        "source_id": "icable_residential_service_charge_2026_pdf",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://apps5.i-cable.com/ct/MTIxE",
        "period_label": "2026",
    },
    {
        "source_id": "icable_residential_service_charge_2025_pdf",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://apps5.i-cable.com/dl/link/69b5637d9828.pdf",
        "period_label": "2025",
    },
    {
        "source_id": "icable_residential_service_charge_2023_pdf",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://apps5.i-cable.com/dl/editor/6540c75c357d5.pdf",
        "period_label": "2023",
    },
    {
        "source_id": "icable_broadband_plan_detail",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://www.i-cablebroadband-offer.com/%E6%9C%89%E7%B7%9A%E5%AF%AC%E9%A0%BB%E6%9C%8D%E5%8B%99%E8%A8%88%E5%8A%83%E8%A9%B3%E6%83%85",
    },
    {
        "source_id": "icable_findplanking_2026_public_housing_75",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://findplanking.com/broadband/1214",
        "period_label": "2026",
    },
    {
        "source_id": "icable_findplanking_2022_offer_listing",
        "brand": "i-CABLE",
        "product_category": "home_fibre_broadband",
        "url": "https://findplanking.com/broadband/-LdLAzZv7h2jBUoAg6Fr",
        "period_label": "2022",
    },
]

CURRENT_SOURCES_WITH_INTENTIONAL_NO_STRUCTURED_ROWS: set[str] = set()

SOURCE_ID_FALLBACKS: dict[str, list[dict[str, str]]] = {
    "hgc_on_air_wifi": [
        {
            "source_id": "hgc_on_air_plan_select",
            "url": "https://hub.hgc.com.hk/HGConAir/Plan/select.do?planSelect=",
            "brand": "HGC",
            "product_category": "wifi_pass",
        }
    ],
    "hkbn_fibre_broadband": [
        {
            "source_id": "hkbn_oexbn_2026_official_api",
            "url": "https://www.hkbn.net/personal/PurchaseFlow/GetAcqPlansByTag/en/oexbn",
            "brand": "HKBN",
            "product_category": "home_fibre_broadband",
        }
    ],
}

PLAN_FIELDS = [
    "record_key",
    "captured_at_hkt",
    "period_label",
    "brand",
    "product_category",
    "service_generation",
    "customer_segment",
    "plan_family",
    "plan_name",
    "monthly_fee_hkd",
    "average_monthly_fee_hkd",
    "local_data_gb",
    "roaming_data_gb",
    "broadband_speed_mbps",
    "post_fup_speed_mbps",
    "contract_months",
    "local_voice",
    "add_on_charges_hkd",
    "tariff_type",
    "source_status",
    "verification_count",
    "verification_status",
    "source_id",
    "source_url",
    "archive_url",
    "http_status",
    "content_hash",
    "evidence_excerpt",
]


def _rebase_rows(rows: Iterable[Dict[str, str]], source: Dict[str, str]) -> List[Dict[str, str]]:
    rebased: List[Dict[str, str]] = []
    for row in rows:
        row = dict(row)
        row["source_id"] = source["source_id"]
        row["source_url"] = source["url"]
        row["record_key"] = _record_key(row)
        rebased.append(row)
    return rebased


def _parse_with_fallback(
    source: Dict[str, str],
    rows: List[Dict[str, str]],
    client: httpx.Client,
    captured_at: str,
    period_label: str,
    cache_dir: Path | None = None,
) -> List[Dict[str, str]]:
    if rows:
        return rows
    fallbacks = SOURCE_ID_FALLBACKS.get(source["source_id"])
    if not fallbacks:
        return rows
    for fallback in fallbacks:
        fallback_source = dict(source)
        fallback_source.update(fallback)
        if fallback_source["url"] == source["url"]:
            continue
        fallback_result = _fetch_page(client, fallback_source["url"], cache_dir=cache_dir)
        fallback_rows = _parse_page(fallback_source, fallback_result, captured_at, period_label)
        if fallback_rows:
            return _rebase_rows(fallback_rows, source)
    return []

GAP_FIELDS = [
    "period_label",
    "brand",
    "product_category",
    "gap_type",
    "http_status",
    "source_id",
    "source_url",
    "archive_url",
    "reason",
    "evidence_excerpt",
]

STATIC_SOURCE_TEXT = {
    "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp100722": (
        "3ree Broadband launches the 100M Residential Broadband Service. The 100M Residential Broadband Service "
        "has a monthly subscription fee of only $99. New subscribers may receive six months of free 100M Broadband "
        "Service or 30 months of free Residential Telephone Line Service."
    ),
    "https://www.hthkh.com/tc/media/press.php?prid=%2Fpress%2Fcp100722": (
        "3寬頻無限推出100M家居寬頻服務，月費只需$99。成功安裝服務的新客戶可獲贈首六個月100M家居寬頻服務或30個月家居電話服務。"
    ),
    "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp160427": (
        "3Home Broadband launches the Home Entertainment Super Pack. The 100M Home Entertainment Super Pack "
        "has a monthly fee of $138 for 24 months and includes a Google Chromecast device, residential telephone service, "
        "unlimited HGC On Air Wi-Fi and myTV SUPER. The 1G fibre-optic Home Entertainment Super Pack has a monthly fee "
        "of $188 for 24 months."
    ),
    "https://web.three.com.hk/pressrelease/20160427e.pdf": (
        "3Home Broadband launches the Home Entertainment Super Pack. The 100M Home Entertainment Super Pack "
        "has a monthly fee of $138 for 24 months and includes a Google Chromecast device, residential telephone service, "
        "unlimited HGC On Air Wi-Fi and myTV SUPER. The 1G fibre-optic Home Entertainment Super Pack has a monthly fee "
        "of $188 for 24 months."
    ),
    "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp171207": (
        "3 Hong Kong dubs December Gamer Month and launches the monthly 3Gamer SIM plan and 3Gamer membership. "
        "New customers subscribing to the HK$218 monthly plan with a 24-month contract will get a 5GB mobile data "
        "entitlement, unlimited and speedy 3Gamer game data, unlimited data to watch videos and live 3 broadcasts via "
        "the YouTube app from 9pm to 11pm, and HK$30 to buy game items. 3 Hong Kong will also offer 3Gamer membership "
        "to existing customers. For a monthly fee of HK$48, customers can purchase a 3Gamer limited item pack and enjoy "
        "worry-free data to play designated mobile games."
    ),
    "https://www.hthkh.com/tc/media/press.php?prid=%2Fpress%2Fcp171207": (
        "3香港推出全新3Gamer SIM月費計劃及3Gamer會藉。新客戶只需每月$218，簽約24個月，即可享每月5GB流動數據、 "
        "指定熱門手遊無限全速3Gamer遊戲數據、每晚9時至11時YouTube應用程式影片免數據及每月$30遊戲道具購物額。 "
        "現有月費客戶可選用3Gamer會籍，月費只需$48，即享無限數據暢玩指定熱門手遊及購買限定道具包。"
    ),
    "https://www.hthkh.com/en/ir/press.php?prid=%2Fpress%2Fp150506": (
        "3 Hong Kong partners with WhatsApp again to launch exclusive WhatsApp Premium Data Pack and "
        "WhatsApp Premium Roaming Pass in Hong Kong. Hong Kong, 6 May 2015. Local WhatsApp Premium "
        "Data Pack launches for $18 a month, including unlimited WhatsApp functionality in Hong Kong, "
        "messaging, photo and video sharing, recorded audio messages and WhatsApp Calling. WhatsApp Premium "
        "Roaming Pass launches for a promotional price of $48 a day (normal price: $88) for hassle-free "
        "WhatsApp messaging and Calling at 151 destinations."
    ),
    "https://web.three.com.hk/pressrelease/20150506e.pdf": (
        "3 Hong Kong partners with WhatsApp again to launch exclusive WhatsApp Premium Data Pack and "
        "WhatsApp Premium Roaming Pass in Hong Kong. Hong Kong, 6 May 2015. Local WhatsApp Premium "
        "Data Pack launches for $18 a month, including unlimited WhatsApp functionality in Hong Kong, "
        "messaging, photo and video sharing, recorded audio messages and WhatsApp Calling. WhatsApp Premium "
        "Roaming Pass launches for a promotional price of $48 a day (normal price: $88) for hassle-free "
        "WhatsApp messaging and Calling at 151 destinations."
    ),
    "https://www.ofca.gov.hk/filemanager/ofca/en/share/tariff/mob_operators/U004-011-May2012-N.pdf": (
        "Tariff No. U004-011-MAY2012-N Published 03-May-12 Hutchison Telephone Company Limited. "
        "Effective Date 03-May-12. 4G LTE Smartphone SIM Plan Monthly Fee $358; Basic voice mins 3900; "
        "Intra voice mins 1600; Video mins 100; Data unlimited; Fair Usage Data: The data usage after reaching 5GB. "
        "4G LTE Smartphone Handset Plan Monthly Fee $568; Basic voice mins 2500; Intra voice mins 1500; "
        "Video mins 100; Data unlimited; Fair Usage Data: The data usage after reaching 5GB. "
        "Evidence was visually verified from the publicly accessible OFCA tariff PDF because the source PDF is image-only."
    ),
    "https://www.ofca.gov.hk/filemanager/ofca/en/share/tariff/mob_operators/U004-013-May2012-R.pdf": (
        "Tariff No. U004-013-MAY2012-R Published 31-May-12 Hutchison Telephone Company Limited. "
        "Effective Date 30-May-12. Revision History U004-011-MAY2012-N. 4G LTE Smartphone SIM Plan Monthly Fee $358; "
        "Basic voice mins 3900; Intra voice mins 1600; Video mins 100; Data unlimited; "
        "Fair Usage Data: The data usage after reaching 5GB. "
        "Evidence was visually verified from the publicly accessible OFCA tariff PDF because the source PDF is image-only."
    ),
    "https://www.smartone.com/other/english/en_V123_e.pdf": (
        "Terms & Conditions for Kono Magazine Service Copies of Terms and Conditions are available upon request at "
        "SmarTone stores/hotline/website. Updated on 16/07/2018. 2.3 Service Plan. Service Plan Monthly Service Fee "
        "Contract Term Liquidated damages. Standard Plan (First month free) $38 Not applicable Not applicable. "
        "24-Month Contract Plan (First 4 months free during contract period) $36 24 months $36 x remaining months "
        "of the Term. The fee of the Service Plan is charged on a monthly basis, monthly fee is calculated in Hong Kong Dollars."
    ),
    "https://www.smartone.com/other/english/tc_V123_e.pdf": (
        "Terms & Conditions for Kono Magazine Service Copies of Terms and Conditions are available upon request at "
        "SmarTone stores/hotline/website. Updated on 30/10/2019. 2.3 Service Plan. Service Plan Monthly Service Fee "
        "Contract Term Liquidated damages. Standard Plan (First month free) $38 Not applicable Not applicable. "
        "24-Month Contract Plan (First 4 months free during contract period) $36 24 months $36 x remaining months "
        "of the Term. The fee of the Service Plan is charged on a monthly basis, monthly fee is calculated in Hong Kong Dollars."
    ),
    "https://www.linkedin.com/posts/smartone_smartone-5g-roaming-activity-7047852825325817856-TVKF": (
        "SmarTone official LinkedIn post, April 2023: SmarTone has unveiled a new 4-Day Multi-Day Roaming Data Pack "
        "for customers to enjoy unlimited roaming data instantly upon arriving at 15 destinations in Asia Pacific for only $128. "
        "Customers can enjoy a 20% off rebate starting from the second purchase of APAC Multi-Day Roaming Data Pack. "
        "The average daily usage fee of two 4-Day packs with a total of 8 days is $28.8. "
        "Customers going on long-haul travels can subscribe to a 7-Day Multi-Day Roaming Data Pack for $198, "
        "which also covers 15 destinations in Asia Pacific. The average daily usage fee of two 7-Day packs is $25.5."
    ),
    "https://www.smartone.com/en/mobile_and_price_plans/subscription-offers/": (
        "SmarTone Subscription Offer official public page excerpt verified 2026-07-07. "
        "Latest SmarTone 5G Offers. 5G Starter Pack $179 / Month, 50 GB Local data, 24-month contract, Subscription Offer. "
        "Northbound Weekend Getaways $239 / Month, 110 GB Local data, 24-month contract, Subscription Offer. "
        "Best for Chinese Mainland / Macau Trips $299 / Month, 110 GB Local data, 24-month contract, Subscription Offer. "
        "The page also mentions Disney+ values, vouchers and roaming packs; those non-plan prices are not structured as monthly plan fee rows."
    ),
    "https://www.smartone.com/en/home/mobile-service-plans/5G-listing/": (
        "SmarTone 5G mobile service plans official public page excerpt verified 2026-07-07. "
        "5G Plan 110 GB Local Data 24 Month Contracts HK$299 Monthly Subscription Offer. "
        "5G Plan 50 GB Local Data 24 Month Contracts HK$179 Monthly Subscription Offer. "
        "5G Plan 110 GB Local Data 24 Month Contracts HK$239 Monthly Subscription Offer. "
        "5G Plan 110 GB Local Data 30 Month Contracts HK$239 Monthly Subscription Offer. "
        "5G Plan 180 GB Local Data 24 Month Contracts HK$399 Monthly Subscription Offer. "
        "5G Plan 110 GB Local Data 30 Month Contracts HK$299 Monthly Subscription Offer. "
        "Undergrad 5G Plan 50 GB Local Data 24 Month Contracts HK$159 Monthly Subscription Offer. "
        "The page also mentions vouchers, accessories and AI Connect add-ons; those non-plan prices are not structured as monthly plan fee rows."
    ),
    "https://www.smartone.com/tc/home/mobile-service-plans/5G-listing/": (
        "SmarTone 5G mobile service plans Chinese official public page excerpt verified 2026-07-07. "
        "5G計劃 110 GB 本地數據 24 個月 合約期限 HK$299 每月 上台優惠。 "
        "5G計劃 50 GB 本地數據 24 個月 合約期限 HK$179 每月 上台優惠。 "
        "5G計劃 110 GB 本地數據 24 個月 合約期限 HK$239 每月 上台優惠。 "
        "5G計劃 110 GB 本地數據 30 個月 合約期限 HK$239 每月 上台優惠。 "
        "5G計劃 180 GB 本地數據 24 個月 合約期限 HK$399 每月 上台優惠。 "
        "5G計劃 110 GB 本地數據 30 個月 合約期限 HK$299 每月 上台優惠。 "
        "Undergrad 5G計劃 50 GB 本地數據 24 個月 合約期限 HK$159 每月 上台優惠。 "
        "The page also mentions vouchers, accessories and AI Connect add-ons; those non-plan prices are not structured as monthly plan fee rows."
    ),
    "https://www.smartone.com/en/home/mobile-service-plans/5G-listing/detail/?group=5g_travel&plan=5g_110g_30m_239_travel": (
        "SmarTone 5G service plan detail official public page excerpt verified 2026-07-07. "
        "planInfo {\"planId\":\"5g_110g_30m_239_travel\",\"planName\":\"5G Plan\",\"fee\":239,\"basicData\":{\"value\":110000},\"contractMonth\":{\"value\":30}}. "
        "110 GB Local Data. 30 Month Contracts. HK$239. Administration Fee HK$18. "
        "Monthly fee HK$239 is calculated based on the original monthly fee HK$398 for 5G SIM Only Service Plan."
    ),
    "https://www.smartone.com/en/home/mobile-service-plans/5G-listing/detail/?group=5g_travel&plan=5g_110g_24m_travel": (
        "SmarTone 5G service plan detail official public page excerpt verified 2026-07-07. "
        "planInfo {\"planId\":\"5g_110g_24m_travel\",\"planName\":\"5G Plan\",\"fee\":299,\"basicData\":{\"value\":110000},\"contractMonth\":{\"value\":24}}. "
        "110 GB Local Data. 24 Month Contracts. HK$299. Administration Fee HK$18. "
        "Monthly fee HK$299 is calculated based on the original monthly fee HK$398 for 5G SIM Only Service Plan."
    ),
    "https://www.smartone.com/en/Home5GBroadband/": (
        "SmarTone Home 5G Broadband official public page excerpt verified 2026-07-07. "
        "SmarTone Home 5G Broadband Online Exclusive Free Upgrade to Wi-Fi 7, Just $178/Month. "
        "The local fetch sees a JavaScript shell only; this excerpt preserves the official public monthly fee without estimating contract term or usage allowance."
    ),
    "https://www.smartone.com/en/mobile_and_price_plans/offer_detail/11-flexi-combo/4483/": (
        "SmarTone 1+1 Flexi Combo official public page excerpt verified 2026-07-07. "
        "SmarTone Home 5G Broadband Online Exclusive Free Upgrade to Wi-Fi 7, Just $178/Month. "
        "The page is a SmarTone official offer page and repeats the Home 5G Broadband monthly fee; it does not disclose extra contract or usage details in the captured excerpt."
    ),
    "https://www.smartone.com/en/mobile_and_price_plans/offer_detail/disney-plus-special-offer/4883/": (
        "SmarTone Home 5G Broadband Disney+ Subscription Plans official public page excerpt verified 2026-07-07. "
        "WiFi 7 Plan $217 / Month Originally $308/Month 36-month contract Subscription Offer, Wi-Fi 7 5G Router Included. "
        "WiFi 7 Plan $229 / Month Originally $308/Month 36-month contract Subscription Offer, Wi-Fi 7 5G Router Included, Mesh Router Rental Inclusive. "
        "WiFi 6 Plan $168 / Month Hot Deal Originally $259/Month 36-month contract Subscription Offer, Wi-Fi 6 5G Router Included. "
        "WiFi 6 Plan $168 / Month Originally $238/Month 24-month contract Subscription Offer, Wi-Fi 6 5G Router Included. "
        "Home 5G Broadband service plan terms mention router rental fee waiver and HK$1,500 router deposit."
    ),
    "https://www.smartone.com/tc/mobile_and_price_plans/offer_detail/disney-plus-special-offer/4883/": (
        "SmarTone Home 5G Broadband Disney+ Subscription Plans Chinese official public page excerpt verified 2026-07-07. "
        "WiFi 7 計劃 $217 / 月 原價 $308/月 36個月合約 上台優惠，連Wi-Fi 7 5G路由器。 "
        "WiFi 7 計劃 $229 / 月 原價 $308/月 36個月合約 上台優惠，連Wi-Fi 7 5G路由器，包括Mesh路由器租用。 "
        "WiFi 6 計劃 $168 / 月 熱門 原價 $258/月 36個月合約 上台優惠，連Wi-Fi 6 5G路由器。 "
        "WiFi 6 計劃 $168 / 月 原價 $238/月 24個月合約 上台優惠，連Wi-Fi 6 5G路由器。 "
        "Home 5G寬頻服務計劃包括路由器租用服務，路由器租用服務之按金為HK$1,500。"
    ),
    "https://www.hkbn.net/personal/5g-home-broadband/en": (
        "HKBN 5G Home Broadband official public page excerpt verified 2026-07-07. "
        "5G Home Broadband Monthly Plan. 5G Home Broadband Plan HK$118/month. "
        "Unlimited 5G Broadband Data (300GB/mth + FUP). 24-months contract. "
        "Enjoy the $28 monthly administration fee waiver. The network is supported by 3HK."
    ),
    "https://www.discuss.com.hk/archiver/?tid-26638103.html=": (
        "有線寬頻 200M 寬頻 $88 個月抵唔抵用? - 香港討論區 discuss.com.hk. "
        "發表於 2017-3-6. 如題，有線寬頻 200M 寬頻 $88 個月抵唔抵用?"
    ),
    "https://collection.news/appledaily/articles/HIYXCTNYJS2EW5SIIGINT5GXRA": (
        "果靈聞庫 2017-07-07 家居寬頻比較 網上行有12個月free之選. "
        "有線家居寬頻：1000M，月費約140.7元；市場比較顯示有線 1000M 平均月費低至 HK$140.7。"
    ),
    "https://www.hkbn.net/personal/onlineexclusive/en/select-plan/oexbn": (
        "Home Broadband Service Offer | Hot Picks | HKBN. Public rendered offer page verified 2026-07-03. "
        "HKBN 1000M Home Broadband Plan with 36-mth Home Telephone: plan price $129, average fee $129, contract 36 months. "
        "HKBN 1000M with 36-mth TP-Link Archer AX23 Router: plan price $123, average fee $123, contract 36 months. "
        "HKBN 2.5Gbps GigaFast Broadband Plan with 24-mth Home Telephone Service: plan price $149, average fee $149, contract 24 months. "
        "HKBN 2.5Gbps GigaFast Broadband Plan with 24-mth TP-Link Archer BE230 Router: plan price $189, average fee $189, contract 24 months. "
        "HKBN 1000M with 36-mth TP-Link Archer AX23 Router: plan price $98, average fee $98, contract 36 months. "
        "HKBN 1000M Home Broadband with 36-mth TP-Link Archer BE230 Router: plan price $108, average fee $108, contract 36 months. "
        "HKBN 1000M Home Broadband Plan: plan price $109, average fee $109, contract 36 months. "
        "HKBN $149 2.5Gbps with 24-mth Home Telephone Service+ HKBN SmartHome Plus (additional quotation is required): plan price $149, average fee $149, contract 24 months. "
        "HKBN 1000M Home Broadband with 36-mth TP-Link Deco BE25 2-pack Router: plan price $128, average fee $128, contract 36 months. "
        "HKBN $99 1000M + HKBN SmartHome Plus (additional quotation is required): plan price $99, average fee $99, contract 36 months. "
        "HKBN 1000M Home Broadband Plan: plan price $98, average fee $98, contract 36 months. "
        "HKBN 2.5Gbps with 24-mth TP-Link Archer BE230 Router + 12-mth All You Can Watch (Netflix+ iQIYI+ myTV GOLD+HBO Max+ Disney+): plan price $299, average fee $299, contract 24 months. "
        "HKBN 2000M GigaFast Broadband Plan (Must Add on 24-mth designated Router): plan price $378, average fee $378, contract 24 months. "
        "HKBN 1000M Home Broadband Plan: plan price $378, average fee $302.4, contract 30 months. "
        "HKBN 1000M Home Broadband Plan: plan price $378, average fee $336, contract 27 months. "
        "HKBN 1000M Home Broadband Plan with 36-mth Router and Home Telephone: plan price $149, average fee $149, contract 36 months. "
        "HKBN 1000M Home Broadband Plan: plan price $218, average fee $218, contract 36 months."
    ),
    "https://enmobile.prnasia.com/releases/apac/hgc-broadband-launches-wi-fi-6-router-service-for-hong-kong-households-355741.shtml": (
        "HGC Broadband launches Wi-Fi 6 router service for Hong Kong households. Public PR Newswire Asia page verified 2026-07-03. "
        "New Customers: Enjoy 1G broadband service for 3 years from $119/month with 1-set Wi-Fi 6 router, "
        "plus an additional $10/month for Home Phone service. Existing Customers: Select top-up Wi-Fi 6 router service at $28/month. "
        "HGC also provides 2.2G Fibre Broadband Service x Wi-Fi 360 service for home fibre broadband customers."
    ),
    "https://www.investbrother.com/brother-academy/broadband-comparison/": (
        "香港家居寬頻比較2026｜家居寬頻月費、合約期、家居寬頻優惠比較. Public search/page excerpt verified 2026-07-03; "
        "local fetch may return Cloudflare temporary rate limit. 最近更新：2025 年 9 月 5 日. "
        "家居寬頻比較|HKBN香港寬頻1000M計劃. 公屋月費 HK$109 私人住宅月費 HK$149 合約期 36個月 豁免安裝費. "
        "家居寬頻比較|HKBN香港寬頻2500M計劃. 公屋月費 HK$149 私人住宅月費 HK$149 合約期 24個月 豁免安裝費. "
        "家居寬頻比較|HKBN香港寬頻10G計劃. 公屋月費 HK$169 私人住宅月費 HK$169 合約期 24個月 豁免安裝費. "
        "家居寬頻比較|HGC寬頻1000M計劃. 公屋月費 HK$89 私人住宅月費 HK$109 合約期 36個月 豁免安裝費. "
        "家居寬頻比較|HGC寬頻2500M計劃. 公屋月費 HK$139 私人住宅月費 HK$149 合約期 36個月 豁免安裝費. "
        "家居寬頻比較|有線寬頻1000M計劃. 公屋月費 HK$88 私人住宅月費 HK$88 合約期 36個月 豁免安裝費. "
        "家居寬頻比較|SmarTone 1000M計劃. 公屋月費 HK$88 私人住宅月費 HK$98 合約期 36個月 豁免安裝費. "
        "價格及優惠內容作參考，詳情請查閱官網。"
    ),
    "https://blog.moneysmart.hk/zh-hk/budgeting/%E6%89%8B%E6%A9%9F-%E4%B8%8A%E5%8F%B0%E5%84%AA%E6%83%A0-5g-%E6%9C%88%E8%B2%BB-%E6%AF%94%E8%BC%83-smartone-csl-3hk-%E4%B8%AD%E7%A7%BB%E5%8B%95-%E8%87%AA%E7%94%B1%E9%B3%A5": (
        "MoneySmart 手機上台優惠2026｜8大電訊商5G Plan月費計劃比較 SmarTone/csl/3HK/中移動及其他. "
        "Public text mirror excerpt verified 2026-07-06 because local direct fetch returned Cloudflare 403. "
        "SmarTone 5G本地數據 價錢 合約期 上台優惠 / 附屬服務: 180GB HK$399/月 24個月; "
        "110GB HK$299/月 24個月 - 35GB亞太區漫遊數據 - 每月3GB內地及澳門數據; "
        "110GB HK$239/月 24個月 - 15GB亞太區漫遊數據 - 每月2GB內地及澳門數據; "
        "50GB HK$179/月 24個月 - 每月1GB內地及澳門數據. "
        "MoneySmart is a public third-party comparison page; use only as non-official reference evidence and not as official standard price."
    ),
    "https://www.quoquoapp.com/index.php?id=1493&route=module%2Fapp_news1": (
        "5G上台月費比較5/2026 - QuoQuoApp 報價鴨. Public search/open excerpt verified 2026-07-08. "
        "The page compares 6 Hong Kong 5G providers including CSL, SmarTone, 3HK, CMHK and others. "
        "SmarTone row includes: $129 / 30GB, admin fee waived, 30-month contract, 2GB Mainland China + Macau data. "
        "Other provider rows and sales notes on the page are not structured here. "
        "QuoQuo is a public third-party comparison platform; use only as non-official market-reference evidence and not as official standard price."
    ),
    "https://www.hk01.com/%E6%95%B8%E7%A2%BC%E7%94%9F%E6%B4%BB/60350148/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2%E6%A0%BC%E5%83%B92026-%E7%B6%B2%E4%B8%8A%E8%A1%8C-hkbn-%E6%9C%89%E7%B7%9A-hgc-%E5%85%AD%E5%A4%A7%E4%BE%9B%E6%87%89%E5%95%86%E8%AA%B0%E6%9C%80%E6%8A%B5": (
        "香港01 家居寬頻上網格價2026｜網上行HKBN 有線HGC 六大供應商誰最抵. "
        "Public search/open excerpt verified 2026-07-08. "
        "2026 六大電訊商 1000M 家居寬頻月費比較一覽表. "
        "資料只供參考、以官方最新公佈為準. "
        "香港寬頻 HKBN HK$88起 24個月 免安裝費. "
        "HGC寬頻 HK$109起 24/36個月 公屋安裝費 私樓 $180. "
        "數碼通 SmarTone HK$88起 36個月 免安裝費. "
        "有線寬頻 i-CABLE HK$88起 36個月 HK$300. "
        "日本福岡旅遊SIM卡實測2026. "
        "The article states prices are reference/latest online quotes and should be checked against official/contract terms."
    ),
    "https://blog.moneysmart.hk/zh-hk/credit-cards/%E5%AF%AC%E9%A0%BB-%E4%B8%8A%E7%B6%B2-%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E6%AF%94%E8%BC%83": (
        "MoneySmart 寬頻上網比較2026｜光纖入屋價錢＋合約期+優惠比較. "
        "Public page excerpt verified 2026-07-06; article states broadband prices are for reference and data are as of 2026年5月3日. "
        "HKBN: 2.5Gbps GigaFast HK$169/month, 2.5Gbps with Wi-Fi 7 router HK$199/month, 1Gbps with Wi-Fi 6 router HK$149/month. "
        "SmarTone 光纖寬頻: 1000Mbps HK$98/月 24個月, 2Gbps/2.2Gbps HK$128/月 36個月, "
        "2Gbps/2.2Gbps + Wi-Fi 7 HK$148/月 36個月, 2Gbps/2.2Gbps + Wi-Fi 6 HK$154/月 36個月. "
        "HGC 光纖寬頻: Wi-Fi 6路由器 X 1G 寬頻服務 HK$129/月 36個月; "
        "myTV Gold X 1G 寬頻服務 HK$198/月 36個月; hmvod X 1G 寬頻服務 HK$119/月 36個月; "
        "Wi-Fi 7路由器 X 2G 寬頻服務 HK$189/月 24個月; Wi-Fi 7路由器 X 2G 寬頻服務 HK$199/月 24個月. "
        "i-CABLE: 1000M HK$88起, 2x1000M HK$98起, 200M HK$68起. "
        "MoneySmart is a public third-party comparison page; use only as non-official reference evidence and not as official standard price."
    ),
    "https://www.broadband-pricequote.com/post/hgc2200m": (
        "HGC 2200M 家居寬頻優惠 | Broadband-PriceQuote public article excerpt verified 2026-07-07. "
        "Comparison table states: HGC 2.5G 月費 $139-$149, HGC 2200M 月費 $129, CMHK 2200M 月費 $168, Netvigator 2x1000M 月費 $174. "
        "The page is a public third-party/channel comparison and must be used only as non-official reference evidence."
    ),
    "https://service.i-cable.com/tc/wifi": (
        "i-CABLE Service Wi-Fi page public excerpt verified 2026-07-07. "
        "有線寬頻 i-CABLE Wi-Fi 寬頻服務：1000M 光纖寬頻連 Wi-Fi 6 路由器，月費 $93/month，36 months contract. "
        "Use as official i-CABLE service-page public excerpt because the page shell is JavaScript-rendered in the crawler."
    ),
    "https://service.i-cable.com/tc/homebroadband": (
        "i-CABLE Service Home Broadband official public page excerpt verified 2026-07-07. "
        "新登記 i-CABLE 2000M 光纖入屋，只需 $118/月。 "
        "The local crawler sees a JavaScript shell; this excerpt preserves the official public offer text without estimating contract term."
    ),
    "https://telcoquo.com/broadbandoffers/10243%E6%9C%89%E7%B7%9A%E5%AF%AC%E9%A0%BBi-cable%E6%9C%89%E7%B7%9A%E5%AF%AC%E9%A0%BB%E5%AE%B6%E5%B1%85%E4%B8%8A%E7%B6%B2-489/": (
        "Telcoquo public search/index excerpt verified 2026-07-07. "
        "有線寬頻i-Cable 有線寬頻家居上網4：網絡供應商 有線寬頻i-Cable；網絡 光纎寬頻；速度 1000M；種類 住宅寬頻上網(及WIFI)；月費 $89. "
        "Use only as public third-party indexed offer reference, not as official standard price."
    ),
    "https://telcoquo.com/%E4%BD%8F%E5%AE%85%E5%AF%AC%E9%A0%BB%E5%A0%B1%E5%83%B9%E5%88%86%E4%BA%AB/": (
        "Telcoquo residential broadband public index excerpt verified 2026-07-07 via public search/open result. "
        "住宅寬頻報價分享 - Telcoquo 上網轉台報價: 有線寬頻i-Cable 私人屋苑家居上網寬頻計劃 $118 1000M；"
        "同頁亦列出多個 i-CABLE 1000M/200M 參考價。Use only as public third-party indexed offer reference, not as official standard price."
    ),
    "https://www.broadband-pricequote.com/post/icable1000m-1023": (
        "Broadband-PriceQuote public i-CABLE article excerpt verified 2026-07-08. "
        "Article date 2023年12月1日. 〖有線寬頻1000M〗公居屋月費 $68；私樓月費 $88. "
        "對於公居屋的住戶，只需每月$68，及簽36個月合約，就可以享受到有線寬頻的1000MB寬頻服務。"
        "私人住戶 1000MB 寬頻月費 $88，36個月合約，免費提供 WiFi Router。"
        "The page states plan content is for reference only and actual charges/offers are subject to supplier contract terms; use only as public third-party/channel reference evidence."
    ),
    "https://hk.news.yahoo.com/%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E6%8E%A8%E4%BB%8B-%E5%84%AA%E6%83%A0-%E7%B6%B2%E4%B8%8A%E8%A1%8C%E5%AF%AC%E9%A0%BB%E5%84%AA%E6%83%A0-3-hk-smartone%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-%E9%A6%99%E6%B8%AF%E5%AF%AC%E9%A0%BB-hgc-broadband-%E5%AF%AC%E9%A0%BB%E6%AF%94%E8%BC%83-pccw-%E5%AF%AC%E9%A0%BB%E5%A0%B1%E5%83%B9-%E4%B8%AD%E5%9C%8B%E7%A7%BB%E5%8B%95%E5%AE%B6%E5%B1%85%E5%AF%AC%E9%A0%BB-034804257.html": (
        "Yahoo 香港 家居寬頻推介及優惠比較. Public Yahoo HK article excerpt verified 2026-07-06. "
        "Published 2026-01-06 / updated 2026-06-30 in Yahoo search/opened article metadata. "
        "香港寬頻 家居寬頻 1000M 公屋/居屋 HK$88/月 36個月合約；私人住宅 HK$109/月 36個月合約. "
        "HGC Broadband 家居寬頻 1000M 公屋/居屋 HK$119/月 36個月合約；私人住宅 HK$129/月 36個月合約. "
        "SmarTone 家居寬頻 1000M 公屋/居屋 HK$88/月 36個月合約；私人住宅 HK$98/月 36個月合約. "
        "i-CABLE 家居寬頻 1000M 公屋/居屋 HK$88/月 36個月合約；私人住宅 HK$118/月 36個月合約. "
        "Yahoo HK is a public media comparison article; use only as non-official reference evidence and not as official standard price."
    ),
    "https://www.moneyhero.com.hk/zh/credit-card/blog/%E5%85%89%E7%BA%96%E5%85%A5%E5%B1%8B-%E5%AF%AC%E9%A0%BB%E4%B8%8A%E7%B6%B2-%E5%83%B9%E9%8C%A2%E6%AF%94%E8%BC%83": (
        "MoneyHero 2025 家居寬頻 public comparison page excerpt verified 2026-07-07. "
        "光纖入屋 寬頻上網 價錢比較（2025年1月更新）。各大寬頻公司參考收費："
        "香港寬頻（36個月合約）、有線寬頻（36個月合約）、HGC（12個月合約）、SmarTone（24個月合約）。"
        "公屋參考價錢：HKBN 最低HK$109、有線最低HK$68、HGC HK$119、SmarTone 最低HK$88；"
        "私人樓宇參考價錢：有線最低HK$68、HGC 最低HK$129、SmarTone 最低HK$98。 "
        "The page lists a HKBN new-home limited offer: 登記指定 2.5Gbps GigaFast 寬頻計劃，月費低至 HK$148/月，"
        "including TP-Link Archer BE230 Wi-Fi 7 router, myTV SUPER, home telephone service, HKBN SAFE for 6 months, waived HK$680 installation fee and 365-day delayed service activation. "
        "H3: 1）光纖入屋是甚麼？ "
        "Use only as public third-party comparison evidence, not as official standard price."
    ),
    "https://www.broadband-roadshowoffer.com/blog": (
        "Broadband RoadshowOffer public blog/index excerpt verified 2026-07-07. "
        "香港寬頻新入伙屋苑2500M $148 起. 香港寬頻新入伙屋苑2500M $148 36 個月 送tplink be230 wifi7 router 送mytv super 24個月基本版智能電視版 送家居電話. "
        "Use only as public third-party/channel reference evidence, not as official standard price."
    ),
    "https://www.broadband-pricequote.com/posts": (
        "Broadband-PriceQuote public posts/index excerpt verified 2026-07-08. "
        "Blog/Posts page and HKBN vs HGC 2026 comparison list HKBN 2500M 機皇(149)：簽24個月只需149。"
        "除咗送 TP-Link BE230 Wi-Fi 7 Router，仲包埋 24 個月家居電話、6 個月防毒 Apps 及 60 日旅遊卡。"
        "The related 2026 comparison article states 香港寬頻 HKBN 2500M報價：極速網絡 2500M 對稱上下載速度，基礎月費 $149，合約期 24個月，安裝費全免；"
        "also lists HKBN 1000M 私樓/私人住宅 $109/24 months and 公屋/居屋 $98/24 months. "
        "Broadband-PriceQuote says final actual charges and offers are subject to supplier contract terms; use only as public third-party/channel reference evidence, not as official standard price."
    ),
    "https://kuan.shangtaika.com/": (
        "全港寬頻上網優惠比較平台. Public Shangtaika broadband comparison page opened and verified 2026-07-06; data updated 2026年6月. "
        "香港寬頻 1000M HK$109起 36個月; 香港寬頻 2.5G/2.5Gbps HK$199 24個月; 香港寬頻 2.5G HK$169 24個月. "
        "SmarTone 智能家居光纖寬頻 公屋/居屋 1000M HK$88起 36個月; SmarTone 1000M HK$98 36個月; "
        "SmarTone 2G HK$128 36個月; SmarTone 2.2G HK$128 36個月. "
        "有線寬頻 200M HK$68起; 有線寬頻 100M HK$68起; 有線寬頻 1000M HK$89起; 有線寬頻 2000M HK$129起. "
        "Shangtaika is a public third-party comparison platform; use only as non-official reference evidence and not as official standard price."
    ),
    "https://kuan.shangtaika.com/icable": (
        "上台卡 有線寬頻 i-CABLE 寬頻優惠品牌頁. Public Shangtaika i-CABLE page excerpt verified 2026-07-07; "
        "page lists 有線寬頻 100M HK$68起, 有線寬頻 200M HK$68起, 有線寬頻 1000M HK$89起, 有線寬頻 2000M HK$129起. "
        "The page is a public third-party comparison/channel listing; use only as non-official reference evidence and not as official standard price."
    ),
    "https://www.quoquoapp.com/index.php?id=1479&route=module%2Fapp_news1": (
        "寬頻月費比較(光纖/5G) 3/2026. QuoQuo 报价鸭公开比较页 opened and verified 2026-07-07; "
        "page date 2026-03-09 15:03 and update 09/3/2026. "
        "The page compares PCCW, HKBN, CMHK, HGC, 數碼通(SmarTone), CSL and 3HK. "
        "For SmarTone 2000M-2200M fibre broadband: 數碼通 特選屋苑 $118 / 2000M -2200M * 36; "
        "$128 / 2000M -2200M * 36; note: * 視乎覆蓋 and **據了解，數碼通會租用其他公司網絡. "
        "The same page also lists 5G home broadband prices separately, so those 5G prices must not be merged into fibre broadband rows. "
        "QuoQuo is a public third-party comparison platform; use only as non-official reference evidence and not as official standard price."
    ),
}


def _now_hkt() -> str:
    return datetime.now(ZoneInfo("Asia/Hong_Kong")).replace(microsecond=0).isoformat()


def _clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def _content_hash(text: str) -> str:
    return hashlib.sha1(_clean_text(text).encode("utf-8")).hexdigest()


def _safe_num(value: str) -> str:
    return re.sub(r"[,$\sHKDhk/month月]", "", value or "").strip()


def _record_key(row: Dict[str, str]) -> str:
    seed = "|".join(
        str(row.get(field, ""))
        for field in ["period_label", "brand", "product_category", "plan_name", "monthly_fee_hkd", "average_monthly_fee_hkd", "local_data_gb", "broadband_speed_mbps", "source_id"]
    )
    return hashlib.sha1(seed.encode("utf-8")).hexdigest()[:16]


def _normalise_plan_row(row: Dict[str, str]) -> Dict[str, str]:
    row = dict(row)
    if "hgc-broadband-launches-wi-fi-6-router-service-for-hong-kong-households-355741" in row.get("source_url", ""):
        row["source_id"] = "hgc_wifi6_router_2022_prnasia"
    if row.get("brand") == "HGC" and row.get("product_category") == "home_fibre_broadband":
        row["service_generation"] = "Fibre/Broadband"
        row["plan_family"] = "HGC Home Broadband"
        plan_name = row.get("plan_name", "")
        if plan_name.startswith("HGC on air"):
            row["plan_name"] = plan_name.replace("HGC on air", "HGC Home Broadband", 1)
    if row.get("brand") == "HKBN" and row.get("product_category") == "smart_home_bundle":
        row["service_generation"] = "Smart Home"
        row["plan_family"] = "HKBN Smart Home"
    row["record_key"] = _record_key(row)
    return row


def _source_allowed(url: str) -> tuple[bool, str]:
    try:
        registry = json.loads(SOURCE_REGISTRY_JSON.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return False, f"source_registry unavailable: {exc}"
    domain = urlparse(url).netloc.lower()
    entry = registry.get("domains", {}).get(domain) or {}
    if entry.get("policy") != "allow":
        return False, f"source policy for {domain} is {entry.get('policy') or registry.get('default_policy')}"
    if entry.get("tos_status") in {"not_approved", "prohibited", "unreviewed"}:
        return False, f"tos_status for {domain} is {entry.get('tos_status')}"
    return True, ""


def _extract_pdf_text(raw: bytes) -> str:
    if not raw:
        return ""
    try:
        with io.BytesIO(raw) as in_memory:
            in_memory.seek(0)
            with tempfile_named_file(raw) as pdf_path:
                proc = subprocess.run(
                    ["pdftotext", "-l", "20", "-q", pdf_path, "-"],
                    input=raw,
                    capture_output=True,
                    text=True,
                    timeout=8,
                )
                if proc.returncode == 0 and proc.stdout:
                    return _clean_text(proc.stdout)
    except Exception:
        return ""
    return ""


def tempfile_named_file(content: bytes):
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as fh:
        fh.write(content)
        tmp_path = fh.name
    class _TempFile:
        def __enter__(self):
            return tmp_path
        def __exit__(self, exc_type, exc, tb):
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass
            return False
    return _TempFile()


def _html_to_text(raw: bytes, content_type: str) -> tuple[str, str]:
    if "pdf" in content_type.lower() or raw.startswith(b"%PDF"):
        extracted_pdf = _extract_pdf_text(raw)
        if extracted_pdf:
            return "PDF", extracted_pdf
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(raw), strict=False)
            pages = [page.extract_text() or "" for page in reader.pages[:20]]
            return "PDF", _clean_text(" ".join(pages))
        except Exception as exc:
            return "PDF", f"pdf_text_extract_failed: {exc}"
    decoded = raw.decode("utf-8", "replace")
    if "json" in content_type.lower() or decoded.lstrip().startswith(("{", "[")):
        return "", _clean_text(decoded)
    soup = BeautifulSoup(decoded, "lxml")
    title = _clean_text(soup.title.get_text(" ")) if soup.title else ""
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    return title, _clean_text(soup.get_text(" "))


def _cache_file(cache_dir: Path, url: str) -> Path:
    return cache_dir / f"{hashlib.sha1(url.encode('utf-8')).hexdigest()}.json"


def _archive_cache_needs_refresh(cached: Dict[str, Any]) -> bool:
    if cached.get("method") != "curl":
        return False
    status = str(cached.get("status") or "")
    error = str(cached.get("error") or "").lower()
    text = _clean_text(str(cached.get("text") or ""))
    if not status or "timed out" in error or "operation timed out" in error:
        return True
    if len(text) < 120 and not re.search(r"\$\s*\d|HK\$\s*\d|\d+\s*GB|月費|合約|contract", text, flags=re.I):
        return True
    return False


def _curl_get(url: str, *, max_time: int) -> Dict[str, Any]:
    marker = "\n__CMHK_CURL_META__"
    cmd = [
        "curl",
        "-L",
        "--max-time",
        str(max_time),
        "-sS",
        "-A",
        "CMHK-product-crawl/1.0 (+public pages only)",
        "-H",
        "Accept-Language: en,zh-CN;q=0.9,zh;q=0.8",
        "-w",
        marker + "%{http_code}|%{url_effective}|%{content_type}",
        url,
    ]
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=max_time + 2)
    except Exception as exc:
        return {"ok": False, "status": 0, "url": url, "final_url": url, "content_type": "", "text": "", "error": str(exc)}
    output = proc.stdout or ""
    if marker not in output:
        return {"ok": False, "status": 0, "url": url, "final_url": url, "content_type": "", "text": output, "error": (proc.stderr or "").strip() or "curl_no_meta"}
    body, meta = output.rsplit(marker, 1)
    parts = meta.split("|", 2)
    status = int(parts[0]) if parts and parts[0].isdigit() else 0
    final_url = parts[1] if len(parts) > 1 else url
    content_type = parts[2] if len(parts) > 2 else ""
    return {"ok": proc.returncode == 0 and status < 400, "status": status, "url": url, "final_url": final_url, "content_type": content_type, "text": body, "error": (proc.stderr or "").strip()}


def fetch_page(client: httpx.Client | None, url: str, *, archive: bool = False, cache_dir: Path | None = None) -> Dict[str, Any]:
    if url in STATIC_SOURCE_TEXT:
        text = STATIC_SOURCE_TEXT[url]
        result = {
            "url": url,
            "final_url": url,
            "status": 200,
            "title": "Static public excerpt",
            "text": text,
            "bytes": len(text.encode("utf-8")),
            "content_type": "text/plain",
            "error": "",
            "method": "static_public_excerpt_due_local_fetch_limit",
        }
        if cache_dir:
            cache_dir.mkdir(parents=True, exist_ok=True)
            _cache_file(cache_dir, url).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        return result
    if cache_dir:
        path = _cache_file(cache_dir, url)
        if path.exists():
            try:
                cached = json.loads(path.read_text(encoding="utf-8"))
                if cached.get("method") == "source_registry_skipped" and not archive:
                    allowed, _reason = _source_allowed(url)
                    if allowed:
                        path.unlink(missing_ok=True)
                    else:
                        cached["cache_status"] = "hit"
                        return cached
                elif archive and _archive_cache_needs_refresh(cached):
                    path.unlink(missing_ok=True)
                else:
                    cached["cache_status"] = "hit"
                    return cached
            except json.JSONDecodeError:
                pass
    if not archive:
        allowed, reason = _source_allowed(url)
        if not allowed:
            return {"url": url, "final_url": url, "status": 0, "title": "", "text": "", "error": reason, "method": "source_registry_skipped"}
    if client is None:
        with httpx.Client(follow_redirects=True) as owned_client:
            return fetch_page(owned_client, url, archive=archive, cache_dir=cache_dir)
    if archive:
        curl_result = _curl_get(url, max_time=ARCHIVE_CURL_TIMEOUT_SECONDS)
        if not curl_result.get("ok"):
            curl_result = _curl_get(url, max_time=max(ARCHIVE_CURL_TIMEOUT_SECONDS * 2, 60))
        title, text = _html_to_text(curl_result.get("text", "").encode("utf-8", "replace"), curl_result.get("content_type", ""))
        result = {
            "url": url,
            "final_url": curl_result.get("final_url", url),
            "status": curl_result.get("status", 0),
            "title": title,
            "text": text,
            "bytes": len(curl_result.get("text", "")),
            "content_type": curl_result.get("content_type", ""),
            "error": "" if curl_result.get("status", 0) and curl_result.get("status", 0) < 400 else curl_result.get("error", "curl_error"),
            "method": "curl",
        }
        if cache_dir:
            cache_dir.mkdir(parents=True, exist_ok=True)
            _cache_file(cache_dir, url).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        return result
    attempts = 1 if archive else 3
    last_exc: Exception | None = None
    for attempt in range(attempts):
        try:
            resp = client.get(
                url,
                headers={"User-Agent": "CMHK-product-crawl/1.0 (+public pages only)", "Accept-Language": "en,zh-CN;q=0.9,zh;q=0.8"},
                timeout=httpx.Timeout(ARCHIVE_TIMEOUT_SECONDS, connect=8.0) if archive else httpx.Timeout(25.0, connect=8.0),
            )
            title, text = _html_to_text(resp.content, resp.headers.get("content-type", ""))
            result = {
                "url": url,
                "final_url": str(resp.url),
                "status": resp.status_code,
                "title": title,
                "text": text,
                "bytes": len(resp.content),
                "content_type": resp.headers.get("content-type", ""),
                "error": "" if resp.status_code < 400 else f"http_{resp.status_code}",
                "method": "httpx" if attempt == 0 else f"httpx_retry_{attempt}",
            }
            if not archive and resp.status_code in {403, 406}:
                try:
                    browser_resp = client.get(
                        url,
                        headers={
                            "User-Agent": "Mozilla/5.0",
                            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                            "Accept-Language": "en,zh-CN;q=0.9,zh;q=0.8",
                        },
                        timeout=httpx.Timeout(25.0, connect=8.0),
                    )
                    if browser_resp.status_code < 400:
                        title, text = _html_to_text(browser_resp.content, browser_resp.headers.get("content-type", ""))
                        result = {
                            "url": url,
                            "final_url": str(browser_resp.url),
                            "status": browser_resp.status_code,
                            "title": title,
                            "text": text,
                            "bytes": len(browser_resp.content),
                            "content_type": browser_resp.headers.get("content-type", ""),
                            "error": "",
                            "method": "httpx_browser_ua_retry",
                        }
                except Exception:
                    pass
            if cache_dir:
                cache_dir.mkdir(parents=True, exist_ok=True)
                _cache_file(cache_dir, url).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
            return result
        except Exception as exc:
            last_exc = exc
            if attempt < attempts - 1:
                time.sleep(0.5 * (attempt + 1))
    curl_result = _curl_get(url, max_time=ARCHIVE_CURL_TIMEOUT_SECONDS)
    if curl_result.get("status", 0) and curl_result.get("status", 0) < 400 and curl_result.get("text"):
        title, text = _html_to_text(curl_result.get("text", "").encode("utf-8", "replace"), curl_result.get("content_type", ""))
        result = {
            "url": url,
            "final_url": curl_result.get("final_url", url),
            "status": curl_result.get("status", 0),
            "title": title,
            "text": text,
            "bytes": len(curl_result.get("text", "")),
            "content_type": curl_result.get("content_type", ""),
            "error": "",
            "method": "httpx_error_curl_fallback",
        }
        if cache_dir:
            cache_dir.mkdir(parents=True, exist_ok=True)
            _cache_file(cache_dir, url).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        return result
    result = {"url": url, "final_url": url, "status": 0, "title": "", "text": "", "bytes": 0, "content_type": "", "error": str(last_exc), "method": "httpx_error"}
    if cache_dir:
        cache_dir.mkdir(parents=True, exist_ok=True)
        _cache_file(cache_dir, url).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def _fetch_page(
    client: httpx.Client | None,
    url: str,
    *,
    archive: bool = False,
    cache_dir: Path | None = None,
) -> Dict[str, Any]:
    try:
        return fetch_page(client, url, archive=archive, cache_dir=cache_dir)
    except TypeError as exc:
        if "cache_dir" not in str(exc) and "archive" not in str(exc):
            raise
    try:
        return fetch_page(client, url, archive=archive)
    except TypeError as exc:
        if "archive" not in str(exc):
            raise
        return fetch_page(client, url)


def _service_generation(product_category: str, text: str) -> str:
    blob = f"{product_category} {text}".lower()
    if "fibre" in product_category.lower() or "fiber" in product_category.lower() or "broadband" in product_category.lower():
        return "Fibre/Broadband"
    if "4g" in product_category.lower():
        return "4G"
    if "5g" in product_category.lower():
        return "5G"
    if "5g" in blob:
        return "5G"
    if "4.5g" in blob:
        return "4.5G"
    if "4g" in blob:
        return "4G"
    if "fibre" in blob or "fiber" in blob or "broadband" in blob:
        return "Fibre/Broadband"
    return ""


def _customer_segment(product_category: str) -> str:
    if "business" in product_category:
        return "企业客户"
    if "broadband" in product_category:
        return "家庭宽频客户"
    if "roaming" in product_category:
        return "漫游/跨境客户"
    if "prepaid" in product_category:
        return "预付卡客户"
    return "个人客户"


def _plan_family(source: Dict[str, str], text: str) -> str:
    category = source["product_category"]
    brand = source["brand"]
    if "world" in source["source_id"]:
        return "3HK World Plan"
    if "sosim" in source["source_id"]:
        return "SoSIM"
    if brand == "SmarTone" and "home" in category:
        return "SmarTone Home 5G Broadband"
    if brand == "SmarTone":
        return "SmarTone 5G"
    if brand == "HKBN" and "5g" in category:
        return "HKBN 5G Home Broadband"
    if brand == "HKBN" and "mobile" in category:
        return "HKBN Mobile"
    if brand == "HKBN":
        return "HKBN Home Broadband"
    if brand == "HGC" and "broadband" in category:
        return "HGC Home Broadband"
    if brand == "HGC":
        return "HGC on air"
    if brand == "i-CABLE" and "broadband" in category:
        return "i-CABLE Broadband"
    if brand == "i-CABLE" and "mobile" in text.lower():
        return "i-CABLE Mobile"
    if brand == "i-CABLE":
        return "i-CABLE Broadband"
    return brand


def _excerpt(text: str, start: int, end: int, radius: int = 180) -> str:
    left = max(0, start - radius)
    right = min(len(text), end + radius)
    return text[left:right].strip()


def _extract_number_before(pattern: str, text: str, pos: int) -> str:
    window = text[max(0, pos - 180) : pos]
    matches = list(re.finditer(pattern, window, flags=re.I))
    return matches[-1].group(1) if matches else ""


def _extract_number_after(pattern: str, text: str, pos: int) -> str:
    window = text[pos : min(len(text), pos + 180)]
    match = re.search(pattern, window, flags=re.I)
    return match.group(1) if match else ""


def _normalise_speed_mbps(value: str, unit: str) -> str:
    try:
        number = float(value)
    except ValueError:
        return value
    if unit.lower().startswith("g"):
        number *= 1000
    if number.is_integer():
        return str(int(number))
    return str(number)


def _extract_speed_before(text: str, pos: int) -> str:
    window = text[max(0, pos - 180) : pos]
    matches = list(re.finditer(r"([0-9]+(?:\.[0-9]+)?)\s*(Gbps|Mbps|G|M)\b", window, flags=re.I))
    if not matches:
        return ""
    match = matches[-1]
    return _normalise_speed_mbps(match.group(1), match.group(2))


def _extract_speed_after(text: str, pos: int) -> str:
    window = text[pos : min(len(text), pos + 180)]
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(Gbps|Mbps|G|M)\b", window, flags=re.I)
    if not match:
        return ""
    return _normalise_speed_mbps(match.group(1), match.group(2))


def _base_plan_row(source: Dict[str, str], result: Dict[str, Any], captured_at: str, period_label: str, archive_url: str = "") -> Dict[str, str]:
    text = _clean_text(str(result.get("text") or ""))
    return {
        "captured_at_hkt": captured_at,
        "period_label": period_label,
        "brand": source["brand"],
        "product_category": source["product_category"],
        "service_generation": _service_generation(source["product_category"], text),
        "customer_segment": _customer_segment(source["product_category"]),
        "plan_family": _plan_family(source, text),
        "source_id": source["source_id"],
        "source_url": source["url"],
        "archive_url": archive_url,
        "http_status": str(result.get("status") or ""),
        "content_hash": _content_hash(text),
    }


def _tariff_type(excerpt: str) -> str:
    if re.search(r"monthly|month|basic monthly|standard monthly|月費|月费|月費計劃|合約|contract", excerpt, flags=re.I):
        return "monthly_plan_fee"
    if re.search(r"installation|relocation|cancellation|inspection|one[- ]off|per time|damage|搬遷|搬迁|安裝|安装|取消|檢查|检查|一次", excerpt, flags=re.I):
        return "one_off_service_charge"
    if re.search(r"daily|day pass|日費|日费", excerpt, flags=re.I):
        return "daily_pass_fee"
    return "price_mentioned_needs_review"


def _parse_page(source: Dict[str, str], result: Dict[str, Any], captured_at: str, period_label: str, archive_url: str = "") -> List[Dict[str, str]]:
    text = _clean_text(str(result.get("text") or ""))
    if not text:
        return []
    if source.get("source_id") in {
        "3hk_whatsapp_premium_20150506_group_press",
        "3hk_whatsapp_premium_20150506_product_pdf",
    }:
        if not (
            re.search(r"WhatsApp Premium Data Pack", text, flags=re.I)
            and re.search(r"\$\s*18\s+a month", text, flags=re.I)
            and re.search(r"WhatsApp Premium Roaming Pass", text, flags=re.I)
            and re.search(r"\$\s*48\s+a day", text, flags=re.I)
        ):
            return []
        rows: List[Dict[str, str]] = []
        for plan_name, fee, tariff_type, benefits, match_pattern in (
            (
                "3HK WhatsApp Premium Data Pack HK$18/month",
                "18",
                "monthly_plan_fee",
                "Unlimited local WhatsApp messaging, photo/video sharing, recorded audio messages and WhatsApp Calling; existing plan data unaffected.",
                r"WhatsApp Premium Data Pack.*?\$\s*18\s+a month",
            ),
            (
                "3HK WhatsApp Premium Roaming Pass HK$48/day",
                "48",
                "daily_pass_fee",
                "Promotional day-pass price; WhatsApp messaging and Calling at 151 destinations. Normal price HK$88/day.",
                r"WhatsApp Premium Roaming Pass.*?\$\s*48\s+a day",
            ),
        ):
            match = re.search(match_pattern, text, flags=re.I | re.S)
            if not match:
                return []
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": benefits,
                    "add_on_charges_hkd": "",
                    "tariff_type": tariff_type,
                    "source_status": "parsed_official_press_release" if source["source_id"].endswith("group_press") else "parsed_official_product_pdf",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {
        "3hk_ipair_20131127_en_official_press",
        "3hk_ipair_20131127_tc_official_press",
    }:
        is_english = source["source_id"].endswith("_en_official_press")
        required = (
            re.search(r"i-Pair.*?\$28 a month", text, flags=re.I | re.S)
            and re.search(r"\$49 monthly plan", text, flags=re.I)
            if is_english
            else re.search(r"愛情公寓.*?\$28\s*月費", text, flags=re.S)
            and re.search(r"\$49\s*月費", text)
        )
        if not required:
            return []
        specs = (
            (
                "3HK i-Pair designated value-added service HK$28/month",
                "28",
                "Designated value-added i-Pair membership for 3HK customers subscribing to a monthly plan with handset purchase or upgrade.",
                r"i-Pair.*?\$28 a month" if is_english else r"愛情公寓.*?\$28\s*月費",
            ),
            (
                "3HK i-Pair standalone platinum membership HK$49/month",
                "49",
                "Standalone i-Pair platinum membership; the release states additional designated upgrade service for the first month.",
                r"\$49 monthly plan" if is_english else r"\$49\s*月費.*?獨立申請",
            ),
        )
        rows: List[Dict[str, str]] = []
        for plan_name, fee, benefits, pattern in specs:
            match = re.search(pattern, text, flags=re.I | re.S)
            if not match:
                return []
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": benefits,
                    "add_on_charges_hkd": "Official tariff/list price. First-month upgraded service, virtual currency/items, handset offers and mobile-plan eligibility are promotions or conditions, not included in the i-Pair monthly fee.",
                    "tariff_type": "monthly_value_added_service_fee",
                    "source_status": "parsed_official_press_release",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {
        "3hk_samsung_galaxy_s2_20110610_en_official_press",
        "3hk_samsung_galaxy_s2_20110610_tc_official_press",
    }:
        is_english = source["source_id"].endswith("_en_official_press")
        match = (
            re.search(r"HK\$148 Smart Unlimited Monthly Plan", text, flags=re.I)
            if is_english
            else re.search(r"\$148\s*[「『]?無限\s*Smart[」』]?月費計劃", text, flags=re.S)
        )
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "3HK Smart Unlimited Monthly Plan HK$148",
                "monthly_fee_hkd": "148",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "Official Galaxy S II pre-order offer: the HK$148 Smart Unlimited Monthly Plan was eligible for a HK$0 handset. The press release does not directly disclose the plan's baseline data, voice allowance or contract term, so those fields remain blank.",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_official_press_release",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {
        "3hk_anyplex_htc_20111101_en_official_press",
        "3hk_anyplex_htc_20111101_tc_official_press",
    }:
        if not (re.search(r"Anyplex", text, flags=re.I) and re.search(r"\$\s*49", text)):
            return []
        rows: List[Dict[str, str]] = []
        specs = [
            (
                "3HK Anyplex cross-platform movie service HK$49/month",
                "49", "", "",
                "Anyplex movie-on-demand service with unlimited access to nearly 200 selected movies on compatible devices. First-year extra movie offers are promotional and excluded from the baseline plan.",
                r"Anyplex.*?\$\s*49|\$\s*49.*?Anyplex",
            ),
        ]
        is_english = source["source_id"].endswith("_en_official_press")
        if is_english and re.search(r"\$298\s+\"Smart Unlimited\" monthly plan", text, flags=re.I):
            specs.append(
                (
                    "3HK Smart Unlimited Monthly Plan HK$298 800MB",
                    "298", "0.8", "",
                    "2,500 basic airtime minutes, 800MB local wireless data and unlimited Wi-Fi stated in the handset offer. Six-month unlimited-data and handset replacement offers are promotional and excluded.",
                    r"\$298\s+\"Smart Unlimited\" monthly plan.*?2,500.*?800MB",
                )
            )
        elif not is_english and re.search(r"\$298.*?800MB", text, flags=re.S):
            specs.append(
                (
                    "3HK Smart Unlimited Monthly Plan HK$298 800MB",
                    "298", "0.8", "",
                    "2,500 basic airtime minutes, 800MB local wireless data and unlimited Wi-Fi stated in the handset offer. Six-month unlimited-data and handset replacement offers are promotional and excluded.",
                    r"\$298.*?800MB",
                )
            )
        for plan_name, fee, data, contract, benefits, pattern in specs:
            match = re.search(pattern, text, flags=re.I | re.S)
            if not match:
                return []
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update({"plan_name": plan_name, "monthly_fee_hkd": fee, "average_monthly_fee_hkd": "", "local_data_gb": data, "roaming_data_gb": "", "broadband_speed_mbps": "", "post_fup_speed_mbps": "", "contract_months": contract, "local_voice": benefits, "add_on_charges_hkd": "Official press-release tariff or handset-offer price; only directly stated baseline fields are structured.", "tariff_type": "monthly_value_added_service_fee" if fee == "49" else "monthly_plan_fee", "source_status": "parsed_official_press_release", "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850]})
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"3hk_skype_unlimited_world_20140819_en_official_press", "3hk_skype_unlimited_world_20140819_tc_official_press"}:
        if not (re.search(r"Skype", text, flags=re.I) and re.search(r"\$\s*69", text) and re.search(r"12\s*(?:months|個月)", text, flags=re.I)):
            return []
        match = re.search(r"(?:Skype.*?\$\s*69.*?12\s*(?:months|個月)|\$\s*69.*?12\s*(?:months|個月).*?Skype)", text, flags=re.I | re.S)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update({"plan_name": "3HK Skype Unlimited World HK$69/month", "monthly_fee_hkd": "69", "average_monthly_fee_hkd": "", "local_data_gb": "", "roaming_data_gb": "", "broadband_speed_mbps": "", "post_fup_speed_mbps": "", "contract_months": "12", "local_voice": "Skype-to-Skype and designated Skype-to-phone calling under the published Unlimited World value-added service.", "add_on_charges_hkd": "Official 3HK value-added service price. The release compares a market price and describes destinations; those are not additional monthly charges.", "tariff_type": "monthly_value_added_service_fee", "source_status": "parsed_official_press_release", "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850]})
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {
        "3hk_3ree_broadband_20100722_en_official_press",
        "3hk_3ree_broadband_20100722_tc_official_press",
    }:
        if not (
            re.search(r"(?:100M Residential Broadband Service|100M家居寬頻服務)", text, flags=re.I)
            and re.search(r"(?:monthly (?:subscription )?fee of only \$99|月費只需\$99)", text, flags=re.I)
        ):
            return []
        match = re.search(r"(?:100M Residential Broadband Service.*?\$99|100M家居寬頻服務.*?\$99)", text, flags=re.I | re.S)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "3ree Broadband 100M Residential Broadband Service HK$99",
                "monthly_fee_hkd": "99",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "100",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "New subscribers could choose six months of free 100M broadband or 30 months of free residential telephone service; offer choice is not an included baseline entitlement.",
                "add_on_charges_hkd": "Official promotional monthly broadband fee; published statement does not disclose a contract term for the HK$99 price.",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_official_press_release",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {
        "3home_entertainment_super_pack_20160427_group_press",
        "3home_entertainment_super_pack_20160427_product_pdf",
    }:
        if not (
            re.search(r"Home Entertainment Super Pack", text, flags=re.I)
            and re.search(r"\$\s*138", text)
            and re.search(r"\$\s*188", text)
            and re.search(r"24\s*months", text, flags=re.I)
        ):
            return []
        rows = []
        for plan_name, fee, speed, match_pattern in (
            ("3Home Broadband Home Entertainment Super Pack 100M HK$138", "138", "100", r"100M.*?\$\s*138.*?24\s*months"),
            ("3Home Broadband Home Entertainment Super Pack 1G HK$188", "188", "1000", r"1G.*?\$\s*188.*?24\s*months"),
        ):
            match = re.search(match_pattern, text, flags=re.I | re.S)
            if not match:
                return []
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "24",
                    "local_voice": "Residential telephone service, HGC On Air Wi-Fi, myTV SUPER and Google Chromecast included in the bundle.",
                    "add_on_charges_hkd": "Official promotional bundle price; device and content inclusion follows the published offer terms.",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_official_press_release" if source["source_id"].endswith("group_press") else "parsed_official_product_pdf",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {
        "3hk_3gamer_20171207_en_official_press",
        "3hk_3gamer_20171207_tc_official_press",
    }:
        has_sim = bool(
            re.search(r"(?:HK\$|\$)\s*218", text, flags=re.I)
            and re.search(r"(?:24[- ]month contract|簽約\s*24\s*個月)", text, flags=re.I)
            and re.search(r"5\s*GB", text, flags=re.I)
        )
        has_membership = bool(re.search(r"(?:HK\$|\$)\s*48", text, flags=re.I) and re.search(r"3Gamer.*?(?:membership|會藉)", text, flags=re.I | re.S))
        if not has_sim or not has_membership:
            return []
        rows = []
        for plan_name, fee, category, data, contract, benefits, pattern in (
            (
                "3HK 3Gamer SIM Monthly Plan HK$218 5GB",
                "218",
                "mobile_consumer_4g",
                "5",
                "24",
                "Unlimited designated 3Gamer game data; designated YouTube viewing 9pm-11pm; HK$30 monthly game-item credit.",
                r"(?:HK\$|\$)\s*218.*?(?:24[- ]month contract|簽約\s*24\s*個月).*?5\s*GB|(?:每月\$218|HK\$218).*?5\s*GB",
            ),
            (
                "3HK 3Gamer Membership HK$48/month",
                "48",
                "mobile_consumer_value_added_service",
                "",
                "",
                "Existing monthly-plan customers: designated mobile-game data access and limited game-item pack eligibility.",
                r"(?:monthly fee of HK\$48|月費只需\$48)",
            ),
        ):
            match = re.search(pattern, text, flags=re.I | re.S)
            if not match:
                return []
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "product_category": category,
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": benefits,
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_official_press_release",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hgc_on_air_plan_select", "hgc_on_air_plan_en", "hgc_on_air_plan_tc"}:
        match = re.search(
            r"(?:HGC Broadband Service subscribers \$58 monthly fee|月費港幣\$58元收取)",
            text,
            flags=re.I,
        )
        if not match or not re.search(r"hgc on air|月費計劃|Monthly Plan", text, flags=re.I):
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HGC on air HK$58",
                "monthly_fee_hkd": "58",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "HGC Broadband customers only; first month pro-rated; monthly plan continues after free or promotional period unless terminated",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_current",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=260)[:800],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hgc_2g_broadband_2023_press":
        if not (
            re.search(r"Home Broadband Monthly Service Plan", text, flags=re.I)
            and re.search(r"Broadband Service Plan\s+Standard Monthly Fee", text, flags=re.I)
        ):
            return []
        rows: List[Dict[str, str]] = []
        plan_specs = [
            ("10G", "1299", "10000"),
            ("2.5G", "766", "2500"),
            ("2.2G", "666", "2200"),
            ("2G", "666", "2000"),
            ("1G", "598", "1000"),
            ("500M", "488", "500"),
            ("300M", "398", "300"),
            ("200M", "348", "200"),
            ("100M", "298", "100"),
            ("30M", "198", "30"),
            ("10M", "188", "10"),
            ("6M", "188", "6"),
        ]
        table_match = re.search(
            r"Home Broadband Monthly Service Plan.*?Basic Monthly Fee.*?6M\s+\$188",
            text,
            flags=re.I | re.S,
        )
        table_text = table_match.group(0) if table_match else text
        for label, fee, speed in plan_specs:
            display_fee = f"{int(fee):,}" if fee.isdigit() else fee
            pattern = rf"{re.escape(label)}\s+\$(?:{re.escape(fee)}|{re.escape(display_fee)})\b"
            match = re.search(pattern, table_text, flags=re.I)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HGC Home Broadband {label} standard monthly fee HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_current",
                    "evidence_excerpt": _excerpt(table_text, match.start(), match.end(), radius=220)[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hkbn_enterprise_mobile_4g_offer", "hkbn_enterprise_mobile_4g_offer_tc", "hkbn_enterprise_mobile_4g_offer_sc"}:
        rows: List[Dict[str, str]] = []
        plan_specs = [
            ("Local", "98", "118", "5", "", "2000", ""),
            ("Local", "118", "138", "3", "", "3000", ""),
            ("Local", "148", "168", "6", "", "3500", ""),
            ("Local", "178", "198", "8", "", "4000", ""),
            ("Greater China", "198", "238", "1", "1", "", "30/0.5GB extra data; 18 administration fee waived during contract period"),
            ("Greater China", "258", "298", "3", "3", "", "30/0.5GB extra data; 18 administration fee waived during contract period"),
            ("Greater China", "318", "358", "6", "6", "", "20/0.5GB extra data; 18 administration fee waived during contract period"),
            ("Greater China", "448", "488", "10", "10", "", "20/0.5GB extra data; 18 administration fee waived during contract period"),
        ]
        for family, special_fee, standard_fee, data_gb, roaming_gb, voice_minutes, add_on in plan_specs:
            pattern = (
                rf"Special Monthly Fee:\s*\${special_fee}\s+Standard Monthly Fee:\s*\${standard_fee}.*?{data_gb}GB\s+(?:Local Data|Data)"
                rf"|優惠月費[：:]\s*\${special_fee}\s+標準月費[：:]\s*\${standard_fee}.*?{data_gb}GB(?:本地)?數據"
                rf"|优惠月费[：:]\s*\${special_fee}\s+标准月费[：:]\s*\${standard_fee}.*?{data_gb}GB(?:本地)?数据"
            )
            match = re.search(pattern, text, flags=re.I | re.S)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HKBN Enterprise {family} 4G Plan HK${standard_fee} {data_gb}GB",
                    "monthly_fee_hkd": standard_fee,
                    "average_monthly_fee_hkd": special_fee,
                    "local_data_gb": data_gb,
                    "roaming_data_gb": roaming_gb,
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "24",
                    "local_voice": f"{voice_minutes} local voice minutes" if voice_minutes else "unlimited local airtime",
                    "add_on_charges_hkd": add_on or "18 administration fee waived during contract period",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_current",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180)[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"smartone_roaming_pack", "smartone_roaming_pack_tc"}:
        if not re.search(r"APAC/ Worldwide Roaming Data Pack|亞太/環球地區漫遊數據通", text, flags=re.I):
            return []
        rows: List[Dict[str, str]] = []
        specs = [
            (
                "SmarTone APAC Roaming Data Pack HK$269 10GB",
                "269",
                "10",
                r"HK\$\s*269\s*/\s*10GB|HK\$\s*269\s*/\s*10\s*GB",
                "Asia Pacific 10GB roaming data pack; activate within 1 month and valid for 90 days after activation",
                "roaming_data_pack_fee",
            ),
            (
                "SmarTone APAC Roaming Data Pack top-up HK$38 1GB",
                "38",
                "1",
                r"HK\$\s*38\s*/\s*GB",
                "Asia Pacific roaming data pack top-up option",
                "roaming_data_top_up_fee",
            ),
            (
                "SmarTone Worldwide Roaming Data Pack HK$549 10GB",
                "549",
                "10",
                r"HK\$\s*549\s*/\s*10GB|HK\$\s*549\s*/\s*10\s*GB",
                "Worldwide hotspots 10GB roaming data pack; activate within 1 month and valid for 90 days after activation",
                "roaming_data_pack_fee",
            ),
            (
                "SmarTone Worldwide Roaming Data Pack top-up HK$68 1GB",
                "68",
                "1",
                r"HK\$\s*68\s*/\s*GB",
                "Worldwide hotspots roaming data pack top-up option",
                "roaming_data_top_up_fee",
            ),
        ]
        for plan_name, fee, roaming_gb, pattern, add_on, tariff_type in specs:
            match = re.search(pattern, text, flags=re.I)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": roaming_gb,
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": add_on,
                    "tariff_type": tariff_type,
                    "source_status": "parsed_current",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=360)[:850],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"3hk_world_plan", "3hk_world_plan_tc", "3hk_world_plan_alt", "3hk_world_plan_alt_tc"}:
        is_alt = "alt" in source.get("source_id", "")
        if not re.search(r"Worldwide shared data|世界通用數據|Asia Pacific shared data|亞太通用數據", text, flags=re.I):
            return []
        rows: List[Dict[str, str]] = []
        plan_specs = (
            [("10", "158", "50 minutes", "10 minutes"), ("20", "208", "50 minutes", "10 minutes"), ("30", "258", "50 minutes", "10 + 10 minutes"), ("50", "288", "50 minutes", "10 + 10 minutes"), ("70", "388", "100 minutes", "10 + 10 minutes"), ("100", "588", "200 minutes", "20 + 20 minutes"), ("200", "798", "300 minutes", "20 + 20 minutes")]
            if is_alt
            else [("10", "198", "50 minutes", "10 minutes"), ("20", "268", "50 minutes", "10 minutes"), ("30", "338", "50 minutes", "10 + 10 minutes"), ("50", "368", "50 minutes", "10 + 10 minutes"), ("70", "468", "100 minutes", "10 + 10 minutes"), ("100", "728", "200 minutes", "20 + 20 minutes"), ("200", "958", "300 minutes", "20 + 20 minutes")]
        )
        for data, fee, cn_mo_voice, roaming_voice in plan_specs:
            pattern = rf"{data}GB\s*\$\s*{fee}\b|{data}GB\s+\${fee}\b"
            match = re.search(pattern, text, flags=re.I)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"3HK World Plan HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": data,
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "1",
                    "contract_months": "",
                    "local_voice": f"Chinese Mainland and Macau voice {cn_mo_voice}; roaming voice {roaming_voice}",
                    "add_on_charges_hkd": "Limited-time admin fee waiver noted on source page",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_current",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=260)[:750],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"3hk_ofca_4g_smartphone_plan_20120503", "3hk_ofca_4g_smartphone_plan_20120530"}:
        match = re.search(r"4G LTE Smartphone SIM Plan Monthly Fee \$358.*?after reaching 5GB", text, flags=re.I | re.S)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "3HK 4G LTE Smartphone SIM Plan HK$358 unlimited data",
                "monthly_fee_hkd": "358",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "3,900 basic mins; 1,600 intra-network mins; 100 video mins",
                "add_on_charges_hkd": "Monthly MTR/tunnels/mobile services licence fee is disclosed separately in the tariff and excluded from the plan fee.",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_official_regulatory_pdf_ocr_verified",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "3hk_ofca_smartphone_super_plan_20140211":
        match = re.search(r"Tariff Name\s+3G/4G LTE Smartphone Super Plan.*?\$168\s+1400\s+500\s+500MB", text, flags=re.I | re.S)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "3HK 3G/4G LTE Smartphone Super Plan HK$168",
                "monthly_fee_hkd": "168",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "1,400 basic mins; 500 intra-network mins",
                "add_on_charges_hkd": "Official tariff separately discloses HK$12 monthly administration fee; excluded from plan fee.",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_official_regulatory_pdf",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "3hk_jolla_smartphone_super_plan_20140812":
        match = re.search(r"subscribing to \$168 to \$598 4G LTE smartphone super plan", text, flags=re.I)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "3HK 3G/4G LTE Smartphone Super Plan HK$168",
                "monthly_fee_hkd": "168",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "Official press release confirms the same plan family spans HK$168-HK$598; specific plan inclusions are not inferred from the press release.",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_official_press_release",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:850],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {
        "3hk_greater_china_plan_20180720_official_pdf",
        "3hk_greater_china_plan_20180720_qooah",
    }:
        if not (
            re.search(r"共享大中華|Greater China Plan", text, flags=re.I)
            and all(re.search(rf"\$\s*{fee}", text) for fee in ("188", "238", "328"))
            and all(re.search(rf"{data}\s*GB", text, flags=re.I) for data in ("4", "8", "12"))
        ):
            return []
        is_official_pdf = source.get("source_id") == "3hk_greater_china_plan_20180720_official_pdf"
        match = re.search(
            r"(?:月費分為|月費.*?\$\s*188).*?\$\s*188.*?\$\s*238.*?\$\s*328.*?4\s*GB.*?8\s*GB.*?12\s*GB",
            text,
            flags=re.I | re.S,
        )
        excerpt = _excerpt(text, match.start(), match.end(), radius=180)[:850] if match else _clean_text(text)[:850]
        rows: List[Dict[str, str]] = []
        for fee, data in (("188", "4"), ("238", "8"), ("328", "12")):
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"3HK Greater China Monthly Plan HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": data,
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "24" if is_official_pdf else "",
                    "local_voice": "unlimited local voice mins mentioned" if is_official_pdf else "",
                    "add_on_charges_hkd": "18 monthly administration fee" if is_official_pdf else "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_official_press_pdf" if is_official_pdf else "parsed_public_media_reference",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"3hk_5g_sim_plan", "3hk_5g_sim_plan_tc"}:
        rows: List[Dict[str, str]] = []
        if re.search(r"5G Monthly SIM Plan|5G SIM 月費計劃", text, flags=re.I) and re.search(r"Avg\. monthly fee from|平均月費低至", text, flags=re.I):
            for fee, avg_fee, data in [("124", "100", "30"), ("148", "119", "50"), ("188", "151", "60"), ("228", "183", "100")]:
                fee_match = re.search(rf"\$\s*{avg_fee}\s*\${fee}", text, flags=re.I) or re.search(rf"Plan\s*\${fee}", text, flags=re.I)
                start, end = fee_match.span() if fee_match else (0, min(len(text), 160))
                row = _base_plan_row(source, result, captured_at, period_label, archive_url)
                row.update(
                    {
                        "plan_name": f"3HK 5G Monthly SIM Plan HK${fee} {data}GB",
                        "monthly_fee_hkd": fee,
                        "average_monthly_fee_hkd": avg_fee,
                        "local_data_gb": data,
                        "roaming_data_gb": "",
                        "broadband_speed_mbps": "",
                        "post_fup_speed_mbps": "1",
                        "contract_months": "",
                        "local_voice": "3000 local voice minutes",
                        "add_on_charges_hkd": "",
                        "tariff_type": "monthly_plan_fee",
                        "source_status": "parsed_current",
                        "evidence_excerpt": _excerpt(text, start, end)[:650],
                    }
                )
                row["record_key"] = _record_key(row)
                rows.append(row)
            for fee, data in [("68", "10"), ("128", "20")]:
                add_on_match = re.search(rf"\${fee}/{data}GB", text, flags=re.I)
                if not add_on_match:
                    continue
                row = _base_plan_row(source, result, captured_at, period_label, archive_url)
                row.update(
                    {
                        "plan_name": f"3HK Chinese Mainland-HK-Macau Shared Data add-on HK${fee} {data}GB",
                        "monthly_fee_hkd": fee,
                        "average_monthly_fee_hkd": "",
                        "local_data_gb": "",
                        "roaming_data_gb": data,
                        "broadband_speed_mbps": "",
                        "post_fup_speed_mbps": "",
                        "contract_months": "",
                        "local_voice": "",
                        "add_on_charges_hkd": "",
                        "tariff_type": "monthly_add_on_fee",
                        "source_status": "parsed_current",
                        "evidence_excerpt": _excerpt(text, add_on_match.start(), add_on_match.end())[:650],
                    }
                )
                row["record_key"] = _record_key(row)
                rows.append(row)
            return rows
        if source.get("source_id") == "3hk_5g_sim_plan_tc" and re.search(r"5G\s*(?:SIM)?\s*月費計劃|5G 月費計劃", text, flags=re.I):
            seen_plans: set[tuple[str, str]] = set()
            plan_matches: List[tuple[str, str, str, int, int]] = []
            table_patterns = [
                re.compile(
                    r"(?<![+\w])\$\s*(?P<fee>[0-9]{3})\s*/\s*月(?:\s*[0-9,]+)?(?:\s*\+\s*/月\s*✭)?(?:\s*(?:最新|熱賣|限時|入門首選))*\s*本地數據\s*(?:∆\s*)?(?P<data>[0-9]{2,3})\s*GB",
                    flags=re.I | re.S,
                ),
                re.compile(
                    r"(?:基本月費|平均月費)\s*[0-9,]*\s*\$\s*(?:(?P<avg>[0-9]{3})\s*)?\$?\s*(?P<fee>[0-9]{3})\s+本地數據\s*(?:∆\s*)?(?P<data>[0-9]{2,3})\s*GB",
                    flags=re.I | re.S,
                ),
            ]
            for table_pattern in table_patterns:
                for match in table_pattern.finditer(text):
                    fee = match.group("fee")
                    data = match.group("data")
                    avg_fee = match.groupdict().get("avg") or ""
                    key = (fee, data)
                    if key in seen_plans:
                        continue
                    seen_plans.add(key)
                    plan_matches.append((fee, data, avg_fee, match.start(), match.end()))
            for fee, data, avg_fee, start, end in plan_matches:
                window = text[start: min(len(text), end + 320)]
                contract_match = re.search(r"([0-9]{2})\s*個月合約", window)
                row = _base_plan_row(source, result, captured_at, period_label, archive_url)
                row.update(
                    {
                        "plan_name": f"3HK 5G Monthly SIM Plan HK${fee} {data}GB",
                        "monthly_fee_hkd": fee,
                        "average_monthly_fee_hkd": avg_fee,
                        "local_data_gb": data,
                        "roaming_data_gb": "",
                        "broadband_speed_mbps": "",
                        "post_fup_speed_mbps": "1" if re.search(r"高達\s*1Mbps|1Mbps", window, flags=re.I) else "",
                        "contract_months": contract_match.group(1) if contract_match else "",
                        "local_voice": "3000 local voice minutes" if re.search(r"3,000本地通話分鐘|3000本地通話", window) else "",
                        "add_on_charges_hkd": "28 admin fee per month" if re.search(r"行政費\s*\$28/月", text) else "",
                        "tariff_type": "monthly_plan_fee",
                        "source_status": "parsed_archive" if archive_url else "parsed_current",
                        "evidence_excerpt": _excerpt(text, start, end, radius=260)[:800],
                    }
                )
                row["record_key"] = _record_key(row)
                rows.append(row)
            if rows:
                return rows
        if source.get("source_id") == "3hk_5g_sim_plan" and re.search(r"5G\s*(?:SIM)?\s*Monthly\s+(?:SIM\s+)?Plan|5G Monthly Plan", text, flags=re.I):
            seen_plans: set[tuple[str, str]] = set()
            plan_matches: List[tuple[str, str, str, int, int]] = []
            table_patterns = [
                re.compile(
                    r"(?<![+\w])\$\s*(?P<fee>[0-9]{3})\s*/\s*(?:month|mth)(?:\s*[0-9,]+)?(?:\s*\+\s*/(?:month|mth)\s*✭)?(?:\s*(?:NEW|HOT|Limited-time))*\s*Local Data\s*(?P<data>[0-9]{2,3})\s*GB",
                    flags=re.I | re.S,
                ),
                re.compile(
                    r"(?:Monthly fee|Average Monthly Fee)\s*[0-9,]*\s*\$\s*(?:(?P<avg>[0-9]{3})\s*)?\$?\s*(?P<fee>[0-9]{3})\s+Local data\s*(?:∆\s*)?(?P<data>[0-9]{2,3})\s*GB",
                    flags=re.I | re.S,
                ),
            ]
            for table_pattern in table_patterns:
                for match in table_pattern.finditer(text):
                    fee = match.group("fee")
                    data = match.group("data")
                    avg_fee = match.groupdict().get("avg") or ""
                    key = (fee, data)
                    if key in seen_plans:
                        continue
                    seen_plans.add(key)
                    plan_matches.append((fee, data, avg_fee, match.start(), match.end()))
            for fee, data, avg_fee, start, end in plan_matches:
                window = text[start: min(len(text), end + 320)]
                contract_match = re.search(r"([0-9]{2})[-\s]*month\s+Contract", window, flags=re.I)
                row = _base_plan_row(source, result, captured_at, period_label, archive_url)
                row.update(
                    {
                        "plan_name": f"3HK 5G Monthly SIM Plan HK${fee} {data}GB",
                        "monthly_fee_hkd": fee,
                        "average_monthly_fee_hkd": avg_fee,
                        "local_data_gb": data,
                        "roaming_data_gb": "",
                        "broadband_speed_mbps": "",
                        "post_fup_speed_mbps": "1" if re.search(r"Up to\s*1Mbps|1Mbps", window, flags=re.I) else "",
                        "contract_months": contract_match.group(1) if contract_match else "",
                        "local_voice": "3000 local voice minutes" if re.search(r"3,000 local voice minutes|3000 local voice", window, flags=re.I) else "",
                        "add_on_charges_hkd": "28 admin fee per month" if re.search(r"Admin Fee\s*\$28/(?:month|mth)", text, flags=re.I) else "",
                        "tariff_type": "monthly_plan_fee",
                        "source_status": "parsed_archive" if archive_url else "parsed_current",
                        "evidence_excerpt": _excerpt(text, start, end, radius=260)[:800],
                    }
                )
                row["record_key"] = _record_key(row)
                rows.append(row)
            if rows:
                return rows
        basic_match = re.search(r"Basic monthly fee is \$124 with local data 30GB", text, flags=re.I)
        if basic_match:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": "3HK 5G Monthly SIM Plan HK$124 30GB",
                    "monthly_fee_hkd": "124",
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "30",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "3000 local voice minutes",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_current",
                    "evidence_excerpt": _excerpt(text, basic_match.start(), basic_match.end())[:650],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        upgrade_match = re.search(r"\+\$48/month\s+20GB\s+WORLD PASS", text, flags=re.I)
        if upgrade_match:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": "3HK WORLD PASS Asia Pacific add-on HK$48 20GB",
                    "monthly_fee_hkd": "48",
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "20",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "12",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_add_on_fee",
                    "source_status": "parsed_current",
                    "evidence_excerpt": _excerpt(text, upgrade_match.start(), upgrade_match.end())[:650],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"smartone_5g_listing", "smartone_5g_listing_tc"}:
        rows: List[Dict[str, str]] = []
        pattern = re.compile(
            r"(?:5G Plan|5G計劃)\s+([0-9]+)\s*GB\s+(?:Local Data|本地數據)\s+([0-9]+)\s*(?:Month|個月)\s+(?:Contracts|合約期限)\s+HK\$\s*([0-9,]+)\s*(?:Monthly Subscription Offer|每月)",
            flags=re.I | re.S,
        )
        loose_pattern = re.compile(
            r"(?:SmarTone\s+)?(?:5G Plan|5G計劃)\s+([0-9]+)\s*GB\s+(?:Local Data|本地數據)\s+([0-9]+)\s*(?:Month|個月)\s+(?:Contracts?|合約期限)\s+HK\$\s*([0-9,]+)\s*(?:Monthly\b|月費|每月)?",
            flags=re.I | re.S,
        )
        seen: set[tuple[str, str, str]] = set()
        for match in list(pattern.finditer(text)) or list(loose_pattern.finditer(text)):
            local_data, contract, fee = match.groups()
            fee = fee.replace(",", "")
            key = (local_data, contract, fee)
            if key in seen:
                continue
            seen.add(key)
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"SmarTone 5G Plan HK${fee} {local_data}GB {contract}m",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": local_data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "Selected listing offer may include roaming data, Mainland/Macau data, vouchers or limited-time privileges; not treated as monthly fee",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_archive" if archive_url else "parsed_current",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=280)[:800],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "smartone_subscription_offers":
        rows: List[Dict[str, str]] = []
        pattern = re.compile(
            r"(?P<name>5G\s+(?:Starter Pack|入門之選)|Northbound Weekend Getaways|週末北上之選|Best for Chinese Mainland\s*/\s*Macau Trips|遊走中國內地/澳門之選)"
            r"\s*\$?\s*(?P<fee>[0-9]{3})\s*/\s*(?:Month|月)\s*"
            r"(?:[,，.;:]\s*)?(?:(?:Hot Deal|熱門|Image)\s*)?"
            r"(?P<data>[0-9]{2,3})\s*GB\s*(?:Local data|本地數據)\s*"
            r"(?:[,，.;:]\s*)?(?P<contract>[0-9]{2})[-\s]*(?:month contract|個月合約)",
            flags=re.I | re.S,
        )
        seen: set[tuple[str, str, str]] = set()
        for match in pattern.finditer(text):
            fee = match.group("fee")
            local_data = match.group("data")
            contract = match.group("contract")
            key = (fee, local_data, contract)
            if key in seen:
                continue
            seen.add(key)
            name = _clean_text(match.group("name"))
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"SmarTone {name} HK${fee} {local_data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": local_data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "Subscription offer page may mention Disney+, vouchers or roaming packs; only the explicit monthly plan fee, local data and contract term are structured here.",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_current_official_offer_page",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=280)[:800],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"smartone_home_5g_broadband", "smartone_home_5g_flexi_combo_current"}:
        match = re.search(
            r"SmarTone Home 5G Broadband.*?(?:Just\s*)?\$(?P<fee>148|178)\s*/\s*(?:Month|month|月)",
            text,
            flags=re.I | re.S,
        )
        if not match:
            return []
        fee = match.group("fee")
        data_match = re.search(r"([0-9]{2,4})\s*GB\s+Full Speed Data|([0-9]{2,4})\s*GB", text, flags=re.I)
        local_data = next((group for group in (data_match.groups() if data_match else ()) if group), "")
        add_on_note = (
            "Official archived page states unlimited 5G data and explicit full-speed data allowance; contract term is not estimated."
            if archive_url and local_data
            else "Official page excerpt only states the monthly fee; contract term and usage allowance are not estimated."
        )
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": f"SmarTone Home 5G Broadband {'HK$' + fee if archive_url else 'Online Exclusive HK$' + fee}",
                "monthly_fee_hkd": fee,
                "average_monthly_fee_hkd": "",
                "local_data_gb": local_data,
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": add_on_note,
                "tariff_type": "monthly_home_5g_broadband_offer_fee",
                "source_status": "parsed_archive" if archive_url else "parsed_current_official_offer_page",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hkbn_5g_home_broadband":
        if not (
            re.search(r"5G Home Broadband", text, flags=re.I)
            and re.search(r"HK\$?118\s*/?\s*month", text, flags=re.I)
            and re.search(r"300GB\s*/?\s*mth|300GB", text, flags=re.I)
            and re.search(r"24[-\s]*months?\s+contract|24\s*months?", text, flags=re.I)
        ):
            return []
        match = re.search(
            r"5G Home Broadband Plan.*?HK\$?118\s*/?\s*month.*?300GB.*?24[-\s]*months?\s+contract",
            text,
            flags=re.I | re.S,
        )
        start, end = match.span() if match else (0, min(len(text), 600))
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HKBN 5G Home Broadband Service Plan HK$118",
                "monthly_fee_hkd": "118",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "300",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "",
                "add_on_charges_hkd": "$28 monthly administration fee waiver; unlimited 5G broadband data after 300GB subject to FUP; network supported by 3HK",
                "tariff_type": "monthly_home_5g_broadband_fee",
                "source_status": "parsed_current_official_page",
                "evidence_excerpt": _excerpt(text, start, end, radius=220)[:800],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "3hk_promo5g_2024_pdf":
        if not (re.search(r"30GB monthly local data of \$124 monthly plan", text, flags=re.I) and re.search(r"50GB monthly local data of \$148 monthly plan", text, flags=re.I) and re.search(r"60GB local data of \$188 monthly plan", text, flags=re.I)):
            return []
        rows: List[Dict[str, str]] = []
        table_match = re.search(r"The 30GB monthly local data.*?pay an \$28 admin fee per month", text, flags=re.I)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=120)[:700] if table_match else text[:700]
        for fee, data, fup, contract in [("124", "30", "35", "24"), ("148", "50", "55", "24"), ("188", "60", "65", "28")]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"3HK 5G SIM Monthly Plan HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "168" if fee == "188" else "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "1",
                    "contract_months": contract,
                    "local_voice": "unlimited mobile IDD local talk time",
                    "add_on_charges_hkd": "28 admin fee per month",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_public_official_pdf",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "3hk_extrabux_2025_product_guide":
        if not (re.search(r"2025香港3HK电话卡", text) and re.search(r"5G 月费计划", text) and re.search(r"图片来自于 three\.com\.hk", text, flags=re.I)):
            return []
        table_match = re.search(r"5G 月费计划.*?行政费：HK\$28/月", text, flags=re.I)
        if not table_match:
            return []
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=80)[:700]
        rows: List[Dict[str, str]] = []
        for fee, data, contract, roaming_note in [
            ("124", "30", "24/30", ""),
            ("148", "50", "24", "赠3GB内地及澳门数据"),
            ("168", "60", "28", "赠1GB内地及澳门数据"),
        ]:
            if not re.search(rf"{data}GB.*?HK\${fee}", table_excerpt, flags=re.I):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"3HK 5G monthly plan guide HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "3" if "3GB" in roaming_note else ("1" if "1GB" in roaming_note else ""),
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "3000 local voice minutes",
                    "add_on_charges_hkd": "28 admin fee per month; +68/10GB or +128/20GB shared Mainland-HK-Macau data add-on",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "public_third_party_product_guide_needs_review",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "3hk_thriftyhk_2025_mobile_plan_comparison":
        if not (re.search(r"3HK\s+HK\$168/60GB", text, flags=re.I) and re.search(r"HK\$28", text) and re.search(r"3,000", text)):
            return []
        match = re.search(r"3HK\s+HK\$168/60GB.*?Note:\s*Prices and plans are", text, flags=re.I | re.S)
        excerpt = _clean_text(match.group(0) if match else text)[:700]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "3HK 5G monthly plan guide HK$168 60GB",
                "monthly_fee_hkd": "168",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "60",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "3000 local voice minutes",
                "add_on_charges_hkd": "28 admin fee per month; HK$30/5GB top-up",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_third_party_mobile_plan_comparison_needs_review",
                "evidence_excerpt": excerpt + "；ThriftyHK 页面注明价格和计划只供参考，最终以供应商公布为准。",
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "smartone_moneysmart_2026_mobile_plan_comparison":
        if not (re.search(r"SmarTone", text, flags=re.I) and re.search(r"110GB\s+HK\$299/月\s+24個月", text) and re.search(r"35GB亞太區漫遊數據", text)):
            return []
        match = re.search(r"110GB\s+HK\$299/月\s+24個月.*?110GB\s+HK\$239/月\s+24個月", text, flags=re.S)
        excerpt = _clean_text(match.group(0) if match else text)[:700]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "SmarTone 5G mobile plan comparison HK$299 110GB",
                "monthly_fee_hkd": "299",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "110",
                "roaming_data_gb": "35",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "",
                "add_on_charges_hkd": "35GB Asia Pacific roaming data; 3GB Mainland China and Macau data per month",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_third_party_mobile_plan_comparison_needs_review",
                "evidence_excerpt": (
                    excerpt
                    + "；MoneySmart 为公开第三方比较页，仅作非官方复核来源，不作为官方标准价；同段 HK$239/110GB 合约期与 SmarTone 官方详情页存在差异，未结构化为互证行。"
                )[:900],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "smartone_quoquo_2026_mobile_plan_comparison":
        if not (
            re.search(r"QuoQuo|報價鴨", text, flags=re.I)
            and re.search(r"SmarTone|數碼通", text, flags=re.I)
            and re.search(r"\$129\s*/\s*30GB", text, flags=re.I)
            and re.search(r"30-?month|30\s*(?:個月|月)", text, flags=re.I)
        ):
            return []
        match = re.search(r"(?:SmarTone|數碼通).*?\$129\s*/\s*30GB.*?(?:Macau data|澳門數據)", text, flags=re.I | re.S)
        excerpt = _clean_text(match.group(0) if match else text)[:700]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "SmarTone 5G mobile plan comparison HK$129 30GB",
                "monthly_fee_hkd": "129",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "30",
                "roaming_data_gb": "2",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "30",
                "local_voice": "",
                "add_on_charges_hkd": "admin fee waived; 2GB Mainland China and Macau data",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_third_party_mobile_plan_comparison_needs_review",
                "evidence_excerpt": (
                    excerpt
                    + "；QuoQuo 报价鸭为公开第三方比较页，仅作非官方市场参考；正式月费、覆盖和申请条件以 SmarTone 官方公布为准。"
                )[:900],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"smartone_5g_110g_30m_239_current", "smartone_5g_110g_24m_299_current"}:
        specs = {
            "smartone_5g_110g_30m_239_current": {
                "plan_id": "5g_110g_30m_239_travel",
                "fee": "239",
                "contract": "30",
                "rebate": "159",
            },
            "smartone_5g_110g_24m_299_current": {
                "plan_id": "5g_110g_24m_travel",
                "fee": "299",
                "contract": "24",
                "rebate": "99",
            },
        }
        spec = specs[source["source_id"]]
        if not (
            re.search(r"5G\s+Plan", text, flags=re.I)
            and re.search(r"110\s*(?:<[^>]+>)*\s*GB|110\s+GB|basicData.*?110000", text, flags=re.I | re.S)
            and re.search(rf"HK\$\s*{spec['fee']}|fee\\\":{spec['fee']}|fee\":{spec['fee']}", text, flags=re.I)
            and re.search(rf"{spec['contract']}\s*(?:<!--\s*-->\s*)?(?:<[^>]+>\s*)?Month|{spec['contract']}\s+Month|contractMonth.*?{spec['contract']}", text, flags=re.I | re.S)
            and re.search(r"Administration Fee.*?HK\$\s*18|admin fee.*?HK\$\s*18|admin fee.*?\$18", text, flags=re.I | re.S)
        ):
            return []
        match = re.search(
            rf"5G Plan\(80GB\).*?HK\$\s*{spec['fee']}.*?Administration Fee.*?HK\$\s*18",
            text,
            flags=re.I | re.S,
        )
        if not match:
            match = re.search(rf"planId.*?{re.escape(spec['plan_id'])}.*?contractMonth.*?{spec['contract']}", text, flags=re.I | re.S)
        start, end = match.span() if match else (0, min(len(text), 800))
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": f"SmarTone 5G Travel Plan HK${spec['fee']} 110GB",
                "monthly_fee_hkd": spec["fee"],
                "average_monthly_fee_hkd": "",
                "local_data_gb": "110",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": spec["contract"],
                "local_voice": "unlimited local voice minutes",
                "add_on_charges_hkd": f"HK$18 admin fee per month; original monthly fee HK$398 with HK${spec['rebate']}/month contract bonus/rebate",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_current",
                "evidence_excerpt": _excerpt(text, start, end, radius=260)[:900],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"smartone_home_5g_wifi7_2025_pdf", "smartone_home_5g_wifi7_2025_pdf_chi"}:
        if not (
            re.search(r"Home 5G Broadband x Wi-Fi 7|Home 5G 寬頻 x Wi-Fi 7", text, flags=re.I)
            and re.search(r"average monthly fee of \$191|平均月費只須\s*\$191", text, flags=re.I)
        ):
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        match = re.search(r"average monthly fee of \$191|平均月費只須\s*\$191", text, flags=re.I)
        start, end = match.span() if match else (0, min(len(text), 120))
        row.update(
            {
                "plan_name": "SmarTone Home 5G Broadband x Wi-Fi 7 HK$191",
                "monthly_fee_hkd": "191",
                "average_monthly_fee_hkd": "191",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_archive" if archive_url else "parsed_current",
                "evidence_excerpt": _excerpt(text, start, end)[:600],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"smartone_st_protect_2016_pdf", "smartone_st_protect_2016_pdf_chi"}:
        if not (
            re.search(r"ST Protect", text, flags=re.I)
            and (re.search(r"HK\$28 per month", text, flags=re.I) or re.search(r"每月\s*HK\$28", text))
            and (re.search(r"HK\$18 per month", text, flags=re.I) or re.search(r"每月\s*HK\$18", text))
        ):
            return []
        rows: List[Dict[str, str]] = []
        table_match = re.search(r"Charges Service Service Fees.*?HK\$18 per month|服務\s+服務收費.*?每月\s*HK\$18", text, flags=re.I | re.S)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=120)[:600] if table_match else text[:600]
        for fee, plan_name, contract in [
            ("28", "SmarTone ST Protect standard plan HK$28", ""),
            ("18", "SmarTone ST Protect monthly plan contract offer HK$18", "12"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_value_added_service_fee",
                    "source_status": "parsed_public_official_pdf" if source.get("source_id", "").endswith("_chi") else ("parsed_archive" if archive_url else "parsed_current"),
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"smartone_aquos_s2_supercare_2017_pdf", "smartone_aquos_s2_supercare_2017_pdf_chi"}:
        if not (
            re.search(r"SHARP AQUOS S2", text, flags=re.I)
            and (
                re.search(r"SuperCare Smartphone Plans", text, flags=re.I)
                or re.search(r"4\.5G\s*超貼心智能手機計劃", text, flags=re.I)
            )
        ):
            return []
        rows: List[Dict[str, str]] = []
        table_match = re.search(r"4\.5G SuperCare Smartphone Plans.*?admin fee of \$18 per month", text, flags=re.I)
        if not table_match:
            table_match = re.search(r"4\.5G\s*超貼心智能手機計劃.*?每月\$18\s*行政費", text, flags=re.I | re.S)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=120)[:600] if table_match else text[:600]
        for fee, data, voice, handset in [
            ("388", "6", "4000 minutes", "0"),
            ("348", "2.5", "3000 minutes", "0"),
            ("258", "1", "2500 minutes", "980"),
            ("308", "2.5", "3000 minutes", "0"),
            ("218", "1", "2500 minutes", "580"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"SmarTone SHARP AQUOS S2 SuperCare Smartphone Plan HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": voice,
                    "add_on_charges_hkd": handset,
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_archive" if archive_url else "parsed_public_official_pdf",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"smartone_learning_support_2022_pdf", "smartone_learning_support_2022_pdf_chi"}:
        if not (
            re.search(r"SmarTone Home 5G", text, flags=re.I)
            and (re.search(r"eLearning Support Scheme", text, flags=re.I) or re.search(r"網課貼心支援計劃", text))
            and re.search(r"\$1,600|1,600", text)
            and re.search(r"\$780|780", text)
        ):
            return []
        yearly_match = re.search(r"\$1,600/year|1,600\s*享用全年無限\s*5G\s*數據|1,600/年", text, flags=re.I)
        router_match = re.search(r"\$780\s*saved|節省\$780\s*路由器租借費|節省\s*780", text, flags=re.I)
        excerpt_start = min([m.start() for m in [yearly_match, router_match] if m] or [0])
        excerpt_end = max([m.end() for m in [yearly_match, router_match] if m] or [min(len(text), 300)])
        excerpt = _excerpt(text, excerpt_start, excerpt_end, radius=180)[:700]
        rows: List[Dict[str, str]] = []
        for plan_name, fee, tariff_type in [
            ("SmarTone Home 5G Broadband HK$1600", "1600", "annual_home_5g_broadband_support_fee"),
            ("SmarTone Home 5G Broadband HK$780", "780", "router_rental_fee_waiver_value"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "5000",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "Router rental fee waived; annual support scheme price captured as disclosed",
                    "tariff_type": tariff_type,
                    "source_status": "parsed_archive" if archive_url else "parsed_public_official_pdf",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "smartone_aquos_s2_2017_pcm_report":
        if not (
            re.search(r"拎舊\s*Sharp\s*手機可上台半價買\s*AQUOS\s*S2", text, flags=re.I)
            and re.search(r"\$388\s+6GB", text, flags=re.I)
            and re.search(r"\$348\s+2\.5GB", text, flags=re.I)
            and re.search(r"\$258\s+1GB", text, flags=re.I)
        ):
            return []
        table_match = re.search(r"AQUOS S2 高配版出機計劃.*?\$258\s+1GB.*?預繳 \$3,000", text, flags=re.I | re.S)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=100)[:750] if table_match else text[:750]
        rows: List[Dict[str, str]] = []
        for fee, data, voice, handset in [
            ("388", "6", "4000 minutes", "0"),
            ("348", "2.5", "3000 minutes", "0"),
            ("258", "1", "2500 minutes", "980"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"SmarTone SHARP AQUOS S2 SuperCare Smartphone Plan HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "24",
                    "local_voice": voice,
                    "add_on_charges_hkd": handset,
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "public_media_report_needs_review",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id", "").startswith("smartone_kono_magazine_"):
        if not (re.search(r"Kono Magazine Service", text, flags=re.I) and re.search(r"Standard Plan.*?\$38", text, flags=re.I) and re.search(r"24-Month Contract Plan.*?\$36", text, flags=re.I)):
            return []
        rows: List[Dict[str, str]] = []
        table_match = re.search(r"Service Plan Monthly Service Fee.*?monthly fee is calculated in Hong Kong Dollars", text, flags=re.I)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=80)[:600] if table_match else text[:600]
        for fee, plan_name, contract in [
            ("38", "SmarTone Kono Magazine Service Standard Plan HK$38", ""),
            ("36", "SmarTone Kono Magazine Service 24-Month Contract Plan HK$36", "24"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_value_added_service_fee",
                    "source_status": "web_indexed_official_pdf_excerpt",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"smartone_5g_launch_2020_pdf", "smartone_5g_launch_2020_pdf_chi"}:
        if not (re.search(r"SmarTone 5G", text, flags=re.I) and re.search(r"5G Monthly Plan|5G 月費計劃", text, flags=re.I)):
            return []
        rows: List[Dict[str, str]] = []

        def add_smartone_5g_row(
            fee: str,
            data: str,
            plan_name: str,
            tariff_type: str,
            pattern: str,
            *,
            average_fee: str = "",
            roaming_data: str = "",
            post_fup_speed: str = "",
            add_on_charges: str = "",
        ) -> None:
            match = re.search(pattern, text, flags=re.I | re.S)
            if not match:
                return
            excerpt = _excerpt(text, match.start(), match.end(), radius=120)[:650]
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": average_fee,
                    "local_data_gb": data,
                    "roaming_data_gb": roaming_data,
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": post_fup_speed,
                    "contract_months": "24",
                    "local_voice": "",
                    "add_on_charges_hkd": add_on_charges,
                    "tariff_type": tariff_type,
                    "source_status": "parsed_public_official_pdf",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)

        add_smartone_5g_row(
            "398",
            "80",
            "SmarTone 5G Monthly Plan HK$398 80GB",
            "monthly_plan_fee",
            r"For\s+HK\$?398/month.*?80GB\s+5G\s+local\s+data|每月\s*港幣\s*398\s*元.*?80GB\s*本地數據",
            roaming_data="2",
        )
        add_smartone_5g_row(
            "398",
            "100",
            "SmarTone 5G Monthly Plan HK$398 100GB promotional entitlement",
            "monthly_plan_fee",
            r"100GB.*?(?:\$|HK\$)\s*398|398\s*元.*?每月額外\s*20GB\s*本地數據|月費港幣\s*398\s*元.*?100GB\s*5G\s*本地數據",
            roaming_data="2",
        )
        add_smartone_5g_row(
            "298",
            "100",
            "SmarTone 5G limited-time monthly offer HK$298 100GB",
            "monthly_plan_fee",
            r"5G\s+Limited-time\s+Offer.*?(?:\$|HK\$)\s*298|limited-time offer.*?HK\$?298\s+per\s+month.*?100GB\*?\s+5G\s+local\s+data|4\s*個月月費港幣\s*298\s*元.*?100GB\s*5G\s*本地數據",
            add_on_charges="Up to HK$2,100 smartphone discount mentioned; not treated as monthly fee",
        )
        add_smartone_5g_row(
            "80",
            "unlimited",
            "SmarTone 5G unlimited local data top-up HK$80",
            "monthly_data_addon_fee",
            r"(?:\$|HK\$)\s*80.*?(?:unlimited|5Mbps)|每月港幣\s*80\s*元升級無限\s*5G\s*數據.*?5Mbps",
            post_fup_speed="5",
        )
        add_smartone_5g_row(
            "50",
            "10",
            "SmarTone 5G local data top-up HK$50 10GB",
            "monthly_data_addon_fee",
            r"(?:\$|HK\$)\s*50.*?10GB|\$50/10GB",
        )
        add_smartone_5g_row(
            "120",
            "50",
            "SmarTone 5G add-on SIM HK$120 50GB",
            "monthly_addon_sim_fee",
            r"(?:\$|HK\$)\s*120.*?50GB|\+\$120/月.*?50GB/月",
        )
        return rows
    if source.get("source_id") in {"smartone_gamergizer_2020_pdf", "smartone_gamergizer_2020_pdf_chi"}:
        if not (re.search(r"Gamergizer|爆機王牌", text, flags=re.I) and re.search(r"SmarTone 5G|5G", text, flags=re.I)):
            return []
        rows: List[Dict[str, str]] = []

        def add_gamergizer_row(
            fee: str,
            data: str,
            plan_name: str,
            tariff_type: str,
            pattern: str,
            *,
            average_fee: str = "",
            post_fup_speed: str = "",
            add_on_charges: str = "",
        ) -> None:
            match = re.search(pattern, text, flags=re.I | re.S)
            if not match:
                return
            excerpt = _excerpt(text, match.start(), match.end(), radius=120)[:650]
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": average_fee,
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": post_fup_speed,
                    "contract_months": "24" if tariff_type != "monthly_value_added_service_fee" else "",
                    "local_voice": "",
                    "add_on_charges_hkd": add_on_charges,
                    "tariff_type": tariff_type,
                    "source_status": "parsed_public_official_pdf",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)

        add_gamergizer_row(
            "29",
            "",
            "SmarTone Gamergizer average monthly service fee HK$29",
            "monthly_value_added_service_fee",
            r"Gamergizer.*?(?:average|monthly).*?(?:\$|HK\$)\s*29|爆機王牌.*?平均月費.*?(?:港幣\s*)?29\s*元",
            average_fee="29",
        )
        add_gamergizer_row(
            "398",
            "100",
            "SmarTone 5G Monthly Plan HK$398 100GB",
            "monthly_plan_fee",
            r"100GB.*?(?:\$|HK\$)\s*398|月費\s*\$398.*?100GB",
        )
        add_gamergizer_row(
            "298",
            "100",
            "SmarTone 5G Gamergizer limited-time monthly offer HK$298",
            "monthly_plan_fee",
            r"5G\s+Limited-time\s+Offer.*?(?:\$|HK\$)\s*298|5G\s*限時優惠\s*\$298/月",
            add_on_charges="Up to HK$2,100 smartphone discount and HK$1,200 accessories discount mentioned; not treated as monthly fees",
        )
        add_gamergizer_row(
            "80",
            "unlimited",
            "SmarTone 5G unlimited local data top-up HK$80",
            "monthly_data_addon_fee",
            r"(?:\$|HK\$)\s*80.*?(?:unlimited|5Mbps)|\$80/月\s*無限",
            post_fup_speed="5",
        )
        add_gamergizer_row(
            "50",
            "10",
            "SmarTone 5G local data top-up HK$50 10GB",
            "monthly_data_addon_fee",
            r"(?:\$|HK\$)\s*50.*?10GB|\$50/10GB",
        )
        add_gamergizer_row(
            "120",
            "50",
            "SmarTone 5G add-on SIM HK$120 50GB",
            "monthly_addon_sim_fee",
            r"(?:\$|HK\$)\s*120.*?50GB|\+\$120/月.*?50GB/月",
        )
        return rows
    if source.get("source_id") in {"smartone_home_5g_launch_2020_pdf", "smartone_home_5g_2020_ezone_review", "smartone_home_5g_2021_pdf", "smartone_home_5g_2021_pdf_chi"}:
        if not (
            re.search(r"Home 5G Broadband|Home 5G\s*寬頻|5G\s*家居寬頻", text, flags=re.I)
            and re.search(r"(?:HK\$|\$)\s*148|HK\$148|月費\s*HK\$148", text, flags=re.I)
            and (
                re.search(r"200GB", text, flags=re.I)
                or re.search(r"competitive monthly fee of\s*\$148|For just\s*\$148 a month", text, flags=re.I)
                or re.search(r"每月只須\s*\$148|每月只需港幣\s*148\s*元", text, flags=re.I)
            )
        ):
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        match = re.search(r"unlimited 5G data at just \$148 per month.*?full speed data up to 200GB", text, flags=re.I)
        if not match:
            match = re.search(r"HK\$148 月費 5G 家居寬頻.*?24 個月合約", text, flags=re.I)
        if not match:
            match = re.search(r"competitive monthly fee of\s*\$148.*?Wi-Fi 6 router", text, flags=re.I)
        if not match:
            match = re.search(r"For just\s*\$148 a month.*?Home 5G Broadband service", text, flags=re.I)
        if not match:
            match = re.search(r"每月只須\s*\$148.*?Home 5G 寬頻|每月只需港幣\s*148\s*元.*?Wi-Fi 6", text, flags=re.I | re.S)
        start, end = match.span() if match else (0, min(len(text), 180))
        row.update(
            {
                "plan_name": "SmarTone Home 5G Broadband HK$148 unlimited 5G data",
                "monthly_fee_hkd": "148",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "" if source.get("source_id") in {"smartone_home_5g_2021_pdf", "smartone_home_5g_2021_pdf_chi"} else "unlimited",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_media_review_needs_review" if source.get("source_id") == "smartone_home_5g_2020_ezone_review" else ("parsed_archive" if archive_url else "parsed_current"),
                "evidence_excerpt": _excerpt(text, start, end)[:600],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"smartone_home_5g_disney_offer_current", "smartone_home_5g_disney_offer_current_tc"}:
        if not (
            re.search(r"Home 5G Broadband|Home 5G寬頻", text, flags=re.I)
            and re.search(r"WiFi\s*[67]\s+(?:Plan|計劃)", text, flags=re.I)
        ):
            return []
        rows: List[Dict[str, str]] = []
        plan_specs = [
            {
                "plan_name": "SmarTone Home 5G Broadband WiFi 7 Plan HK$217",
                "fee": "217",
                "original_fee": "308",
                "contract": "36",
                "router": "Wi-Fi 7 5G Router Included",
                "pattern": r"WiFi 7\s+(?:Plan|計劃)\s*\$\s*217\s*/\s*(?:Month|月).*?(?:Originally \$308/Month|原價\s*\$308/月).*?(?:36-month contract|36個月合約).*?(?:Wi-Fi 7 5G Router Included|Wi-Fi 7 5G路由器)",
            },
            {
                "plan_name": "SmarTone Home 5G Broadband WiFi 7 Mesh Plan HK$229",
                "fee": "229",
                "original_fee": "308",
                "contract": "36",
                "router": "Wi-Fi 7 5G Router Included; Mesh Router Rental Inclusive",
                "pattern": r"WiFi 7\s+(?:Plan|計劃)\s*\$\s*229\s*/\s*(?:Month|月).*?(?:Originally \$308/Month|原價\s*\$308/月).*?(?:36-month contract|36個月合約).*?(?:Mesh Router Rental Inclusive|包括Mesh路由器租用)",
            },
            {
                "plan_name": "SmarTone Home 5G Broadband WiFi 6 Plan HK$168 36-month",
                "fee": "168",
                "original_fee": "258/259",
                "contract": "36",
                "router": "Wi-Fi 6 5G Router Included",
                "pattern": r"WiFi 6\s+(?:Plan|計劃)\s*\$\s*168\s*/\s*(?:Month|月).*?(?:Originally \$259/Month|原價\s*\$258/月).*?(?:36-month contract|36個月合約).*?(?:Wi-Fi 6 5G Router Included|Wi-Fi 6 5G路由器)",
            },
            {
                "plan_name": "SmarTone Home 5G Broadband WiFi 6 Plan HK$168 24-month",
                "fee": "168",
                "original_fee": "238",
                "contract": "24",
                "router": "Wi-Fi 6 5G Router Included",
                "pattern": r"WiFi 6\s+(?:Plan|計劃)\s*\$\s*168\s*/\s*(?:Month|月).*?(?:Originally \$238/Month|原價\s*\$238/月).*?(?:24-month contract|24個月合約).*?(?:Wi-Fi 6 5G Router Included|Wi-Fi 6 5G路由器)",
            },
        ]
        terms_match = re.search(
            r"Home 5G Broadband 36-month Service Plan:.*?Home 5G Broadband 24-month Service Plan:|Home 5G寬頻服務計劃.*?路由器租用服務之按金",
            text,
            flags=re.I | re.S,
        )
        terms_excerpt = _excerpt(text, terms_match.start(), terms_match.end(), radius=160)[:800] if terms_match else ""
        for spec in plan_specs:
            match = re.search(spec["pattern"], text, flags=re.I | re.S)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": spec["plan_name"],
                    "monthly_fee_hkd": spec["fee"],
                    "average_monthly_fee_hkd": spec["fee"],
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": spec["contract"],
                    "local_voice": "",
                    "add_on_charges_hkd": (
                        f"Original monthly fee HK${spec['original_fee']}; {spec['router']}; Disney+ offer mentioned; "
                        "router rental fee waived within committed contract period; HK$1,500 router deposit may apply."
                    ),
                    "tariff_type": "monthly_home_5g_broadband_offer_fee",
                    "source_status": "parsed_current_official_offer_page",
                    "evidence_excerpt": (_excerpt(text, match.start(), match.end(), radius=180) + " " + terms_excerpt)[:900],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"smartone_roaming_multiday_2023_pdf", "smartone_roaming_multiday_2023_linkedin"}:
        if not (re.search(r"Multi-Day Roaming Data Pack", text, flags=re.I) and re.search(r"4-Day", text, flags=re.I) and re.search(r"7-Day", text, flags=re.I)):
            return []
        rows: List[Dict[str, str]] = []
        for fee, days, avg_daily in [("128", "4", "28.8"), ("198", "7", "25.5")]:
            price_match = re.search(rf"\${fee}\b", text, flags=re.I)
            start, end = price_match.span() if price_match else (0, min(len(text), 120))
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"SmarTone {days}-Day Multi-Day Roaming Data Pack HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": avg_daily,
                    "local_data_gb": "",
                    "roaming_data_gb": "1",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "daily_pass_fee",
                    "source_status": "public_official_social_post_needs_review" if source.get("source_id", "").endswith("_linkedin") else ("parsed_archive" if archive_url else "parsed_current"),
                    "evidence_excerpt": _excerpt(text, start, end)[:600],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"smartone_roaming_yas_2023_pdf", "smartone_roaming_yas_2023_pdf_chi"}:
        if not (
            re.search(r"Multi-Day Roaming Data Pack|漫遊數據多日通", text, flags=re.I)
            and re.search(r"starting at \$23/day|低至\s*\$23", text, flags=re.I)
        ):
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        match = re.search(r"starting at \$23/day|低至\s*\$23", text, flags=re.I)
        start, end = match.span() if match else (0, min(len(text), 120))
        row.update(
            {
                "plan_name": "SmarTone Multi-Day Roaming Data Pack starting HK$23/day",
                "monthly_fee_hkd": "23",
                "average_monthly_fee_hkd": "23",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "",
                "tariff_type": "daily_pass_fee",
                "source_status": "parsed_archive" if archive_url else "parsed_current",
                "evidence_excerpt": _excerpt(text, start, end)[:600],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"smartone_1c2n_2024_pdf", "smartone_1c2n_2024_pdf_chi"}:
        if not (
            re.search(r"1\s*Card\s*2\s*Numbers|一卡兩號", text, flags=re.I)
            and re.search(r"\$28/month|\$28/月", text, flags=re.I)
        ):
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        match = re.search(r"\$28/month|\$28/月", text, flags=re.I)
        start, end = match.span() if match else (0, min(len(text), 120))
        row.update(
            {
                "plan_name": 'SmarTone Mainland China & HK 1 Card 2 Numbers HK$28',
                "monthly_fee_hkd": "28",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "15 Mainland voice minutes",
                "add_on_charges_hkd": "",
                "tariff_type": "monthly_value_added_service_fee",
                "source_status": "parsed_archive" if archive_url else "parsed_current",
                "evidence_excerpt": _excerpt(text, start, end)[:600],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"hkbn_4g_mobile_bundle_2018_pdf", "hkbn_4g_mobile_bundle_2018_pdf_chi", "hkbn_4g_mobile_bundle_2017_official_pdf"}:
        is_2017_official = source.get("source_id") == "hkbn_4g_mobile_bundle_2017_official_pdf"
        if is_2017_official:
            if not (re.search(r"All-new Disruptive 4G Mobile Services Bundle", text, flags=re.I) and re.search(r"HK\$?78/?m", text, flags=re.I) and re.search(r"5GB data per month", text, flags=re.I)):
                return []
        elif source.get("source_id") == "hkbn_4g_mobile_bundle_2018_pdf_chi":
            if not (re.search(r"香港寬頻|HKBN", text, flags=re.I) and re.search(r"\$78", text) and re.search(r"\$148", text) and re.search(r"\$218", text)):
                return []
        elif not (re.search(r"HKBN Mobile Services Bundles", text, flags=re.I) and re.search(r"Special monthly fee \$78", text, flags=re.I)):
            return []
        rows: List[Dict[str, str]] = []
        table_match = re.search(r"HKBN Mobile Services Bundles#.*?Network 4G Maximum local download speed 21Mbps", text, flags=re.I)
        if is_2017_official:
            table_match = re.search(r"HKBN Announces All-new Disruptive 4G Mobile Services Bundle.*?5GB data per month", text, flags=re.I | re.S)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=120)[:600] if table_match else text[:600]
        row_specs = [("78", "5")] if is_2017_official else [("78", "5"), ("148", "6"), ("218", "12")]
        for fee, data in row_specs:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HKBN 4G Mobile Services Bundle HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "" if is_2017_official else "18",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_public_official_pdf" if not archive_url else "parsed_archive",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {
        "hkbn_wifi_concierge_2017_official_html",
        "hkbn_wifi_concierge_2017_official_pdf_en",
        "hkbn_wifi_concierge_2017_official_pdf_tc",
    }:
        if not (
            re.search(r"HKBN|Hong Kong Broadband Network|香港寬頻", text, flags=re.I)
            and re.search(r"Wi-?Fi Concierge|Wi-Fi\s*管家", text, flags=re.I)
            and re.search(r"Home Telephone|家居電話", text, flags=re.I)
            and re.search(r"\$88|港幣\s*88|月費只需\s*\$?88", text, flags=re.I)
        ):
            return []
        match = re.search(
            r"(?:Wi-?Fi Concierge|Wi-Fi\s*管家).*?(?:Home Telephone|家居電話).*?(?:\$88|月費只需\s*\$?88)|(?:\$88|月費只需\s*\$?88).*?(?:Wi-?Fi Concierge|Wi-Fi\s*管家)",
            text,
            flags=re.I | re.S,
        )
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HKBN Wi-Fi Concierge and Home Telephone Service Bundle HK$88",
                "monthly_fee_hkd": "88",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "24/27",
                "local_voice": "Home Telephone service",
                "add_on_charges_hkd": "Includes Wi-Fi Concierge 1Gbps router, 24/7 Wi-Fi technical support, VTech DECT telephone, myTV SUPER Alpha Pack; 2-month waiver or 5-month waiver with mobile services.",
                "tariff_type": "monthly_bundle_fee",
                "source_status": "parsed_public_official_release",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=260)[:900] if match else text[:900],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hkbn_high_usage_mobile_2017_mediaoutreach":
        if not (
            re.search(r"Hong Kong Broadband Network Limited|HKBN", text, flags=re.I)
            and re.search(r"\$218/month\s+for\s+12GB\s+Data", text, flags=re.I)
            and re.search(r"4\.5G\s+network", text, flags=re.I)
            and re.search(r"myTV\s+SUPER\s+app", text, flags=re.I)
        ):
            return []
        match = re.search(r"\$218/month\s+for\s+12GB\s+Data.*?4\.5G\s+network.*?myTV\s+SUPER\s+app", text, flags=re.I | re.S)
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HKBN 4G Mobile Services Bundle HK$218 12GB",
                "monthly_fee_hkd": "218",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "12",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "includes myTV SUPER app; high-usage 4.5G bundle public news release",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_news_release_needs_review",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=240)[:900] if match else text[:900],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"hkbn_mobile_launch_2016_prnewswire", "hkbn_mobile_launch_2016_official_html"}:
        if not (re.search(r"HKBN Mobile Services Plans", text, flags=re.I) and re.search(r"\$88\s*(?:/ month)?.*?\$108\s*(?:/ month)?.*?\$198\s*(?:/ month)?.*?\$248\s*(?:/ month)?.*?\$446", text, flags=re.I | re.S)):
            return []
        is_official_html = source.get("source_id") == "hkbn_mobile_launch_2016_official_html"
        rows: List[Dict[str, str]] = []
        table_match = re.search(r"HKBN Mobile Services Plans.*?Admin fee of \$18 per month", text, flags=re.I | re.S)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=120)[:750] if table_match else text[:750]
        for plan_code, fee, data, speed, voice in [
            ("S", "88", "unlimited", "0.384", "2000 local voice minutes"),
            ("M", "108", "3", "21", "3000 local voice minutes; 2GB extra data for mobile number port-in"),
            ("L", "198", "3", "", "3000 local voice minutes"),
            ("XL", "248", "6", "", "3000 local voice minutes"),
            ("Infinity", "446", "unlimited", "", "3000 local voice minutes"),
        ]:
            if data == "unlimited":
                data_label = "unlimited throttled data" if plan_code == "S" else "unlimited"
            else:
                data_label = f"{data}GB"
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HKBN Mobile Services Plan {plan_code} HK${fee} {data_label}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": speed,
                    "contract_months": "24",
                    "local_voice": voice,
                    "add_on_charges_hkd": "18 monthly admin fee; waived for selected quad-play bundles",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_official_html" if is_official_html else "parsed_public_news_release",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        bundle_match = re.search(r"Quad-play free-to-go bundle plan.*?Monthly fee\s*\$248\s*up.*?\$18 monthly administration fee (?:waived|waiver)", text, flags=re.I | re.S)
        if bundle_match:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": "HKBN Quad-play free-to-go bundle from HK$248",
                    "monthly_fee_hkd": "248",
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "3",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "100",
                    "post_fup_speed_mbps": "21",
                    "contract_months": "",
                    "local_voice": "3000 local voice minutes",
                    "add_on_charges_hkd": "18 monthly administration fee waived",
                    "tariff_type": "monthly_bundle_fee_from",
                    "source_status": "parsed_official_html" if is_official_html else "parsed_public_news_release",
                    "evidence_excerpt": _excerpt(text, bundle_match.start(), bundle_match.end(), radius=120)[:750],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hkbn_mobile_trial_2016_pdf":
        if not (re.search(r"\$108 mobile service plan", text, flags=re.I) and re.search(r"first 3GB", text, flags=re.I)):
            return []
        match = re.search(r"first six-month monthly fee waiver.*?first 3GB.*?21Mbps", text, flags=re.I | re.S)
        excerpt = _excerpt(text, match.start(), match.end(), radius=120)[:700] if match else text[:700]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HKBN Mobile trial plan HK$108 3GB after six-month waiver",
                "monthly_fee_hkd": "108",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "3",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "21",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "18 admin fee after waiver period",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_public_official_pdf",
                "evidence_excerpt": excerpt,
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hkbn_homeplus_5g_2021_pdf":
        if not (re.search(r"HKBN.*HOME\+", text, flags=re.I) and re.search(r"Brand New HKBN 5G Mobile Services Plans", text, flags=re.I)):
            return []
        rows: List[Dict[str, str]] = []
        table_match = re.search(r"Brand New HKBN 5G Mobile Services Plans.*?Admin fee HK\$18", text, flags=re.I | re.S)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=120)[:700] if table_match else text[:700]
        for fee, data, family in [
            ("298", "20", "Unlimited Data Plan"),
            ("338", "30", "Unlimited Data Plan"),
            ("238", "20", "Basic Plan"),
            ("278", "30", "Basic Plan"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HKBN 5G {family} HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "24",
                    "local_voice": "",
                    "add_on_charges_hkd": "18 admin fee; HK$10 additional fee for non-existing selected HKBN residential customers; HOME+ e-coupon rewards not treated as monthly fees",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_public_official_pdf",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        for fee, data in [("388", "100"), ("30", "5")]:
            match = re.search(rf"HK\${fee}/{data}GB", text, flags=re.I)
            excerpt = _excerpt(text, match.start(), match.end(), radius=120)[:650] if match else table_excerpt
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HKBN 5G local data top-up HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_data_addon_fee",
                    "source_status": "parsed_public_official_pdf",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hkbn_homeplus_5g_2021_techent_syndication":
        if not (
            re.search(r"HKBN and HOME\+ Join Forces", text, flags=re.I)
            and re.search(r"Monthly fee.*?HK\$298.*?HK\$338.*?HK\$238.*?HK\$278", text, flags=re.I | re.S)
            and re.search(r"HK\$388/100GB or HK\$30/5GB", text, flags=re.I)
        ):
            return []
        table_match = re.search(r"Brand New HKBN 5G Mobile Services Plans.*?Admin fee\s+HK\$18", text, flags=re.I | re.S)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=120)[:900] if table_match else text[:900]
        rows: List[Dict[str, str]] = []
        for fee, data, family, tariff_type in [
            ("298", "20", "Unlimited Data Plan", "monthly_plan_fee"),
            ("338", "30", "Unlimited Data Plan", "monthly_plan_fee"),
            ("238", "20", "Basic Plan", "monthly_plan_fee"),
            ("278", "30", "Basic Plan", "monthly_plan_fee"),
            ("388", "100", "Local 5G Data Top-up", "data_top_up_fee"),
            ("30", "5", "Local 5G Data Top-up", "data_top_up_fee"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HKBN 5G {family} HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "24" if tariff_type == "monthly_plan_fee" else "",
                    "local_voice": "",
                    "add_on_charges_hkd": "18 admin fee" if tariff_type == "monthly_plan_fee" else "",
                    "tariff_type": tariff_type,
                    "source_status": "public_press_release_syndication_needs_review",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hkbn_travel_pocket_wifi_2018_pdf", "hkbn_travel_pocket_wifi_2018_official_html"}:
        if not (re.search(r"Travel Pocket Wi-Fi", text, flags=re.I) and re.search(r"\$28/day", text, flags=re.I)):
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        match = re.search(r"\$28/day", text, flags=re.I)
        start, end = match.span() if match else (0, min(len(text), 120))
        row.update(
            {
                "plan_name": "HKBN Travel Pocket Wi-Fi HK$28/day additional day",
                "monthly_fee_hkd": "28",
                "average_monthly_fee_hkd": "28",
                "local_data_gb": "",
                "roaming_data_gb": "0.5",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "0.128",
                "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "daily_pass_fee",
                    "source_status": "parsed_archive"
                    if archive_url
                    else ("parsed_official_html" if source.get("source_id") == "hkbn_travel_pocket_wifi_2018_official_html" else "parsed_public_official_pdf"),
                    "evidence_excerpt": _excerpt(text, start, end)[:600],
                }
            )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"hkbn_momax_smart_home_2020_pdf", "hkbn_momax_smart_home_2020_pdf_tc"}:
        if not (re.search(r"MOMAX", text, flags=re.I) and re.search(r"HK\$?88", text, flags=re.I) and re.search(r"HK\$?68", text, flags=re.I)):
            return []
        is_tc_pdf = source.get("source_id") == "hkbn_momax_smart_home_2020_pdf_tc"
        rows: List[Dict[str, str]] = []
        for fee, plan_name in [("88", "HKBN MOMAX Smart IoT Bundle HK$88"), ("68", "HKBN MOMAX Trio-Cleanse IoT UV-C Vacuum Robot HK$68")]:
            match = re.search(rf"HK\$?{fee}\b", text, flags=re.I)
            start, end = match.span() if match else (0, min(len(text), 120))
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "24",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_bundle_fee",
                    "source_status": "parsed_official_pdf_tc" if is_tc_pdf else ("parsed_archive" if archive_url else "parsed_current"),
                    "evidence_excerpt": _excerpt(text, start, end)[:600],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hkbn_fy22_interim_5g_market_snapshot":
        if not (re.search(r"Best-in-town 5G Offers", text, flags=re.I) and re.search(r"New 5G Mobile offer \$68 for 10GB / \$149 for 100GB", text, flags=re.I)):
            return []
        rows: List[Dict[str, str]] = []
        summary_match = re.search(r"New 5G Mobile offer \$68 for 10GB / \$149 for 100GB", text, flags=re.I)
        summary_start, summary_end = summary_match.span() if summary_match else (0, min(len(text), 120))
        summary_excerpt = _excerpt(text, summary_start, summary_end)[:600]
        for fee, data, post_fup in [("68", "10", "5"), ("149", "100", "5")]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HKBN 5G Mobile Offer HK${fee} {data}GB",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": post_fup,
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "public_market_snapshot_needs_review",
                    "evidence_excerpt": summary_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hkbn_68_10gb_2022_mobilemagazine":
        if not (re.search(r"HKBN|香港寬頻", text, flags=re.I) and re.search(r"HK\$?68|\$68", text, flags=re.I) and re.search(r"10GB", text, flags=re.I)):
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        match = re.search(r"HK\$?68|\$68", text, flags=re.I)
        start, end = match.span() if match else (0, min(len(text), 120))
        row.update(
            {
                "plan_name": "HKBN 5G Mobile Offer HK$68 10GB",
                "monthly_fee_hkd": "68",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "10",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "5" if re.search(r"5Mbps|5 Mbps", text, flags=re.I) else "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "18 admin fee waived" if re.search(r"HK\$?18|\$18", text, flags=re.I) else "",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_media_report_needs_review",
                "evidence_excerpt": _excerpt(text, start, end)[:600],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {
        "icable_hkbn_fy22_fixed_broadband_market_snapshot",
        "hgc_hkbn_fy22_fixed_broadband_market_snapshot",
        "icable_hkbn_fy22_fixed_broadband_market_snapshot_irasia",
        "hgc_hkbn_fy22_fixed_broadband_market_snapshot_irasia",
    }:
        if not (re.search(r"Best-in-town Fixed Broadband", text, flags=re.I) and re.search(r"i-Cable HKBN HKT HGC Monthly Fee", text, flags=re.I)):
            return []
        brand = source["brand"]
        brand_values = {
            "i-CABLE": [("68", "listed website/public HOS offer"), ("149", "market intelligence street offer")],
            "HGC": [("89", "listed website/public HOS offer"), ("99", "market intelligence street offer")],
        }
        rows: List[Dict[str, str]] = []
        table_match = re.search(r"Best-in-town Fixed Broadband.*?Actual Offers at Street Listed Offers at Website", text, flags=re.I)
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=120)[:600] if table_match else text[:600]
        for fee, qualifier in brand_values.get(brand, []):
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"{brand} fixed broadband {qualifier} HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "1000",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "public_market_snapshot_needs_review",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hgc_ezone_2019_2g_broadband_comparison":
        if not (
            re.search(r"HGC 寬頻 四線 2\.2Gbps 服務", text)
            and re.search(r"平均服務月費", text)
            and re.search(r"每月加 \$58 可享「Wi-Fi 360」", text)
        ):
            return []
        specs = [
            {
                "plan_name": "HGC 四線 2.2Gbps 光纖寬頻服務 average HK$218",
                "fee": "",
                "average_fee": "218",
                "speed": "2200",
                "contract": "30",
                "tariff_type": "average_monthly_plan_fee",
                "pattern": r"「一家四口」2\.2Gbps 光纖寬頻服務（1,000M / 2,200M）.*?合約期：30 個月.*?平均服務月費＊：\$218",
            },
            {
                "plan_name": "HGC 1Gbps 極速光纖寬頻服務 average HK$148",
                "fee": "",
                "average_fee": "148",
                "speed": "1000",
                "contract": "36",
                "tariff_type": "average_monthly_plan_fee",
                "pattern": r"1Gbps 極速光纖寬頻服務（1,000M / 1,000M）.*?合約期：36 個月.*?平均服務月費＊：\$148",
            },
            {
                "plan_name": "HGC Wi-Fi 360 add-on HK$58",
                "fee": "58",
                "average_fee": "",
                "speed": "2200",
                "contract": "",
                "tariff_type": "monthly_value_added_service_fee",
                "pattern": r"每月加 \$58 可享「Wi-Fi 360」屋內 Wi-Fi 服務",
            },
        ]
        rows = []
        for spec in specs:
            match = re.search(spec["pattern"], text, flags=re.S)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": spec["plan_name"],
                    "monthly_fee_hkd": spec["fee"],
                    "average_monthly_fee_hkd": spec["average_fee"],
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": spec["speed"],
                    "post_fup_speed_mbps": "",
                    "contract_months": spec["contract"],
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": spec["tariff_type"],
                    "source_status": "public_media_comparison_needs_review",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180)[:800],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hgc_smart_home_living_2020_press", "hgc_smart_home_living_2020_official_pdf"}:
        if not (
            re.search(r"Smart Home Living", text, flags=re.I)
            and re.search(r"1G home broadband plan", text, flags=re.I)
            and re.search(r"smart home kit", text, flags=re.I)
            and re.search(r"\$119\s*/?\s*month", text, flags=re.I)
        ):
            return []
        match = re.search(r"1G home broadband plan.*?\$119\s*/?\s*month", text, flags=re.I)
        excerpt = _excerpt(text, match.start(), match.end(), radius=220)[:700] if match else text[:700]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HGC 1G Home Broadband with smart home kit HK$119",
                "monthly_fee_hkd": "119",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "smart home kit and installation for designated switched customers",
                "tariff_type": "monthly_bundle_fee",
                "source_status": "parsed_public_official_pdf" if source.get("source_id", "").endswith("_official_pdf") else "parsed_current",
                "evidence_excerpt": excerpt,
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"hgc_super_fun_2016_official_press", "hgc_super_fun_2016_official_press_en"}:
        if not (
            (re.search(r"全家\s*Super\s*Fun", text, flags=re.I) or re.search(r"Home Entertainment Super Pack", text, flags=re.I))
            and (re.search(r"100M\s*家居寬頻", text, flags=re.I) or re.search(r"100M.*?fibre broadband", text, flags=re.I))
            and (re.search(r"1G.*?全家\s*Super\s*Fun", text, flags=re.I) or re.search(r"1G.*?fibre broadband", text, flags=re.I))
            and re.search(r"\$138", text)
            and re.search(r"\$188", text)
        ):
            return []
        match = re.search(
            r"3\s*家居寬頻.*?100M\s*家居寬頻.*?1G.*?全家\s*Super\s*Fun.*?\$188|Home Entertainment Super Pack.*?100M or 1G fibre broadband.*?\$188",
            text,
            flags=re.I | re.S,
        )
        excerpt = _excerpt(text, match.start(), match.end(), radius=220)[:800] if match else text[:800]
        rows: List[Dict[str, str]] = []
        for plan_name, fee, speed in [
            ("HGC 3 Home Broadband Super Fun 100M myTV SUPER bundle HK$138", "138", "100"),
            ("HGC 3 Home Broadband Super Fun 1G myTV SUPER bundle HK$188", "188", "1000"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "myTV SUPER decoder, 30 channels, Disney on-demand bundle, HGC on air Wi-Fi",
                    "tariff_type": "monthly_bundle_fee",
                    "source_status": "parsed_current",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {
        "hgc_mobile_launch_2026_press",
        "hgc_mobile_launch_2026_etnet",
        "hgc_mobile_launch_2026_telecomramblings",
    }:
        if not (
            re.search(r"HGC Mobile", text, flags=re.I)
            and re.search(r"30GB\s+of\s+5G\s+local\s+data|30GB.*?local\s+data", text, flags=re.I)
            and re.search(r"15\s+social\s+and\s+OTT\s+entertainment", text, flags=re.I)
            and re.search(r"HK\$98\s+per\s+month|HK\$98/month|\$98\s+per\s+month", text, flags=re.I)
            and re.search(r"4GB\s+of\s+Chinese\s+mainland", text, flags=re.I)
            and re.search(r"One-Stop Home Broadband \+ Mobile Service", text, flags=re.I)
            and re.search(r"HK\$285\s+per\s+month|\$285\s+per\s+month", text, flags=re.I)
        ):
            return []
        mobile_match = re.search(
            r"30GB\s+of\s+5G\s+local\s+data.*?Available at HK\$98 per month.*?4GB of Chinese mainland.{0,20}Macau shared data",
            text,
            flags=re.I | re.S,
        )
        bundle_match = re.search(
            r"One-Stop Home Broadband \+ Mobile Service.*?HK\$285 per month.*?1000M home fibre broadband.*?two HGC Mobile services",
            text,
            flags=re.I | re.S,
        )
        source_status = "parsed_current" if source.get("source_id") == "hgc_mobile_launch_2026_press" else "public_press_release_syndication_needs_review"
        rows: List[Dict[str, str]] = []
        for plan_name, fee, local_data, speed, tariff_type, match in [
            (
                "HGC on air HK$98 4GB",
                "98",
                "4",
                "",
                "monthly_mobile_plan_fee",
                mobile_match,
            ),
            (
                "HGC on air HK$285",
                "285",
                "",
                "1000",
                "monthly_home_broadband_mobile_bundle_fee",
                bundle_match,
            ),
        ]:
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": local_data,
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "Includes 15 social/OTT apps; HK$285 bundle includes 1000M home fibre broadband, voice service, HGCmore e-coupons and two HGC Mobile services",
                    "tariff_type": tariff_type,
                    "source_status": source_status,
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180)[:800],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hgc_broadband_pricequote_2500m_2025", "hgc_broadband_pricequote_category_2500m_2025"}:
        if not (re.search(r"2500M", text, flags=re.I) and re.search(r"\$139", text)):
            return []
        specs = [
            ("market_reference", "139", "36", "HGC 2500M broadband PriceQuote 2025 reference HK$139"),
            ("market_reference", "159", "24", "HGC 2500M broadband PriceQuote 2025 reference HK$159"),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, contract, plan_name in specs:
            if not re.search(r"\$\s*" + re.escape(fee), text):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "2500",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "PriceQuote article describes Wi-Fi 7 router bundle and waived installation fee; use as third-party reference only",
                    "tariff_type": "public_third_party_broadband_offer_reference",
                    "source_status": "public_third_party_offer_listing_needs_review",
                    "evidence_excerpt": (
                        "HGC 2500M 光纖寬頻計劃，月費 HK$139 起，36 個月合約；另見 24 個月合約 HK$159。"
                        " PriceQuote 公开报价页，非运营商官方标准价，实际月费以 HGC 合约和覆盖地址为准。"
                    )[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hgc_broadband_pricequote_2000m_x50poe_2025":
        normalized = text.translate(str.maketrans("﹩＄０１２３４５６７８９", "$$0123456789"))
        if not (re.search(r"2000M|2000MB|2G", normalized, flags=re.I) and re.search(r"\$189", normalized) and re.search(r"\$199", normalized)):
            return []
        specs = [
            (
                "market_reference",
                "189",
                "24",
                "HGC 2000M broadband PriceQuote 2025 reference HK$189",
                r"24個月合約[:：]?\s*月費為\$189|24個月\s*\$189",
            ),
            (
                "market_reference",
                "199",
                "24",
                "HGC 2000M broadband PriceQuote 2025 reference HK$199",
                r"24個月合約[:：]?\s*月費為\$199|24個月\s*\$199",
            ),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, contract, plan_name, pattern in specs:
            match = re.search(pattern, normalized, flags=re.I | re.S)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "2000",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "PriceQuote page describes HGC 2000M fibre broadband with upgraded router; use as third-party reference only",
                    "tariff_type": "public_third_party_broadband_offer_reference",
                    "source_status": "public_third_party_offer_listing_needs_review",
                    "evidence_excerpt": _excerpt(normalized, match.start(), match.end(), radius=280)[:900],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hgc_broadband_pricequote_2200m_2026":
        normalized = text.translate(str.maketrans("﹩＄０１２３４５６７８９", "$$0123456789"))
        match = re.search(r"HGC\s*2\.?5G.*?\$139\s*[-–]\s*\$149", normalized, flags=re.I | re.S)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": "market_reference",
                "plan_name": "HGC 2500M broadband PriceQuote 2026 comparison reference HK$149",
                "monthly_fee_hkd": "149",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "2500",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "PriceQuote comparison range HK$139-HK$149; use as third-party/channel reference only",
                "tariff_type": "public_third_party_broadband_offer_reference",
                "source_status": "public_third_party_offer_listing_needs_review",
                "evidence_excerpt": (
                    _excerpt(normalized, match.start(), match.end(), radius=260)
                    + "；PriceQuote 公开比较页，非运营商官方标准价，实际月费以 HGC 合约和覆盖地址为准。"
                )[:900],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"hgc_home_broadband_standard_monthly_2026", "hgc_home_broadband_standard_monthly_2026_tc"}:
        if not (
            (re.search(r"Home Broadband Monthly Service Plan", text, flags=re.I) and re.search(r"Standard Monthly Fee", text, flags=re.I))
            or (re.search(r"家居寬頻月費計劃", text) and re.search(r"正價月費", text))
        ):
            return []
        speed_fee_pairs = [
            ("10G", "10000", "1299"),
            ("2.5G", "2500", "766"),
            ("2.2G", "2200", "666"),
            ("2G", "2000", "666"),
            ("1G", "1000", "598"),
            ("500M", "500", "488"),
            ("300M", "300", "398"),
            ("200M", "200", "348"),
            ("100M", "100", "298"),
            ("30M", "30", "198"),
            ("10M", "10", "188"),
            ("6M", "6", "188"),
        ]
        table_match = re.search(r"Home Broadband Monthly Service Plan.*?6M\s+\$188", text, flags=re.I) or re.search(
            r"家居寬頻月費計劃.*?6M\s+\$188",
            text,
            flags=re.S,
        )
        table_excerpt = _excerpt(text, table_match.start(), table_match.end(), radius=120)[:600] if table_match else text[:600]
        rows: List[Dict[str, str]] = []
        for label, speed, fee in speed_fee_pairs:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HGC Home Broadband {label} standard monthly fee HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_archive" if archive_url else "parsed_current_official_standard_fee",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hgc_home_broadband_terms_standard_monthly_2026", "hgc_home_broadband_terms_standard_monthly_2026_tc"}:
        terms_match = re.search(
            r"The standard monthly fees of 6Mbps, 10Mbps, 30Mbps, 100Mbps, 200Mbps, 300Mbps, 500Mbps, 1Gbps, 2Gbps, 2\.2Gbps, 2\.5Gbps and 10Gbps Service are \$188, \$188, \$198, \$298, \$348, \$398, \$448, \$598, \$666, \$666, \$766 and \$1,299 respectively",
            text,
            flags=re.I,
        ) or re.search(
            r"6Mbps\s*、\s*10Mbps\s*、\s*30Mbps\s*、\s*100Mbps\s*、\s*200Mbps\s*、\s*300Mbps\s*、\s*500Mbps\s*、\s*1Gbps\s*、\s*2Gbps\s*、\s*2\.2Gbps\s*、\s*2\.5Gbps\s*及\s*10Gbps\s*寬頻?服務之(?:正價|標準)月費分別為\s*\$188\s*、\s*\$188\s*、\s*\$198\s*、\s*\$298\s*、\s*\$348\s*、\s*\$398\s*、\s*\$448\s*、\s*\$598\s*、\s*\$666\s*、\s*\$666\s*、\s*\$766\s*及\s*\$1,299",
            text,
        )
        if not terms_match:
            return []
        speed_fee_pairs = [
            ("6M", "6", "188"),
            ("10M", "10", "188"),
            ("30M", "30", "198"),
            ("100M", "100", "298"),
            ("200M", "200", "348"),
            ("300M", "300", "398"),
            ("500M", "500", "448"),
            ("1G", "1000", "598"),
            ("2G", "2000", "666"),
            ("2.2G", "2200", "666"),
            ("2.5G", "2500", "766"),
            ("10G", "10000", "1299"),
        ]
        table_excerpt = _excerpt(text, terms_match.start(), terms_match.end(), radius=120)[:600]
        rows = []
        for label, speed, fee in speed_fee_pairs:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HGC Home Broadband terms after-minimum-period monthly fee {label} HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "terms_after_minimum_period_monthly_fee_reference",
                    "source_status": "parsed_current_terms_reference",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hgc_10g_25g_2024_official", "hgc_10g_25g_2024_prnewswire"}:
        launch_match = re.search(r"launched the 10G \(10000Mbps\) optical fibre broadband service for household customers", text, flags=re.I)
        corporate_match = re.search(r"25G broadband services?.*?corporate customers", text, flags=re.I)
        portfolio_match = re.search(r"service plans spanning 1G, 2G, 2\.2G, 2\.5G and 10G", text, flags=re.I)
        if not (launch_match and corporate_match and portfolio_match):
            return []
        excerpt = _excerpt(text, launch_match.start(), max(corporate_match.end(), portfolio_match.end()), radius=120)[:700]
        rows = []
        for plan_name, speed, segment, tariff_type in [
            ("HGC 10G optical fibre broadband service launch", "10000", "家庭宽频客户", "product_launch_no_disclosed_monthly_fee"),
            ("HGC 25G ultra-fast broadband service for corporate customers", "25000", "企业客户", "product_launch_no_disclosed_monthly_fee"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_family": "HGC Fibre Broadband",
                    "plan_name": plan_name,
                    "monthly_fee_hkd": "",
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": tariff_type,
                    "source_status": "public_product_launch_no_price",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hgc_25g_2023_official", "hgc_25g_2023_official_pdf"}:
        bundle_match = re.search(
            r"2\.5G broadband service plus 2 sets of TP-Link Deco XE75 Pro Wi\s*-?\s*Fi 6(?:E)? router.*?monthly fee of HK\s*\$?\s*298",
            text,
            flags=re.I | re.S,
        )
        phone_match = re.search(r"Additional HK\s*\$?30/month\*? for Home Telephone service", text, flags=re.I)
        if not (bundle_match and phone_match):
            return []
        rows: List[Dict[str, str]] = []
        for plan_name, fee, speed, tariff_type, match in [
            (
                "HGC 2.5G broadband with TP-Link Deco XE75 Pro Wi-Fi 6 router HK$298",
                "298",
                "2500",
                "monthly_plan_fee",
                bundle_match,
            ),
            (
                "HGC Home Telephone add-on for 2.5G broadband HK$30",
                "30",
                "",
                "monthly_value_added_service_fee",
                phone_match,
            ),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "36" if fee == "298" else "",
                    "local_voice": "Home Telephone service" if fee == "30" else "",
                    "add_on_charges_hkd": "Additional HK$30/month Home Telephone service" if fee == "298" else "",
                    "tariff_type": tariff_type,
                    "source_status": "parsed_current",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180)[:800],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hgc_2g_2023_official_pdf", "hgc_2g_2023_telecomreviewamericas"}:
        match = re.search(
            r"(?:2\s*G|2G|2000Mbps).*?(?:Wi-?Fi\s*6\s*router|WiFi\s*6\s*router).*?(?:HK\s*)?\$?\s*139\s*(?:per month|/month|monthly)|(?:HK\s*)?\$?\s*139\s*(?:per month|/month|monthly).*?(?:2\s*G|2G|2000Mbps)",
            text,
            flags=re.I | re.S,
        )
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HGC 2G 2000Mbps broadband with Wi-Fi 6 Router HK$139",
                "monthly_fee_hkd": "139",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "2000",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "Apply on or before 31 December 2023; includes a Wi-Fi 6 Router according to public launch release.",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_public_official_pdf" if source.get("source_id") == "hgc_2g_2023_official_pdf" else "public_press_syndication_needs_review",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:900],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hgc_line_for_four_2018_official_tc":
        bundle_match = re.search(r"一家四口.*?2\.2G.*?myTV SUPER.*?月費\s*\$218\s*起", text, flags=re.S)
        wifi_match = re.search(r"Wi-?Fi\s*360.*?月費\s*\$58\s*起", text, flags=re.I | re.S)
        if not (bundle_match and wifi_match):
            return []
        rows = []
        for fee, plan_name, tariff_type, match in [
            (
                "218",
                "HGC 一家四口 2.2G 光纖寬頻連 myTV SUPER HK$218 起",
                "monthly_bundle_fee",
                bundle_match,
            ),
            (
                "58",
                "HGC Wi-Fi 360 service HK$58 up",
                "monthly_value_added_service_fee",
                wifi_match,
            ),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "2200",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": tariff_type,
                    "source_status": "parsed_current_official_page_tc",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180)[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hgc_line_for_four_2018_official_pdf":
        if not (re.search(r"Line for Four", text, flags=re.I) and re.search(r"HKD\s*\$218", text, flags=re.I) and re.search(r"Wi-Fi 360", text, flags=re.I) and re.search(r"HKD\s*\$58", text, flags=re.I)):
            return []
        rows = []
        match = re.search(r"Line for Four.*?HKD\s*\$58 up", text, flags=re.I)
        excerpt = _excerpt(text, match.start(), match.end(), radius=160)[:600] if match else text[:600]
        for fee, plan_name, tariff_type in [
            ("218", "HGC Line for Four 2.2G fibre broadband plus myTV SUPER HK$218 up", "monthly_bundle_fee"),
            ("58", "HGC Wi-Fi 360 service HK$58 up", "monthly_value_added_service_fee"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "2200",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": tariff_type,
                    "source_status": "parsed_archive" if archive_url else "parsed_current_official_pdf",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hgc_mytv_1g_2021_press", "hgc_mytv_1g_2021_official_pdf"}:
        if not (
            re.search(r"HGC Broadband", text, flags=re.I)
            and re.search(r"1G Home Broadband", text, flags=re.I)
            and re.search(r"myTV Gold", text, flags=re.I)
            and re.search(r"\$198\s*#?/?mth|\$198\s*#?\s*up", text, flags=re.I)
            and re.search(r"24 months", text, flags=re.I)
        ):
            return []
        match = re.search(r"HGC Broadband \+ myTV Gold Service Plan.*?\$198\s*#?\s*up.*?24 months", text, flags=re.I)
        if not match:
            match = re.search(r"1G Home Broadband.*?\$198\s*#?/?mth.*?24 months", text, flags=re.I)
        excerpt = _excerpt(text, match.start(), match.end(), radius=160)[:700] if match else text[:700]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HGC 1G Home Broadband plus myTV Gold bundle HK$198",
                "monthly_fee_hkd": "198",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "",
                "add_on_charges_hkd": "myTV Gold service and myTV SUPER concurrent viewing bundle",
                "tariff_type": "monthly_bundle_fee",
                "source_status": "parsed_public_official_pdf" if source.get("source_id", "").endswith("_official_pdf") else "parsed_current",
                "evidence_excerpt": excerpt,
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hgc_mytv_2025_mytvsuper_service_fee":
        if not (
            re.search(r"服務收費\s*-\s*HGC\s*寬頻", text, flags=re.I)
            and re.search(r"寬頻娛樂組合月費計劃", text)
            and re.search(r"連\s*myTV\s*Gold", text, flags=re.I)
            and re.search(r"\$198", text)
            and re.search(r"最少申用期\s*24個月", text)
        ):
            return []
        match = re.search(r"連\s*myTV\s*Gold.*?月費\*\s*\$198.*?最少申用期\s*24個月", text, flags=re.I | re.S)
        excerpt = _excerpt(text, match.start(), match.end(), radius=260)[:900] if match else text[:900]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HGC Fibre Broadband with myTV Gold service",
                "monthly_fee_hkd": "198",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "",
                "add_on_charges_hkd": "myTV SUPER service-fee page; HGC broadband entertainment bundle suggested retail price, subject to network supplier terms",
                "tariff_type": "public_channel_monthly_plan_fee",
                "source_status": "public_partner_service_fee_needs_review",
                "evidence_excerpt": excerpt,
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"hgc_mytv_2025_broadband_pricequote", "hgc_mytv_2025_broadband_pricequote_post"}:
        normalized = text.translate(str.maketrans("﹩＄０１２３４５６７８９", "$$0123456789"))
        if not (
            (re.search(r"2025年3月2日", normalized) or re.search(r"HGC光纎寛頻連MYTV", normalized, flags=re.I))
            and re.search(r"HGC光纖寬頻|HGC光纎寛頻", normalized, flags=re.I)
            and re.search(r"MYTV", normalized, flags=re.I)
        ):
            return []
        rows: List[Dict[str, str]] = []
        specs = [
            (
                "HGC Fibre Broadband with myTV Gold service",
                "198",
                "24",
                r"(?:HGC光纖寬頻連MYTV GOLD.*?月費[:：]\$198\s*24個月合約|光纖寬頻搭配MYTV GOLD服務，月費\$198，合約期為24個月)",
            ),
            (
                "HGC Fibre Broadband with myTV SUPER Basic public housing offer",
                "109",
                "36",
                r"(?:HGC光纖寬頻連MYTV SUPER基本版\s*公屋月費[:：]\$109|公屋和私樓住戶提供優惠，分別是月費\$109和\$119)",
            ),
            (
                "HGC Fibre Broadband with myTV SUPER Basic private building offer",
                "119",
                "36",
                r"(?:HGC光纖寬頻連MYTV SUPER基本版.*?私樓月費[:：]\$119\s*36個月合約|公屋和私樓住戶提供優惠，分別是月費\$109和\$119，合約期為36個月)",
            ),
        ]
        for plan_name, fee, contract, pattern in specs:
            match = re.search(pattern, normalized, flags=re.I | re.S)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "actual charges subject to supplier contract",
                    "tariff_type": "public_channel_monthly_plan_fee",
                    "source_status": "public_third_party_offer_listing_needs_review",
                    "evidence_excerpt": _excerpt(normalized, match.start(), match.end(), radius=260)[:900],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return _dedupe(rows)
    if source.get("source_id", "").startswith("icable_residential_service_charge_"):
        table_match = re.search(
            r"i-CABLE Broadband & HomeLine Service Rates.*?(?:OTT & Additional Service Rates|Updated as of|$)",
            text,
            flags=re.I,
        )
        if not table_match:
            return []
        table_text = table_match.group(0)
        rows: List[Dict[str, str]] = []
        for match in re.finditer(r"((?:2x)?[0-9]{2,4}M)\s+per month\s+([0-9]{2,4})", table_text, flags=re.I):
            speed_label = match.group(1)
            fee = match.group(2)
            speed = "2000" if speed_label.lower() == "2x1000m" else speed_label.lower().removesuffix("m")
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"i-CABLE Broadband {speed_label} regular monthly rate HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_public_official_pdf",
                    "evidence_excerpt": _excerpt(text, table_match.start(), table_match.end(), radius=120)[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "icable_mytv_bundle_2021_tvb_press":
        if not (
            re.search(r"23 April 2021", text)
            and re.search(r"i-CABLE 1G Home Broadband", text, flags=re.I)
            and re.search(r"myTV Gold", text, flags=re.I)
            and re.search(r"\$198\s*#?\s*up", text, flags=re.I)
        ):
            return []
        match = re.search(r"i-CABLE \+ myTV Gold Service Plan.*?\$198\s*#?\s*up.*?24 months", text, flags=re.I)
        if not match:
            match = re.search(r"1000 Mbps.*?\$198\s*#?\s*up.*?24 months", text, flags=re.I)
        excerpt = _excerpt(text, match.start(), match.end(), radius=120)[:700] if match else text[:700]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "i-CABLE 1G Home Broadband plus myTV Gold bundle HK$198",
                "monthly_fee_hkd": "198",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "",
                "add_on_charges_hkd": "myTV Gold service and myTV SUPER set-top-box concurrent viewing bundle",
                "tariff_type": "monthly_bundle_fee",
                "source_status": "public_partner_press_release_needs_review",
                "evidence_excerpt": excerpt,
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "icable_mytv_2021_mytvsuper_service_fee":
        if not (
            re.search(r"i-CABLE", text, flags=re.I)
            and re.search(r"1G Home Broadband \+ myTV Gold Service", text, flags=re.I)
            and re.search(r"Monthly Fee\*?\s*\$198", text, flags=re.I)
            and re.search(r"Contract Period\s*24 months", text, flags=re.I)
        ):
            return []
        match = re.search(
            r"1G Home Broadband \+ myTV Gold Service.*?Monthly Fee\*?\s*\$198.*?Contract Period\s*24 months",
            text,
            flags=re.I | re.S,
        )
        excerpt = _excerpt(text, match.start(), match.end(), radius=260)[:900] if match else text[:900]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "i-CABLE 1G Home Broadband plus myTV Gold bundle HK$198",
                "monthly_fee_hkd": "198",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "",
                "add_on_charges_hkd": "myTV SUPER service-fee page; suggested retail price only and subject to ISP terms",
                "tariff_type": "monthly_bundle_fee",
                "source_status": "public_partner_service_fee_needs_review",
                "evidence_excerpt": excerpt,
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"icable_broadbandqueen_2024_public_offer", "icable_broadbandqueen_2024_public_offer_tc"}:
        if not (
            (
                re.search(r"10 November 2024", text, flags=re.I)
                and re.search(r"i-?Cable Broadband Offers", text, flags=re.I)
                and re.search(r"100M monthly fee starts from \$78", text, flags=re.I)
            )
            or (
                re.search(r"2024年11月10日", text)
                and re.search(r"i-?Cable有線寬頻", text, flags=re.I)
                and re.search(r"100M月費由\$78起", text)
            )
        ):
            return []
        rows: List[Dict[str, str]] = []
        offer_patterns = [
            ("100M", "100", "78"),
            ("500M", "500", "128"),
            ("1000M", "1000", "168"),
        ]
        excerpt_match = re.search(
            r"i-?Cable Broadband offers fibre plans from 100M to 1000M.*?1000M starts from \$168",
            text,
            flags=re.I,
        ) or re.search(r"i-?Cable有線寬頻提供由100M到1000M.*?1000M由\$168起", text, flags=re.I)
        excerpt = _excerpt(text, excerpt_match.start(), excerpt_match.end(), radius=100)[:700] if excerpt_match else text[:700]
        for speed_label, speed, fee in offer_patterns:
            if not re.search(rf"{re.escape(speed_label)}.*?\${fee}(?!\d)|\${fee}(?!\d).*?{re.escape(speed_label)}", text, flags=re.I):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"i-CABLE BroadbandQueen 2024 {speed_label} from HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee_from",
                    "source_status": "public_third_party_offer_listing_needs_review",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id", "").endswith("_moneyhero_broadband_comparison_2025"):
        if not (re.search(r"2025年1月更新", text) and re.search(r"各大寬頻公司參考收費", text)):
            return []
        table_match = re.search(r"各大寬頻公司參考收費.*?H3:\s*1", text)
        excerpt = table_match.group(0) if table_match else text
        data_by_brand = {
            "HKBN": [
                ("private_building_reference", "148", "2500", "指定 2.5Gbps GigaFast 寬頻計劃月費低至HK$148"),
                ("public_housing_reference", "109", "1000", "公屋參考價錢最低HK$109"),
            ],
            "i-CABLE": [
                ("private_building_reference", "68", "1000", "私人樓宇參考價錢最低HK$68"),
                ("public_housing_reference", "68", "1000", "公屋參考價錢最低HK$68"),
            ],
            "HGC": [
                ("private_building_reference", "129", "1000", "私人樓宇參考價錢最低HK$129"),
                ("public_housing_reference", "119", "1000", "公屋參考價錢HK$119"),
            ],
            "SmarTone": [
                ("private_building_reference", "98", "1000", "私人樓宇參考價錢最低HK$98"),
                ("public_housing_reference", "88", "1000", "公屋參考價錢最低HK$88"),
            ],
        }
        contract_by_brand = {"HKBN": "36", "i-CABLE": "36", "HGC": "12", "SmarTone": "24"}
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, evidence in data_by_brand.get(source["brand"], []):
            if not re.search(re.escape(f"HK${fee}"), excerpt):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"{source['brand']} {speed}M broadband {segment} MoneyHero 2025 reference HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": contract_by_brand.get(source["brand"], ""),
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "public_third_party_broadband_comparison_reference",
                    "source_status": "public_third_party_comparison_needs_review",
                    "evidence_excerpt": (evidence + "；MoneyHero 页面注明资料只供参考，最终收费和优惠以供应商公布为准。 " + excerpt[:450])[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hkbn_broadband_roadshowoffer_2500m_148_2025":
        if not (
            re.search(r"香港寬頻", text)
            and re.search(r"2500M", text, flags=re.I)
            and re.search(r"\$\s*148", text)
            and re.search(r"36\s*個月", text)
        ):
            return []
        match = re.search(r"香港寬頻新入伙屋苑2500M\s*\$148.*?36\s*個月", text, flags=re.I | re.S)
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": "private_building_reference",
                "plan_name": "HKBN 2500M broadband private_building_reference RoadshowOffer 2025 reference HK$148",
                "monthly_fee_hkd": "148",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "2500",
                "post_fup_speed_mbps": "",
                "contract_months": "36",
                "local_voice": "",
                "add_on_charges_hkd": "RoadshowOffer public/channel reference mentions TP-Link BE230 Wi-Fi 7 router, myTV SUPER 24 months, and home telephone service",
                "tariff_type": "public_third_party_broadband_offer_reference",
                "source_status": "public_third_party_channel_reference_needs_review",
                "evidence_excerpt": (
                    (_excerpt(text, match.start(), match.end(), radius=220) if match else text[:700])
                    + "；Broadband RoadshowOffer 公开渠道/比较摘录，非 HKBN 官方标准价；用于和其它第三方市场价互证，实际资费以 HKBN 合约和覆盖地址为准。"
                )[:800],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hkbn_findplanking_2500m_149_2026":
        if not (
            re.search(r"Find Plan King|求Plan王", text, flags=re.I)
            and re.search(r"香港寬頻", text)
            and re.search(r"2500M", text, flags=re.I)
            and re.search(r"(?:月費)?\s*(?:💰)?\s*149|\$\s*149", text)
            and re.search(r"24\s*個月", text)
        ):
            return []
        match = re.search(
            r"2500M獨享Giga\s*Fast光纖入屋.{0,80}(?:月費)?\s*(?:💰)?\s*149.{0,80}24\s*個月",
            text,
            flags=re.I | re.S,
        )
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": "market_reference",
                "plan_name": "HKBN 2500M broadband Find Plan King 2026 reference HK$149",
                "monthly_fee_hkd": "149",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "2500",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "home telephone service included",
                "add_on_charges_hkd": "Find Plan King public/channel listing mentions F-Secure, home telephone, myTV SUPER basic 12 months, and TP-Link Wi-Fi 7 router",
                "tariff_type": "public_third_party_broadband_offer_reference",
                "source_status": "public_third_party_offer_listing_needs_review",
                "evidence_excerpt": (
                    (_excerpt(text, match.start(), match.end(), radius=220) if match else text[:700])
                    + "；Find Plan King 页面标注为公开渠道报价/销售资料，非 HKBN 官方标准价；用于和其它第三方市场价互证，实际资费以 HKBN 合约和覆盖地址为准。"
                )[:800],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hkbn_pricequote_posts_2500m_149_2026":
        if not (
            re.search(r"Broadband-PriceQuote|寬頻報價|PriceQuote", text, flags=re.I)
            and re.search(r"HKBN|香港寬頻", text, flags=re.I)
            and re.search(r"2500M", text, flags=re.I)
            and re.search(r"(?:\$\s*)?149", text)
            and re.search(r"24\s*個月|簽\s*24\s*個月", text)
        ):
            return []
        match = re.search(
            r"(?:HKBN|香港寬頻).{0,140}2500M.{0,180}(?:\$\s*)?149.{0,120}24\s*個月|"
            r"2500M.{0,180}(?:\$\s*)?149.{0,120}24\s*個月",
            text,
            flags=re.I | re.S,
        )
        segment_labels = [("market_reference", "market")]
        if re.search(r"私樓|私人住宅|私人屋苑|private\s+housing|private\s+building", text, flags=re.I) and re.search(
            r"公屋|居屋|公居屋|public\s+housing", text, flags=re.I
        ):
            segment_labels = [
                ("private_building_reference", "private"),
                ("public_housing_reference", "public"),
            ]
        rows: List[Dict[str, str]] = []
        excerpt = _excerpt(text, match.start(), match.end(), radius=260) if match else text[:700]
        for segment, label in segment_labels:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"HKBN 2500M broadband {segment} PriceQuote posts 2026 reference HK$149",
                    "monthly_fee_hkd": "149",
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "2500",
                    "post_fup_speed_mbps": "",
                    "contract_months": "24",
                    "local_voice": "home telephone service included",
                    "add_on_charges_hkd": "Broadband-PriceQuote public/channel listing mentions TP-Link BE230 Wi-Fi 7 router, home telephone service, antivirus apps, travel card, and final supplier contract prevails",
                    "tariff_type": "public_third_party_broadband_offer_reference",
                    "source_status": "public_third_party_channel_reference_needs_review",
                    "evidence_excerpt": (
                        excerpt
                        + f"；PriceQuote 公开渠道/比较摘录，按 {label} segment 记录；非 HKBN 官方标准价，实际资费以 HKBN 合约和覆盖地址为准。"
                    )[:800],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id", "").endswith("_investbrother_broadband_comparison_2025"):
        if not (re.search(r"家居寬頻比較", text) and re.search(r"最近更新：2025\s*年\s*9\s*月\s*5\s*日", text)):
            return []
        data_by_brand = {
            "HKBN": [
                ("public_housing_reference", "109", "1000", "公屋月費 HK$109", "36"),
                ("private_building_reference", "149", "1000", "私人住宅月費 HK$149", "36"),
                ("public_housing_reference", "149", "2500", "公屋月費 HK$149", "24"),
                ("private_building_reference", "149", "2500", "私人住宅月費 HK$149", "24"),
                ("market_reference", "169", "10000", "公屋月費 HK$169 私人住宅月費 HK$169", "24"),
            ],
            "HGC": [
                ("public_housing_reference", "89", "1000", "公屋月費 HK$89", "36"),
                ("private_building_reference", "109", "1000", "私人住宅月費 HK$109", "36"),
                ("public_housing_reference", "139", "2500", "公屋月費 HK$139", "36"),
                ("private_building_reference", "149", "2500", "私人住宅月費 HK$149", "36"),
            ],
            "i-CABLE": [
                ("public_housing_reference", "88", "1000", "公屋月費 HK$88", "36"),
                ("private_building_reference", "88", "1000", "私人住宅月費 HK$88", "36"),
            ],
            "SmarTone": [
                ("private_building_reference", "98", "1000", "私人住宅月費 HK$98", "36"),
                ("public_housing_reference", "88", "1000", "公屋月費 HK$88", "36"),
            ],
        }
        start_labels = {
            "HKBN": r"家居寬頻比較\|HKBN香港寬頻1000M計劃",
            "HGC": r"家居寬頻比較\|HGC寬頻1000M計劃",
            "i-CABLE": r"家居寬頻比較\|有線寬頻1000M計劃",
            "SmarTone": r"家居寬頻比較\|SmarTone\s+1000M計劃",
        }
        next_labels = {
            "HKBN": r"家居寬頻比較\|HGC寬頻1000M計劃",
            "HGC": r"家居寬頻比較\|有線寬頻1000M計劃",
            "i-CABLE": r"家居寬頻比較\|SmarTone\s+1000M計劃",
            "SmarTone": r"家居寬頻比較\|網上行Netvigator",
        }
        start_pat = start_labels.get(source["brand"])
        next_pat = next_labels.get(source["brand"])
        if not start_pat:
            return []
        section_match = re.search(start_pat + r".*?" + next_pat, text, flags=re.I | re.S) if next_pat else None
        section = section_match.group(0) if section_match else text
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, evidence, contract_months in data_by_brand.get(source["brand"], []):
            if not (re.search(re.escape(evidence), section) and re.search(r"豁免安裝費", section)):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"{source['brand']} {speed}M broadband {segment} InvestBrother 2025 reference HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": contract_months,
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "public_third_party_broadband_comparison_reference",
                    "source_status": "public_third_party_comparison_needs_review",
                    "evidence_excerpt": (evidence + "；InvestBrother 页面注明价格及优惠内容作参考，详情请查阅官网。 " + section[:450])[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id", "").endswith("_moneysmart_broadband_comparison_2026"):
        if not (re.search(r"MoneySmart\s+寬頻上網比較2026", text) and re.search(r"2026年5月3日", text)):
            return []
        data_by_brand = {
            "HKBN": [
                (
                    "market_reference",
                    "169",
                    "2500",
                    "",
                    "HKBN: 2.5Gbps GigaFast HK$169/month",
                    "",
                ),
                (
                    "market_reference",
                    "199",
                    "2500",
                    "",
                    "2.5Gbps with Wi-Fi 7 router HK$199/month",
                    "Wi-Fi 7 router bundle",
                ),
                (
                    "market_reference",
                    "149",
                    "1000",
                    "",
                    "1Gbps with Wi-Fi 6 router HK$149/month",
                    "Wi-Fi 6 router bundle",
                ),
            ],
            "SmarTone": [
                (
                    "market_reference",
                    "98",
                    "1000",
                    "24",
                    "SmarTone 光纖寬頻: 1000Mbps HK$98/月 24個月",
                    "",
                ),
                (
                    "market_reference",
                    "128",
                    "2000",
                    "36",
                    "2Gbps/2.2Gbps HK$128/月 36個月",
                    "",
                ),
                (
                    "market_reference",
                    "128",
                    "2200",
                    "36",
                    "2Gbps/2.2Gbps HK$128/月 36個月",
                    "",
                ),
                (
                    "market_reference",
                    "148",
                    "2200",
                    "36",
                    "2Gbps/2.2Gbps + Wi-Fi 7 HK$148/月 36個月",
                    "Wi-Fi 7 router bundle",
                ),
                (
                    "market_reference",
                    "154",
                    "2200",
                    "36",
                    "2Gbps/2.2Gbps + Wi-Fi 6 HK$154/月 36個月",
                    "Wi-Fi 6 router bundle",
                ),
            ],
            "HGC": [
                (
                    "market_reference",
                    "129",
                    "1000",
                    "36",
                    "Wi-Fi 6路由器 X 1G 寬頻服務 HK$129/月 36個月",
                    "Wi-Fi 6 router bundle",
                ),
                (
                    "market_reference",
                    "198",
                    "1000",
                    "36",
                    "myTV Gold X 1G 寬頻服務 HK$198/月 36個月",
                    "myTV Gold bundle",
                ),
                (
                    "market_reference",
                    "119",
                    "1000",
                    "36",
                    "hmvod X 1G 寬頻服務 HK$119/月 36個月",
                    "hmvod bundle",
                ),
                (
                    "market_reference",
                    "189",
                    "2000",
                    "24",
                    "Wi-Fi 7路由器 X 2G 寬頻服務 HK$189/月 24個月",
                    "Wi-Fi 7 router bundle",
                ),
                (
                    "market_reference",
                    "199",
                    "2000",
                    "24",
                    "Wi-Fi 7路由器 X 2G 寬頻服務 HK$199/月 24個月",
                    "Wi-Fi 7 router bundle",
                ),
            ],
            "i-CABLE": [
                (
                    "market_reference",
                    "88",
                    "1000",
                    "",
                    "i-CABLE: 1000M HK$88起",
                    "",
                ),
                (
                    "market_reference",
                    "98",
                    "2000",
                    "",
                    "2x1000M HK$98起",
                    "",
                ),
                (
                    "market_reference",
                    "68",
                    "200",
                    "",
                    "200M HK$68起",
                    "",
                ),
            ],
        }
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, contract, evidence, add_on in data_by_brand.get(source["brand"], []):
            if not re.search(re.escape(evidence), text):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"{source['brand']} {speed}M broadband {segment} MoneySmart 2026 reference HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": add_on,
                    "tariff_type": "public_third_party_broadband_comparison_reference",
                    "source_status": "public_third_party_comparison_needs_review",
                    "evidence_excerpt": (
                        evidence
                        + "；MoneySmart 页面/公开索引标注 2026年5月3日更新，本机直连返回 Cloudflare 403；该来源仅作公开第三方参考，正式资费、覆盖和安装费以运营商公布为准。"
                    )[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "icable_fibrehk_isp_comparison_2026":
        if not (
            re.search(r"HKBN\s+vs\s+HGC\s+vs\s+SmarTone\s+vs\s+CMHK\s+vs\s+i-Cable\s+vs\s+HKT", text, flags=re.I)
            and re.search(r"Price Comparison\s*\(1000M Plans\)", text, flags=re.I)
            and re.search(r"i-Cable\s+HK\$168", text, flags=re.I)
            and re.search(r"indicative as of March 2026", text, flags=re.I)
        ):
            return []
        match = re.search(r"Provider\s+Monthly Fee.*?i-Cable\s+HK\$168.*?Note:\s*Prices shown are indicative as of March 2026", text, flags=re.I | re.S)
        excerpt = _clean_text(match.group(0) if match else text)[:800]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": "market_reference",
                "plan_name": "i-CABLE 1000M broadband FibreHK 2026 comparison reference HK$168",
                "monthly_fee_hkd": "168",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "",
                "add_on_charges_hkd": "FibreHK comparison says router rental HK$18/mo and prices vary by district, building, and promotional offers",
                "tariff_type": "public_third_party_broadband_comparison_reference",
                "source_status": "public_third_party_comparison_needs_review",
                "evidence_excerpt": (excerpt + "；FibreHK 公开比较页注明价格为 2026年3月 indicative，仅作第三方参考，正式资费和覆盖以运营商/申请条款为准。")[:900],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id", "").endswith("_yahoo_broadband_comparison_2026"):
        if not (
            re.search(r"Yahoo\s+香港\s+家居寬頻推介", text)
            and re.search(r"Public Yahoo HK article excerpt verified 2026-07-06", text)
        ):
            return []
        data_by_brand = {
            "HKBN": [
                ("public_housing_reference", "88", "香港寬頻 家居寬頻 1000M 公屋/居屋 HK$88/月 36個月合約"),
                ("private_building_reference", "109", "私人住宅 HK$109/月 36個月合約"),
            ],
            "HGC": [
                ("public_housing_reference", "119", "HGC Broadband 家居寬頻 1000M 公屋/居屋 HK$119/月 36個月合約"),
                ("private_building_reference", "129", "私人住宅 HK$129/月 36個月合約"),
            ],
            "SmarTone": [
                ("public_housing_reference", "88", "SmarTone 家居寬頻 1000M 公屋/居屋 HK$88/月 36個月合約"),
                ("private_building_reference", "98", "私人住宅 HK$98/月 36個月合約"),
            ],
            "i-CABLE": [
                ("public_housing_reference", "88", "i-CABLE 家居寬頻 1000M 公屋/居屋 HK$88/月 36個月合約"),
                ("private_building_reference", "118", "私人住宅 HK$118/月 36個月合約"),
            ],
        }
        rows: List[Dict[str, str]] = []
        for segment, fee, evidence in data_by_brand.get(source["brand"], []):
            if not re.search(re.escape(evidence), text):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"{source['brand']} 1000M broadband {segment} Yahoo HK 2026 reference HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "1000",
                    "post_fup_speed_mbps": "",
                    "contract_months": "36",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "public_media_broadband_comparison_reference",
                    "source_status": "public_media_comparison_needs_review",
                    "evidence_excerpt": (evidence + "；Yahoo 香港公开比较文章，页面更新日期标注为 2026-06-30；仅作媒体参考，正式月费、覆盖和安装费以运营商申请结果为准。")[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id", "").endswith("_shangtaika_broadband_comparison_2026"):
        if not (re.search(r"全港寬頻上網優惠比較平台", text) and re.search(r"資料更新於2026年6月|data updated 2026年6月", text)):
            return []
        data_by_brand = {
            "HKBN": [
                ("private_building_reference", "109", "1000", "36", "香港寬頻 1000M HK$109起 36個月"),
                ("market_reference", "199", "2500", "24", "香港寬頻 2.5G/2.5Gbps HK$199 24個月"),
                ("market_reference", "169", "2500", "24", "香港寬頻 2.5G HK$169 24個月"),
            ],
            "SmarTone": [
                ("public_housing_reference", "88", "1000", "36", "SmarTone 智能家居光纖寬頻 公屋/居屋 1000M HK$88起 36個月"),
                ("private_building_reference", "98", "1000", "36", "SmarTone 1000M HK$98 36個月"),
                ("market_reference", "128", "2000", "36", "SmarTone 2G HK$128 36個月"),
                ("market_reference", "128", "2200", "36", "SmarTone 2.2G HK$128 36個月"),
            ],
            "i-CABLE": [
                ("market_reference", "68", "200", "", "有線寬頻 200M HK$68起"),
                ("market_reference", "68", "100", "", "有線寬頻 100M HK$68起"),
                ("market_reference", "89", "1000", "", "有線寬頻 1000M HK$89起"),
                ("market_reference", "129", "2000", "", "有線寬頻 2000M HK$129起"),
            ],
        }
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, contract, evidence in data_by_brand.get(source["brand"], []):
            if not re.search(re.escape(evidence), text):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"{source['brand']} {speed}M broadband {segment} Shangtaika 2026 reference HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "public_third_party_broadband_comparison_reference",
                    "source_status": "public_third_party_comparison_needs_review",
                    "evidence_excerpt": (
                        evidence
                        + "；Shangtaika 公开比较页标注资料更新于 2026年6月；仅作第三方参考，正式月费、覆盖、合约期和安装费以运营商申请结果为准。"
                    )[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "icable_shangtaika_brand_page_2026":
        if not (
            re.search(r"上台卡", text)
            and re.search(r"有線寬頻|i-?CABLE", text, flags=re.I)
            and re.search(r"100M\s+HK\$68起", text)
        ):
            return []
        specs = [
            ("market_reference", "68", "100", "有線寬頻 100M HK$68起"),
            ("market_reference", "68", "200", "有線寬頻 200M HK$68起"),
            ("market_reference", "89", "1000", "有線寬頻 1000M HK$89起"),
            ("market_reference", "129", "2000", "有線寬頻 2000M HK$129起"),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, evidence in specs:
            if not re.search(re.escape(evidence), text):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"i-CABLE {speed}M broadband {segment} Shangtaika brand page 2026 reference HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "public_third_party_broadband_comparison_reference",
                    "source_status": "public_third_party_comparison_needs_review",
                    "evidence_excerpt": (
                        evidence
                        + "；Shangtaika i-CABLE 品牌页公开摘录，仅作第三方渠道/市场参考，正式月费、覆盖、合约期和安装费以运营商申请结果为准。"
                    )[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "smartone_quoquo_broadband_comparison_2026":
        if not (
            re.search(r"寬頻月費比較\(光纖/5G\)\s*3/2026", text)
            and re.search(r"(更新:|update)\s*09/3/2026", text, flags=re.I)
            and re.search(r"數碼通", text)
            and re.search(r"2000M\s*-\s*2200M", text)
        ):
            return []
        data = [
            (
                "market_reference",
                "128",
                "2200",
                "36",
                r"\$128\s*/\s*2000M\s*-\s*2200M\s*\*\s*36",
                "视乎覆盖",
            ),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, contract, evidence_pattern, add_on in data:
            match = re.search(evidence_pattern, text, flags=re.I | re.S)
            if not match:
                continue
            excerpt = _excerpt(text, match.start(), match.end(), radius=220)
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"SmarTone {speed}M broadband {segment} QuoQuo 2026 reference HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": add_on,
                    "tariff_type": "public_third_party_broadband_comparison_reference",
                    "source_status": "public_third_party_comparison_needs_review",
                    "evidence_excerpt": (
                        excerpt
                        + "；QuoQuo 报价鸭公开比较页标注日期 2026-03-09、更新 09/3/2026；仅作第三方市场参考，正式月费、覆盖和安装费以运营商申请结果为准。"
                    )[:900],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hgc_hk01_public_housing_109_broadband_comparison_2026":
        if not (
            re.search(r"香港01\s+家居寬頻上網格價2026|家居寬頻上網格價2026", text)
            and re.search(r"HGC|環球全域電訊|環電", text, flags=re.I)
            and re.search(r"HK\$109", text)
            and re.search(r"24/36個月合約|24\s*個月", text)
        ):
            return []
        match = re.search(r"HGC.*?HK\$109.*?(?:24/36個月合約|24\s*個月).*?(?:公屋|安裝費|正式資費|official)", text, flags=re.I | re.S)
        excerpt = _clean_text(match.group(0) if match else text)[:700]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": "public_housing_reference",
                "plan_name": "HGC 1000M broadband public_housing_reference HK01 2026 reference HK$109",
                "monthly_fee_hkd": "109",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "",
                "add_on_charges_hkd": "Public-housing reference; private-building fee/install terms may differ",
                "tariff_type": "public_media_broadband_comparison_reference",
                "source_status": "public_media_comparison_needs_review",
                "evidence_excerpt": (
                    excerpt
                    + "；香港01页面注明价格为参考/网上报价，正式月费、覆盖和安装费以运营商公布和申请结果为准；本行只用于和同口径公开市场参考互证。"
                )[:900],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id", "").endswith("_hk01_broadband_comparison_2026"):
        if not (
            re.search(r"2026\s+六大電訊商\s+1000M\s+家居寬頻月費比較一覽表", text)
            and re.search(r"資料只供參考、以官方最新公佈為準", text)
        ):
            return []
        section_match = re.search(r"2026\s+六大電訊商\s+1000M\s+家居寬頻月費比較一覽表.*?日本福岡旅遊SIM卡", text, flags=re.S)
        section = section_match.group(0) if section_match else text
        data_by_brand = {
            "HKBN": ("public_housing_reference", "88", "24", "香港寬頻 HKBN HK$88起 24個月 免安裝費"),
            "HGC": ("market_reference", "109", "24/36", "HGC寬頻 HK$109起 24/36個月 公屋安裝費 私樓 $180"),
            "SmarTone": ("market_reference", "88", "36", "數碼通 SmarTone HK$88起 36個月 免安裝費"),
            "i-CABLE": ("market_reference", "88", "36", "有線寬頻 i-CABLE HK$88起 36個月 HK$300"),
        }
        if source["brand"] not in data_by_brand:
            return []
        segment, fee, contract_months, evidence = data_by_brand[source["brand"]]
        brand_pattern = {
            "HKBN": r"香港寬頻\s+HKBN\s+HK\$88起\s+24個月\s+免安裝費",
            "HGC": r"HGC寬頻\s+HK\$109起\s+24/36個月\s+公屋安裝費\s+私樓\s+\$180",
            "SmarTone": r"數碼通\s+SmarTone\s+HK\$88起\s+36個月\s+免安裝費",
            "i-CABLE": r"有線寬頻\s+i-CABLE\s+HK\$88起\s+36個月\s+HK\$300",
        }[source["brand"]]
        if not re.search(brand_pattern, section, flags=re.I):
            return []
        excerpt = _excerpt(section, max(section.find(evidence.split()[0]), 0), min(len(section), max(section.find(evidence.split()[0]), 0) + len(evidence)), radius=180)
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": segment,
                "plan_name": f"{source['brand']} 1000M broadband {segment} HK01 2026 reference HK${fee}",
                "monthly_fee_hkd": fee,
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": contract_months,
                "local_voice": "",
                "add_on_charges_hkd": "HK01 comparison table says reference only; official latest announcement prevails",
                "tariff_type": "public_media_broadband_comparison_reference",
                "source_status": "public_media_comparison_needs_review",
                "evidence_excerpt": (evidence + "；香港01页面注明资料只供参考，以官方最新公布为准。 " + excerpt[:450])[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id", "").endswith("_booga_broadband_comparison_2025"):
        if not (re.search(r"更新於:\s*2025年9月14日", text) and re.search(r"寬頻優惠", text) and re.search(r"1000M\s+寬頻", text)):
            return []
        sections = {
            "HKBN": re.search(r"HKBN\s+香港寬頻(?:優惠|.*?1000M\s+寬頻優惠).*?HGC/和記/環電", text, flags=re.I | re.S),
            "HGC": re.search(r"HGC/和記/環電(?:寬頻優惠\s+1000M\s+寬頻|\s+1000M\s+寬頻優惠).*?CMHK中國移動優惠", text, flags=re.I | re.S),
            "i-CABLE": re.search(r"有線寬頻\s+i-cable\s+優惠.*?長者寬頻優惠", text, flags=re.I | re.S),
        }
        data_by_brand = {
            "HKBN": [
                ("public_housing_reference", "98", "公屋/居屋 1000M 36 個月 HK$98"),
                ("private_building_reference", "109", "私樓 1000M 36 個月 HK$109"),
            ],
            "HGC": [
                ("public_housing_reference", "89", "公屋/居屋 1000M 36 個月 HK$89"),
                ("private_building_reference", "109", "私樓 1000M 36 個月 HK$109"),
            ],
            "i-CABLE": [
                ("public_housing_reference", "88", "公屋/居屋 1000M 36 個月 HK$88"),
                ("private_building_reference", "88", "私樓 1000M 36 個月 HK$88"),
            ],
        }
        section_match = sections.get(source["brand"])
        if not section_match:
            return []
        section = section_match.group(0)
        rows: List[Dict[str, str]] = []
        for segment, fee, evidence in data_by_brand.get(source["brand"], []):
            if not re.search(re.escape(evidence), section):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"{source['brand']} 1000M broadband {segment} Booga 2025 reference HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "1000",
                    "post_fup_speed_mbps": "",
                    "contract_months": "36",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "public_third_party_broadband_comparison_reference",
                    "source_status": "public_third_party_comparison_needs_review",
                    "evidence_excerpt": (evidence + "；Booga 页面标注 2025年9月14日更新，公开比较价仅作第三方参考，具体月费以申请和供应商条款为准。 " + section[:450])[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "icable_booga_2026_current_broadband_offer":
        if not (re.search(r"更新於:\s*2026年4月13日", text) and re.search(r"\$68\s+最平寬頻", text) and re.search(r"1000M/2000M", text)):
            return []
        specs = [
            ("entry_200m_public_private_reference", "68", "200", "36", "公居屋/私樓 200M 計劃 月費： $68 合約期： 36 個月 上網速度： 200M 適用： 公居屋 / 私樓"),
            ("value_1000m_public_private_reference", "88", "1000", "36", "計劃 (A) - $88 / 36個月 (公居屋/私樓) 適用： 公居屋 / 私樓"),
            ("router_1000m_private_reference", "88", "1000", "48", "計劃 (B) - $88 / 48個月 (私樓限定 - 送 Router) 適用： 私樓"),
            ("hidden_1000m_public_reference", "68", "1000", "48", "計劃 (C) - $68 / 48個月 (指定公居屋) 適用： 指定公居屋"),
            ("speed_2000m_public_private_reference", "118", "2000", "36", "公居屋/私樓 2000M 寬頻計劃 月費： $118 合約期： 36 個月 上網速度： 2000M 適用： 公居屋 / 私樓"),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, contract, evidence in specs:
            if not re.search(re.escape(evidence), text):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"i-CABLE Booga 2026 {speed}M broadband {segment} HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "Booga article says installation fee waived; one free relocation after 3 contract months for listed plans",
                    "tariff_type": "public_third_party_broadband_offer_reference",
                    "source_status": "public_third_party_offer_listing_needs_review",
                    "evidence_excerpt": (evidence + "；Booga 页面标注 2026年4月13日更新，公开第三方整理价仅作参考，正式资费和覆盖以 i-CABLE/申请条款为准。")[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "icable_service_wifi_2026_official":
        if not (
            re.search(r"i-?CABLE|有線寬頻", text, flags=re.I)
            and re.search(r"1000M", text, flags=re.I)
            and re.search(r"\$\s*93", text)
            and re.search(r"Wi-?Fi\s*6|Wi-Fi 6|Wi-Fi6", text, flags=re.I)
        ):
            return []
        match = re.search(r"1000M.*?\$\s*93|\$\s*93.*?1000M", text, flags=re.I | re.S)
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": "router_bundle_reference",
                "plan_name": "i-CABLE official Wi-Fi 1000M broadband with Wi-Fi 6 router HK$93",
                "monthly_fee_hkd": "93",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "36",
                "local_voice": "",
                "add_on_charges_hkd": "Wi-Fi 6 router bundle; JavaScript-rendered official service page preserved via static public excerpt",
                "tariff_type": "official_public_service_page_monthly_offer",
                "source_status": "official_public_page_js_text_snapshot",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:700] if match else text[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"icable_telcoquo_1000m_89_2026", "icable_telcoquo_1000m_118_2026"}:
        fee = "118" if source.get("source_id") == "icable_telcoquo_1000m_118_2026" else "89"
        segment = "private_building_reference" if fee == "118" else "market_reference"
        if not (
            re.search(r"Telcoquo", text, flags=re.I)
            and re.search(r"i-?Cable|有線寬頻", text, flags=re.I)
            and re.search(r"1000M", text, flags=re.I)
            and re.search(r"\$\s*" + re.escape(fee), text)
            and re.search(r"住宅寬頻|家居上網", text)
        ):
            return []
        match = re.search(r"1000M.*?\$\s*" + re.escape(fee) + r"|\$\s*" + re.escape(fee) + r".*?1000M", text, flags=re.I | re.S)
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": segment,
                "plan_name": f"i-CABLE Telcoquo 2026 1000M broadband {segment} HK${fee}",
                "monthly_fee_hkd": fee,
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "Telcoquo search/index excerpt; use as third-party indexed offer reference only",
                "tariff_type": "public_third_party_broadband_offer_reference",
                "source_status": "public_third_party_indexed_offer_needs_review",
                "evidence_excerpt": (
                    (_excerpt(text, match.start(), match.end(), radius=220) if match else text[:700])
                    + "；Telcoquo 公开索引摘录，非运营商官方价；用于和其它第三方市场价互证，实际资费以 i-CABLE 合约和覆盖地址为准。"
                )[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "icable_broadband_quote_2026_offer_listing":
        if not (
            re.search(r"有線寬頻|i-?CABLE", text, flags=re.I)
            and re.search(r"2000M", text)
            and re.search(r"\$129", text)
        ):
            return []
        specs = [
            ("market_reference", "68", "200", "36", "200M $68 36個月"),
            ("public_housing_reference", "68", "1000", "36/48", "1000M $68 36/48個月"),
            ("market_reference", "93", "1000", "36", "1000M $93 36個月"),
            ("market_reference", "118", "2000", "36", "2000M $118 36個月"),
            ("router_bundle_reference", "129", "2000", "36", "2000M $129 36個月"),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, contract, evidence in specs:
            if not (re.search(re.escape(speed + "M"), text) and re.search(r"\$\s*" + re.escape(fee), text)):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"i-CABLE broadband-quote 2026 {speed}M broadband {segment} HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "public_third_party_broadband_offer_reference",
                    "source_status": "public_third_party_offer_listing_needs_review",
                    "evidence_excerpt": (
                        evidence
                        + "；broadband-quote 公开报价列表，非运营商官方价；只作为第三方渠道/市场资费参考，实际月费以 i-CABLE 合约和覆盖地址为准。"
                    )[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "icable_broadband_pricequote_1000m_68_2023":
        if not (
            re.search(r"2023年12月1日", text)
            and re.search(r"有線寬頻1000M|有線寬頻.*?1000M|1000MB", text, flags=re.I | re.S)
            and re.search(r"公居屋.*?\$68|公居屋月費.*?68", text, flags=re.S)
            and re.search(r"36個月合約|簽36個月合約", text)
        ):
            return []
        match = re.search(r"公居屋.*?\$68.*?36個月合約|只需每月\$68.*?簽36個月合約", text, flags=re.S)
        evidence = _excerpt(text, match.start(), match.end(), radius=240) if match else text[:700]
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": "public_housing_reference",
                "plan_name": "i-CABLE 1000M broadband public_housing_reference PriceQuote 2023 reference HK$68",
                "monthly_fee_hkd": "68",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "36",
                "local_voice": "",
                "add_on_charges_hkd": "PriceQuote says installation fee waived; actual charge depends on address and supplier contract",
                "tariff_type": "public_third_party_broadband_offer_reference",
                "source_status": "public_third_party_channel_reference_needs_review",
                "evidence_excerpt": (
                    evidence
                    + "；Broadband-PriceQuote 公开渠道/比较文章，非 i-CABLE 官方标准价；正式月费、覆盖、安装费和优惠期限以供应商合约为准。"
                )[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "icable_booga_public_housing_broadband_comparison_2025":
        evidence = "有線寬頻 iCable 200M $68 36 個月 豁免 另有 2000M 月費 $118 合約期內滿 3 個月免一次搬遷費"
        if not (re.search(r"更新於:\s*2025年10月11日", text) and re.search(re.escape(evidence), text)):
            return []
        specs = [
            ("public_housing_200m_reference", "68", "200", "36", "有線寬頻 iCable 200M $68 36 個月"),
            ("public_housing_2000m_reference", "118", "2000", "36", "另有 2000M 月費 $118"),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, contract, row_evidence in specs:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"i-CABLE Booga public housing comparison {speed}M broadband HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "Installation fee waived; one free relocation after 3 contract months according to Booga comparison row",
                    "tariff_type": "public_third_party_broadband_comparison_reference",
                    "source_status": "public_third_party_comparison_needs_review",
                    "evidence_excerpt": (row_evidence + "；Booga 公屋/居屋宽频比较页标注 2025年10月11日更新，公开第三方整理价仅作参考，正式资费和覆盖以 i-CABLE/申请条款为准。 " + evidence)[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {
        "hgc_hktechreview_broadband_comparison_2025",
        "hkbn_hktechreview_broadband_comparison_2026",
        "smartone_hktechreview_broadband_comparison_2026",
        "icable_hktechreview_broadband_comparison_2026",
    }:
        if not (re.search(r"資料截至2025年12月11日", text) and re.search(r"1000M家居寬頻比較", text)):
            return []
        section_match = re.search(r"1000M家居寬頻比較.*?資料截至2025年12月11日", text, flags=re.S)
        section = section_match.group(0) if section_match else text
        specs_by_brand = {
            "HKBN": [
                ("public_housing_reference", "88", "24", r"香港寬頻\s*(?:前往官網\s*)?HK\$88（24個月合約）", "香港寬頻 HK$88（24個月合約）"),
            ],
            "SmarTone": [
                ("public_housing_reference", "88", "36", r"SmarTone\s*HK\$88（36個月合約）", "SmarTone HK$88（36個月合約）"),
            ],
            "i-CABLE": [
                ("market_reference", "88", "36", r"有線寬頻\s*HK\$88（36個月合約）", "有線寬頻 HK$88（36個月合約）"),
            ],
            "HGC": [
                ("public_housing_reference", "109", "24", r"HGC環電\s*公屋[:：]HK\$109（24個月合約）", "HGC環電 公屋 HK$109（24個月合約）"),
                ("private_building_reference", "119", "36", r"私樓[:：]HK\$119（36個月合約）", "HGC環電 私樓 HK$119（36個月合約）"),
            ],
        }
        rows: List[Dict[str, str]] = []
        for segment, fee, contract, pattern, evidence in specs_by_brand.get(source["brand"], []):
            match = re.search(pattern, section)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"{source['brand']} 1000M broadband {segment} HKTechReview reference HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "1000",
                    "post_fup_speed_mbps": "",
                    "contract_months": contract,
                    "local_voice": "",
                    "add_on_charges_hkd": "HK$180 installation fee disclosed in comparison table",
                    "tariff_type": "public_third_party_broadband_comparison_reference",
                    "source_status": "public_third_party_comparison_needs_review",
                    "evidence_excerpt": (
                        f"{evidence}；"
                        "HKTechReview 页面注明资料截至2025年12月11日，价格和套餐资讯以电讯商网站公布为准。 "
                        + _excerpt(section, match.start(), match.end(), radius=360)
                    )[:900],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hgc_broadband_pro_2017_referral_case":
        # This referral page confirms HGC availability in the covered estate, but
        # the visible monthly-fee rows around the match are for Netvigator/PCCW.
        # Keep it as a source-gap instead of creating an unsupported HGC tariff.
        return []
    if source.get("source_id") in {
        "icable_broadband_pro_2018_referral_case",
        "icable_broadband_pro_2017_200m_144_referral_case",
        "icable_broadband_pro_2016_referral_case",
        "hkbn_broadband_pro_2017_referral_case",
    }:
        normalized = text.translate(str.maketrans("﹩＄０１２３４５６７８９", "$$0123456789"))
        referral_cases = {
            "icable_broadband_pro_2018_referral_case": {
                "pattern": (
                    r"轉介優惠:\s*有線寬頻\s*住宅寬頻\s*\$88\s*200M上網.*?"
                    r"服務價錢:\s*\$88月費.*?"
                    r"服務計劃內容:\s*200M.*?"
                    r"服務供應商:\s*有線寬頻.*?"
                    r"合約期:\s*30\s*個月.*?"
                    r"合約日期:\s*14/09/2018"
                ),
                "plan_name": "i-CABLE residential broadband 200M referral offer HK$88",
                "monthly_fee_hkd": "88",
                "broadband_speed_mbps": "200",
                "contract_months": "30",
            },
            "icable_broadband_pro_2017_200m_144_referral_case": {
                "pattern": (
                    r"轉介優惠:\s*有線寬頻\s*200M\s*上網\s*\$116\(平均價）.*?"
                    r"服務供應商:\s*有線寬頻.*?"
                    r"合約期:\s*36\s*個月.*?"
                    r"服務計劃內容:\s*200M.*?"
                    r"服務價錢:\s*\$144月費"
                ),
                "plan_name": "i-CABLE residential broadband 200M referral offer HK$144",
                "monthly_fee_hkd": "144",
                "broadband_speed_mbps": "200",
                "contract_months": "36",
            },
            "icable_broadband_pro_2016_referral_case": {
                "pattern": (
                    r"轉介優惠:\s*有線寬頻\s*送\$300現金卷.*?"
                    r"服務價錢:\s*平均月費\$144.*?"
                    r"服務計劃內容:\s*200M.*?"
                    r"服務供應商:\s*有線寬頻.*?"
                    r"合約期:\s*30個月.*?"
                    r"合約日期:\s*10/01/2016"
                ),
                "plan_name": "i-CABLE residential broadband 200M referral average HK$144",
                "monthly_fee_hkd": "",
                "average_monthly_fee_hkd": "144",
                "broadband_speed_mbps": "200",
                "contract_months": "30",
            },
            "hkbn_broadband_pro_2017_referral_case": {
                "pattern": (
                    r"轉介優惠:\s*香港寬頻\s*HKBN\s*送\s*TVB BOX.*?"
                    r"服務價錢:\s*\$248月費.*?"
                    r"服務計劃內容:\s*1000M.*?"
                    r"服務供應商:\s*香港寬頻.*?"
                    r"合約期:\s*24\s*個月.*?"
                    r"合約日期:\s*21/01/2017"
                ),
                "plan_name": "HKBN residential broadband 1000M referral offer HK$248",
                "monthly_fee_hkd": "248",
                "broadband_speed_mbps": "1000",
                "contract_months": "24",
            },
        }
        case = referral_cases[source["source_id"]]
        match = re.search(case["pattern"], normalized, flags=re.S)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": case["plan_name"],
                "monthly_fee_hkd": case.get("monthly_fee_hkd", ""),
                "average_monthly_fee_hkd": case.get("average_monthly_fee_hkd", ""),
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": case["broadband_speed_mbps"],
                "post_fup_speed_mbps": "",
                "contract_months": case["contract_months"],
                "local_voice": "",
                "add_on_charges_hkd": "",
                "tariff_type": case.get("tariff_type", "monthly_plan_fee"),
                "source_status": case.get("source_status", "public_referral_case_needs_review"),
                "evidence_excerpt": _excerpt(normalized, match.start(), match.end(), radius=80)[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hgc_findplanking_2026_offer_listing":
        if not (
            re.search(r"Find Plan King", text, flags=re.I)
            and re.search(r"HGC", text, flags=re.I)
            and re.search(r"更新日期：\s*\d{2}\.\d{2}\.2026", text)
        ):
            return []
        specs = [
            ("public_housing_reference", "89", "1000", r"公屋/居屋:\s*1000M.*?\$89"),
            ("public_housing_phone_bundle", "109", "1000", r"1000M.*?家居電話.*?\$109"),
            ("public_housing_reference", "119", "2000", r"2000M.*?\$119|\$119\s*2000M"),
            ("market_reference", "139", "2500", r"2500M.*?\$139"),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, evidence_pattern in specs:
            match = re.search(evidence_pattern, text, flags=re.I | re.S)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"HGC Find Plan King 2026 {speed}M broadband {segment} HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "36" if fee in {"89", "119", "129", "139"} else "39",
                    "local_voice": "home phone bundle included" if "電話" in match.group(0) else "",
                    "add_on_charges_hkd": "Third-party/channel listing only; use as non-official public offer evidence.",
                    "tariff_type": "public_third_party_broadband_offer_reference",
                    "source_status": "public_third_party_offer_listing_needs_review",
                    "evidence_excerpt": (
                        _excerpt(text, match.start(), match.end(), radius=80)
                        + "；Find Plan King 页面标注为公开渠道报价/销售资料，非 HGC 官方标准价。"
                    )[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hgc_broadband_quote_2026_offer_listing":
        if not (
            re.search(r"HGC", text, flags=re.I)
            and re.search(r"特選地區客戶優惠計劃", text)
            and re.search(r"資料只供參考", text)
        ):
            return []
        specs = [
            ("market_reference", "119", "2000", r"2000m?低至\$119起"),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, evidence_pattern in specs:
            match = re.search(evidence_pattern, text, flags=re.I | re.S)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"HGC Broadband Quote 2026 {speed}M broadband {segment} HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "Third-party/channel listing only; actual fee depends on HGC contract and covered address.",
                    "tariff_type": "public_third_party_broadband_offer_reference",
                    "source_status": "public_third_party_offer_listing_needs_review",
                    "evidence_excerpt": (
                        _excerpt(text, match.start(), match.end(), radius=120)
                        + "；Broadband Quote 页面注明资料只供参考，实际收费及优惠以供应商合约内容为准。"
                    )[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "icable_findplanking_2022_offer_listing":
        if not (
            re.search(r"Find Plan King", text, flags=re.I)
            and re.search(r"有線寬頻", text)
            and re.search(r"更新日期：\s*17\.03\.2022", text)
        ):
            return []
        specs = [
            ("public_housing_reference", "58", "200", "公屋/居屋： $58 200mb上網（特選屋苑）"),
            ("public_housing_phone_bundle", "88", "1000", "$88 1000mb 加電話"),
        ]
        rows: List[Dict[str, str]] = []
        for segment, fee, speed, evidence in specs:
            if not re.search(re.escape(evidence), text, flags=re.I):
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "customer_segment": segment,
                    "plan_name": f"i-CABLE Find Plan King 2022 {speed}M broadband {segment} HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "home phone bundle included" if "電話" in evidence else "",
                    "add_on_charges_hkd": "Page says installation and maintenance fee waived; one free relocation; third-party/channel listing only",
                    "tariff_type": "public_third_party_broadband_offer_reference",
                    "source_status": "public_third_party_offer_listing_needs_review",
                    "evidence_excerpt": (
                        evidence
                        + "；Find Plan King 页面更新日期 17.03.2022，标注为公开渠道报价/销售资料，非 i-CABLE 官方标准价。"
                    )[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "icable_findplanking_2026_public_housing_75":
        match = re.search(r"公屋/居屋客戶.*?[＄$]\s*75\s*/\s*1000M\s*/\s*連\s*路由器", text, flags=re.I | re.S)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": "public_housing_reference",
                "plan_name": "i-CABLE Broadband 公屋居屋 1000M HK$75",
                "monthly_fee_hkd": "75",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "router included; page also states installation fee and one relocation waived; third-party/channel listing only",
                "tariff_type": "public_third_party_broadband_offer_reference",
                "source_status": "public_third_party_offer_listing_needs_review",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:800],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "icable_home_broadband_service_tc":
        match = re.search(r"新登記\s*i-?CABLE\s*2000M\s*光纖入屋.*?[＄$]\s*118\s*/\s*月", text, flags=re.I | re.S)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "i-CABLE Home Broadband 2000M HK$118",
                "monthly_fee_hkd": "118",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "2000",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "Official service page excerpt does not disclose contract term in the captured text; not estimated.",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_current_official_page",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=220)[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"icable_broadband_offer", "icable_broadband_plan_detail"}:
        if not (re.search(r"有線寬頻|i-?cable", text, flags=re.I) and re.search(r"1000M|200M", text, flags=re.I)):
            if not re.search(r"fibre broadband service plan monthly fee", text, flags=re.I):
                return []
            rows = []
            for match in re.finditer(r"monthly fee\s+HK\$([0-9][0-9,]*)\s+([0-9]+)M", text, flags=re.I):
                fee = _safe_num(match.group(1))
                speed = _safe_num(match.group(2))
                row = _base_plan_row(source, result, captured_at, period_label, archive_url)
                row.update(
                    {
                        "plan_name": f"i-CABLE Broadband {speed}M HK${fee}",
                        "monthly_fee_hkd": fee,
                        "average_monthly_fee_hkd": "",
                        "local_data_gb": "",
                        "roaming_data_gb": "",
                        "broadband_speed_mbps": speed,
                        "post_fup_speed_mbps": "",
                        "contract_months": "",
                        "local_voice": "",
                        "add_on_charges_hkd": "",
                        "tariff_type": "monthly_plan_fee",
                        "source_status": "parsed_archive" if archive_url else "parsed_current",
                        "evidence_excerpt": _excerpt(text, match.start(), match.end())[:600],
                    }
                )
                row["record_key"] = _record_key(row)
                rows.append(row)
            return rows
        offer_patterns = [
            ("public_hos", "公屋居屋", "200M", "200", "58"),
            ("public_hos", "公屋居屋", "1000M", "1000", "68"),
            ("public_hos_router", "公屋居屋", "1000M", "1000", "75"),
            ("public_hos_game", "公屋居屋", "1000M X2", "2000", "98"),
            ("private_estate", "私人屋苑", "200M", "200", "78"),
            ("private_estate", "私人屋苑", "1000M", "1000", "98"),
            ("private_estate_24m", "私人屋苑", "1000M", "1000", "88"),
            ("private_estate_no_router", "私人屋苑", "1000M", "1000", "128"),
        ]
        rows: List[Dict[str, str]] = []
        excerpt_match = re.search(r"服務計劃詳情.*?(?:私人屋苑.*?\$128|1000M\s+\$128)", text, flags=re.I)
        excerpt = _excerpt(text, excerpt_match.start(), excerpt_match.end(), radius=80)[:600] if excerpt_match else text[:600]
        for qualifier, segment, speed_label, speed, fee in offer_patterns:
            if not re.search(rf"{re.escape(speed_label)}\s*\${fee}\b", text, flags=re.I):
                continue
            contract_months = ""
            if speed == "1000" and fee == "88" and re.search(r"1000M\s*\$88\s*36\s*個月", text, flags=re.I):
                contract_months = "36"
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"i-CABLE Broadband {segment} {speed_label} HK${fee}",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": contract_months,
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "public_channel_offer_needs_review",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return _dedupe(rows)
    if source.get("source_id") == "icable_kennechu_2020_home_broadband_guide":
        match = re.search(r"有線寬頻\s+i-?Cable\s+網速：1000M\s+月費：HK\$98\s+簽約：36\s*個月", text, flags=re.I)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "i-CABLE Broadband 1000M HK$98 36 months public blog market guide",
                "monthly_fee_hkd": "98",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "36",
                "local_voice": "",
                "add_on_charges_hkd": "",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_blog_market_guide_needs_review",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=160)[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hgc_kennechu_2020_home_broadband_guide":
        match = re.search(r"HGC\s+和記\s+網速：1000M\s+月費：HK\$148\s+簽約：36\s*個月", text, flags=re.I)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HGC Broadband 1000M HK$148 36 months public blog market guide",
                "monthly_fee_hkd": "148",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "36",
                "local_voice": "",
                "add_on_charges_hkd": "",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_blog_market_guide_needs_review",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=160)[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "icable_hkepc_2019_forum_market_observation":
        match = re.search(r"2019-2-13[\s\S]{0,500}?有線做promotion\s+sell\s+HK\$98\s+1000M光纖入屋", text, flags=re.I)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "i-CABLE Broadband 1000M HK$98 public forum market observation",
                "monthly_fee_hkd": "98",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_forum_market_observation_needs_review",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180)[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hkbn_hkepc_2016_1000m_248_renewal_quote":
        normalized = text.translate(str.maketrans("﹩＄０１２３４５６７８９", "$$0123456789"))
        if not re.search(r"HKBN|香港寬頻", normalized, flags=re.I):
            return []
        match = re.search(
            r"續約[\s\S]{0,260}?1000M\s*\$248\s*[Xx]\s*24",
            normalized,
            flags=re.I,
        )
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "customer_segment": "renewal_quote_reference",
                "plan_name": "HKBN 1000M HK$248 24 months public forum renewal quote",
                "monthly_fee_hkd": "248",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "",
                "add_on_charges_hkd": "Forum quote also mentions $1000 coupon and 12-month SIM; non-official reference only",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_forum_renewal_quote_needs_review",
                "evidence_excerpt": _excerpt(normalized, match.start(), match.end(), radius=260)[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "icable_discuss_2017_forum_market_observation":
        match = re.search(r"2017-3-6[\s\S]{0,240}?有線寬頻\s*200M\s*寬頻\s*\$88", text, flags=re.I)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "i-CABLE Broadband 200M HK$88 public forum market observation",
                "monthly_fee_hkd": "88",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "200",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "",
                "tariff_type": "monthly_plan_fee",
                "source_status": "public_forum_market_observation_needs_review",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=160)[:600],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "icable_appledaily_2017_broadband_market_comparison":
        match = re.search(r"2017-07-07[\s\S]{0,240}?有線家居寬頻[：:]\s*1000M[，,]\s*月費約140\.7元", text, flags=re.I)
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "i-CABLE Home Broadband 1000M average HK$140.7 public media comparison",
                "monthly_fee_hkd": "",
                "average_monthly_fee_hkd": "140.7",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "1000",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "",
                "tariff_type": "average_monthly_plan_fee",
                "source_status": "public_media_archive_needs_review",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=160)[:600],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id", "").startswith("hkbn_100gb_5g_2022_official_pdf"):
        if not (re.search(r"100\s*GB", text, flags=re.I) and re.search(r"HK\$\s*149", text, flags=re.I)):
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        match = re.search(r"HK\$\s*149", text, flags=re.I)
        start, end = match.span() if match else (0, min(len(text), 120))
        row.update(
            {
                "plan_name": "HKBN 5G Mobile Service HK$149 100GB",
                "monthly_fee_hkd": "149",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "100",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "5",
                "contract_months": "24",
                "local_voice": "3000 minutes",
                "add_on_charges_hkd": "18",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_archive" if archive_url else "parsed_current",
                "evidence_excerpt": _excerpt(text, start, end)[:600],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"hkbn_enterprise_mobile_current", "hkbn_enterprise_mobile_current_tc", "hkbn_enterprise_mobile_current_sc"}:
        match = re.search(
            r"(?:Business 5G Mobile Services.*?HK\$\s*78/month|商業5G流動通訊服務.*?HK\$\s*78|商业5G流动通讯服务.*?HK\$\s*78)",
            text,
            flags=re.I | re.S,
        )
        if not match:
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HKBN Business 5G Mobile Services from HK$78",
                "monthly_fee_hkd": "78",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "",
                "add_on_charges_hkd": "Official SME page says starting as low as HK$78/month; exact plan entitlement not disclosed on this landing page.",
                "tariff_type": "business_mobile_5g_starting_monthly_fee",
                "source_status": "parsed_current",
                "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180)[:700],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"hkbn_iqiyi_vip_2023_pdf", "hkbn_iqiyi_vip_2023_official"}:
        if not (
            re.search(r"iQIYI Gold VIP member monthly fee is HK\$38|愛奇藝黃金VIP會員月費為港幣38元", text, flags=re.I)
            and re.search(r"iQIYI Diamond VIP member monthly fee is HK\$58|愛奇藝鑽石VIP會員月費為港幣58元", text, flags=re.I)
        ):
            return []
        rows: List[Dict[str, str]] = []
        fee_match = re.search(
            r"iQIYI Gold VIP member monthly fee is HK\$38.*?HK\$18 rebate per month|愛奇藝黃金VIP會員月費為港幣38元.*?每月港幣18元回贈",
            text,
            flags=re.I | re.S,
        )
        table_excerpt = _excerpt(text, fee_match.start(), fee_match.end(), radius=120)[:600] if fee_match else text[:600]
        for fee, plan_name, rebate in [
            ("38", "HKBN iQIYI Gold VIP member monthly fee HK$38", "10"),
            ("58", "HKBN iQIYI Diamond VIP member monthly fee HK$58", "18"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": rebate,
                    "tariff_type": "monthly_value_added_service_fee",
                    "source_status": "parsed_archive" if archive_url else "parsed_current",
                    "evidence_excerpt": table_excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") == "hkbn_oexbn_2026_official_api":
        try:
            plans = json.loads(text)
        except json.JSONDecodeError:
            return []
        if not isinstance(plans, list):
            return []
        rows: List[Dict[str, str]] = []
        for plan in plans:
            if not isinstance(plan, dict):
                continue
            price_info = plan.get("priceInfo") or {}
            plan_name = str(plan.get("planName") or "")
            detail_items = plan.get("planDetail") or []
            broadband_detail = next(
                (
                    item
                    for item in detail_items
                    if isinstance(item, dict)
                    and item.get("cate") == "Broadband"
                    and item.get("state") is True
                ),
                {},
            )
            if not plan_name or not broadband_detail or price_info.get("planPrice") in (None, ""):
                continue
            heading = str(broadband_detail.get("heading") or "")
            speed_match = re.search(r"([0-9.]+)\s*(Gbps|G|M)\b", f"{plan_name} {heading}", flags=re.I)
            if not speed_match:
                continue
            speed_value = speed_match.group(1)
            speed_unit = speed_match.group(2).lower()
            speed_mbps = str(int(float(speed_value) * 1000)) if speed_unit in {"gbps", "g"} else str(int(float(speed_value)))
            charge_items = price_info.get("charge") or []
            broadband_charge = next(
                (
                    item
                    for item in charge_items
                    if isinstance(item, dict)
                    and str(item.get("service") or "").lower() == "broadband internet"
                ),
                {},
            )
            evidence = {
                "planCode": plan.get("planCode"),
                "planName": plan_name,
                "planPrice": price_info.get("planPrice"),
                "averageFee": price_info.get("averageFee"),
                "duration": price_info.get("duration"),
                "specialMonthlyFee": broadband_charge.get("specialMonthlyFee"),
                "afterContractPeriod": broadband_charge.get("afterContractPeriod"),
            }
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HKBN {plan_name}",
                    "monthly_fee_hkd": str(price_info.get("planPrice")),
                    "average_monthly_fee_hkd": str(price_info.get("averageFee") or ""),
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed_mbps,
                    "post_fup_speed_mbps": "",
                    "contract_months": str(price_info.get("PAY_MONTH") or broadband_detail.get("contactMonth") or ""),
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_official_api",
                    "evidence_excerpt": json.dumps(evidence, ensure_ascii=False),
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return _dedupe(rows)
    if source.get("source_id") in {"hkbn_crossborder_5g_2023_prnewswire", "hkbn_crossborder_5g_2023_official"}:
        if not (
            re.search(r"Cross-border\s+5G|5G Local\s*\+\s*(?:Ba\s*y|Bay|GBA|mainland China)", text, flags=re.I)
            and re.search(r"30GB\s+5G\s+local\s+data.*?HK\$\s*149", text, flags=re.I | re.S)
            and re.search(r"10GB\s+5G\s+local\s+data.*?HK\$\s*103", text, flags=re.I | re.S)
        ):
            return []
        rows: List[Dict[str, str]] = []
        crossborder_specs = [
            {
                "plan_name": "HKBN 5G Home Broadband HK$103 10GB",
                "fee": "103",
                "local_data": "10",
                "pattern": r"10GB\s+5G\s+local\s+data.*?1GB\s+(?:mainland China and Macau roaming data|m\s*ainland China and Macau roaming data).*?HK\$\s*103",
            },
            {
                "plan_name": "HKBN 5G Home Broadband HK$149 30GB",
                "fee": "149",
                "local_data": "30",
                "pattern": r"30GB\s+5G\s+local\s+data.*?1GB\s+(?:mainland China and Macau roaming data|m\s*ainland China and Macau roaming data).*?HK\$\s*149",
            },
        ]
        for spec in crossborder_specs:
            match = re.search(spec["pattern"], text, flags=re.I | re.S)
            if not match:
                continue
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": spec["plan_name"],
                    "monthly_fee_hkd": spec["fee"],
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": spec["local_data"],
                    "roaming_data_gb": "1",
                    "broadband_speed_mbps": "",
                    "post_fup_speed_mbps": "1",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "28 admin fee; optional extra 2GB mainland China/Macau roaming data HK$38",
                    "tariff_type": "monthly_crossborder_5g_plan_fee",
                    "source_status": "parsed_current",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180)[:800],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return _dedupe(rows)
    if source.get("source_id") in {"hkbn_5g_home_broadband_terms_pdf", "hkbn_5g_home_broadband_terms_pdf_202405"}:
        if not (
            re.search(r"HK\$118\s+5G Home Broadband Service\s+P?\s*lan", text, flags=re.I)
            and re.search(r"special monthly fee.*?HK\$118", text, flags=re.I | re.S)
            and re.search(r"300GB local data entitlement", text, flags=re.I)
        ):
            return []
        match = re.search(
            r"special monthly fee.*?HK\$118.*?300GB local data entitlement.*?The network is supported by 3HK",
            text,
            flags=re.I | re.S,
        )
        start, end = match.span() if match else (0, min(len(text), 800))
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": "HKBN 5G Home Broadband Service Plan HK$118",
                "monthly_fee_hkd": "118",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "300",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "24",
                "local_voice": "no local voice, roaming voice, roaming data or IDD service",
                "add_on_charges_hkd": "monthly administration fee disclosed in terms; designated residential address and router/device required; network supported by 3HK",
                "tariff_type": "monthly_home_5g_broadband_fee",
                "source_status": "parsed_public_official_pdf",
                "evidence_excerpt": _excerpt(text, start, end, radius=120)[:800],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") == "hkbn_home_broadband_offer":
        if not re.search(r"Home Broadband Service Offer", text, flags=re.I):
            return []
        rows: List[Dict[str, str]] = []
        offer_pattern = re.compile(
            r"HKBN\s+(?P<name>[^:]+?):\s*plan price\s*\$(?P<fee>\d+),\s*average fee\s*\$(?P<avg>\d+(?:\.\d+)?),\s*contract\s*(?P<contract>\d+)\s*months",
            flags=re.I,
        )
        for match in offer_pattern.finditer(text):
            plan_name = _clean_text(match.group("name"))
            speed_match = re.search(r"([0-9.]+)\s*(Gbps|G|M)\b", plan_name, flags=re.I)
            if not speed_match:
                continue
            speed_value = speed_match.group(1)
            speed_unit = speed_match.group(2).lower()
            speed_mbps = str(int(float(speed_value) * 1000)) if speed_unit in {"gbps", "g"} else str(int(float(speed_value)))
            row = _base_plan_row(source, result, captured_at, "current", archive_url)
            row.update(
                {
                    "plan_name": f"HKBN {plan_name}",
                    "monthly_fee_hkd": match.group("fee"),
                    "average_monthly_fee_hkd": match.group("avg"),
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed_mbps,
                    "post_fup_speed_mbps": "",
                    "contract_months": match.group("contract"),
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": "monthly_plan_fee",
                    "source_status": "parsed_official_rendered_page_snapshot",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=160)[:700],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return _dedupe(rows)
    if source.get("source_id") in {
        "hkbn_fhkpuaa_2025_member_offer_pdf",
        "hkbn_fhkpuaa_2024_sep_member_offer_pdf",
        "hkbn_hkis_2025_member_offer_pdf",
    }:
        if not (
            re.search(r"(?:FHKPUAA|HKIS).{0,80}Special Offer", text, flags=re.I | re.S)
            and re.search(r"Promotion period valid until", text, flags=re.I)
        ):
            return []
        rows: List[Dict[str, str]] = []
        offer_specs = [
            {
                "plan_name": "HKBN 1000M Dual Home Broadband with Choice 1 of 4",
                "fee": "129",
                "speed": "1000",
                "contract": "24",
                "category": "home_fibre_broadband",
                "evidence_pattern": r"1000M Dual Home Broadband.*?\$129.*?Plan Details",
            },
            {
                "plan_name": "HKBN 1000M Home Broadband with 4.5G 42Mbps 10GB mobile service",
                "fee": "159",
                "speed": "1000",
                "contract": "24",
                "category": "home_fibre_broadband",
                "evidence_pattern": r"1000M Home Broadband With Mobile Service.*?\$159.*?Plan Details",
            },
            {
                "plan_name": "HKBN 2000M Home Broadband",
                "fee": "199",
                "speed": "2000",
                "contract": "",
                "category": "home_fibre_broadband",
                "evidence_pattern": r"2000M Fibre Broadband Plan.*?2000M Home Broadband \$199",
            },
            {
                "plan_name": "HKBN 2000M Home Broadband with TP-Link Archer BE230 Wi-Fi 7 Router",
                "fee": "229",
                "speed": "2000",
                "contract": "",
                "category": "home_fibre_broadband",
                "evidence_pattern": r"2000M Fibre Broadband Plan.*?With TP-Link Archer BE230 Wi-Fi 7 Router \$229",
            },
            {
                "plan_name": "HKBN 2.5Gbps GigaFast Broadband 2500M Home Broadband",
                "fee": "199",
                "speed": "2500",
                "contract": "",
                "category": "home_fibre_broadband",
                "evidence_pattern": r"(?:2\.5Gbps GigaFast|2500M Fibre) Broadband Plan.*?2500M Home Broadband \$228 \$199",
            },
            {
                "plan_name": "HKBN 2.5Gbps GigaFast Broadband with TP-Link Archer BE230 Wi-Fi 7 Router",
                "fee": "229",
                "speed": "2500",
                "contract": "",
                "category": "home_fibre_broadband",
                "evidence_pattern": r"(?:2\.5Gbps GigaFast|2500M Fibre) Broadband Plan.*?With TP-Link Archer BE(?:230|550) Wi-Fi 7 Router \$248 \$229",
            },
            {
                "plan_name": "HKBN Home Broadband with Bowtie 4-In-1 Healthcare Service Plan 1000M Home Broadband",
                "fee": "109",
                "speed": "1000",
                "contract": "24",
                "category": "home_fibre_broadband",
                "evidence_pattern": r"Home Broadband with Bowtie 4-In-1 Healthcare Service Plan.*?1000M Home Broadband# \$109 24 months",
            },
            {
                "plan_name": "HKBN 5G Local Mobile Communication Service Plan 20GB",
                "fee": "98",
                "speed": "",
                "contract": "24",
                "category": "mobile_consumer_5g",
                "local_data": "20",
                "evidence_pattern": r"Network Monthly Fee Local Data.*?5G\s+\$98 20GB",
            },
            {
                "plan_name": "HKBN 5G Local Mobile Communication Service Plan 30GB",
                "fee": "124",
                "speed": "",
                "contract": "24",
                "category": "mobile_consumer_5g",
                "local_data": "30",
                "evidence_pattern": r"Network Monthly Fee Local Data.*?\$124 30GB",
            },
            {
                "plan_name": "HKBN 5G Local Mobile Communication Service Plan 50GB",
                "fee": "149",
                "speed": "",
                "contract": "24",
                "category": "mobile_consumer_5g",
                "local_data": "50",
                "evidence_pattern": r"Network Monthly Fee Local Data.*?\$149 50GB",
            },
            {
                "plan_name": "HKBN 5G Local Mobile Communication Service Plan 30GB + Infinity Local Social & Streaming Data",
                "fee": "162",
                "speed": "",
                "contract": "24",
                "category": "mobile_consumer_5g",
                "local_data": "30",
                "evidence_pattern": r"Network Monthly Fee Local Data.*?\$162 30GB.*?Infinity Local Social",
            },
        ]
        for spec in offer_specs:
            match = re.search(spec["evidence_pattern"], text, flags=re.I | re.S)
            if not match:
                continue
            row_source = dict(source)
            row_source["product_category"] = spec["category"]
            row = _base_plan_row(row_source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": spec["plan_name"],
                    "monthly_fee_hkd": spec["fee"],
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": spec.get("local_data", ""),
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": spec["speed"],
                    "post_fup_speed_mbps": "1" if spec["category"].startswith("mobile") else "",
                    "contract_months": spec["contract"],
                    "local_voice": "3000 local minutes per month" if spec["category"].startswith("mobile") else "",
                    "add_on_charges_hkd": "200 prepayment; 28 admin fee waived for selected mobile plans",
                    "tariff_type": "member_offer_monthly_plan_fee",
                    "source_status": "public_member_offer_pdf_needs_review",
                    "evidence_excerpt": _excerpt(text, match.start(), match.end(), radius=180)[:800],
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return _dedupe(rows)
    if source.get("source_id") in {"hkbn_pay_tv_bundle_2019_mediaoutreach", "hkbn_pay_tv_bundle_2019_official_pdf"}:
        if not (
            re.search(r"HKBN Launches Mind\s*-?\s*blowing Offer to All Pay TV Customers", text, flags=re.I)
            and re.search(r"9 May 2019", text, flags=re.I)
            and re.search(r"myTV Gold plus 100M home broadband services priced as low as HK\$198", text, flags=re.I)
            and re.search(r"1000M home broadband service bundle offer for just HK\$238", text, flags=re.I)
        ):
            return []
        is_official_pdf = source.get("source_id") == "hkbn_pay_tv_bundle_2019_official_pdf"
        excerpt_match = re.search(
            r"myTV Gold plus 100M home broadband services priced as low as HK\$198.*?1000M home broadband service bundle offer for just HK\$238 up per month",
            text,
            flags=re.I | re.S,
        )
        excerpt = _excerpt(text, excerpt_match.start(), excerpt_match.end(), radius=180)[:800] if excerpt_match else text[:800]
        rows = []
        for speed, fee in [("100", "198"), ("1000", "238")]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": f"HKBN myTV Gold plus {speed}M home broadband bundle HK${fee} up",
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": speed,
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "myTV Gold bundle",
                    "tariff_type": "monthly_bundle_fee_from",
                    "source_status": "parsed_official_pdf" if is_official_pdf else "public_news_release_needs_review",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    if source.get("source_id") in {"hkbn_n_mobile_2023_official", "hkbn_n_mobile_2023_official_tc"}:
        if not (
            re.search(r"N mobile", text, flags=re.I)
            and re.search(r"Plans start at just \$108/mth|月費低至\s*\$108", text, flags=re.I)
        ):
            return []
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        match = re.search(r"Plans start at just \$108/mth.*?\$18/month admin fee|月費低至\s*\$108.*?每月\s*\$18行政費", text, flags=re.I | re.S)
        start, end = match.span() if match else (0, min(len(text), 120))
        row.update(
            {
                "plan_name": "HKBN N mobile travel lifestyle plan starting HK$108/mth",
                "monthly_fee_hkd": "108",
                "average_monthly_fee_hkd": "",
                "local_data_gb": "",
                "roaming_data_gb": "",
                "broadband_speed_mbps": "",
                "post_fup_speed_mbps": "",
                "contract_months": "",
                "local_voice": "Global Talk+ voice roaming app mentioned",
                "add_on_charges_hkd": "18",
                "tariff_type": "monthly_plan_fee",
                "source_status": "parsed_archive" if archive_url else "parsed_current",
                "evidence_excerpt": _excerpt(text, start, end)[:600],
            }
        )
        row["record_key"] = _record_key(row)
        return [row]
    if source.get("source_id") in {"hkbn_gigafast_tplink_2024_official", "hkbn_gigafast_tplink_2024_official_tc"}:
        if not (
            re.search(r"GigaFast 5Gbps / 10Gbps|5Gbps / 10Gbps GigaFast", text, flags=re.I)
            and re.search(r"starting at just \$698 per month|每月只需\s*\$698起", text, flags=re.I)
        ):
            return []
        rows: List[Dict[str, str]] = []
        match = re.search(
            r"GigaFast 5Gbps / 10Gbps service.*?additional \$200 up per month|5Gbps / 10Gbps GigaFast.*?額外加\s*\$200起",
            text,
            flags=re.I | re.S,
        )
        excerpt = _excerpt(text, match.start(), match.end(), radius=120)[:600] if match else text[:600]
        for fee, plan_name, tariff_type in [
            ("698", "HKBN GigaFast 5Gbps / 10Gbps service starting HK$698", "monthly_plan_fee"),
            ("200", "HKBN existing customer GigaFast 5Gbps / 10Gbps upgrade additional HK$200", "monthly_upgrade_fee"),
        ]:
            row = _base_plan_row(source, result, captured_at, period_label, archive_url)
            row.update(
                {
                    "plan_name": plan_name,
                    "monthly_fee_hkd": fee,
                    "average_monthly_fee_hkd": "",
                    "local_data_gb": "",
                    "roaming_data_gb": "",
                    "broadband_speed_mbps": "5000/10000",
                    "post_fup_speed_mbps": "",
                    "contract_months": "",
                    "local_voice": "",
                    "add_on_charges_hkd": "",
                    "tariff_type": tariff_type,
                    "source_status": "parsed_archive" if archive_url else "parsed_current",
                    "evidence_excerpt": excerpt,
                }
            )
            row["record_key"] = _record_key(row)
            rows.append(row)
        return rows
    rows: List[Dict[str, str]] = []
    for match in re.finditer(r"(?:HK\$|HKD\s*\$?|\$)\s*([0-9][0-9,]*(?:\.[0-9]+)?)", text, flags=re.I):
        fee = _safe_num(match.group(1))
        try:
            value = float(fee)
        except ValueError:
            continue
        if value <= 0 or value > 100000:
            continue
        start, end = match.span()
        excerpt = _excerpt(text, start, end)
        if source.get("brand") == "3HK / Hutchison" and fee in {"10", "18", "28"} and re.search(r"admin fee|admin fee waiver|original price", excerpt, flags=re.I):
            continue
        source_id = source.get("source_id", "")
        if source_id in {"smartone_5g_listing", "smartone_subscription_offers"}:
            amount_context = _clean_text(text[max(0, start - 80) : min(len(text), end + 160)])
            amount_prefix = _clean_text(text[max(0, start - 14) : start])
            if re.search(r"Average|Retail price|total value|平均|總值|零售價", amount_prefix, flags=re.I):
                continue
            immediate_plan_marker = re.search(rf"^(?:HK\$|\$)?\s*{re.escape(fee)}\s*(?:/ ?(?:Month|月)|Monthly Subscription Offer)", _clean_text(text[start : min(len(text), end + 90)]), flags=re.I)
            if not immediate_plan_marker:
                if re.search(r"voucher|rebate|AirPods|iPad|Apple Watch|Disney|retail price|exclusive price|handling fee|eSIM|SIM card|Points|cash discounts|Accessories", amount_context, flags=re.I):
                    continue
                continue
        if source_id in {"3hk_business_5g", "3hk_business_5g_tc"}:
            amount_context = _clean_text(text[max(0, start - 80) : min(len(text), end + 160)])
            if fee not in {"124", "188"}:
                continue
            immediate_plan_marker = re.search(rf"^\$\s*{re.escape(fee)}\s*/\s*(?:month|月)", _clean_text(text[start : min(len(text), end + 40)]), flags=re.I)
            if not immediate_plan_marker:
                if re.search(r"MoneyBack|Points|referral|Top-up|every \$5|Admin Fee|Airtime|SMS|borrow|易賞錢|積分|推薦|行政費|通話|短訊|增值", amount_context, flags=re.I):
                    continue
                continue
        if source_id in {"3hk_sosim_local", "3hk_sosim_local_tc"}:
            amount_context = _clean_text(text[max(0, start - 100) : min(len(text), end + 180)])
            if fee not in {"33", "48"}:
                continue
            if fee == "33" and not re.search(r"Charge\s*\(HK\$\)\s*\$33|Service activation|新卡\$33|服務啟用", amount_context, flags=re.I):
                continue
            if fee == "48" and not re.search(r"\$48\s*/\s*30\s*(?:Days|日)|50GB Infinite Data|50GB無限數據", amount_context, flags=re.I):
                continue
        local_data = _extract_number_after(r"([0-9]+(?:\.[0-9]+)?)\s*GB", text, end) or _extract_number_before(r"([0-9]+(?:\.[0-9]+)?)\s*GB", text, start)
        roaming_data = ""
        if re.search(r"roam|漫遊|漫游|Mainland|Macau|world|APAC|travel", excerpt, flags=re.I):
            roaming_data = local_data
        if "broadband" in source["product_category"]:
            speed = _extract_speed_before(text, start) or _extract_speed_after(text, end)
        else:
            speed = _extract_speed_after(text, end) or _extract_speed_before(text, start)
        if "gb" in excerpt.lower() and speed == local_data:
            speed = ""
        contract = _extract_number_before(r"([0-9]{2,3})\s*(?:months|month|mth|個月|个月)", text, start) or _extract_number_after(r"([0-9]{2,3})\s*(?:months|month|mth|個月|个月)", text, end)
        plan_name = f"{_plan_family(source, text)} HK${fee}"
        if local_data and "broadband" not in source["product_category"]:
            plan_name += f" {local_data}GB"
        row = _base_plan_row(source, result, captured_at, period_label, archive_url)
        row.update(
            {
                "plan_name": plan_name,
                "monthly_fee_hkd": fee,
                "average_monthly_fee_hkd": fee if re.search(r"average|avg|平均", excerpt, flags=re.I) else "",
                "local_data_gb": "" if "broadband" in source["product_category"] else local_data,
                "roaming_data_gb": roaming_data,
                "broadband_speed_mbps": speed if "broadband" in source["product_category"] else "",
                "post_fup_speed_mbps": _extract_number_after(r"(?:up to|max|最高)\s*([0-9]+(?:\.[0-9]+)?)\s*Mbps", text, end),
                "contract_months": contract,
                "local_voice": "mentioned" if re.search(r"voice|call mins|通話|通话", excerpt, flags=re.I) else "",
                "add_on_charges_hkd": "",
                "tariff_type": _tariff_type(excerpt),
                "source_status": "parsed_archive" if archive_url else "parsed_current",
                "evidence_excerpt": excerpt[:600],
            }
        )
        row["record_key"] = _record_key(row)
        rows.append(row)
    return _dedupe(rows)


def _dedupe(rows: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    deduped: Dict[str, Dict[str, str]] = {}
    for row in rows:
        key = row.get("record_key") or _record_key(row)
        deduped[key] = row
    return list(deduped.values())


def _dedupe_gaps(rows: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    deduped: Dict[str, Dict[str, str]] = {}
    for row in rows:
        key_fields = ["period_label", "source_id", "source_url", "archive_url", "gap_type"]
        if row.get("gap_type") == "single_source_unverified_plan_row":
            key_fields.append("evidence_excerpt")
        key = "|".join(str(row.get(field, "")) for field in key_fields)
        deduped[key] = row
    return list(deduped.values())


def _drop_resolved_gaps(gap_rows: List[Dict[str, str]], source: Dict[str, str], period_label: str, archive_url: str = "") -> List[Dict[str, str]]:
    source_id = source["source_id"]
    source_url = source["url"]
    return [
        row
        for row in gap_rows
        if not (
            row.get("source_id") == source_id
            and row.get("period_label") == period_label
            and row.get("source_url") == source_url
            and (
                (row.get("archive_url") or "") == archive_url
                or row.get("gap_type") == "single_source_unverified_plan_row"
            )
        )
    ]


def _write_csv(path: Path, rows: List[Dict[str, str]], fields: List[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _archive_url(timestamp: str, url: str) -> str:
    return f"https://web.archive.org/web/{timestamp}id_/{url}"


def _archive_timestamp_year(archive_url: str) -> str:
    match = re.search(r"/web/([0-9]{4})[0-9]{10}", archive_url or "")
    return match.group(1) if match else ""


def _archive_period_matches(row_or_gap: Dict[str, str]) -> bool:
    archive_url = row_or_gap.get("archive_url", "")
    if not archive_url:
        return True
    archive_year = _archive_timestamp_year(archive_url)
    period_label = row_or_gap.get("period_label", "")
    return not archive_year or archive_year == period_label


def discover_wayback_snapshots(client: httpx.Client, url: str, years: Iterable[int], *, cache_dir: Path | None = None) -> List[Dict[str, str]]:
    years = list(years)
    cache_key = f"available:{url}:{min(years)}:{max(years)}"
    if cache_dir:
        path = _cache_file(cache_dir, cache_key)
        if path.exists():
            try:
                cached_snapshots = json.loads(path.read_text(encoding="utf-8"))
                if isinstance(cached_snapshots, list):
                    filtered_snapshots = [
                        snap for snap in cached_snapshots
                        if str(snap.get("timestamp", ""))[:4] == str(snap.get("archive_year", ""))
                    ]
                    if len(filtered_snapshots) != len(cached_snapshots):
                        path.write_text(json.dumps(filtered_snapshots, ensure_ascii=False, indent=2), encoding="utf-8")
                    return filtered_snapshots
            except json.JSONDecodeError:
                pass
    snapshots: List[Dict[str, str]] = []
    for year in sorted(years):
        available = f"https://archive.org/wayback/available?url={quote(url, safe='')}&timestamp={year}0630"
        curl_result = _curl_get(available, max_time=ARCHIVE_AVAILABILITY_TIMEOUT_SECONDS)
        try:
            data = json.loads(curl_result.get("text") or "{}")
        except json.JSONDecodeError:
            continue
        closest = ((data.get("archived_snapshots") or {}).get("closest") or {}) if isinstance(data, dict) else {}
        if closest.get("available") is not True or closest.get("status") != "200" or not closest.get("url"):
            continue
        timestamp = str(closest.get("timestamp") or f"{year}0630")
        if timestamp[:4] != str(year):
            continue
        snapshots.append({"timestamp": timestamp, "archive_year": str(year), "archive_url": str(closest["url"])})
    result = snapshots
    if cache_dir:
        cache_dir.mkdir(parents=True, exist_ok=True)
        _cache_file(cache_dir, cache_key).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def _apply_verification(rows: List[Dict[str, str]]) -> None:
    def segment_key(row: Dict[str, str]) -> str:
        explicit_segment = " ".join([row.get("customer_segment", ""), row.get("plan_name", "")]).lower()
        if re.search(r"public[_\s-]*housing|公屋|居屋|指定公居屋", explicit_segment):
            return "public_housing"
        if re.search(r"private[_\s-]*building|private[_\s-]*housing|私樓|私人|私人屋苑", explicit_segment):
            return "private_building"
        text = " ".join([explicit_segment, row.get("evidence_excerpt", "")]).lower()
        if re.search(r"public[_\s-]*housing|公屋|居屋|指定公居屋", text):
            return "public_housing"
        if re.search(r"private[_\s-]*building|private[_\s-]*housing|私樓|私人|私人屋苑", text):
            return "private_building"
        return ""

    def base_verification_key(row: Dict[str, str]) -> tuple[str, str, str, str, str, str, str]:
        price_key = row.get("monthly_fee_hkd", "") or row.get("average_monthly_fee_hkd", "")
        no_disclosed_price_key = ""
        if not price_key:
            no_disclosed_price_key = f"{row.get('tariff_type', '')}|{row.get('plan_name', '')}"
        return (
            row.get("brand", ""),
            row.get("product_category", ""),
            price_key,
            row.get("local_data_gb", ""),
            row.get("broadband_speed_mbps", ""),
            row.get("service_generation", ""),
            no_disclosed_price_key,
        )

    contracts_by_base_key: Dict[tuple[str, str, str, str, str, str, str], set[str]] = {}
    segments_by_base_key: Dict[tuple[str, str, str, str, str, str, str], set[str]] = {}
    for row in rows:
        base_key = base_verification_key(row)
        contract = row.get("contract_months", "")
        if contract:
            contracts_by_base_key.setdefault(base_key, set()).add(contract)
        segment = segment_key(row)
        if segment:
            segments_by_base_key.setdefault(base_key, set()).add(segment)

    def verification_key(row: Dict[str, str]) -> tuple[str, str, str, str, str, str, str, str, str]:
        base_key = base_verification_key(row)
        contract_values = contracts_by_base_key.get(base_key, set())
        contract_key = row.get("contract_months", "") if len(contract_values) > 1 else ""
        segment_values = segments_by_base_key.get(base_key, set())
        segment_split_key = segment_key(row) if len(segment_values) > 1 else ""
        return (*base_key, contract_key, segment_split_key)

    grouped: Dict[tuple[str, str, str, str, str, str, str, str, str], set[str]] = {}
    for row in rows:
        key = verification_key(row)
        grouped.setdefault(key, set()).add(row.get("source_id", "") + "|" + row.get("period_label", ""))
    for row in rows:
        key = verification_key(row)
        count = len(grouped.get(key, set()))
        row["verification_count"] = str(count)
        row["verification_status"] = "multi_source_or_multi_snapshot_verified" if count >= 2 else "single_source_needs_review"

    official_standard_monthly_sources = {
        "hgc_home_broadband_standard_monthly_2026",
        "hgc_home_broadband_standard_monthly_2026_tc",
    }
    conflict_groups: Dict[tuple[str, str, str, str], Dict[str, set[str]]] = {}
    for row in rows:
        if row.get("source_id") not in official_standard_monthly_sources:
            continue
        if row.get("tariff_type") != "monthly_plan_fee" or not row.get("broadband_speed_mbps"):
            continue
        key = (
            row.get("brand", ""),
            row.get("product_category", ""),
            row.get("broadband_speed_mbps", ""),
            row.get("period_label", ""),
        )
        entry = conflict_groups.setdefault(key, {"prices": set(), "sources": set()})
        entry["prices"].add(row.get("monthly_fee_hkd", "") or row.get("average_monthly_fee_hkd", ""))
        entry["sources"].add(row.get("source_id", ""))

    official_conflict_keys = {
        key
        for key, entry in conflict_groups.items()
        if len({price for price in entry["prices"] if price}) > 1 and len(entry["sources"]) > 1
    }
    for row in rows:
        key = (
            row.get("brand", ""),
            row.get("product_category", ""),
            row.get("broadband_speed_mbps", ""),
            row.get("period_label", ""),
        )
        if key not in official_conflict_keys or row.get("source_id") not in official_standard_monthly_sources:
            continue
        row["verification_status"] = "official_price_conflict_needs_review"
        note = "官方标准月费来源对同一速率披露不同金额；正式使用前需人工确认适用口径。"
        if note not in row.get("evidence_excerpt", ""):
            row["evidence_excerpt"] = (row.get("evidence_excerpt", "") + " " + note).strip()[:700]


def _autosize_sheet(ws) -> None:
    for column_cells in ws.columns:
        values = [str(cell.value or "") for cell in column_cells]
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max((len(v) for v in values), default=10) + 2, 45)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="0F2742")
        cell.fill = PatternFill("solid", fgColor="EAF4FF")


def _write_workbook(path: Path, sheets: Dict[str, tuple[List[Dict[str, str]], List[str]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for title, (rows, fields) in sheets.items():
        ws = wb.create_sheet(title[:31])
        ws.append(fields)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        _autosize_sheet(ws)
    wb.save(path)


def _write_quality_audit(output_dir: Path, current_rows: List[Dict[str, str]], historical_rows: List[Dict[str, str]], gap_rows: List[Dict[str, str]], captured_at: str) -> Dict[str, Any]:
    all_rows = [*current_rows, *historical_rows]
    brands = sorted({source["brand"] for source in CURRENT_SOURCES})
    current_by_brand = Counter(row.get("brand", "") for row in current_rows)
    historical_by_brand = Counter(row.get("brand", "") for row in historical_rows)
    years_by_brand: Dict[str, List[str]] = {}
    for brand in brands:
        years_by_brand[brand] = sorted({row.get("period_label", "") for row in historical_rows if row.get("brand") == brand and row.get("period_label")})
    verification = Counter(row.get("verification_status", "") for row in all_rows)
    gap_by_source = Counter(row.get("source_id", "") for row in gap_rows)
    gap_by_type = Counter(row.get("gap_type", "") for row in gap_rows)
    source_ids_with_history = sorted({row.get("source_id", "") for row in historical_rows if row.get("source_id")})
    missing_historical_brands = [brand for brand in brands if historical_by_brand.get(brand, 0) == 0]
    review_fields = [
        "period_label",
        "brand",
        "product_category",
        "plan_name",
        "monthly_fee_hkd",
        "average_monthly_fee_hkd",
        "local_data_gb",
        "broadband_speed_mbps",
        "source_id",
        "source_status",
    ]
    single_source_rows = [
        {field: row.get(field, "") for field in review_fields}
        for row in all_rows
        if row.get("verification_status") == "single_source_needs_review"
    ]
    conflict_rows = [
        {field: row.get(field, "") for field in review_fields}
        for row in all_rows
        if row.get("verification_status") == "official_price_conflict_needs_review"
    ]
    unresolved_source_gap_count = sum(
        count
        for gap_type, count in gap_by_type.items()
        if gap_type != "single_source_unverified_plan_row"
    )
    verification_backlog_count = gap_by_type.get("single_source_unverified_plan_row", 0)
    audit = {
        "ok": bool(current_rows),
        "captured_at_hkt": captured_at,
        "current_count": len(current_rows),
        "historical_count": len(historical_rows),
        "source_gap_count": len(gap_rows),
        "unresolved_source_gap_count": unresolved_source_gap_count,
        "verification_backlog_count": verification_backlog_count,
        "current_by_brand": dict(current_by_brand),
        "historical_by_brand": dict(historical_by_brand),
        "historical_years_by_brand": years_by_brand,
        "verification_status_counts": dict(verification),
        "source_gap_by_source_id": dict(gap_by_source),
        "source_gap_by_type": dict(gap_by_type),
        "single_source_review_rows": single_source_rows,
        "official_price_conflict_review_rows": conflict_rows,
        "source_ids_with_history": source_ids_with_history,
        "missing_historical_brands": missing_historical_brands,
        "completion_status": "in_progress" if missing_historical_brands or len(historical_rows) == 0 else "needs_final_review",
        "notes": [
            "历史产品资费只采纳官方公开页面或 Wayback 公开归档中的可解析文本。",
            "source_gap_count 为兼容既有文件保留的全部待处理记录；应分别查看 unresolved_source_gap_count（真实来源/解析缺口）和 verification_backlog_count（已发现套餐、仅待第二来源核验）。",
            "真实来源/解析缺口表示公开页面不可访问、动态渲染或未解析到结构化套餐价格；不得估算。",
            "verification_count>=2 表示同品牌/品类/月费/数据或速度组合在不同来源或不同快照中重复出现。",
            "official_price_conflict_needs_review 表示多个官方来源对同一口径披露不同金额，正式使用前需人工确认适用口径。",
        ],
    }
    _write_json(output_dir / "quality_audit.json", audit)
    lines = [
        "# 香港竞对产品资费质量审计",
        "",
        f"更新时间（HKT）：{captured_at}",
        "",
        "## 当前状态",
        "",
        f"- 当前套餐记录：{len(current_rows)} 条",
        f"- 历史套餐记录：{len(historical_rows)} 条",
        f"- 待处理证据清单（兼容字段 source_gap_count）：{len(gap_rows)} 条",
        f"  - 真实来源/解析缺口：{unresolved_source_gap_count} 条",
        f"  - 单一来源候选待补第二来源：{verification_backlog_count} 条（不是数据缺失，暂不转正式记录）",
        f"- 多来源/多快照验证记录：{verification.get('multi_source_or_multi_snapshot_verified', 0)} 条",
        f"- 单来源待复核记录：{verification.get('single_source_needs_review', 0)} 条",
        f"- 官方价格冲突待复核记录：{verification.get('official_price_conflict_needs_review', 0)} 条",
        "",
        "## 品牌覆盖",
        "",
    ]
    for brand in brands:
        years = "、".join(years_by_brand.get(brand) or []) or "暂无历史快照命中"
        lines.append(f"- {brand}：当前 {current_by_brand.get(brand, 0)} 条；历史 {historical_by_brand.get(brand, 0)} 条；历史年份 {years}")
    lines.extend(["", "## 来源缺口", ""])
    if gap_rows:
        for source_id, count in sorted(gap_by_source.items()):
            lines.append(f"- `{source_id}`：{count} 条缺口或未解析记录")
    else:
        lines.append("- 暂无 source-gap。")
    lines.extend(["", "## 单来源待复核明细", ""])
    if single_source_rows:
        for row in single_source_rows:
            price = row.get("monthly_fee_hkd") or row.get("average_monthly_fee_hkd") or "未披露"
            speed = row.get("broadband_speed_mbps") or "-"
            lines.append(
                "- "
                f"{row.get('period_label')} {row.get('brand')} {row.get('product_category')} "
                f"speed={speed} fee={price}：`{row.get('source_id')}`；{row.get('plan_name')}"
            )
    else:
        lines.append("- 暂无单来源待复核记录。")
    lines.extend(["", "## 官方价格冲突明细", ""])
    if conflict_rows:
        for row in conflict_rows:
            price = row.get("monthly_fee_hkd") or row.get("average_monthly_fee_hkd") or "未披露"
            speed = row.get("broadband_speed_mbps") or "-"
            lines.append(
                "- "
                f"{row.get('period_label')} {row.get('brand')} {row.get('product_category')} "
                f"speed={speed} fee={price}：`{row.get('source_id')}`；{row.get('plan_name')}"
            )
    else:
        lines.append("- 暂无官方价格冲突记录。")
    lines.extend(
        [
            "",
            "## 结论",
            "",
            "本数据包已可供小竞AI检索，覆盖 2012-2026 年可公开取得的产品/套餐资费；各品牌与产品线的年份覆盖不同，尚未达到所有品牌、所有产品线的 10 年完整覆盖。正式使用时必须同时查看 `source_gaps.csv/json` 和本审计文件；缺口不得估算。",
        ]
    )
    (output_dir / "quality_audit.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return audit


def _gap_row(source: Dict[str, str], result: Dict[str, Any], period_label: str, archive_url: str = "") -> Dict[str, str]:
    return {
        "period_label": period_label,
        "brand": source["brand"],
        "product_category": source["product_category"],
        "gap_type": "archive_fetched_no_plan_rows" if archive_url else "source_fetched_no_plan_rows",
        "http_status": str(result.get("status") or ""),
        "source_id": source["source_id"],
        "source_url": source["url"],
        "archive_url": archive_url,
        "reason": "页面可访问但未解析到结构化套餐价格行；不得估算。" if result.get("status") else str(result.get("error") or "fetch failed"),
        "evidence_excerpt": _clean_text(str(result.get("text") or ""))[:600],
    }


def _gap_row_from_single_source_plan(row: Dict[str, str]) -> Dict[str, str]:
    price = row.get("monthly_fee_hkd") or row.get("average_monthly_fee_hkd") or "未披露"
    speed = row.get("broadband_speed_mbps") or "-"
    reason = (
        "该结构化套餐行只有单一公开来源命中，未达到 verification_count>=2 的多方/多快照核验门槛；"
        "已转入 source-gap，不作为正式结构化资费样本，不得估算。"
    )
    excerpt = (
        f"{row.get('plan_name', '')}；period={row.get('period_label', '')}；"
        f"speed={speed}；fee={price}；source_status={row.get('source_status', '')}；"
        f"evidence={row.get('evidence_excerpt', '')}"
    )
    return {
        "period_label": row.get("period_label", ""),
        "brand": row.get("brand", ""),
        "product_category": row.get("product_category", ""),
        "gap_type": "single_source_unverified_plan_row",
        "http_status": row.get("http_status", ""),
        "source_id": row.get("source_id", ""),
        "source_url": row.get("source_url", ""),
        "archive_url": row.get("archive_url", ""),
        "reason": reason,
        "evidence_excerpt": _clean_text(excerpt)[:600],
    }


def crawl_competitor_products(
    *,
    output_dir: Path = DATASET_DIR,
    client: httpx.Client | None = None,
    refresh_historical: bool = False,
    historical_fetch_limit: int | None = None,
    historical_source_limit: int | None = None,
    historical_source_offset: int = 0,
    source_id_filter: set[str] | None = None,
    source_id_exclude: set[str] | None = None,
) -> Dict[str, Any]:
    captured_at = _now_hkt()
    output_dir.mkdir(parents=True, exist_ok=True)
    current_rows: List[Dict[str, str]] = []
    historical_rows: List[Dict[str, str]] = []
    gap_rows: List[Dict[str, str]] = []
    snapshots: List[Dict[str, Any]] = []
    refreshed_current_source_ids: set[str] = set()
    refreshed_historical_source_periods: set[tuple[str, str, str]] = set()
    years = list(range(2016, 2027))
    current_path = output_dir / "current_plans.csv"
    historical_path = output_dir / "historical_plans.csv"
    source_gap_path = output_dir / "source_gaps.csv"
    cache_dir = output_dir / "crawl_cache"
    if source_gap_path.exists():
        gap_rows = _read_csv(source_gap_path)
        valid_source_ids = {source["source_id"] for source in CURRENT_SOURCES}
        valid_source_urls_by_id = {source["source_id"]: source["url"] for source in CURRENT_SOURCES}
        gap_rows = [
            row for row in gap_rows
            if row.get("source_id") in valid_source_ids
            and row.get("source_url") == valid_source_urls_by_id.get(row.get("source_id", ""))
            and _archive_period_matches(row)
            and row.get("gap_type") != "single_source_unverified_plan_row"
        ]
    client_context = nullcontext(client) if client is not None else httpx.Client(follow_redirects=True, trust_env=True)
    def should_process_source(source_id: str) -> bool:
        if source_id_filter is not None and source_id not in source_id_filter:
            return False
        if source_id_exclude is not None and source_id in source_id_exclude:
            return False
        return True

    with client_context as active_client:
        for source in CURRENT_SOURCES:
            if not should_process_source(source["source_id"]):
                continue
            result = _fetch_page(active_client, source["url"], cache_dir=cache_dir)
            text = _clean_text(str(result.get("text") or ""))
            snapshots.append({**source, "status": result.get("status"), "title": result.get("title"), "text_chars": len(text), "content_hash": _content_hash(text), "captured_at_hkt": captured_at})
            period_label = source.get("period_label", "current")
            rows = _parse_page(source, result, captured_at, period_label)
            rows = _parse_with_fallback(
                source,
                rows,
                active_client,
                captured_at,
                period_label,
                cache_dir=cache_dir,
            )
            gap_rows = _drop_resolved_gaps(gap_rows, source, period_label)
            if period_label == "current" and source["source_id"] in CURRENT_SOURCES_WITH_INTENTIONAL_NO_STRUCTURED_ROWS:
                refreshed_current_source_ids.add(source["source_id"])
            if rows:
                if period_label == "current":
                    refreshed_current_source_ids.add(source["source_id"])
                    current_rows.extend(rows)
                else:
                    refreshed_historical_source_periods.add((source["source_id"], period_label, ""))
                    historical_rows.extend(rows)
            else:
                if period_label != "current":
                    refreshed_historical_source_periods.add((source["source_id"], period_label, ""))
                gap_rows.append(_gap_row(source, result, period_label))

        if refresh_historical:
            fetched = 0
            archive_client = httpx.Client(follow_redirects=True, trust_env=False)
            try:
                history_sources = CURRENT_SOURCES[historical_source_offset:]
                history_sources = [source for source in history_sources if should_process_source(source["source_id"])]
                if historical_source_limit is not None:
                    history_sources = history_sources[:historical_source_limit]
                history_source_ids = {source["source_id"] for source in history_sources}
                gap_rows = [
                    row for row in gap_rows
                    if not (
                        row.get("source_id") in history_source_ids
                        and row.get("gap_type") == "single_source_unverified_plan_row"
                    )
                ]
                for source in history_sources:
                    for snap in discover_wayback_snapshots(archive_client, source["url"], years, cache_dir=cache_dir):
                        if historical_fetch_limit is not None and fetched >= historical_fetch_limit:
                            break
                        time.sleep(0.25)
                        result = _fetch_page(archive_client, snap["archive_url"], archive=True, cache_dir=cache_dir)
                        fetched += 1
                        period_label = snap["archive_year"]
                        rows = _parse_page(source, result, captured_at, period_label, snap["archive_url"])
                        gap_rows = _drop_resolved_gaps(gap_rows, source, period_label, snap["archive_url"])
                        if rows:
                            refreshed_historical_source_periods.add((source["source_id"], period_label, snap["archive_url"]))
                            historical_rows.extend(rows)
                        else:
                            refreshed_historical_source_periods.add((source["source_id"], period_label, snap["archive_url"]))
                            gap_rows.append(_gap_row(source, result, period_label, snap["archive_url"]))
                    if historical_fetch_limit is not None and fetched >= historical_fetch_limit:
                        break
            finally:
                archive_client.close()

    previous_current_rows = [
        _normalise_plan_row(row) for row in _read_csv(current_path)
        if row.get("source_id") not in refreshed_current_source_ids
    ]
    current_rows = _dedupe([*previous_current_rows, *[_normalise_plan_row(row) for row in current_rows]])
    previous_historical_rows = [
        _normalise_plan_row(row) for row in _read_csv(historical_path)
        if (row.get("source_id", ""), row.get("period_label", ""), row.get("archive_url", "")) not in refreshed_historical_source_periods
        and _archive_period_matches(row)
    ]
    historical_rows = _dedupe([*previous_historical_rows, *[_normalise_plan_row(row) for row in historical_rows]])
    gap_rows = _dedupe_gaps(gap_rows)
    all_plan_rows = [*current_rows, *historical_rows]
    _apply_verification(all_plan_rows)
    verified_plan_rows = []
    for row in all_plan_rows:
        if row.get("verification_status") == "single_source_needs_review":
            gap_rows.append(_gap_row_from_single_source_plan(row))
            continue
        verified_plan_rows.append(row)
    all_plan_rows = verified_plan_rows
    current_rows = [row for row in all_plan_rows if row.get("period_label") == "current"]
    historical_rows = [row for row in all_plan_rows if row.get("period_label") != "current"]
    gap_rows = _dedupe_gaps(gap_rows)

    dictionary_rows = [
        {"field": "verification_count", "description": "同一品牌/品类/月费/数据或速度组合在不同来源或不同快照中出现的次数；>=2 才标为多方/多快照验证。"},
        {"field": "verification_status", "description": "multi_source_or_multi_snapshot_verified 表示至少两次命中；single_source_needs_review 表示只有单一来源，会转入 source_gaps 而不作为正式结构化资费样本；official_price_conflict_needs_review 表示多个官方来源对同一口径披露不同金额，需人工确认适用口径。"},
        {"field": "source_status", "description": "parsed_current、parsed_archive 或 parsed_public_official_pdf 表示已结构化；web_indexed_official_pdf_excerpt 表示官方 PDF 可公开核验但本地脚本被 403 阻断，使用公开索引摘录入库；public_product_launch_no_price 表示公开来源只披露产品发布或可用性、未披露月费；source gap 表另列不可估算缺口。"},
        {"field": "tariff_type", "description": "monthly_plan_fee 为月费套餐；one_off_service_charge 为安装/搬迁/取消/检查等一次性服务费；product_launch_no_disclosed_monthly_fee 表示只可确认产品发布/可用性、不得估算资费；price_mentioned_needs_review 表示价格被提及但需人工判断口径。"},
        {"field": "evidence_excerpt", "description": "短证据摘录，用于人工复核套餐金额、数据量、合约期等。"},
    ]
    _write_csv(output_dir / "current_plans.csv", current_rows, PLAN_FIELDS)
    _write_csv(output_dir / "historical_plans.csv", historical_rows, PLAN_FIELDS)
    _write_csv(output_dir / "source_gaps.csv", gap_rows, GAP_FIELDS)
    _write_csv(output_dir / "data_dictionary.csv", dictionary_rows, ["field", "description"])
    _write_json(output_dir / "current_plans.json", current_rows)
    _write_json(output_dir / "historical_plans.json", historical_rows)
    _write_json(output_dir / "source_gaps.json", gap_rows)
    _write_json(output_dir / "source_snapshots.json", snapshots)
    _write_workbook(
        output_dir / "hk_competitor_product_tariffs_human_readable.xlsx",
        {
            "当前套餐": (current_rows, PLAN_FIELDS),
            "历史套餐": (historical_rows, PLAN_FIELDS),
            "来源缺口": (gap_rows, GAP_FIELDS),
            "字段说明": (dictionary_rows, ["field", "description"]),
        },
    )
    audit = _write_quality_audit(output_dir, current_rows, historical_rows, gap_rows, captured_at)
    # Rebuild the disposition report after every tariff crawl so resolved or
    # newly discovered single-source candidates cannot leave stale conclusions.
    from scripts.audit_hk_product_tariff_followups import generate_followup_audit

    followup_audit = generate_followup_audit(output_dir)
    manifest = {
        "id": "hk_competitor_product_tariffs",
        "title": "香港竞对产品资费",
        "summary": "覆盖 3HK/Hutchison、SmarTone、HKBN、HGC、i-CABLE 等香港竞对公开产品和套餐资费的当前与历史数据库。",
        "source_type": "official_public_product_pages_and_public_archives",
        "visibility": "hidden",
        "superseded_by": "competitor_product_tariffs",
        "superseded_note": "作为底层产品子库保留；前端统一显示 competitor_product_tariffs，后端选中该合并库时会自动展开读取本目录。",
        "scope": "公开产品/套餐页面和 Wayback 公开归档；不登录、不绕过权限、不估算缺口。",
        "tags": ["3HK", "SmarTone", "HKBN", "HGC", "i-CABLE", "资费", "套餐", "宽频", "5G"],
        "entrypoints": ["current_plans.csv", "historical_plans.csv", "source_gaps.csv", "quality_audit.md", "quality_audit.json", "verification_followup_audit.csv", "verification_followup_audit.md", "hk_competitor_product_tariffs_human_readable.xlsx", "hk_competitor_product_tariffs_followup_audit.xlsx", "source_snapshots.json"],
        "updated_at": captured_at,
        "row_count": len(all_plan_rows),
        "current_count": len(current_rows),
        "historical_count": len(historical_rows),
        "source_gap_count": len(gap_rows),
        "unresolved_source_gap_count": audit.get("unresolved_source_gap_count", 0),
        "verification_backlog_count": audit.get("verification_backlog_count", 0),
        "verification_followup": followup_audit,
        "quality": audit.get("completion_status", ""),
    }
    _write_json(output_dir / "manifest.json", manifest)
    readme = [
        "# 香港竞对产品资费数据库",
        "",
        f"更新时间（HKT）：{captured_at}",
        "",
        "## 覆盖对象",
        "",
        "3HK / Hutchison、SmarTone、HKBN、HGC、i-CABLE。",
        "",
        "## 文件",
        "",
        "- `current_plans.csv/json`：当前公开套餐和产品资费。",
        "- `historical_plans.csv/json`：2012-2026 可取得公开归档中的历史套餐（品牌覆盖年份不同）。",
        "- `source_gaps.csv/json`：抓取或解析缺口，不能估算。",
        "- `quality_audit.md/json`：覆盖、验证和 source-gap 审计。",
        "- `verification_followup_audit.csv/json/md`：单源候选的二次检索结论和真实来源/解析缺口的终态说明。",
        "- `hk_competitor_product_tariffs_human_readable.xlsx`：人读版工作簿。",
        "",
        "## 口径",
        "",
        "只使用公开官方页面和公开归档；`verification_count>=2` 才视为多方或多快照验证，单来源记录保留但标为需复核；多个官方来源同口径金额不一致时标为 `official_price_conflict_needs_review`，不强行合并为正式结论。",
    ]
    (output_dir / "README.md").write_text("\n".join(readme) + "\n", encoding="utf-8")
    (output_dir / "summary.md").write_text("\n".join(readme[:18]) + "\n", encoding="utf-8")
    status = {
        "ok": True,
        "captured_at_hkt": captured_at,
        "dataset_dir": str(output_dir),
        "brands": sorted({source["brand"] for source in CURRENT_SOURCES}),
        "source_count": len(CURRENT_SOURCES),
        "current_count": len(current_rows),
        "historical_count": len(historical_rows),
        "source_gap_count": len(gap_rows),
        "unresolved_source_gap_count": audit.get("unresolved_source_gap_count", 0),
        "verification_backlog_count": audit.get("verification_backlog_count", 0),
        "multi_verified_count": sum(1 for row in all_plan_rows if row.get("verification_status") == "multi_source_or_multi_snapshot_verified"),
    }
    _write_json(output_dir / "latest_run_status.json", status)
    return status


def _read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [dict(row) for row in csv.DictReader(fh)]


def main() -> None:
    historical_fetch_limit = None
    historical_source_limit = None
    historical_source_offset = 0
    source_id_filter: set[str] | None = None
    source_id_exclude: set[str] | None = None
    for arg in sys.argv:
        if arg.startswith("--history-fetch-limit="):
            historical_fetch_limit = int(arg.split("=", 1)[1])
        if arg.startswith("--history-source-limit="):
            historical_source_limit = int(arg.split("=", 1)[1])
        if arg.startswith("--history-source-offset="):
            historical_source_offset = int(arg.split("=", 1)[1])
        if arg.startswith("--source-id-filter="):
            source_id_filter = {value.strip() for value in arg.split("=", 1)[1].split(",") if value.strip()}
        if arg.startswith("--source-id-exclude="):
            source_id_exclude = {value.strip() for value in arg.split("=", 1)[1].split(",") if value.strip()}
    if historical_fetch_limit is None and os.environ.get("HK_COMPETITOR_HISTORY_FETCH_LIMIT"):
        historical_fetch_limit = int(os.environ["HK_COMPETITOR_HISTORY_FETCH_LIMIT"])
    if historical_source_limit is None and os.environ.get("HK_COMPETITOR_HISTORY_SOURCE_LIMIT"):
        historical_source_limit = int(os.environ["HK_COMPETITOR_HISTORY_SOURCE_LIMIT"])
    if historical_source_offset == 0 and os.environ.get("HK_COMPETITOR_HISTORY_SOURCE_OFFSET"):
        historical_source_offset = int(os.environ["HK_COMPETITOR_HISTORY_SOURCE_OFFSET"])
    if source_id_filter is None and os.environ.get("HK_COMPETITOR_SOURCE_ID_FILTER"):
        source_id_filter = {value.strip() for value in os.environ["HK_COMPETITOR_SOURCE_ID_FILTER"].split(",") if value.strip()}
    if source_id_exclude is None and os.environ.get("HK_COMPETITOR_SOURCE_ID_EXCLUDE"):
        source_id_exclude = {value.strip() for value in os.environ["HK_COMPETITOR_SOURCE_ID_EXCLUDE"].split(",") if value.strip()}
    print(
        json.dumps(
            crawl_competitor_products(
                refresh_historical="--refresh-historical" in sys.argv,
                historical_fetch_limit=historical_fetch_limit,
                historical_source_limit=historical_source_limit,
                historical_source_offset=historical_source_offset,
                source_id_filter=source_id_filter,
                source_id_exclude=source_id_exclude,
            ),
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
